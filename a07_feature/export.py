from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


_AMOUNT_COLUMNS = {
    "Belop",
    "A07_Belop",
    "A07",
    "GL_Belop",
    "GL_Sum",
    "Diff",
    "IB",
    "Endring",
    "UB",
    "Score",
}
_INTEGER_COLUMNS = {"AntallKontoer"}
_BOOL_COLUMNS = {"WithinTolerance"}
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")


def _normalise_dataframe(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    return df.copy()


def _serialise_value(value: object, *, numeric: bool, integer: bool, boolean: bool) -> object:
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if boolean:
        return "Ja" if bool(value) else "Nei"

    if isinstance(value, Decimal):
        return int(value) if integer else float(value)

    if numeric:
        try:
            return int(float(value)) if integer else float(value)
        except Exception:
            return value

    return value


def _prepare_sheet(df: pd.DataFrame | None) -> pd.DataFrame:
    work = _normalise_dataframe(df)
    if work.empty:
        return work

    for column in work.columns:
        numeric = column in _AMOUNT_COLUMNS or column in _INTEGER_COLUMNS
        integer = column in _INTEGER_COLUMNS
        boolean = column in _BOOL_COLUMNS
        if not (numeric or boolean):
            continue
        work[column] = [
            _serialise_value(value, numeric=numeric, integer=integer, boolean=boolean)
            for value in work[column].tolist()
        ]
    return work


def _fit_columns(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
            except Exception:
                value = ""
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max(12, max_len + 2), 42)


def _style_sheet(ws) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    headers = [cell.value for cell in ws[1]]
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL

    for row in ws.iter_rows(min_row=2):
        for cell, header in zip(row, headers):
            if header in _AMOUNT_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
            elif header in _INTEGER_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = "0"

    _fit_columns(ws)


def export_a07_workbook(
    out_path: str | Path,
    *,
    overview_df: pd.DataFrame | None,
    reconcile_df: pd.DataFrame | None,
    mapping_df: pd.DataFrame | None,
    control_statement_df: pd.DataFrame | None = None,
    suggestions_df: pd.DataFrame | None = None,
    unmapped_df: pd.DataFrame | None = None,
) -> Path:
    target = Path(out_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("Kontroll", _prepare_sheet(overview_df)),
        ("Avstemming", _prepare_sheet(reconcile_df)),
        ("Mapping", _prepare_sheet(mapping_df)),
    ]
    if control_statement_df is not None and not control_statement_df.empty:
        sheets.append(("Kontrolloppstilling", _prepare_sheet(control_statement_df)))
    if suggestions_df is not None and not suggestions_df.empty:
        sheets.append(("Forslag", _prepare_sheet(suggestions_df)))
    if unmapped_df is not None and not unmapped_df.empty:
        sheets.append(("Umappede", _prepare_sheet(unmapped_df)))

    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = load_workbook(target)
    for ws in workbook.worksheets:
        _style_sheet(ws)
    workbook.save(target)
    return target


__all__ = ["export_a07_workbook"]
