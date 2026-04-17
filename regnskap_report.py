"""regnskap_report.py — Eksporter årsregnskap til Excel, HTML og PDF.

Bruker regnskap_data for strukturer og nokkeltall_engine for UB-oppslag.
HTML er print-optimert (A4 portrett). Excel er profesjonelt stylet.
PDF genereres via playwright (headless Chromium).
"""
from __future__ import annotations

import html as _html_escape
import json
from pathlib import Path
from typing import Any

import pandas as pd

from regnskap_data import (
    RS_STRUCTURE,
    BS_STRUCTURE,
    BS_EIENDELER,
    BS_EK_GJELD,
    NOTE_SPECS,
    NOTE_REFS,
    PRINSIPP_DEFAULT,
    PRINSIPP_DEFAULTS,
    ub_lookup,
    fmt_amount,
    eval_auto_row,
    build_cf_rows,
    get_notes_for_framework,
    build_note_numbers,
)


# ---------------------------------------------------------------------------
# CSS (A4, profesjonell stil, print-optimert)
# ---------------------------------------------------------------------------

_CSS = """\
* { margin:0; padding:0; box-sizing:border-box; }
@page { size: A4 portrait; margin: 20mm 20mm 18mm 24mm; }
body {
    font-family: "Segoe UI", "Calibri", system-ui, sans-serif;
    font-size: 11pt;
    color: #1a1a2e;
    background: #f4f6f9;
    line-height: 1.5;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
}
.page {
    background: white;
    max-width: 700px;
    margin: 24px auto;
    padding: 36px 40px;
    border-radius: 6px;
    box-shadow: 0 1px 6px rgba(0,0,0,0.10);
}
/* Cover */
.cover {
    display: flex; flex-direction: column; justify-content: center;
    min-height: 70vh;
}
.cover-company { font-size: 28pt; font-weight: 700; color: #1a2e5a; margin-bottom: 12px; }
.cover-title   { font-size: 20pt; font-weight: 300; color: #4472C4; margin-bottom: 40px; }
.cover-year    { font-size: 14pt; font-weight: 600; color: #555; margin-bottom: 8px; }
.cover-line    { border-top: 3px solid #4472C4; margin: 36px 0 36px 0; }
/* Page header */
.page-header {
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    border-bottom: 2px solid #4472C4;
    padding-bottom: 8px;
    margin-bottom: 24px;
}
.page-header-company { font-size: 10pt; color: #888; }
.page-title { font-size: 15pt; font-weight: 700; color: #1a2e5a; }
/* Statement table */
.stmt-table {
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 16px;
}
.stmt-table th {
    font-size: 10pt;
    font-weight: 600;
    text-align: right;
    padding: 6px 10px 6px 6px;
    color: #4472C4;
    border-bottom: 1.5px solid #4472C4;
}
.stmt-table th.lbl { text-align: left; }
.stmt-table th.note-ref { text-align: center; }
.stmt-table td {
    padding: 4px 10px 4px 0;
    font-size: 11pt;
}
.stmt-table td.amt { text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }
.stmt-table td.lbl { width: 50%; }
.stmt-table tr.section-hdr td {
    font-size: 9.5pt; font-weight: 600; color: #4472C4;
    padding-top: 14px; padding-bottom: 4px;
    border-bottom: none;
    text-transform: uppercase; letter-spacing: 0.4px;
}
.stmt-table tr.sum-row td {
    font-weight: 700;
    border-top: 1px solid #aab5cc;
    border-bottom: 2px solid #1a2e5a;
    background: #f0f4f9;
}
.stmt-table tr.major-sum td {
    font-weight: 700;
    font-size: 11.5pt;
    border-top: 1.5px solid #1a2e5a;
    border-bottom: 3px double #1a2e5a;
    background: #e8eef7;
}
.stmt-table tr.normal-row td { border-bottom: 1px solid #eff1f5; }
.stmt-table tr.normal-row:hover td { background: #fafbfc; }
/* Note reference column */
.stmt-table td.note-ref { text-align: center; width: 56px; white-space: nowrap; }
.stmt-table td.note-ref a {
    color: #4472C4; font-weight: 600; font-size: 9.5pt;
    text-decoration: none; border-bottom: 1px dotted #4472C4;
}
.stmt-table td.note-ref a:hover { color: #1a2e5a; }
/* Column header year badge */
.col-year { color: #1a2e5a; font-size: 11pt; font-weight: 700; }
/* Note section */
.note-block { margin-bottom: 10px; }
.note-title {
    font-size: 12pt; font-weight: 700; color: #1a2e5a;
    margin: 28px 0 12px 0;
    padding-bottom: 4px;
    border-bottom: 1.5px solid #4472C4;
}
.note-title:first-child { margin-top: 0; }
.note-section-hdr {
    font-size: 9.5pt; font-weight: 600; color: #4472C4;
    text-transform: uppercase; letter-spacing: 0.3px;
    margin: 16px 0 6px 0;
}
.note-table {
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 10px;
}
.note-table td {
    padding: 4px 10px 4px 0;
    font-size: 11pt;
    border-bottom: 1px solid #f0f2f5;
}
.note-table td.lbl { width: 55%; color: #2c3e50; }
.note-table td.val { text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }
.note-table td.auto-val { color: #1a56a0; font-weight: 600; text-align: right; }
.note-text { font-size: 11pt; line-height: 1.65; white-space: pre-wrap; color: #2c3e50; }
.sep-line { border-top: 1px solid #dde2ea; margin: 12px 0; }
/* Signature section — Maestro-inspired compact 3-col layout */
.sig-section {
    margin-top: 28px;
}
.sig-place-date {
    font-size: 10.5pt; color: #333;
    text-align: center;
    margin-bottom: 4px;
}
.sig-board-label {
    font-size: 10.5pt; color: #333;
    text-align: center;
    margin-bottom: 20px;
}
.sig-grid {
    display: flex;
    flex-wrap: wrap;
    justify-content: center;
    gap: 24px 32px;
}
.sig-item {
    flex: 0 0 calc(33.333% - 22px);
    min-width: 140px;
    text-align: center;
}
.sig-line {
    border-top: 1px solid #1a2e5a;
    margin-top: 28px;
    padding-top: 4px;
}
.sig-name {
    font-size: 10.5pt; font-weight: 600; color: #1a2e5a;
}
.sig-role {
    font-size: 9.5pt; color: #666; margin-top: 1px;
}
/* Footer — pushed to bottom of page in print */
.page-footer {
    margin-top: 36px;
    padding-top: 8px;
    border-top: 1px solid #dde2ea;
    font-size: 8.5pt;
    color: #aaa;
    text-align: right;
}
/* ---- Print overrides (MUST be last to win cascade) ---- */
@media print {
    body { background: white; }
    .page {
        box-shadow: none; margin: 0; padding: 0;
        max-width: none; border-radius: 0;
        display: flex; flex-direction: column;
        min-height: 100vh;
    }
    .page-footer { margin-top: auto; }
    .cover { min-height: 0; padding-top: 200px; }
    .no-print { display: none !important; }
    .page-break { page-break-before: always; }
    .stmt-table, .note-table { page-break-inside: auto; }
    .stmt-table tr, .note-table tr { page-break-inside: avoid; }
    .stmt-table thead { display: table-header-group; }
    .note-block { page-break-inside: avoid; }
    .sig-item { page-break-inside: avoid; }
    .page-header { page-break-after: avoid; }
    .note-title { page-break-after: avoid; }
}
"""


# ---------------------------------------------------------------------------
# HTML builders
# ---------------------------------------------------------------------------

def _e(text: Any) -> str:
    return _html_escape.escape(str(text or ""), quote=False)


def _stmt_table_html(
    structure: list[tuple],
    ub: dict[int, float],
    ub_prev: dict[int, float] | None,
    year: str,
    year_prev: str,
    has_prev: bool,
    note_refs: dict[int, tuple[int, str]] | None = None,
) -> str:
    prev_th = f'<th class="amt">{_e(year_prev or "Fjorår")}</th>' if has_prev else ""
    total_cols = (3 if has_prev else 2) + 1  # +1 for note column
    html = f'<table class="stmt-table"><thead><tr>'
    html += (f'<th class="lbl"></th>'
             f'<th class="note-ref" style="font-size:10pt;color:#4472C4">Note</th>'
             f'<th class="amt col-year">{_e(year or "I år")}</th>'
             f'{prev_th}'
             f'</tr></thead><tbody>')

    if note_refs is None:
        note_refs = NOTE_REFS
    for entry in structure:
        regnr, label, level, is_sum, is_header = entry
        indent = "\u00a0" * (level * 4)

        if regnr is None:
            html += f'<tr class="section-hdr"><td colspan="{total_cols}">{_e(indent + label)}</td></tr>'
            continue

        val = ub.get(regnr)
        val_prev = ub_prev.get(regnr) if ub_prev else None

        if val is None and val_prev is None and not is_sum:
            continue

        val_str = fmt_amount(val) if val is not None else "–"
        prev_td = f'<td class="amt">{fmt_amount(val_prev) if val_prev is not None else "–"}</td>' if has_prev else ""

        note_ref = note_refs.get(regnr)
        if note_ref:
            note_num, note_id = note_ref
            note_td = (f'<td class="note-ref">'
                       f'<a href="#note_{note_id}">Note {note_num}</a></td>')
        else:
            note_td = '<td class="note-ref"></td>'

        row_class = "major-sum" if (is_sum and level == 0) else ("sum-row" if is_sum else "normal-row")
        html += (
            f'<tr class="{row_class}">'
            f'<td class="lbl">{_e(indent + label)}</td>'
            f'{note_td}'
            f'<td class="amt">{_e(val_str)}</td>'
            f'{prev_td}</tr>'
        )

    html += "</tbody></table>"
    return html


def _cf_table_html(
    ub: dict[int, float],
    ub_prev: dict[int, float] | None,
) -> str:
    rows = build_cf_rows(ub, ub_prev)
    html = '<table class="stmt-table"><thead><tr>'
    html += '<th class="lbl"></th><th class="amt">Beløp</th></tr></thead><tbody>'
    for idx, (label, val, is_sum, is_hdr) in enumerate(rows):
        if is_hdr:
            html += f'<tr class="section-hdr"><td colspan="2">{_e(label)}</td></tr>'
        elif not label.strip():
            html += '<tr><td colspan="2">&nbsp;</td></tr>'
        else:
            val_str = fmt_amount(val) if val is not None else "–"
            row_class = "major-sum" if is_sum else "normal-row"
            html += (
                f'<tr class="{row_class}">'
                f'<td class="lbl">{_e(label)}</td>'
                f'<td class="amt">{_e(val_str)}</td></tr>'
            )
    html += "</tbody></table>"
    return html


def _note_html(
    note_id: str,
    note_label: str,
    spec: list | None,
    note_data: dict[str, str],
    ub: dict[int, float],
    ub_prev: dict[int, float] | None,
) -> str:
    html = f'<div class="note-block"><h3 class="note-title" id="note_{note_id}">{_e(note_label)}</h3>'

    if spec is None:
        # Free text (regnskapsprinsipper or custom)
        tekst = note_data.get("tekst") or (PRINSIPP_DEFAULT if note_id == "regnskapsprinsipper" else "")
        if tekst:
            html += f'<p class="note-text">{_e(tekst)}</p>'
        html += '</div>'
        return html

    html += '<table class="note-table"><tbody>'
    for row in spec:
        rtype = row["type"]
        if rtype == "header":
            html += f'<tr><td colspan="2" class="note-section-hdr">{_e(row["label"])}</td></tr>'
        elif rtype == "sep":
            html += '<tr><td colspan="2"><div class="sep-line"></div></td></tr>'
        elif rtype == "auto":
            val = eval_auto_row(row, ub, ub_prev)
            val_str = fmt_amount(val) if val is not None else "–"
            html += (
                f'<tr><td class="lbl">{_e(row["label"])}</td>'
                f'<td class="auto-val">{_e(val_str)}</td></tr>'
            )
        elif rtype == "field":
            key = row.get("key", "")
            val = note_data.get(key, "")
            html += (
                f'<tr><td class="lbl">{_e(row["label"])}</td>'
                f'<td class="val">{_e(val) if val else "&nbsp;"}</td></tr>'
            )
    html += "</tbody></table></div>"
    return html


def _page_header_html(company: str, title: str) -> str:
    return (
        f'<div class="page-header">'
        f'<span class="page-title">{_e(title)}</span>'
        f'<span class="page-header-company">{_e(company)}</span>'
        f'</div>'
    )


# ---------------------------------------------------------------------------
# HTML full report
# ---------------------------------------------------------------------------

def build_report_html(
    rl_df: pd.DataFrame,
    *,
    notes_data: dict[str, dict[str, str]] | None = None,
    client: str = "",
    year: str = "",
    framework: str = "",
    custom_notes: list[tuple[str, str]] | None = None,
    include_cf: bool = True,
    signatories: list[dict[str, str]] | None = None,
) -> str:
    """Bygg komplett HTML-rapport (årsregnskap)."""
    if notes_data is None:
        notes_data = {}

    ub = ub_lookup(rl_df, "UB")
    has_prev = "UB_fjor" in rl_df.columns
    ub_prev = ub_lookup(rl_df, "UB_fjor") if has_prev else None
    year_prev = str(int(year) - 1) if year and year.isdigit() else ("Fjorår" if has_prev else "")

    # Build note list and refs based on framework
    fw_notes = get_notes_for_framework(framework) if framework else NOTE_SPECS
    custom = [(nid, lbl, None) for nid, lbl in (custom_notes or [])]
    all_notes = fw_notes + custom
    _, note_refs = build_note_numbers(all_notes)

    company_e = _e(client or "")

    # Cover
    cover = (
        f'<div class="page cover">'
        f'<div class="cover-company">{company_e}</div>'
        f'<div class="cover-title">Årsregnskap</div>'
        f'<div class="cover-line"></div>'
        f'<div class="cover-year">{_e(year)}</div>'
        f'</div>'
    )

    # Resultatregnskap
    rs_html = (
        f'<div class="page page-break">'
        + _page_header_html(client, "Resultatregnskap")
        + _stmt_table_html(RS_STRUCTURE, ub, ub_prev, year, year_prev, has_prev,
                           note_refs=note_refs)
        + f'<div class="page-footer">Alle beløp i NOK</div>'
        f'</div>'
    )

    # Balanse + signatur (signatur plasseres rett under balansen)
    sig_inline = ""
    if signatories:
        sig_items = ""
        for s in signatories:
            sig_items += (
                f'<div class="sig-item">'
                f'<div class="sig-line"></div>'
                f'<div class="sig-name">{_e(s.get("navn", ""))}</div>'
                f'<div class="sig-role">{_e(s.get("rolle", ""))}</div>'
                f'</div>'
            )
        place_date = f"__________________, den ____.____.{_e(year)}"
        board_label = f"Styret i {_e(client)}" if client else ""
        sig_inline = (
            f'<div class="sig-section">'
            f'<p class="sig-place-date">{place_date}</p>'
            + (f'<p class="sig-board-label">{board_label}</p>' if board_label else "")
            + f'<div class="sig-grid">{sig_items}</div>'
            f'</div>'
        )

    # Balanse side 1: Eiendeler
    bs_eiendeler_html = (
        f'<div class="page page-break">'
        + _page_header_html(client, "Balanse")
        + _stmt_table_html(BS_EIENDELER, ub, ub_prev, year, year_prev, has_prev,
                           note_refs=note_refs)
        + f'<div class="page-footer">Alle beløp i NOK</div>'
        f'</div>'
    )
    # Balanse side 2: EK + Gjeld + signatur
    bs_ek_gjeld_html = (
        f'<div class="page page-break">'
        + _page_header_html(client, "Balanse")
        + _stmt_table_html(BS_EK_GJELD, ub, ub_prev, year, year_prev, has_prev,
                           note_refs=note_refs)
        + sig_inline
        + f'<div class="page-footer">Alle beløp i NOK</div>'
        f'</div>'
    )
    bs_html = bs_eiendeler_html + "\n" + bs_ek_gjeld_html

    # Kontantstrøm
    if has_prev and include_cf:
        cf_html = (
            f'<div class="page page-break">'
            + _page_header_html(client, "Kontantstrøm (indirekte metode)")
            + _cf_table_html(ub, ub_prev)
            + f'<div class="page-footer">Alle beløp i NOK</div>'
            f'</div>'
        )
    else:
        cf_html = ""

    # Noter — each note wrapped in note-block for page-break-inside:avoid
    notes_html_parts = []
    for note_id, note_label, spec in all_notes:
        nd = notes_data.get(note_id) or {}
        notes_html_parts.append(
            _note_html(note_id, note_label, spec, nd, ub, ub_prev)
        )

    notes_page = (
        f'<div class="page page-break">'
        + _page_header_html(client, "Noter til årsregnskapet")
        + "\n".join(notes_html_parts)
        + f'<div class="page-footer">Alle beløp i NOK</div>'
        f'</div>'
    )

    return f"""<!DOCTYPE html>
<html lang="no">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Årsregnskap — {_e(client)} {_e(year)}</title>
<style>{_CSS}</style>
</head>
<body>
{cover}
{rs_html}
{bs_html}
{cf_html}
{notes_page}
</body>
</html>"""


def save_report_html(
    path: str | Path,
    rl_df: pd.DataFrame,
    *,
    notes_data: dict | None = None,
    client: str = "",
    year: str = "",
    framework: str = "",
    custom_notes: list[tuple[str, str]] | None = None,
    include_cf: bool = True,
    signatories: list[dict[str, str]] | None = None,
) -> str:
    """Lagre HTML-rapport til fil og returner filstien."""
    html = build_report_html(rl_df, notes_data=notes_data, client=client, year=year,
                             framework=framework, custom_notes=custom_notes,
                             include_cf=include_cf, signatories=signatories)
    out = Path(path)
    out.write_text(html, encoding="utf-8")
    return str(out)


def save_report_pdf(
    path: str | Path,
    rl_df: pd.DataFrame,
    *,
    notes_data: dict | None = None,
    client: str = "",
    year: str = "",
    framework: str = "",
    custom_notes: list[tuple[str, str]] | None = None,
    include_cf: bool = True,
    signatories: list[dict[str, str]] | None = None,
) -> str:
    """Lagre PDF-rapport via playwright og returner filstien."""
    import tempfile
    html = build_report_html(rl_df, notes_data=notes_data, client=client, year=year,
                             framework=framework, custom_notes=custom_notes,
                             include_cf=include_cf, signatories=signatories)

    with tempfile.NamedTemporaryFile(suffix=".html", mode="w", encoding="utf-8",
                                     delete=False) as tmp:
        tmp.write(html)
        tmp_path = tmp.name

    out = Path(path)
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.goto(f"file:///{Path(tmp_path).as_posix()}")
            page.wait_for_load_state("networkidle")
            page.pdf(
                path=str(out),
                format="A4",
                margin={"top": "15mm", "bottom": "15mm", "left": "18mm", "right": "15mm"},
                print_background=True,
            )
            browser.close()
    finally:
        Path(tmp_path).unlink(missing_ok=True)

    return str(out)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def save_report_excel(
    path: str | Path,
    rl_df: pd.DataFrame,
    *,
    notes_data: dict | None = None,
    client: str = "",
    year: str = "",
    framework: str = "",
    custom_notes: list[tuple[str, str]] | None = None,
    include_cf: bool = True,
    signatories: list[dict[str, str]] | None = None,
) -> str:
    """Lagre profesjonelt Excel-årsregnskap og returner filstien."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    if notes_data is None:
        notes_data = {}

    ub = ub_lookup(rl_df, "UB")
    has_prev = "UB_fjor" in rl_df.columns
    ub_prev = ub_lookup(rl_df, "UB_fjor") if has_prev else None
    year_prev = str(int(year) - 1) if year and year.isdigit() else ("Fjorår" if has_prev else "")

    # Resolve notes for framework
    fw_notes = get_notes_for_framework(framework) if framework else NOTE_SPECS
    custom = [(nid, lbl, None) for nid, lbl in (custom_notes or [])]
    all_notes = fw_notes + custom
    _, xl_note_refs = build_note_numbers(all_notes)

    # --- Style helpers ---
    DARK_BLUE  = "1A2E5A"
    MID_BLUE   = "4472C4"
    LIGHT_BLUE = "E8EEF7"
    PALE_BLUE  = "F0F4F9"
    SECTION_BG = "EFF3F8"
    HDR_FG     = "FFFFFF"

    def _font(bold=False, size=10, color="1A1A2E", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border(top=None, bottom=None):
        thin = Side(style="thin", color="AABBCC")
        med  = Side(style="medium", color=DARK_BLUE)
        dbl  = Side(style="double", color=DARK_BLUE)
        return Border(
            top={"thin": thin, "medium": med, "double": dbl, None: Side()}.get(top, Side()),
            bottom={"thin": thin, "medium": med, "double": dbl, None: Side()}.get(bottom, Side()),
        )

    NUM_FMT = '#,##0;-#,##0;"-"'  # thousands sep, dash for zero

    def _apply_row(ws, row_num, cells: list, font=None, fill=None, border=None, alignment=None):
        for col_num, value in enumerate(cells, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment
        return row_num + 1

    def _setup_stmt_sheet(ws, title: str) -> None:
        ws.title = title
        ws.sheet_view.showGridLines = False
        # Column widths: A=Post, B=I år, C=Fjor, D=Note
        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 10
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_margins.left = 0.87
        ws.page_margins.right = 0.79

    def _write_stmt_header(ws, title: str, row: int) -> int:
        # Company name
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=client)
        c.font = _font(bold=True, size=13, color=DARK_BLUE)
        row += 1
        # Report title
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = _font(bold=False, size=11, color=MID_BLUE)
        row += 1
        ws.row_dimensions[row].height = 6
        row += 1
        # Column headers: Post | I år | Fjor | Note
        hdr_vals = ["", year or "I år", year_prev if has_prev else "", "Note"]
        for col, val in enumerate(hdr_vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = _font(bold=True, size=10, color=DARK_BLUE)
            c.alignment = Alignment(horizontal="center" if col == 4 else
                                    ("right" if col > 1 else "left"))
            c.border = _border(bottom="medium")
        row += 1
        return row

    def _write_stmt_rows(ws, structure, row: int) -> int:
        for entry in structure:
            regnr, label, level, is_sum, is_header = entry
            indent = "  " * level

            if regnr is None:
                ws.merge_cells(f"A{row}:D{row}")
                c = ws.cell(row=row, column=1, value=label)
                c.font = _font(bold=True, size=9, color=MID_BLUE, italic=False)
                c.fill = _fill(SECTION_BG)
                c.alignment = Alignment(horizontal="left", indent=level)
                ws.row_dimensions[row].height = 15
                row += 1
                continue

            val = ub.get(regnr)
            val_prev = ub_prev.get(regnr) if ub_prev else None

            if val is None and val_prev is None and not is_sum:
                continue

            major_sum = is_sum and level == 0
            font = _font(bold=is_sum, size=10 if not major_sum else 10.5,
                         color=DARK_BLUE if major_sum else "1A1A2E")
            fill = _fill(LIGHT_BLUE) if major_sum else (_fill(PALE_BLUE) if is_sum else PatternFill())
            top_border = "medium" if major_sum else ("thin" if is_sum else None)
            bottom_border = "double" if major_sum else ("medium" if is_sum else None)
            brd = _border(top=top_border, bottom=bottom_border)

            lbl_cell = ws.cell(row=row, column=1, value=indent + label)
            lbl_cell.font = font
            lbl_cell.fill = fill
            lbl_cell.border = brd

            num_cell = ws.cell(row=row, column=2, value=val)
            num_cell.font = font
            num_cell.fill = fill
            num_cell.border = brd
            num_cell.number_format = NUM_FMT
            num_cell.alignment = Alignment(horizontal="right")

            if has_prev:
                prev_cell = ws.cell(row=row, column=3, value=val_prev)
                prev_cell.font = _font(bold=is_sum, size=10)
                prev_cell.fill = fill
                prev_cell.border = brd
                prev_cell.number_format = NUM_FMT
                prev_cell.alignment = Alignment(horizontal="right")

            # Note reference column (D)
            note_ref = xl_note_refs.get(regnr)
            note_cell = ws.cell(row=row, column=4,
                                value=f"Note {note_ref[0]}" if note_ref else "")
            note_cell.font = _font(bold=False, size=9,
                                   color=MID_BLUE if note_ref else "1A1A2E")
            note_cell.fill = fill
            note_cell.border = brd
            note_cell.alignment = Alignment(horizontal="center")

            row += 1

        return row

    def _write_note_sheet(ws, note_id: str, note_label: str, spec: list | None) -> None:
        ws.title = f"Note — {note_label}"[:31]
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 22
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        nd = notes_data.get(note_id) or {}
        row = 1

        # Header
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=client)
        c.font = _font(bold=True, size=13, color=DARK_BLUE)
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=note_label)
        c.font = _font(bold=True, size=11, color=MID_BLUE)
        row += 2

        if spec is None:
            # Free text
            tekst = nd.get("tekst") or PRINSIPP_DEFAULT
            for line in tekst.split("\n"):
                c = ws.cell(row=row, column=1, value=line)
                c.font = _font(size=10)
                c.alignment = Alignment(wrap_text=True)
                ws.merge_cells(f"A{row}:B{row}")
                row += 1
            return

        for spec_row in spec:
            rtype = spec_row["type"]
            if rtype == "header":
                ws.merge_cells(f"A{row}:B{row}")
                c = ws.cell(row=row, column=1, value=spec_row["label"])
                c.font = _font(bold=True, size=9, color=MID_BLUE)
                c.fill = _fill(SECTION_BG)
                c.alignment = Alignment(horizontal="left")
                ws.row_dimensions[row].height = 15
                row += 1
            elif rtype == "sep":
                for col in (1, 2):
                    ws.cell(row=row, column=col).border = _border(bottom="thin")
                row += 1
            elif rtype == "auto":
                val = eval_auto_row(spec_row, ub, ub_prev)
                lbl_c = ws.cell(row=row, column=1, value="  " + spec_row["label"])
                lbl_c.font = _font(size=10)
                val_c = ws.cell(row=row, column=2, value=val)
                val_c.font = _font(bold=True, size=10, color=MID_BLUE)
                val_c.number_format = NUM_FMT
                val_c.alignment = Alignment(horizontal="right")
                row += 1
            elif rtype == "field":
                key = spec_row.get("key", "")
                raw = nd.get(key, "")
                # Try numeric
                try:
                    num = float(raw.replace("\u202f", "").replace(" ", "").replace(",", "."))
                    val_write = num
                    num_fmt = NUM_FMT
                except (ValueError, AttributeError):
                    val_write = raw
                    num_fmt = "@"
                lbl_c = ws.cell(row=row, column=1, value="  " + spec_row["label"])
                lbl_c.font = _font(size=10)
                val_c = ws.cell(row=row, column=2, value=val_write)
                val_c.font = _font(size=10)
                val_c.number_format = num_fmt
                val_c.alignment = Alignment(horizontal="right" if num_fmt != "@" else "left")
                row += 1

    # --- Build workbook ---
    wb = Workbook()

    # RS sheet
    ws_rs = wb.active
    _setup_stmt_sheet(ws_rs, "Resultatregnskap")
    r = _write_stmt_header(ws_rs, "Resultatregnskap", 1)
    _write_stmt_rows(ws_rs, RS_STRUCTURE, r)

    # BS sheet
    ws_bs = wb.create_sheet("Balanse")
    _setup_stmt_sheet(ws_bs, "Balanse")
    r = _write_stmt_header(ws_bs, "Balanse", 1)
    _write_stmt_rows(ws_bs, BS_STRUCTURE, r)

    # CF sheet (only if prev year data and enabled)
    if has_prev and ub_prev and include_cf:
        ws_cf = wb.create_sheet("Kontantstrøm")
        ws_cf.sheet_view.showGridLines = False
        ws_cf.column_dimensions["A"].width = 46
        ws_cf.column_dimensions["B"].width = 18
        ws_cf.page_setup.orientation = ws_cf.ORIENTATION_PORTRAIT
        ws_cf.page_setup.paperSize = ws_cf.PAPERSIZE_A4
        row = 1
        ws_cf.merge_cells(f"A{row}:B{row}")
        c = ws_cf.cell(row=row, column=1, value=client)
        c.font = _font(bold=True, size=13, color=DARK_BLUE)
        row += 1
        ws_cf.merge_cells(f"A{row}:B{row}")
        c = ws_cf.cell(row=row, column=1, value="Kontantstrøm (indirekte metode)")
        c.font = _font(size=11, color=MID_BLUE)
        row += 2
        for label, val, is_sum, is_hdr in build_cf_rows(ub, ub_prev):
            if is_hdr:
                ws_cf.merge_cells(f"A{row}:B{row}")
                c = ws_cf.cell(row=row, column=1, value=label)
                c.font = _font(bold=True, size=9, color=MID_BLUE)
                c.fill = _fill(SECTION_BG)
                row += 1
            else:
                c = ws_cf.cell(row=row, column=1, value=label)
                c.font = _font(bold=is_sum, size=10)
                if is_sum:
                    c.fill = _fill(PALE_BLUE)
                v = ws_cf.cell(row=row, column=2, value=val)
                v.font = _font(bold=is_sum, size=10)
                v.number_format = NUM_FMT
                v.alignment = Alignment(horizontal="right")
                if is_sum:
                    v.fill = _fill(PALE_BLUE)
                    v.border = _border(top="thin", bottom="medium")
                    c.border = _border(top="thin", bottom="medium")
                row += 1

    # Note sheets
    for note_id, note_label, spec in all_notes:
        ws_note = wb.create_sheet()
        _write_note_sheet(ws_note, note_id, note_label, spec)

    out = Path(path)
    wb.save(str(out))
    return str(out)
