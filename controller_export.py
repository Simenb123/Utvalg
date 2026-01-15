"""controller_export.py

Excel-eksport.

Denne modulen er et stabilt (bakoverkompatibelt) eksport-lag som brukes både
fra GUI og i testene.

Støttede kall:

1) export_to_excel(path, df)
   -> lager en arbeidsbok med arket "Data".

2) export_to_excel(path, {"Ark": df, "Ark2": df2})
   -> lager en arbeidsbok med flere ark.

3) export_to_excel(path, Utvalg=df_sample, Grunnlag=df_population, ...)
   -> genererer en revisjonsvennlig rapport (flere ark).

Rapport-eksporten inkluderer i tillegg et bilagsgrunnlag (bilagsnivå) som
markerer hva som er valgt spesifikt / tilfeldig, samt nøkkeltall som tar hensyn
til at spesifikk utvelgelse reduserer restpopulasjonen.

NB: Under pytest åpnes aldri filutforsker automatisk.
"""

from __future__ import annotations

import logging
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import load_workbook

from excel_formatting import polish_sheet

logger = logging.getLogger(__name__)

try:
    # Optional: used to include *all* transactions per bilag (motposter etc.)
    from session import session  # type: ignore
except Exception:  # pragma: no cover
    session = None  # type: ignore


class DataControllerExport:
    """Legacy controller-style export used elsewhere in the app."""

    def export(self, df: pd.DataFrame, file_path: str) -> None:
        if df is None or df.empty:
            raise ValueError("Ingen data å eksportere.")

        df.to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active
        ws.title = "Export"
        polish_sheet(ws)
        wb.save(file_path)


def export_to_excel(  # noqa: C901 - deliberate, user-facing convenience wrapper
    path: str | os.PathLike,
    df_or_mapping: pd.DataFrame | Mapping[str, pd.DataFrame] | None = None,
    sample_df: pd.DataFrame | None = None,
    *,
    sheets: Mapping[str, pd.DataFrame] | None = None,
    meta: Mapping[str, Any] | None = None,
    auto_filename: bool | None = None,
    open_folder: bool | None = None,
    filename_prefix: str | None = None,
    **kwargs: Any,
) -> str:
    """Export DataFrames to Excel.

    Returns the final file path as string (important for tests and GUI).

    Args:
        path:
            Chosen save path (file or directory).
        df_or_mapping:
            Either a DataFrame (single export) or mapping of sheet->DataFrame.
        sample_df:
            Legacy alias for df_or_mapping (kept for backwards compatibility).
        sheets:
            Optional mapping of sheet->DataFrame.
        meta:
            Optional meta/forutsetninger for report export.
        auto_filename:
            If True, ignores the chosen filename and generates a generic
            timestamp-based filename in the chosen directory.
        open_folder:
            If True, opens file explorer after saving.
        filename_prefix:
            Prefix used for auto_filename.
        kwargs:
            Convenience: allows passing DataFrames (as sheets) and meta values
            directly. DataFrames are treated as sheets, other values as meta.
    """

    # Coerce inputs
    sheets_dict: dict[str, pd.DataFrame] = {}
    meta_dict: dict[str, Any] = dict(meta or {})

    sheets_dict.update(_coerce_export_inputs(df_or_mapping, sample_df, sheets, kwargs))
    # Anything left in kwargs that is *not* a DataFrame is meta
    for k, v in kwargs.items():
        if isinstance(v, pd.DataFrame):
            continue
        meta_dict[str(k)] = v

    # Detect report export (Utvalg + Grunnlag provided)
    utvalg_key = _find_key_case_insensitive(sheets_dict, "Utvalg")
    grunnlag_key = _find_key_case_insensitive(sheets_dict, "Grunnlag")
    is_utvalg_report = utvalg_key is not None and grunnlag_key is not None

    # Defaults that make sense in GUI, but never annoy in tests
    if auto_filename is None:
        auto_filename = bool(is_utvalg_report) and not _running_under_pytest()
    if open_folder is None:
        open_folder = bool(is_utvalg_report) and not _running_under_pytest()

    output_path = _resolve_output_path(
        path,
        auto_filename=bool(auto_filename),
        filename_prefix=filename_prefix or "Eksport utvalg",
    )

    # If this looks like a report export, build the standard report sheets.
    if is_utvalg_report:
        utvalg_df = sheets_dict[utvalg_key]  # type: ignore[index]
        grunnlag_df = sheets_dict[grunnlag_key]  # type: ignore[index]

        tx_key = _find_key_case_insensitive(sheets_dict, "Bilagtransaksjoner")
        df_transactions = sheets_dict.get(tx_key) if tx_key else None

        sheets_dict = _build_utvalg_report_sheets(
            utvalg_df=utvalg_df,
            grunnlag_df=grunnlag_df,
            df_transactions=df_transactions,
            meta=meta_dict,
        )

    _write_sheets_to_excel(output_path, sheets_dict)

    if open_folder:
        _open_in_file_explorer(output_path)

    return str(output_path)


def _coerce_export_inputs(
    df_or_mapping: pd.DataFrame | Mapping[str, pd.DataFrame] | None,
    sample_df: pd.DataFrame | None,
    sheets: Mapping[str, pd.DataFrame] | None,
    kwargs: Mapping[str, Any],
) -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {}

    if sheets:
        for k, v in sheets.items():
            if isinstance(v, pd.DataFrame):
                out[str(k)] = v

    if isinstance(df_or_mapping, pd.DataFrame):
        out.setdefault("Data", df_or_mapping)
    elif isinstance(df_or_mapping, Mapping):
        for k, v in df_or_mapping.items():
            if isinstance(v, pd.DataFrame):
                out[str(k)] = v

    if sample_df is not None and isinstance(sample_df, pd.DataFrame):
        out.setdefault("Data", sample_df)

    # DataFrames passed as kwargs are treated as sheets.
    for k, v in kwargs.items():
        if isinstance(v, pd.DataFrame):
            out[str(k)] = v

    return out


def _find_key_case_insensitive(d: Mapping[str, Any], target: str) -> str | None:
    t = target.strip().lower()
    for k in d.keys():
        if str(k).strip().lower() == t:
            return str(k)
    return None


def _safe_sheet_name(name: str, used: set[str]) -> str:
    # Excel sheet name rules: max 31 chars, no : \ / ? * [ ]
    n = re.sub(r"[:\\/?*\[\]]", "_", name)
    n = n.strip()
    if not n:
        n = "Sheet"
    n = n[:31]

    base = n
    i = 2
    while n in used:
        suffix = f"_{i}"
        n = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(n)
    return n


def _write_sheets_to_excel(path: Path, sheets_dict: Mapping[str, pd.DataFrame]) -> None:
    if not sheets_dict:
        raise ValueError("Ingen data å eksportere.")

    used: set[str] = set()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets_dict.items():
            safe = _safe_sheet_name(name, used)
            df.to_excel(writer, sheet_name=safe, index=False)

    # Apply formatting (polish) after writing.
    wb = load_workbook(path)
    for ws in wb.worksheets:
        polish_sheet(ws)
    wb.save(path)


def _running_under_pytest() -> bool:
    return "PYTEST_CURRENT_TEST" in os.environ


def _resolve_output_path(path: str | os.PathLike, auto_filename: bool, filename_prefix: str) -> Path:
    p = Path(path)

    # If a directory is given, always auto-name.
    if p.exists() and p.is_dir():
        auto_filename = True

    if auto_filename:
        base_dir = p if p.is_dir() else p.parent
        base_dir.mkdir(parents=True, exist_ok=True)
        fname = _format_timestamp_filename(filename_prefix)
        candidate = base_dir / fname
        return _ensure_unique_filename(candidate)

    # Ensure .xlsx suffix
    if p.suffix.lower() != ".xlsx":
        p = p.with_suffix(".xlsx")
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


def _format_timestamp_filename(prefix: str) -> str:
    # "Eksport utvalg dd.mm.yyyy HH.MM.xlsx" (no colon)
    ts = datetime.now().strftime("%d.%m.%Y %H.%M")
    return f"{prefix} {ts}.xlsx"


def _ensure_unique_filename(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    for i in range(2, 10_000):
        cand = parent / f"{stem} ({i}){suffix}"
        if not cand.exists():
            return cand
    return path


def _open_in_file_explorer(path: Path) -> None:
    try:
        if sys.platform.startswith("win"):
            # Select file in explorer
            subprocess.run(["explorer", "/select,", str(path)], check=False)
        elif sys.platform == "darwin":
            subprocess.run(["open", "-R", str(path)], check=False)
        else:
            subprocess.run(["xdg-open", str(path.parent)], check=False)
    except Exception:  # pragma: no cover
        logger.exception("Could not open file explorer")


# ------------------------- Report helpers (Utvalg + Grunnlag) -------------------------


def _build_utvalg_report_sheets(
    *,
    utvalg_df: pd.DataFrame,
    grunnlag_df: pd.DataFrame,
    df_transactions: pd.DataFrame | None,
    meta: Mapping[str, Any],
) -> dict[str, pd.DataFrame]:
    utvalg = utvalg_df.copy()
    grunnlag = grunnlag_df.copy()

    # Ensure UtvalgNr exists (important for traceability)
    if "UtvalgNr" not in utvalg.columns:
        utvalg.insert(0, "UtvalgNr", range(1, len(utvalg) + 1))

    bilag_col_utvalg = _find_column_case_insensitive(utvalg, "Bilag")
    if bilag_col_utvalg is None:
        bilag_col_utvalg = "Bilag" if "Bilag" in utvalg.columns else utvalg.columns[0]

    # Build UtvalgType mapping (Spesifikk / Tilfeldig / Valgfritt)
    utvalg_type = _build_utvalgtype_map(utvalg, bilag_col=bilag_col_utvalg)
    if "UtvalgType" not in utvalg.columns:
        utvalg["UtvalgType"] = utvalg[bilag_col_utvalg].astype(str).map(utvalg_type).fillna("")

    # Try to derive tolerable error from utvalg interval, if not already provided.
    meta_mut = dict(meta)
    te = _try_extract_tolerable_error_from_utvalg(utvalg)
    if te is not None and ("Tolererbar feil" not in meta_mut or meta_mut.get("Tolererbar feil") in (None, "")):
        meta_mut["Tolererbar feil"] = te

    # Transactions for selected bilag: prefer explicit df_transactions; else try full dataset; else grunnlag.
    tx_source = df_transactions
    if tx_source is None and session is not None:
        try:
            tx_source = getattr(session, "dataset", None)
        except Exception:
            tx_source = None
    if tx_source is None:
        tx_source = grunnlag

    bilag_to_utvalgnr = {
        str(b): int(n)
        for b, n in zip(
            utvalg[bilag_col_utvalg].astype(str).tolist(),
            pd.to_numeric(utvalg["UtvalgNr"], errors="coerce").fillna(0).astype(int).tolist(),
        )
    }
    bilag_keys = set(bilag_to_utvalgnr.keys())

    bilag_col_tx = _find_column_case_insensitive(tx_source, "Bilag")
    tx = tx_source.copy()
    if bilag_col_tx is not None:
        tx_b = tx[bilag_col_tx].astype(str)
        tx = tx.loc[tx_b.isin(bilag_keys)].copy()
        tx.insert(0, "UtvalgNr", tx_b.map(bilag_to_utvalgnr).astype("Int64"))
        tx.sort_values(["UtvalgNr"] + ([bilag_col_tx] if bilag_col_tx in tx.columns else []), inplace=True, kind="mergesort")
    else:
        tx.insert(0, "UtvalgNr", pd.Series([pd.NA] * len(tx), dtype="Int64"))

    # Bilagsgrunnlag (bilag-level population)
    bilagsgrunnlag = _build_bilagsgrunnlag(grunnlag)
    if not bilagsgrunnlag.empty:
        bilagsgrunnlag["IUtvalg"] = bilagsgrunnlag["Bilag"].astype(str).isin(bilag_keys).map(
            {True: "Ja", False: "Nei"}
        )
        bilagsgrunnlag["UtvalgType"] = bilagsgrunnlag["Bilag"].astype(str).map(utvalg_type).fillna("")

    assumptions_df = _build_forutsetninger(meta_mut)
    summary_df = _build_oppsummering(
        utvalg=utvalg,
        grunnlag=grunnlag,
        bilagsgrunnlag=bilagsgrunnlag,
        meta=meta_mut,
        utvalg_type_map=utvalg_type,
    )

    return {
        "Oppsummering": summary_df,
        "Forutsetninger": assumptions_df,
        "Utvalg": utvalg,
        "Bilagtransaksjoner": tx,
        "Bilagsgrunnlag": bilagsgrunnlag,
        "Grunnlag": grunnlag,
    }


def _find_column_case_insensitive(df: pd.DataFrame, name: str) -> str | None:
    t = name.strip().lower()
    for c in df.columns:
        if str(c).strip().lower() == t:
            return str(c)
    return None


def _build_utvalgtype_map(utvalg: pd.DataFrame, *, bilag_col: str) -> dict[str, str]:
    group_col = _find_column_case_insensitive(utvalg, "Gruppe")
    interval_col = _find_column_case_insensitive(utvalg, "Intervall")

    out: dict[str, str] = {}
    for _, row in utvalg.iterrows():
        b = str(row.get(bilag_col, ""))
        g = str(row.get(group_col, "")) if group_col else ""
        i = str(row.get(interval_col, "")) if interval_col else ""
        gl = g.strip().lower()
        il = i.strip().lower()

        if "valgfri" in gl:
            t = "Valgfritt"
        elif gl.startswith("spes") or "spesifikk" in gl:
            t = "Spesifikk"
        elif ">=" in il:
            t = "Spesifikk"
        else:
            t = "Tilfeldig"

        if b:
            out[b] = t
    return out


def _try_extract_tolerable_error_from_utvalg(utvalg: pd.DataFrame) -> float | None:
    interval_col = _find_column_case_insensitive(utvalg, "Intervall")
    if interval_col is None:
        return None

    vals = utvalg[interval_col].dropna().astype(str)
    extracted: list[float] = []
    for v in vals:
        m = re.search(r">=\s*([0-9\s\u00A0.,-]+)", v)
        if not m:
            continue
        num = _parse_number(m.group(1))
        if num is None:
            continue
        extracted.append(abs(float(num)))

    if not extracted:
        return None
    return float(min(extracted))


def _parse_number(text: str) -> float | None:
    s = str(text).strip()
    if not s:
        return None
    s = s.replace("\u00A0", " ").replace(" ", "")
    # Norwegian style: 1 234 567,89
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    m = re.match(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _amount_series_from_df(df: pd.DataFrame) -> pd.Series:
    for col in ("SumBeløp", "SumBelop", "Beløp", "Belop", "Amount", "amount"):
        if col in df.columns:
            return pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    return pd.Series([0.0] * len(df))


def _build_bilagsgrunnlag(grunnlag: pd.DataFrame) -> pd.DataFrame:
    bilag_col = _find_column_case_insensitive(grunnlag, "Bilag")
    if bilag_col is None:
        return pd.DataFrame(columns=["Bilag", "AntallTransaksjoner", "SumBeløp"])

    # Prefer transaction amount column in grunnlag
    amount_col = None
    for c in ("Beløp", "Belop", "Amount", "SumBeløp", "SumBelop"):
        if c in grunnlag.columns:
            amount_col = c
            break
    if amount_col is None:
        return pd.DataFrame(columns=["Bilag", "AntallTransaksjoner", "SumBeløp"])

    date_col = _find_column_case_insensitive(grunnlag, "Dato")
    text_col = _find_column_case_insensitive(grunnlag, "Tekst")

    g = grunnlag.copy()
    g[amount_col] = pd.to_numeric(g[amount_col], errors="coerce").fillna(0.0)

    gb = g.groupby(bilag_col, dropna=False)
    data: dict[str, Any] = {
        "AntallTransaksjoner": gb.size(),
        "SumBeløp": gb[amount_col].sum(),
    }
    if date_col is not None and date_col in g.columns:
        data["Dato"] = gb[date_col].min()
    if text_col is not None and text_col in g.columns:
        data["Tekst"] = gb[text_col].first()

    out = pd.DataFrame(data).reset_index().rename(columns={bilag_col: "Bilag"})

    # Order columns
    cols = ["Bilag"]
    for c in ("Dato", "Tekst"):
        if c in out.columns:
            cols.append(c)
    cols += ["AntallTransaksjoner", "SumBeløp"]
    # plus any optional added later
    out = out[cols]
    return out


def _build_forutsetninger(meta: Mapping[str, Any]) -> pd.DataFrame:
    preferred_keys = [
        "Risiko",
        "Sikkerhet",
        "Tolererbar feil",
        "Metode",
        "Antall grupper (k)",
        "Utvalgsstørrelse",
        "Retning",
        "Beløp (netto) fra/til",
    ]

    rows: list[dict[str, Any]] = []
    used: set[str] = set()

    for k in preferred_keys:
        if k in meta:
            rows.append({"Felt": k, "Verdi": meta[k]})
            used.add(k)

    for k in sorted(meta.keys()):
        if k in used:
            continue
        rows.append({"Felt": k, "Verdi": meta[k]})

    if not rows:
        rows = [{"Felt": "", "Verdi": ""}]
    return pd.DataFrame(rows)


def _build_oppsummering(
    *,
    utvalg: pd.DataFrame,
    grunnlag: pd.DataFrame,
    bilagsgrunnlag: pd.DataFrame,
    meta: Mapping[str, Any],
    utvalg_type_map: Mapping[str, str],
) -> pd.DataFrame:
    grunnlag_amounts = _amount_series_from_df(grunnlag)
    utvalg_amounts = _amount_series_from_df(utvalg)

    n_rows_grunnlag = int(len(grunnlag))
    n_bilag_grunnlag = _nunique(grunnlag, "Bilag")
    n_konto_grunnlag = _nunique(grunnlag, "Konto")

    n_bilag_utvalg = _nunique(utvalg, "Bilag")

    sum_net_grunnlag = float(grunnlag_amounts.sum())
    sum_abs_grunnlag = float(grunnlag_amounts.abs().sum())
    sum_net_utvalg = float(utvalg_amounts.sum())
    sum_abs_utvalg = float(utvalg_amounts.abs().sum())

    share = (n_bilag_utvalg / n_bilag_grunnlag) if n_bilag_grunnlag else 0.0

    # Split: spesifikk / tilfeldig / restpop
    bilag_col_utvalg = _find_column_case_insensitive(utvalg, "Bilag") or "Bilag"
    selected_bilag = set(utvalg[bilag_col_utvalg].astype(str).tolist()) if bilag_col_utvalg in utvalg.columns else set()
    spes_bilag = {b for b, t in utvalg_type_map.items() if t == "Spesifikk"}
    valgfri_bilag = {b for b, t in utvalg_type_map.items() if t == "Valgfritt"}
    tilfeldig_bilag = selected_bilag - spes_bilag - valgfri_bilag

    # Restpopulasjon etter spesifikk (inkluderer de som senere ble trukket tilfeldig)
    pop_bilag = set(bilagsgrunnlag["Bilag"].astype(str).tolist()) if (bilagsgrunnlag is not None and not bilagsgrunnlag.empty) else set()
    spes_in_pop = spes_bilag & pop_bilag
    rest_bilag = pop_bilag - spes_in_pop

    abs_net_rest_basis = 0.0
    if bilagsgrunnlag is not None and not bilagsgrunnlag.empty and "SumBeløp" in bilagsgrunnlag.columns:
        s = pd.to_numeric(bilagsgrunnlag.loc[bilagsgrunnlag["Bilag"].astype(str).isin(rest_bilag), "SumBeløp"], errors="coerce").fillna(0.0)
        abs_net_rest_basis = float(s.abs().sum())

    te_from_utvalg = _try_extract_tolerable_error_from_utvalg(utvalg)

    rows: list[dict[str, Any]] = [
        {"Felt": "Eksportert", "Verdi": datetime.now().strftime("%d.%m.%Y %H:%M:%S")},
        {"Felt": "Antall rader i grunnlag", "Verdi": n_rows_grunnlag},
        {"Felt": "Antall bilag i grunnlag", "Verdi": n_bilag_grunnlag},
        {"Felt": "Antall kontoer i grunnlag", "Verdi": n_konto_grunnlag},
        {"Felt": "Sum beløp i grunnlag", "Verdi": sum_net_grunnlag},
        {"Felt": "Sum absolutt beløp i grunnlag", "Verdi": sum_abs_grunnlag},
        {"Felt": "Antall bilag i utvalg", "Verdi": n_bilag_utvalg},
        {"Felt": "Sum beløp i utvalg", "Verdi": sum_net_utvalg},
        {"Felt": "Sum absolutt beløp i utvalg", "Verdi": sum_abs_utvalg},
        {"Felt": "Utvalgsandel (bilag)", "Verdi": share},
        # Nytt: split og restpop-basis
        {"Felt": "Antall bilag i utvalg (totalt)", "Verdi": int(len(selected_bilag))},
        {"Felt": "Antall bilag spesifikk utvelgelse", "Verdi": int(len(spes_bilag))},
        {"Felt": "Antall bilag restpopulasjon (etter spesifikk)", "Verdi": int(len(rest_bilag))},
        {"Felt": "|Netto restpopulasjon| (basis)", "Verdi": abs_net_rest_basis},
        {"Felt": "Antall bilag tilfeldig utvalg", "Verdi": int(len(tilfeldig_bilag))},
    ]

    if te_from_utvalg is not None:
        rows.append({"Felt": "Tolererbar feil (fra utvalg)", "Verdi": float(te_from_utvalg)})

    # Include meta as-is at the end (useful for audit trail)
    for k, v in meta.items():
        # Avoid duplicate keys already present in rows
        if any(r["Felt"] == k for r in rows):
            continue
        rows.append({"Felt": k, "Verdi": v})

    return pd.DataFrame(rows)


def _nunique(df: pd.DataFrame, col: str) -> int:
    if df is None or df.empty:
        return 0
    c = _find_column_case_insensitive(df, col)
    if c is None or c not in df.columns:
        return 0
    return int(df[c].nunique())
