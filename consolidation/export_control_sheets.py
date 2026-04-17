"""Control sheets for consolidation export."""

from __future__ import annotations

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from consolidation.models import CompanyTB, CurrencyDetail, EliminationJournal, RunResult

_HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
_SUM_FILL = PatternFill("solid", fgColor="F3F6F9")
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_AMOUNT_FMT = "#,##0.00;[Red]-#,##0.00"


def _excel_col(idx: int) -> str:
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(val)
        return 0.0 if pd.isna(f) else f
    except (ValueError, TypeError):
        return 0.0


def build_valutakontroll(
    wb: Workbook,
    currency_details: list[CurrencyDetail],
) -> None:
    if not currency_details:
        return

    ws = wb.create_sheet("Valutakontroll")

    headers = [
        "Selskap",
        "Valuta",
        "Regnr",
        "Regnskapslinje",
        "Type",
        "Beloep foer omregning",
        "Kurs brukt",
        "Kursregel",
        "Beloep etter omregning",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER

    for row_idx, cd in enumerate(currency_details, start=2):
        ws.cell(row=row_idx, column=1, value=cd.company_name).border = _BORDER
        ws.cell(row=row_idx, column=2, value=cd.currency).border = _BORDER
        ws.cell(row=row_idx, column=3, value=cd.regnr).border = _BORDER
        ws.cell(row=row_idx, column=4, value=cd.regnskapslinje).border = _BORDER
        ws.cell(row=row_idx, column=5, value=cd.line_type).border = _BORDER
        c = ws.cell(row=row_idx, column=6, value=cd.amount_before)
        c.number_format = _AMOUNT_FMT
        c.border = _BORDER
        c = ws.cell(row=row_idx, column=7, value=cd.rate)
        c.number_format = "0.0000"
        c.border = _BORDER
        ws.cell(row=row_idx, column=8, value=cd.rate_rule).border = _BORDER
        c = ws.cell(row=row_idx, column=9, value=cd.amount_after)
        c.number_format = _AMOUNT_FMT
        c.border = _BORDER

    ws.freeze_panes = "A2"
    col_widths = [22, 8, 8, 30, 10, 22, 12, 12, 22]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[_excel_col(i)].width = w


def build_saldobalanse_alle(
    wb: Workbook,
    account_details: pd.DataFrame | None,
) -> None:
    if account_details is None or account_details.empty:
        return

    ws = wb.create_sheet("Saldobalanse alle")

    headers = [
        "Selskap",
        "Konto",
        "Kontonavn",
        "Regnr",
        "Regnskapslinje",
        "IB",
        "Bevegelse",
        "UB",
        "Valuta",
        "Kurs brukt",
        "Kursregel",
        "Beloep foer omregning",
        "Beloep etter omregning",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER

    _amt_cols = {6, 7, 8, 12, 13}
    _rate_col = 10

    for row_idx, (_, row) in enumerate(account_details.iterrows(), start=2):
        regnr_raw = row.get("regnr")
        regnr_val = int(regnr_raw) if pd.notna(regnr_raw) else ""
        vals = [
            row.get("selskap", ""),
            row.get("konto", ""),
            row.get("kontonavn", ""),
            regnr_val,
            row.get("regnskapslinje", ""),
            _safe_float(row.get("ib")),
            _safe_float(row.get("netto")),
            _safe_float(row.get("ub_original")),
            row.get("valuta", ""),
            row.get("kurs", 1.0),
            row.get("kursregel", ""),
            _safe_float(row.get("ub_original")),
            _safe_float(row.get("ub")),
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            if col_idx in _amt_cols:
                cell.number_format = _AMOUNT_FMT
            elif col_idx == _rate_col:
                cell.number_format = "0.0000"

    ws.freeze_panes = "A2"
    col_widths = [22, 10, 30, 8, 28, 16, 16, 16, 8, 12, 12, 18, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[_excel_col(i)].width = w


def build_kontrollark(
    wb: Workbook,
    run_result: RunResult,
    companies: list[CompanyTB],
    eliminations: list[EliminationJournal],
    *,
    client: str | None = None,
    year: str | None = None,
) -> None:
    ws = wb.create_sheet("Kontrollark")

    rows = [
        ("Klient", client or ""),
        ("Aar", year or ""),
        ("Kjoert", datetime.fromtimestamp(run_result.run_at).strftime("%d.%m.%Y %H:%M:%S")),
        ("Run ID", run_result.run_id),
        ("Antall selskaper", len(run_result.company_ids)),
        ("Selskaper", ", ".join(c.name for c in companies if c.company_id in run_result.company_ids)),
        ("Antall elimineringer", len(eliminations)),
        ("Antall EK-bilag", sum(1 for j in eliminations if j.kind == "equity_method")),
        ("", ""),
    ]

    for j in eliminations:
        status = "Balansert" if j.is_balanced else f"UBALANSE ({j.net:.2f})"
        rows.append((f"Eliminering: {j.name}", status))

    rows.append(("", ""))
    rows.append(("Resultat-hash (SHA256)", run_result.result_hash))

    if run_result.warnings:
        rows.append(("", ""))
        rows.append(("Advarsler", ""))
        for w in run_result.warnings:
            rows.append(("", w))

    for row_idx, (key, value) in enumerate(rows, start=1):
        cell_key = ws.cell(row=row_idx, column=1, value=key)
        cell_key.font = Font(bold=True) if key else Font()
        ws.cell(row=row_idx, column=2, value=value)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
