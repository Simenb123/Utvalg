"""consolidation.export -- Excel-eksport for konsolidering.

Produserer en arbeidsbok med:
  1. Konsernoppstilling (hoveddataark)
  2. Elimineringer (per-journal breakdown)
  3. TB - {selskap} (ett ark per selskap, mapped)
  4. Kontrollark (metadata, balansesjekk, hash)
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from consolidation.elimination import journals_to_dataframe
from consolidation.control_rows import append_control_rows
from consolidation.models import CompanyTB, CurrencyDetail, EliminationJournal, RunResult

logger = logging.getLogger(__name__)

# Stiler (same as analyse_regnskapsoppstilling_excel.py)
_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
_HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
_SUM_FILL = PatternFill("solid", fgColor="F3F6F9")
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_AMOUNT_FMT = "#,##0.00;[Red]-#,##0.00"


def _excel_col(idx: int) -> str:
    """1-indexed kolonne -> bokstav (1=A, 27=AA)."""
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def build_consolidation_workbook(
    result_df: pd.DataFrame,
    companies: list[CompanyTB],
    eliminations: list[EliminationJournal],
    mapped_tbs: dict[str, pd.DataFrame],
    run_result: RunResult,
    *,
    client: str | None = None,
    year: str | None = None,
    parent_company_id: str = "",
    regnr_to_name: dict[int, str] | None = None,
    hide_zero: bool = False,
) -> Workbook:
    """Bygg komplett konsoliderings-arbeidsbok."""
    wb = Workbook()

    _build_konsernoppstilling(
        wb,
        result_df,
        client=client,
        year=year,
        companies=companies,
        parent_company_id=parent_company_id,
        hide_zero=hide_zero,
    )
    company_names = {c.company_id: c.name for c in companies}
    _build_elimineringer(wb, eliminations, company_names=company_names)
    # Sortér: parent først, resten alfabetisk
    companies_sorted = sorted(
        companies,
        key=lambda c: (0 if c.company_id == parent_company_id else 1, c.name),
    )
    _build_company_sheets(wb, companies_sorted, mapped_tbs, regnr_to_name=regnr_to_name, hide_zero=hide_zero)
    _build_valutakontroll(wb, run_result.currency_details)
    _build_saldobalanse_alle(wb, run_result.account_details)
    _build_kontrollark(wb, run_result, companies, eliminations, client=client, year=year)

    return wb


def save_consolidation_workbook(
    path: str | Path,
    *,
    result_df: pd.DataFrame,
    companies: list[CompanyTB],
    eliminations: list[EliminationJournal],
    mapped_tbs: dict[str, pd.DataFrame],
    run_result: RunResult,
    client: str | None = None,
    year: str | None = None,
    parent_company_id: str = "",
    regnr_to_name: dict[int, str] | None = None,
    hide_zero: bool = False,
) -> str:
    """Bygg og lagre arbeidsbok. Returnerer filstien."""
    wb = build_consolidation_workbook(
        result_df, companies, eliminations, mapped_tbs, run_result,
        client=client, year=year,
        parent_company_id=parent_company_id,
        regnr_to_name=regnr_to_name,
        hide_zero=hide_zero,
    )
    p = Path(path)
    if p.suffix.lower() != ".xlsx":
        p = p.with_suffix(".xlsx")
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(p))
    logger.info("Saved consolidation workbook -> %s", p)
    return str(p)


# ---------------------------------------------------------------------------
# Ark 1: Konsernoppstilling
# ---------------------------------------------------------------------------

def _build_konsernoppstilling(
    wb: Workbook,
    result_df: pd.DataFrame,
    *,
    client: str | None = None,
    year: str | None = None,
    companies: list[CompanyTB] | None = None,
    parent_company_id: str = "",
    hide_zero: bool = False,
) -> None:
    augmented = append_control_rows(result_df)
    if augmented is not None:
        result_df = augmented
    ws = wb.active
    ws.title = "Konsernoppstilling"

    # Tittel
    title_parts = ["Konsernoppstilling"]
    if client:
        title_parts.append(str(client))
    if year:
        title_parts.append(str(year))
    title = " - ".join(title_parts)

    # Finn alle kolonner
    meta_cols = {"regnr", "regnskapslinje", "sumpost", "formel"}
    data_cols = [c for c in result_df.columns if c not in meta_cols]
    company_names = {
        c.name for c in (companies or [])
        if getattr(c, "name", None)
    }
    parent_name = next(
        (c.name for c in (companies or []) if c.company_id == parent_company_id and c.name),
        "",
    )
    aggregate_order = ["Mor", "Doetre", "sum_foer_elim", "eliminering", "konsolidert"]
    company_cols = [c for c in data_cols if c in company_names]
    other_cols = [c for c in data_cols if c not in company_names and c not in aggregate_order]

    ordered_company_cols: list[str] = []
    if parent_name and parent_name in company_cols:
        ordered_company_cols.append(parent_name)
    ordered_company_cols.extend(
        sorted(
            (c for c in company_cols if c != parent_name),
            key=lambda name: str(name).lower(),
        )
    )
    ordered_data_cols = (
        ordered_company_cols
        + other_cols
        + [c for c in aggregate_order if c in data_cols]
    )
    all_headers = ["Nr", "Regnskapslinje"] + ordered_data_cols
    total_cols = len(all_headers)
    last_col_letter = _excel_col(total_cols)

    # Tittelrad
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws["A1"].fill = _TITLE_FILL

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    # Headers (rad 4)
    header_row = 4
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

    # Data
    data_row = header_row + 1
    for _, row in result_df.iterrows():
        regnr = int(row.get("regnr", 0))
        is_sum = bool(row.get("sumpost", False))

        # Filtrer null-linjer (samme logikk som GUI)
        if hide_zero and not is_sum:
            data_vals = [_safe_float(row.get(dc)) for dc in ordered_data_cols]
            if all(abs(v) < 0.005 for v in data_vals):
                continue

        values = [regnr, str(row.get("regnskapslinje", "") or "")]
        for dc in ordered_data_cols:
            values.append(_safe_float(row.get(dc)))

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.border = _BORDER
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
            if col_idx >= 3:
                cell.number_format = _AMOUNT_FMT
            if is_sum:
                cell.font = Font(bold=True)
                cell.fill = _SUM_FILL

        data_row += 1

    # Freeze + filter
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_col_letter}{max(data_row - 1, 4)}"

    # Kolonnebredder
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    for i in range(3, total_cols + 1):
        ws.column_dimensions[_excel_col(i)].width = 16


# ---------------------------------------------------------------------------
# Ark 2: Elimineringer
# ---------------------------------------------------------------------------

_KIND_LABELS = {"manual": "Manuell", "from_suggestion": "Forslag", "template": "Template"}


def _build_elimineringer(
    wb: Workbook,
    eliminations: list[EliminationJournal],
    company_names: dict[str, str] | None = None,
) -> None:
    ws = wb.create_sheet("Elimineringer")

    if not eliminations:
        ws["A1"] = "Ingen elimineringer registrert."
        return

    name_map = company_names or {}
    headers = ["Bilag", "Type", "Regnr", "Selskap", "Beloep", "Beskrivelse"]
    row = 1

    for journal in eliminations:
        # Journalnavn som header
        label = journal.display_label
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, size=12)
        kind_label = _KIND_LABELS.get(journal.kind, journal.kind)
        ws.cell(row=row, column=2, value=kind_label)
        balanced_text = "Balansert" if journal.is_balanced else f"UBALANSE ({journal.net:.2f})"
        ws.cell(row=row, column=4, value=balanced_text)
        row += 1

        # Kolonneheaders
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.border = _BORDER
        row += 1

        # Linjer
        for line in journal.lines:
            ws.cell(row=row, column=1, value=label).border = _BORDER
            ws.cell(row=row, column=2, value=kind_label).border = _BORDER
            ws.cell(row=row, column=3, value=line.regnr).border = _BORDER
            company_display = name_map.get(line.company_id, line.company_id[:16])
            ws.cell(row=row, column=4, value=company_display).border = _BORDER
            c = ws.cell(row=row, column=5, value=line.amount)
            c.number_format = _AMOUNT_FMT
            c.border = _BORDER
            ws.cell(row=row, column=6, value=line.description).border = _BORDER
            row += 1

        row += 1  # blank linje mellom journaler

    for col, w in {"A": 20, "B": 12, "C": 10, "D": 20, "E": 16, "F": 30}.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Ark 3..N: TB per selskap
# ---------------------------------------------------------------------------

def _build_company_sheets(
    wb: Workbook,
    companies: list[CompanyTB],
    mapped_tbs: dict[str, pd.DataFrame],
    *,
    regnr_to_name: dict[int, str] | None = None,
    hide_zero: bool = False,
) -> None:
    _col_headers = {
        "konto": "Konto", "kontonavn": "Kontonavn",
        "regnr": "Regnr", "rl_navn": "Regnskapslinje",
        "ib": "IB", "netto": "Bevegelse", "ub": "UB",
    }
    # Kanonisk kolonnerekkefølge: metadata, deretter IB | Netto | UB
    _col_order = ["konto", "kontonavn", "regnr", "rl_navn", "ib", "netto", "ub"]
    _amount_cols = {"ib", "netto", "ub"}

    for company in companies:
        tb = mapped_tbs.get(company.company_id)
        if tb is None or tb.empty:
            continue

        sheet_name = f"TB - {company.name}"[:31]  # Excel 31-char limit
        ws = wb.create_sheet(sheet_name)

        # Bygg regnskapslinje-navn kolonne dersom regnr finnes
        has_regnr = "regnr" in tb.columns
        show_cols = [c for c in _col_order if c in tb.columns or c == "rl_navn"]
        if not has_regnr:
            show_cols = [c for c in show_cols if c not in ("regnr", "rl_navn")]

        # Headers
        for col_idx, col in enumerate(show_cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=_col_headers.get(col, col))
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.border = _BORDER

        # Data
        row_idx = 2
        for _, row in tb.iterrows():
            # Filter: skip zero-lines when hide_zero is active
            if hide_zero:
                data_vals = [float(row.get(c, 0) or 0) for c in _amount_cols if c in tb.columns]
                if all(abs(v) < 0.005 for v in data_vals):
                    continue
            for col_idx, col in enumerate(show_cols, start=1):
                if col == "rl_navn":
                    # Slaa opp regnskapslinje-navn fra regnr
                    regnr_raw = row.get("regnr")
                    try:
                        rn = int(regnr_raw) if pd.notna(regnr_raw) else None
                    except (ValueError, TypeError):
                        rn = None
                    val = (regnr_to_name or {}).get(rn, "") if rn is not None else ""
                else:
                    val = row.get(col)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = _BORDER
                if col in _amount_cols:
                    cell.number_format = _AMOUNT_FMT
            row_idx += 1

        ws.freeze_panes = "A2"

        # Kolonnebredder
        for i, col in enumerate(show_cols, start=1):
            if col == "kontonavn":
                ws.column_dimensions[_excel_col(i)].width = 35
            elif col == "rl_navn":
                ws.column_dimensions[_excel_col(i)].width = 30
            elif col in _amount_cols:
                ws.column_dimensions[_excel_col(i)].width = 16
            else:
                ws.column_dimensions[_excel_col(i)].width = 12


# ---------------------------------------------------------------------------
# Ark: Valutakontroll
# ---------------------------------------------------------------------------

def _build_valutakontroll(
    wb: Workbook,
    currency_details: list[CurrencyDetail],
) -> None:
    """Bygg kontrollark for valutaomregning: en rad per selskap x regnr."""
    if not currency_details:
        return

    ws = wb.create_sheet("Valutakontroll")

    headers = [
        "Selskap", "Valuta", "Regnr", "Regnskapslinje", "Type",
        "Beloep foer omregning", "Kurs brukt", "Kursregel",
        "Beloep etter omregning",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER

    for row_idx, cd in enumerate(currency_details, start=2):
        ws.cell(row=row_idx, column=1, value=cd.company_name).border = _BORDER
        ws.cell(row=row_idx, column=2, value=cd.currency).border = _BORDER
        ws.cell(row=row_idx, column=3, value=cd.regnr).border = _BORDER
        ws.cell(row=row_idx, column=4, value=cd.regnskapslinje).border = _BORDER
        ws.cell(row=row_idx, column=5, value=cd.line_type).border = _BORDER
        c = ws.cell(row=row_idx, column=6, value=cd.amount_before)
        c.number_format = _AMOUNT_FMT
        c.border = _BORDER
        c = ws.cell(row=row_idx, column=7, value=cd.rate)
        c.number_format = "0.0000"
        c.border = _BORDER
        ws.cell(row=row_idx, column=8, value=cd.rate_rule).border = _BORDER
        c = ws.cell(row=row_idx, column=9, value=cd.amount_after)
        c.number_format = _AMOUNT_FMT
        c.border = _BORDER

    ws.freeze_panes = "A2"

    col_widths = [22, 8, 8, 30, 10, 22, 12, 12, 22]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[_excel_col(i)].width = w


# ---------------------------------------------------------------------------
# Ark: Saldobalanse alle (flat per-konto per-selskap)
# ---------------------------------------------------------------------------

def _build_saldobalanse_alle(
    wb: Workbook,
    account_details: pd.DataFrame | None,
) -> None:
    """Bygg flatt kontrollark med en rad per konto per selskap.

    Kolonner: Selskap, Konto, Kontonavn, Regnr, Regnskapslinje,
              IB, Bevegelse, UB, Valuta, Kurs brukt, Kursregel,
              Beloep foer omregning, Beloep etter omregning.

    ``Beloep foer/etter omregning`` refererer til UB-kolonnen som
    er det eneste feltet motoren bruker for aggregering.
    """
    if account_details is None or account_details.empty:
        return

    ws = wb.create_sheet("Saldobalanse alle")

    headers = [
        "Selskap", "Konto", "Kontonavn", "Regnr", "Regnskapslinje",
        "IB", "Bevegelse", "UB",
        "Valuta", "Kurs brukt", "Kursregel",
        "Beloep foer omregning", "Beloep etter omregning",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER

    _amt_cols = {6, 7, 8, 12, 13}  # 1-indexed columns with amount format
    _rate_col = 10

    for row_idx, (_, row) in enumerate(account_details.iterrows(), start=2):
        regnr_raw = row.get("regnr")
        regnr_val = int(regnr_raw) if pd.notna(regnr_raw) else ""
        vals = [
            row.get("selskap", ""),
            row.get("konto", ""),
            row.get("kontonavn", ""),
            regnr_val,
            row.get("regnskapslinje", ""),
            _safe_float(row.get("ib")),
            _safe_float(row.get("netto")),
            _safe_float(row.get("ub_original")),
            row.get("valuta", ""),
            row.get("kurs", 1.0),
            row.get("kursregel", ""),
            _safe_float(row.get("ub_original")),
            _safe_float(row.get("ub")),
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            if col_idx in _amt_cols:
                cell.number_format = _AMOUNT_FMT
            elif col_idx == _rate_col:
                cell.number_format = "0.0000"

    ws.freeze_panes = "A2"

    col_widths = [22, 10, 30, 8, 28, 16, 16, 16, 8, 12, 12, 18, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[_excel_col(i)].width = w


def _safe_float(val) -> float:
    """Konverter til float, returnerer 0.0 for None/NaN."""
    if val is None:
        return 0.0
    try:
        f = float(val)
        return 0.0 if pd.isna(f) else f
    except (ValueError, TypeError):
        return 0.0


# ---------------------------------------------------------------------------
# Ark N+1: Kontrollark
# ---------------------------------------------------------------------------

def _build_kontrollark(
    wb: Workbook,
    run_result: RunResult,
    companies: list[CompanyTB],
    eliminations: list[EliminationJournal],
    *,
    client: str | None = None,
    year: str | None = None,
) -> None:
    ws = wb.create_sheet("Kontrollark")

    rows = [
        ("Klient", client or ""),
        ("Aar", year or ""),
        ("Kjoert", datetime.fromtimestamp(run_result.run_at).strftime("%d.%m.%Y %H:%M:%S")),
        ("Run ID", run_result.run_id),
        ("Antall selskaper", len(run_result.company_ids)),
        ("Selskaper", ", ".join(c.name for c in companies if c.company_id in run_result.company_ids)),
        ("Antall elimineringer", len(eliminations)),
        ("", ""),
    ]

    # Balansesjekk per journal
    for j in eliminations:
        status = "Balansert" if j.is_balanced else f"UBALANSE ({j.net:.2f})"
        rows.append((f"Eliminering: {j.name}", status))

    rows.append(("", ""))
    rows.append(("Resultat-hash (SHA256)", run_result.result_hash))

    # Advarsler
    if run_result.warnings:
        rows.append(("", ""))
        rows.append(("Advarsler", ""))
        for w in run_result.warnings:
            rows.append(("", w))

    for row_idx, (key, value) in enumerate(rows, start=1):
        cell_key = ws.cell(row=row_idx, column=1, value=key)
        cell_key.font = Font(bold=True) if key else Font()
        ws.cell(row=row_idx, column=2, value=value)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_float(v: object) -> float:
    try:
        f = float(v)
        return f if f == f else 0.0  # NaN check
    except (TypeError, ValueError):
        return 0.0
