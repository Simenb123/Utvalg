from __future__ import annotations

from pathlib import Path
from typing import Dict
import os
import sys
import tempfile
import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def export_strata_and_sample(xlsx_path: str, strata_df: pd.DataFrame, sample_df: pd.DataFrame) -> None:
    """
    Eksporter to ark: 'Strata' og 'Trekk' med norsk formateringsvennlig data.
    Lar Excel gjøre visningen; vi holder oss enkle for kompatibilitet.
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Strata"
    for r in dataframe_to_rows(strata_df, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet("Trekk")
    for r in dataframe_to_rows(sample_df, index=False, header=True):
        ws2.append(r)

    wb.save(xlsx_path)


def export_temp_excel(sheets: Dict[str, pd.DataFrame], prefix: str = "Utvalg_") -> str:
    """
    Generisk eksport av flere DataFrames til én midlertidig Excel-fil.

    - `sheets`: dict {arknavn -> DataFrame}
    - `prefix`: filnavnprefiks, f.eks. "Analyser_"
    Returnerer full sti til den lagrede .xlsx-filen (str),
    og forsøker å åpne den i standard Excel-viser på plattformen.
    """
    # Sørg for en stabil temp-mappe for Utvalg-filer
    tmpdir = Path(tempfile.gettempdir()) / "Utvalg"
    tmpdir.mkdir(parents=True, exist_ok=True)

    # Lag et nogenlunde unikt filnavn
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}{stamp}.xlsx"
    path = tmpdir / filename

    # Hvis det mot formodning ikke er noen ark, lag et lite info-ark
    if not sheets:
        sheets = {
            "Info": pd.DataFrame(
                {"Melding": ["Ingen data å eksportere. (export_temp_excel fikk et tomt 'sheets'-dictionary.)"]}
            )
        }

    wb = Workbook()
    first = True

    for sheet_name, df in sheets.items():
        title = str(sheet_name).strip() or "Ark"
        # Excel tillater maks 31 tegn i arknavn
        title = title[:31]

        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)

        if isinstance(df, pd.DataFrame):
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
        else:
            # Fallback dersom noen sender inn noe annet enn DataFrame
            ws.append(["export_temp_excel: Ikke-støttet sheet-type", type(df).__name__])

    wb.save(path)

    # Forsøk å åpne filen automatisk (som i export_utils.export_selection_to_excel)
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            os.system(f"open '{path}'")
        else:
            os.system(f"xdg-open '{path}' >/dev/null 2>&1 &")
    except Exception:
        # Åpning er "best effort" – selve eksporten er viktigst.
        pass

    return str(path)
