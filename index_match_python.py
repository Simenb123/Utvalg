#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Index/Match sammenligning av to Excel-filer – med GUI for fil- og kolonnevalg

Denne versjonen er laget for å fungere på tvers av ulike selskaper/filer ved at du
kan **bla til filer** og **velge hvilke kolonner** som tilsvarer Kundenr/Beløp/Kundenavn
via en enkel Tkinter-GUI. Den støtter også CLI-bruk (uten GUI) og en innebygd demo.

Viktige punkter
- absdiff == toleranse regnes som **Match** (inkluderende)
- Trykk bare «Kjør» i VS Code/IDLE: GUI åpnes automatisk
- Alternativt: `python index_match_python.py --gui` (eller CLI-flagg som før)

Avhengigheter
    pip install pandas openpyxl
På Linux for GUI: `sudo apt-get install python3-tk`
"""
from __future__ import annotations
import argparse
import json
import math
import os
import sys
import datetime as dt
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.worksheet.worksheet import Worksheet

# Tkinter for GUI
try:
    import tkinter as tk
    from tkinter import filedialog, ttk, messagebox
    TK_OK = True
except Exception:
    TK_OK = False

SETTINGS_FILE = Path("index_match_settings.json")
DEFAULT_TOLERANSE = 1.0

@dataclass
class Settings:
    file1: Optional[str] = None
    file2: Optional[str] = None
    header1: Optional[int] = None
    header2: Optional[int] = None
    fil1_var1: Optional[str] = None  # kundenr
    fil1_var2: Optional[str] = None  # beløp
    fil1_var3: Optional[str] = None  # kundenavn
    fil2_var1: Optional[str] = None
    fil2_var2: Optional[str] = None
    fil2_var3: Optional[str] = None
    toleranse: float = DEFAULT_TOLERANSE

    @classmethod
    def load(cls) -> "Settings":
        if SETTINGS_FILE.exists():
            try:
                data = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
                return cls(**data)
            except Exception:
                pass
        return cls()

    def save(self) -> None:
        SETTINGS_FILE.write_text(json.dumps(self.__dict__, ensure_ascii=False, indent=2), encoding="utf-8")

# ---------- Nytt: robust auto-oppdagelse for mange varianter ----------

def _normalize(s: str) -> str:
    return (str(s) if s is not None else "").strip().lower() \
        .replace(" ", "").replace(":", "").replace("-", "").replace("_", "") \
        .replace("ø", "o").replace("å", "a").replace("æ", "ae")

CANDS_K = {
    "kundenr","kundenummer","kundenr.","customerid","custid",
    "accountnumber","accountno","accountid","account","orderaccountaccountnumber"
}
CANDS_B = {
    "belop","beløp","belopkr","amount","total","totalamount",
    "subtotalamount","balance","accountbalance","balanceamount","totalamountbase"
}
CANDS_N = {
    "kundenavn","kunde","navn","customer","customername","accountname",
    "orderaccountname","orderaccountourreference","yourreference"
}

def guess_columns(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    norm_map = {_normalize(c): c for c in df.columns}
    def find(cands: set) -> Optional[str]:
        # Eksakt først
        for k, orig in norm_map.items():
            if k in cands:
                return orig
        # Startswith/delsøk for felt som "Account: Name" → accountname
        for k, orig in norm_map.items():
            if any(k.startswith(c) or c in k for c in cands):
                return orig
        return None
    return find(CANDS_K), find(CANDS_B), find(CANDS_N)

# ---------- I/O og sammenligning ----------

def load_excel_with_header(path: Path, header_row: int) -> pd.DataFrame:
    if header_row is None or header_row < 1:
        raise ValueError("header_row må være 1-basert og >= 1")
    if not Path(path).exists():
        raise FileNotFoundError(f"Finner ikke fil: {path}")
    df = pd.read_excel(path, header=header_row - 1, dtype=object)
    df.columns = [str(c).replace("\n", " ").replace("\r", " ").strip() for c in df.columns]
    return df

def summarize_by_kundenr(df: pd.DataFrame, col_kundenr: str, col_belop: str, col_kundenavn: str) -> pd.DataFrame:
    out = df[[col_kundenr, col_belop, col_kundenavn]].copy()
    out[col_kundenr] = out[col_kundenr].astype(str).str.strip()
    out[col_belop] = pd.to_numeric(out[col_belop], errors="coerce").fillna(0.0)
    out[col_kundenavn] = out[col_kundenavn].astype(str).where(out[col_kundenavn].notna(), "")
    grp = (
        out.groupby(col_kundenr, dropna=False)
           .agg({col_belop: "sum", col_kundenavn: "last"})
           .reset_index()
           .rename(columns={col_kundenr: "kundenr", col_belop: "belop", col_kundenavn: "kundenavn"})
    )
    return grp

def compare_frames(f1: pd.DataFrame, f2: pd.DataFrame, toleranse: float) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, float]]:
    merged = pd.merge(f1, f2, on="kundenr", how="outer", suffixes=("_1", "_2"))
    merged["belop_1"] = merged["belop_1"].fillna(0.0)
    merged["belop_2"] = merged["belop_2"].fillna(0.0)
    merged["kundenavn_1"] = merged["kundenavn_1"].fillna("")
    merged["kundenavn_2"] = merged["kundenavn_2"].fillna("")

    merged["diff"] = merged["belop_1"] - merged["belop_2"]
    merged["absdiff"] = merged["diff"].abs()

    def status_row(r) -> str:
        if math.isclose(r["absdiff"], 0.0, abs_tol=toleranse):
            return "Match"  # inkluderer absdiff == toleranse
        if r["belop_1"] != 0 and r["belop_2"] == 0:
            return "Kun i Systemfil 1"
        if r["belop_1"] == 0 and r["belop_2"] != 0:
            return "Kun i Systemfil 2"
        return "Avvik"

    merged["status"] = merged.apply(status_row, axis=1)

    totals = {
        "antall_match": int((merged["status"] == "Match").sum()),
        "antall_avvik": int((merged["status"] != "Match").sum()),
        "antall_kun1": int((merged["status"] == "Kun i Systemfil 1").sum()),
        "antall_kun2": int((merged["status"] == "Kun i Systemfil 2").sum()),
        "total_avvik": float(merged.loc[merged["status"] != "Match", "absdiff"].sum()),
        "sum_belop_1": float(merged["belop_1"].sum()),
        "sum_belop_2": float(merged["belop_2"].sum()),
    }

    resultater = merged[[
        "kundenr", "kundenavn_1", "kundenavn_2", "status", "belop_1", "belop_2", "diff"
    ]].copy()
    avvik = resultater[resultater["status"] != "Match"].copy()
    return resultater, avvik, totals

# ---------- Excel-ut ----------

def autofit_columns(ws: Worksheet) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
            except Exception:
                v = ""
            max_len = max(max_len, len(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max_len + 2)

def write_excel(resultater: pd.DataFrame, avvik: pd.DataFrame, outfile: Path,
                file1_name: str, file2_name: str, totals: Dict[str, float]) -> None:
    wb = Workbook()
    ws_res: Worksheet = wb.active
    ws_res.title = "Resultater"
    ws_avv: Worksheet = wb.create_sheet("Avvik")

    header_fill_blue = PatternFill("solid", fgColor="BDD7EE")
    header_fill_pink = PatternFill("solid", fgColor="FFC0CB")
    summary_fill = PatternFill("solid", fgColor="FFF2CC")
    bold = Font(bold=True)
    num_style = NamedStyle(name="num_style"); num_style.number_format = "#,##0.00"
    if "num_style" not in wb.named_styles:
        wb.add_named_style(num_style)

    ws_res["A1"] = "Systemfil 1:"; ws_res["B1"] = file1_name
    ws_res["A2"] = "Systemfil 2:"; ws_res["B2"] = file2_name
    ws_res["D1"] = "Antall match:"; ws_res["E1"] = totals["antall_match"]
    ws_res["D2"] = "Antall avvik:"; ws_res["E2"] = totals["antall_avvik"]
    ws_res["F1"] = "Kun i fil 1:"; ws_res["G1"] = totals["antall_kun1"]
    ws_res["F2"] = "Kun i fil 2:"; ws_res["G2"] = totals["antall_kun2"]
    ws_res["I1"] = "Sum beløp fil1:"; ws_res["J1"] = totals["sum_belop_1"]
    ws_res["I2"] = "Sum beløp fil2:"; ws_res["J2"] = totals["sum_belop_2"]
    ws_res["I3"] = "Totalt avvik:";  ws_res["J3"] = totals["total_avvik"]

    for cell in ["B1","B2","E1","E2","G1","G2","J1","J2","J3"]:
        ws_res[cell].fill = summary_fill
        if cell.startswith("J"): ws_res[cell].style = num_style
        else: ws_res[cell].font = bold

    headers = ["KUNDENR","KUNDENAVN (fil1)","KUNDENAVN (fil2)","STATUS","BELØP FIL 1","BELØP FIL 2","DIFF"]
    start_row = 5
    for col_idx, h in enumerate(headers, start=1):
        c = ws_res.cell(row=start_row, column=col_idx, value=h)
        c.font = bold; c.fill = header_fill_blue

    for r_idx, (_, row) in enumerate(resultater.iterrows(), start=start_row+1):
        ws_res.cell(row=r_idx, column=1, value=row["kundenr"]).alignment = Alignment(horizontal="left")
        ws_res.cell(row=r_idx, column=2, value=row["kundenavn_1"])
        ws_res.cell(row=r_idx, column=3, value=row["kundenavn_2"])
        ws_res.cell(row=r_idx, column=4, value=row["status"])
        c5 = ws_res.cell(row=r_idx, column=5, value=float(row["belop_1"]))
        c6 = ws_res.cell(row=r_idx, column=6, value=float(row["belop_2"]))
        c7 = ws_res.cell(row=r_idx, column=7, value=float(row["diff"]))
        c5.style = c6.style = c7.style = num_style

    last_data_row = start_row + 1 + len(resultater)
    sum_row = last_data_row + 1
    ws_res.cell(row=sum_row, column=5, value=f"=SUM(E{start_row+1}:E{last_data_row})").style = num_style
    ws_res.cell(row=sum_row, column=6, value=f"=SUM(F{start_row+1}:F{last_data_row})").style = num_style
    ws_res.cell(row=sum_row, column=7, value=f"=SUM(G{start_row+1}:G{last_data_row})").style = num_style

    for col in range(5, 8):
        ws_res.cell(row=sum_row, column=col).font = bold
        ws_res.cell(row=sum_row, column=col).fill = summary_fill

    ws_res.freeze_panes = ws_res["A6"]

    for col_idx, h in enumerate(headers, start=1):
        c = ws_avv.cell(row=start_row, column=col_idx, value=h)
        c.font = bold; c.fill = header_fill_pink
    for r_idx, (_, row) in enumerate(avvik.iterrows(), start=start_row+1):
        ws_avv.cell(row=r_idx, column=1, value=row["kundenr"]).alignment = Alignment(horizontal="left")
        ws_avv.cell(row=r_idx, column=2, value=row["kundenavn_1"])
        ws_avv.cell(row=r_idx, column=3, value=row["kundenavn_2"])
        ws_avv.cell(row=r_idx, column=4, value=row["status"])
        c5 = ws_avv.cell(row=r_idx, column=5, value=float(row["belop_1"]))
        c6 = ws_avv.cell(row=r_idx, column=6, value=float(row["belop_2"]))
        c7 = ws_avv.cell(row=r_idx, column=7, value=float(row["diff"]))
        c5.style = c6.style = c7.style = num_style

    ws_avv.freeze_panes = ws_avv["A6"]
    autofit_columns(ws_res); autofit_columns(ws_avv)
    outfile.parent.mkdir(parents=True, exist_ok=True)
    wb.save(outfile)

# ---------- Kjerne ----------

def _run_core(file1: Path, header1: int, col1_k: str, col1_b: str, col1_n: str,
              file2: Path, header2: int, col2_k: str, col2_b: str, col2_n: str,
              toleranse: float, out: Optional[Path] = None) -> Path:
    df1 = load_excel_with_header(file1, header1)
    df2 = load_excel_with_header(file2, header2)
    sum1 = summarize_by_kundenr(df1, col1_k, col1_b, col1_n)
    sum2 = summarize_by_kundenr(df2, col2_k, col2_b, col2_n)
    resultater, avvik, totals = compare_frames(sum1, sum2, toleranse)

    if out is None or not str(out).strip():
        outname = f"IndexMatch_resultater_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        out = Path.cwd() / outname
    write_excel(resultater, avvik, out, Path(file1).name, Path(file2).name, totals)
    return out

# ---------- GUI: fil + kolonner ----------

def run_gui_picker(settings: Settings) -> Tuple[Path, Path, int, int, str, str, str, str, str, str]:
    if not TK_OK:
        raise SystemExit("GUI ikke tilgjengelig (Tkinter). Installer Tk eller bruk CLI.")

    root = tk.Tk()
    root.title("Index/Match – velg filer og kolonner")

    # Vars
    v_file1 = tk.StringVar(value=settings.file1 or "")
    v_file2 = tk.StringVar(value=settings.file2 or "")
    v_h1 = tk.IntVar(value=settings.header1 or 1)
    v_h2 = tk.IntVar(value=settings.header2 or 1)

    cols1: List[str] = []
    cols2: List[str] = []

    # Combobox-vars
    v1_k = tk.StringVar(); v1_b = tk.StringVar(); v1_n = tk.StringVar()
    v2_k = tk.StringVar(); v2_b = tk.StringVar(); v2_n = tk.StringVar()

    def browse(var: tk.StringVar):
        path = filedialog.askopenfilename(title="Velg Excel-fil", filetypes=[("Excel", "*.xlsx")])
        if path:
            var.set(path)
            try_load_headers()

    def try_load_headers():
        nonlocal cols1, cols2
        # Les bare hvis fil og headerrad er satt
        def load_cols(path_str: str, h: int) -> List[str]:
            p = Path(path_str)
            if p.exists() and h >= 1:
                try:
                    df = load_excel_with_header(p, h)
                    return df.columns.tolist()
                except Exception:
                    return []
            return []
        cols1 = load_cols(v_file1.get(), v_h1.get())
        cols2 = load_cols(v_file2.get(), v_h2.get())
        cb1_k["values"] = cb1_b["values"] = cb1_n["values"] = cols1
        cb2_k["values"] = cb2_b["values"] = cb2_n["values"] = cols2
        # Auto-guess
        try:
            if cols1:
                df1 = pd.DataFrame(columns=cols1)
                g1 = guess_columns(df1)
                if g1[0]: v1_k.set(g1[0])
                if g1[1]: v1_b.set(g1[1])
                if g1[2]: v1_n.set(g1[2])
            if cols2:
                df2 = pd.DataFrame(columns=cols2)
                g2 = guess_columns(df2)
                if g2[0]: v2_k.set(g2[0])
                if g2[1]: v2_b.set(g2[1])
                if g2[2]: v2_n.set(g2[2])
        except Exception:
            pass

    def go():
        if not Path(v_file1.get()).exists() or not Path(v_file2.get()).exists():
            messagebox.showerror("Feil", "Velg gyldige filer for både Systemfil 1 og 2.")
            return
        if not all([v1_k.get(), v1_b.get(), v1_n.get(), v2_k.get(), v2_b.get(), v2_n.get()]):
            messagebox.showerror("Feil", "Velg kolonner for Kundenr/Beløp/Kundenavn i begge filer.")
            return
        root.destroy()

    # Layout
    pad = {"padx": 8, "pady": 4}
    frm = ttk.Frame(root); frm.pack(fill="both", expand=True, **pad)

    # Rad: filer
    ttk.Label(frm, text="Systemfil 1").grid(row=0, column=0, sticky="w", **pad)
    e1 = ttk.Entry(frm, textvariable=v_file1, width=60); e1.grid(row=0, column=1, **pad)
    ttk.Button(frm, text="Bla...", command=lambda: browse(v_file1)).grid(row=0, column=2, **pad)
    ttk.Label(frm, text="Header-rad 1").grid(row=1, column=0, sticky="w", **pad)
    ttk.Spinbox(frm, from_=1, to=50, textvariable=v_h1, width=6, command=try_load_headers).grid(row=1, column=1, sticky="w", **pad)

    ttk.Label(frm, text="Systemfil 2").grid(row=2, column=0, sticky="w", **pad)
    e2 = ttk.Entry(frm, textvariable=v_file2, width=60); e2.grid(row=2, column=1, **pad)
    ttk.Button(frm, text="Bla...", command=lambda: browse(v_file2)).grid(row=2, column=2, **pad)
    ttk.Label(frm, text="Header-rad 2").grid(row=3, column=0, sticky="w", **pad)
    ttk.Spinbox(frm, from_=1, to=50, textvariable=v_h2, width=6, command=try_load_headers).grid(row=3, column=1, sticky="w", **pad)

    # Kolonnevalg – fil 1
    ttk.Label(frm, text="KUNDENR (fil1)").grid(row=4, column=0, sticky="w", **pad)
    cb1_k = ttk.Combobox(frm, textvariable=v1_k, state="readonly"); cb1_k.grid(row=4, column=1, **pad)
    ttk.Label(frm, text="BELØP (fil1)").grid(row=5, column=0, sticky="w", **pad)
    cb1_b = ttk.Combobox(frm, textvariable=v1_b, state="readonly"); cb1_b.grid(row=5, column=1, **pad)
    ttk.Label(frm, text="KUNDENAVN (fil1)").grid(row=6, column=0, sticky="w", **pad)
    cb1_n = ttk.Combobox(frm, textvariable=v1_n, state="readonly"); cb1_n.grid(row=6, column=1, **pad)

    # Kolonnevalg – fil 2
    ttk.Label(frm, text="KUNDENR (fil2)").grid(row=7, column=0, sticky="w", **pad)
    cb2_k = ttk.Combobox(frm, textvariable=v2_k, state="readonly"); cb2_k.grid(row=7, column=1, **pad)
    ttk.Label(frm, text="BELØP (fil2)").grid(row=8, column=0, sticky="w", **pad)
    cb2_b = ttk.Combobox(frm, textvariable=v2_b, state="readonly"); cb2_b.grid(row=8, column=1, **pad)
    ttk.Label(frm, text="KUNDENAVN (fil2)").grid(row=9, column=0, sticky="w", **pad)
    cb2_n = ttk.Combobox(frm, textvariable=v2_n, state="readonly"); cb2_n.grid(row=9, column=1, **pad)

    ttk.Button(frm, text="Kjør", command=go).grid(row=10, column=1, sticky="e", **pad)

    try_load_headers()
    root.mainloop()

    return (Path(v_file1.get()), Path(v_file2.get()), v_h1.get(), v_h2.get(),
            v1_k.get(), v1_b.get(), v1_n.get(), v2_k.get(), v2_b.get(), v2_n.get())

# ---------- CLI ----------

def _is_tty() -> bool:
    try:
        return sys.stdin.isatty() and sys.stdout.isatty()
    except Exception:
        return False

def parse_args(settings: Settings) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Index/Match sammenligning av to Excel-filer")
    p.add_argument("--file1", type=str, default=os.getenv("IM_FILE1", settings.file1))
    p.add_argument("--file2", type=str, default=os.getenv("IM_FILE2", settings.file2))
    p.add_argument("--header1", type=int, default=int(os.getenv("IM_HEADER1", settings.header1 or 1)))
    p.add_argument("--header2", type=int, default=int(os.getenv("IM_HEADER2", settings.header2 or 1)))
    p.add_argument("--col1-kundenr", dest="col1_k", type=str, default=os.getenv("IM_COL1_K", settings.fil1_var1 or ""))
    p.add_argument("--col1-belop", dest="col1_b", type=str, default=os.getenv("IM_COL1_B", settings.fil1_var2 or ""))
    p.add_argument("--col1-kundenavn", dest="col1_n", type=str, default=os.getenv("IM_COL1_N", settings.fil1_var3 or ""))
    p.add_argument("--col2-kundenr", dest="col2_k", type=str, default=os.getenv("IM_COL2_K", settings.fil2_var1 or ""))
    p.add_argument("--col2-belop", dest="col2_b", type=str, default=os.getenv("IM_COL2_B", settings.fil2_var2 or ""))
    p.add_argument("--col2-kundenavn", dest="col2_n", type=str, default=os.getenv("IM_COL2_N", settings.fil2_var3 or ""))
    p.add_argument("--toleranse", type=float, default=float(os.getenv("IM_TOLERANSE", settings.toleranse)))
    p.add_argument("--out", type=str, default=os.getenv("IM_OUT", ""))
    p.add_argument("--demo", action="store_true")
    p.add_argument("--picker", action="store_true")
    p.add_argument("--auto-columns", action="store_true")
    p.add_argument("--gui", action="store_true", help="Full GUI for fil- og kolonnevalg")
    return p.parse_args()

# ---------- Demo ----------

def _make_demo_files(tmpdir: Path) -> Tuple[Path, Path, int, int, Tuple[str, str, str], Tuple[str, str, str]]:
    tmpdir.mkdir(parents=True, exist_ok=True)
    df1 = pd.DataFrame({"Account: Account number": ["100","101","102"],
                        "Account: Name": ["Alpha","Beta","Gamma"],
                        "Total amount": [1000, 1999, 3000]})
    df2 = pd.DataFrame({"Account: Account number": ["100","101","104"],
                        "Account: Name": ["Alpha AS","Beta","Epsilon"],
                        "Total amount": [1000, 2000, 500]})
    f1 = tmpdir/"demo1.xlsx"; f2 = tmpdir/"demo2.xlsx"
    df1.to_excel(f1, index=False); df2.to_excel(f2, index=False)
    return f1, f2, 1, 1, ("Account: Account number","Total amount","Account: Name"), ("Account: Account number","Total amount","Account: Name")

def run_demo_and_tests():
    tmp = Path("/mnt/data/indexmatch_demo"); f1, f2, h1, h2, c1, c2 = _make_demo_files(tmp)
    out = _run_core(f1, h1, c1[0], c1[1], c1[2], f2, h2, c2[0], c2[1], c2[2], DEFAULT_TOLERANSE)
    assert Path(out).exists(), "Output ble ikke skrevet"
    print("Selvtest OK. Output:", out)

# ---------- Main ----------

def main() -> None:
    settings = Settings.load()
    args = parse_args(settings)

    if args.demo:
        run_demo_and_tests(); return

    if args.gui:
        (p1, p2, h1, h2, c1k, c1b, c1n, c2k, c2b, c2n) = run_gui_picker(settings)
    else:
        # Filvalg
        p1 = Path(args.file1) if args.file1 else None
        p2 = Path(args.file2) if args.file2 else None
        if args.picker and TK_OK:
            # Bare velg filer (kolonner via auto/CLI)
            root = tk.Tk(); root.withdraw()
            if not p1 or not p1.exists():
                f = filedialog.askopenfilename(title="Velg Systemfil 1", filetypes=[("Excel","*.xlsx")])
                p1 = Path(f) if f else None
            if not p2 or not p2.exists():
                f = filedialog.askopenfilename(title="Velg Systemfil 2", filetypes=[("Excel","*.xlsx")])
                p2 = Path(f) if f else None
        if not p1 or not p2:
            missing = ", ".join([n for n,v in {"Systemfil 1":p1, "Systemfil 2":p2}.items() if not v])
            raise SystemExit(
                "Mangler filer (" + missing + ")\n"
                "- Angi gyldige stier med --file1/--file2\n"
                "- bruk --picker for filvalg eller --gui for fil + kolonner."
            )

        h1 = int(args.header1 or settings.header1 or 1)
        h2 = int(args.header2 or settings.header2 or 1)

        c1k = args.col1_k or settings.fil1_var1
        c1b = args.col1_b or settings.fil1_var2
        c1n = args.col1_n or settings.fil1_var3
        c2k = args.col2_k or settings.fil2_var1
        c2b = args.col2_b or settings.fil2_var2
        c2n = args.col2_n or settings.fil2_var3

        if args.auto_columns or not all([c1k, c1b, c1n, c2k, c2b, c2n]):
            df1_prev = load_excel_with_header(p1, h1)
            df2_prev = load_excel_with_header(p2, h2)
            g1 = guess_columns(df1_prev); g2 = guess_columns(df2_prev)
            c1k = c1k or g1[0]; c1b = c1b or g1[1]; c1n = c1n or g1[2]
            c2k = c2k or g2[0]; c2b = c2b or g2[1]; c2n = c2n or g2[2]

    # Sjekk kolonner
    required = {"fil1 kundenr": c1k, "fil1 beløp": c1b, "fil1 kundenavn": c1n,
                "fil2 kundenr": c2k, "fil2 beløp": c2b, "fil2 kundenavn": c2n}
    missing_cols = [k for k,v in required.items() if not v]
    if missing_cols:
        raise SystemExit("Mangler kolonner: " + ", ".join(missing_cols) + "\nBruk --gui eller oppgi --col*-* flagg.")

    # Kjør
    out_path = Path(args.out) if args.out else None
    outfile = _run_core(p1, h1, c1k, c1b, c1n, p2, h2, c2k, c2b, c2n, float(args.toleranse), out_path)

    # Lagre preferanser
    settings.file1 = str(p1); settings.file2 = str(p2)
    settings.header1 = h1; settings.header2 = h2
    settings.fil1_var1 = c1k; settings.fil1_var2 = c1b; settings.fil1_var3 = c1n
    settings.fil2_var1 = c2k; settings.fil2_var2 = c2b; settings.fil2_var3 = c2n
    settings.toleranse = float(args.toleranse)
    settings.save()

    print("\nFerdig. Resultat-fil lagret:", outfile)

if __name__ == "__main__":
    # Når du trykker på 'Kjør' i VS Code/IDLE, åpner vi GUI direkte.
    try:
        settings = Settings.load()
        (p1, p2, h1, h2, c1k, c1b, c1n, c2k, c2b, c2n) = run_gui_picker(settings)
        out_path = None
        outfile = _run_core(p1, h1, c1k, c1b, c1n, p2, h2, c2k, c2b, c2n, DEFAULT_TOLERANSE, out_path)
        print("\nFerdig. Resultat-fil lagret:", outfile)
    except Exception as e:
        # Faller tilbake til CLI-modus om GUI feiler (f.eks. headless miljø)
        print("GUI feilet eller ble avbrutt (", e, ") – prøver CLI-oppstart…")
        main()
