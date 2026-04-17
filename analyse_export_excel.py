"""analyse_export_excel.py — Felles Excel-eksport for MVA, Lønn, Skatt.

Eksporterer data fra ttk.Treeview-tabeller til et formatert Excel-ark.
Brukes av page_mva, page_lonn og page_skatt.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)


def export_trees_to_excel(
    filepath: str | Path,
    sheets: list[dict],
    *,
    title: str = "",
    client: str = "",
    year: str = "",
) -> None:
    """Eksporter én eller flere Treeview-tabeller til Excel.

    Args:
        filepath: Sti til .xlsx-fil (overskrives).
        sheets: Liste over ark-beskrivelser:
            {
              "title": str,          # Arknavn
              "heading": str,        # Overskrift i arket
              "columns": [str, ...], # Kolonnenavn
              "rows": [              # Rader
                  {"values": [str|float, ...], "bold": bool, "bg": str|None}
              ]
            }
        title: Rapport-tittel (vises øverst på første ark).
        client: Klientnavn (metadata).
        year: Regnskapsår (metadata).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl ikke installert. Kan ikke eksportere til Excel.")
        return

    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    # Farger
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
    ALT_FILL    = PatternFill("solid", fgColor="F2F7FD")

    def _thin() -> Border:
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    for sheet_def in sheets:
        sname = str(sheet_def.get("title", "Ark"))[:31]
        ws = wb.create_sheet(title=sname)

        heading = str(sheet_def.get("heading", sname))
        columns: list[str] = sheet_def.get("columns", [])
        rows: list[dict] = sheet_def.get("rows", [])

        row_cursor = 1

        # Metadata-rad øverst
        if title or client or year:
            meta = f"{title}"
            if client:
                meta += f"  |  {client}"
            if year:
                meta += f"  |  {year}"
            ws.cell(row_cursor, 1, meta).font = Font(italic=True, color="888888")
            row_cursor += 1

        # Overskrift
        ws.merge_cells(start_row=row_cursor, start_column=1,
                       end_row=row_cursor, end_column=max(len(columns), 1))
        h_cell = ws.cell(row_cursor, 1, heading)
        h_cell.font = Font(bold=True, size=13, color="FFFFFF")
        h_cell.fill = HEADER_FILL
        h_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
        ws.row_dimensions[row_cursor].height = 22
        row_cursor += 1

        # Kolonneheader
        for ci, col in enumerate(columns, 1):
            c = ws.cell(row_cursor, ci, col)
            c.font = Font(bold=True, color="1F4E79")
            c.fill = SUBHDR_FILL
            c.alignment = Alignment(horizontal="center")
            c.border = _thin()
        row_cursor += 1

        # Data-rader
        for ridx, rdef in enumerate(rows):
            vals = rdef.get("values", [])
            bold = rdef.get("bold", False)
            bg   = rdef.get("bg")   # hex string without #, or None
            sep  = rdef.get("sep", False)

            if sep:
                row_cursor += 1
                continue

            fill = (PatternFill("solid", fgColor=bg) if bg
                    else (ALT_FILL if ridx % 2 else None))

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row_cursor, ci, val)
                c.font = Font(bold=bold)
                c.border = _thin()
                if fill:
                    c.fill = fill
                # Justering: tall høyrestilles
                if ci > 1 and isinstance(val, (int, float)):
                    c.alignment = Alignment(horizontal="right")
                    c.number_format = '#,##0.00'
                elif ci > 1:
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.alignment = Alignment(horizontal="left", indent=1)

            row_cursor += 1

        # Kolonnebredder
        col_widths: list[float] = [0.0] * len(columns)
        for row_idx in range(1, row_cursor):
            for ci in range(1, len(columns) + 1):
                val = ws.cell(row_idx, ci).value
                w = len(str(val)) if val is not None else 0
                if ci - 1 < len(col_widths):
                    col_widths[ci - 1] = max(col_widths[ci - 1], w)
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2, 12), 60)

    wb.save(str(filepath))
    log.info("Excel eksportert: %s", filepath)


def _resolve_visible_columns(tree: Any) -> tuple[list[str], list[int]]:
    """Returner (synlige kolonne-ID-er, deres posisjoner i tree["columns"]).

    Respekterer `displaycolumns`. Når `displaycolumns` er `"#all"` eller
    ikke satt, brukes alle kolonner i original rekkefølge.
    """
    all_columns = list(tree["columns"])

    try:
        display = tree["displaycolumns"]
    except Exception:
        display = "#all"

    if isinstance(display, str):
        display_list: list[str] = [display]
    else:
        try:
            display_list = [str(x) for x in display]
        except Exception:
            display_list = ["#all"]

    use_all = (
        not display_list
        or display_list == ["#all"]
        or "#all" in display_list
    )

    if use_all:
        visible = list(all_columns)
    else:
        known = set(all_columns)
        visible = [c for c in display_list if c in known]
        if not visible:
            visible = list(all_columns)

    positions = [all_columns.index(c) for c in visible]
    return visible, positions


def treeview_to_sheet(
    tree: Any,
    *,
    title: str,
    heading: str | None = None,
    bold_tags: tuple[str, ...] = ("header", "sum", "sumline", "sumline_major", "sumline_total"),
    bg_tags: dict[str, str] | None = None,
) -> dict:
    """Konverter en ttk.Treeview til et sheet-dict for export_trees_to_excel.

    Respekterer `displaycolumns`: bare kolonner som faktisk er synlige i
    Treeview-en havner i Excel-eksporten, i samme rekkefølge som i GUI-en.
    Overskriftstekst hentes fra `tree.heading(col)["text"]`.

    Args:
        tree: ttk.Treeview-widget.
        title: Arknavn.
        heading: Overskrift i arket (default: samme som title).
        bold_tags: Tags som skal gi fet tekst.
        bg_tags: {tag: hex_color} for bakgrunnsfarge.
    """
    if bg_tags is None:
        bg_tags = {}

    if heading is None:
        heading = title

    try:
        visible_cols, positions = _resolve_visible_columns(tree)
        col_headers = [tree.heading(c)["text"] for c in visible_cols]
    except Exception:
        return {"title": title, "heading": heading, "columns": [], "rows": []}

    rows: list[dict] = []
    try:
        children = tree.get_children("")
    except Exception:
        children = ()

    for iid in children:
        try:
            vals = list(tree.item(iid, "values"))
            tags = tree.item(iid, "tags") or ()
        except Exception:
            continue

        bold = any(t in bold_tags for t in tags)
        bg = next((bg_tags[t] for t in tags if t in bg_tags), None)

        filtered = [vals[i] if i < len(vals) else "" for i in positions]
        parsed: list[Any] = [_try_parse_number(str(v)) for v in filtered]

        rows.append({"values": parsed, "bold": bold, "bg": bg})

    return {
        "title":   title,
        "heading": heading,
        "columns": col_headers,
        "rows":    rows,
    }


# Alias beholdes for Analyse-eksport som bruker ...dict-navnet.
treeview_to_sheet_dict = treeview_to_sheet


def _try_parse_number(s: str) -> Any:
    """Prøv å konvertere norsk-formatert tallstreng til float."""
    clean = s.replace("\u202f", "").replace("\xa0", "").replace(" ", "")
    # Norsk format: 1 234,56
    if "," in clean and clean.replace(",", "").replace("-", "").replace(".", "").isdigit():
        try:
            return float(clean.replace(".", "").replace(",", "."))
        except ValueError:
            pass
    # Forsøk direkte
    try:
        return float(clean.replace(",", "."))
    except ValueError:
        return s


def open_save_dialog(
    title: str = "Lagre Excel",
    default_filename: str = "rapport.xlsx",
    master: Any = None,
) -> str | None:
    """Vis filvelger-dialog og returner valgt filsti (eller None)."""
    try:
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            parent=master,
            title=title,
            defaultextension=".xlsx",
            filetypes=[("Excel-fil", "*.xlsx"), ("Alle filer", "*.*")],
            initialfile=default_filename,
        )
        return path if path else None
    except Exception:
        return None


def export_and_open(
    filepath: str,
    sheets: list[dict],
    *,
    title: str = "",
    client: str = "",
    year: str = "",
) -> None:
    """Eksporter og åpne filen etterpå."""
    import subprocess, sys
    export_trees_to_excel(filepath, sheets, title=title, client=client, year=year)
    try:
        if sys.platform == "win32":
            subprocess.Popen(["start", "", filepath], shell=True)
        else:
            subprocess.Popen(["open" if sys.platform == "darwin" else "xdg-open", filepath])
    except Exception:
        pass
