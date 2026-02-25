from __future__ import annotations

"""excel_sheet_guess.py

Best-effort heuristikk for å velge "riktig" Excel-ark.

Målet er å fjerne friksjon i UI ved å auto-velge arket som mest sannsynlig
inneholder hovedbok/bilagsjournal (dvs. kolonner som Konto, Bilag, Beløp osv.).

Dette er *heuristikk* – hvis vi ikke finner noe som ligner på en hovedbok,
returnerer vi None og lar UI falle tilbake til første ark.
"""

from pathlib import Path

from column_names import make_safe_unique_column_names
from header_detection import detect_header_row
from ml_map_utils import suggest_mapping


def _is_nonempty_cell(v: object) -> bool:
    if v is None:
        return False
    try:
        if isinstance(v, str):
            return v.strip() != ""
        s = str(v).strip()
        return s != ""
    except Exception:
        return False


def _is_excel_path(path: Path) -> bool:
    return path.suffix.lower() in {".xlsx", ".xlsm"}


def guess_best_excel_sheet(
    path: Path,
    *,
    scan_rows: int = 80,
    scan_max_cols: int = 120,
) -> str | None:
    """Best-effort: gjett hvilket Excel-ark som er "hovedbok".

    Vi scorer hvert ark ved å:
      1) lese en liten sample (begrenset rader/kolonner)
      2) detektere hvilken rad som ser ut som header
      3) bruke `ml_map_utils.suggest_mapping()` på header-raden

    Arket med høyest score velges. Scoren vektlegger spesielt at vi finner
    de tre "minste" feltene Konto, Bilag og Beløp.
    """

    # Import lokalt for å unngå å dra inn openpyxl på alle kall.
    import openpyxl

    if not _is_excel_path(path):
        return None
    if not path.exists() or not path.is_file():
        return None

    scan_rows = max(1, int(scan_rows))
    scan_max_cols = max(1, int(scan_max_cols))

    required = {"Konto", "Bilag", "Beløp"}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        best_sheet: str | None = None
        best_score: int = -1

        for sheet in wb.sheetnames:
            try:
                ws = wb[sheet]
            except Exception:
                continue

            rows: list[list[object]] = []
            for row in ws.iter_rows(
                min_row=1,
                max_row=scan_rows,
                min_col=1,
                max_col=scan_max_cols,
                values_only=True,
            ):
                rows.append(list(row))

            if not rows:
                continue

            # Trim trailing tomme kolonner basert på hele sample
            last_col = 0
            for r in rows:
                for idx, val in enumerate(r, start=1):
                    if _is_nonempty_cell(val):
                        last_col = max(last_col, idx)
            if last_col <= 0:
                continue
            rows = [r[:last_col] for r in rows]

            idx0 = detect_header_row(rows)
            header_idx0 = int(idx0) if idx0 is not None else 0
            header_idx0 = max(0, min(header_idx0, len(rows) - 1))
            header_raw = rows[header_idx0]

            headers = make_safe_unique_column_names(header_raw, placeholder_prefix="kol")

            try:
                guess = suggest_mapping(headers) or {}
            except Exception:
                guess = {}

            required_hits = sum(1 for k in required if k in guess)
            score = required_hits * 100 + len(guess)

            # Litt bonus for "menneskelige" headers (ikke bare kolX)
            real_headers = sum(1 for h in headers if h and not h.lower().startswith("kol"))
            score += min(real_headers, 20)

            if score > best_score:
                best_score = score
                best_sheet = sheet

        # Ikke returner et tilfeldig ark hvis vi ikke fant noe som ligner hovedbok
        if best_score <= 0:
            return None
        return best_sheet
    finally:
        try:
            wb.close()
        except Exception:
            pass
