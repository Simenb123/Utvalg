"""Fast, robust dataset builder for common ledger files.

This module contains a small helper used by the Dataset UI:
- Read Excel/CSV with a selected sheet and header row
- Apply a mapping from source column names -> canonical column names
- Coerce common types (dates + amounts)

Notes
-----
The test-suite expects `build_from_file()` to expose both the canonical
(column names as provided in `mapping` keys, e.g. "Konto") and lowercase
aliases (e.g. "konto") for backwards compatibility.

For the interactive application we avoid the extra duplicate columns by
calling `build_from_file(..., include_lowercase_aliases=False)`.
"""

from __future__ import annotations

from dataclasses import dataclass
import math
from pathlib import Path
import re
from typing import Any, Dict, Iterable, Optional, Union

import pandas as pd

from column_names import make_safe_unique_column_names
from header_detection import detect_header_row


@dataclass(frozen=True)
class _ReadOptions:
    sheet_name: Optional[Union[str, int]]
    header_row: int  # 0-indexed


def _coerce_header_row(header_row: int) -> int:
    """UI uses 1-indexed header rows; pandas uses 0-indexed."""
    try:
        n = int(header_row)
    except Exception:
        return 0
    return max(0, n - 1)


def _norm_header(h: object) -> str:
    if h is None:
        return ""
    s = str(h).strip()
    s = s.replace("\u00a0", " ")
    return s.strip().lower()


def _guess_excel_sheet_and_header(path: Path, *, expected_cols: list[str]) -> Optional[_ReadOptions]:
    """Try to guess which worksheet + header row contains `expected_cols`.

    Returns a _ReadOptions with 0-indexed header row or None if no reasonable
    guess is found.
    """
    import openpyxl

    expected = {_norm_header(c) for c in expected_cols if c}
    if not expected:
        return None

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None

    best_sheet: Optional[str] = None
    best_header_idx0: int = 0
    best_match: int = -1

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # sample first ~100 rows
        rows: list[list[object]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 100:
                break
            rows.append(list(row))

        if not rows:
            continue

        guess_idx0: int
        try:
            guess = detect_header_row(rows)
            guess_idx0 = int(guess) if guess is not None else 0
        except Exception:
            guess_idx0 = 0

        guess_idx0 = max(0, min(guess_idx0, len(rows) - 1))

        header_raw = rows[guess_idx0]
        header_safe = make_safe_unique_column_names(header_raw, placeholder_prefix="kol")
        header_norm = {_norm_header(c) for c in header_safe if c is not None}
        match = len(expected & header_norm)

        if match > best_match:
            best_match = match
            best_sheet = sheet
            best_header_idx0 = guess_idx0

    if best_sheet is None:
        return None

    # Require at least one expected column match to avoid picking random sheets
    if best_match <= 0:
        return None

    return _ReadOptions(sheet_name=best_sheet, header_row=best_header_idx0)


def _nullish_mask(series: pd.Series) -> pd.Series:
    """True for values that should be treated as missing."""
    mask = series.isna()
    try:
        s = series.astype(str).str.strip().str.lower()
        mask = mask | s.isin({"", "nan", "none", "null"})
    except Exception:
        pass
    return mask


_amount_keep_re = re.compile(r"[^0-9,\.-]")


def _parse_amount(value: Any) -> Optional[float]:
    """Parse Norwegian/Excel-ish amount formats.

    Supports:
    - parentheses negatives: (34,50)
    - trailing minus: 100-
    - thousand separators (space, NBSP, '.' or ',')
    """
    if value is None:
        return None

    if isinstance(value, float) and math.isnan(value):
        return None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    s = str(value).strip()
    if not s:
        return None

    s_low = s.lower()
    if s_low in {"nan", "none", "null"}:
        return None

    # normalise spaces and unicode minus
    s = s.replace("\u00a0", " ").replace(" ", "")
    s = s.replace("\u2212", "-")

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]

    if s.endswith("-"):
        neg = True
        s = s[:-1]

    if s.startswith("-"):
        neg = True
        s = s[1:]

    # keep only digits and separators
    s = _amount_keep_re.sub("", s)
    if not s:
        return None

    # Determine decimal separator
    if "," in s and "." in s:
        # decimal is the last occurring separator
        if s.rfind(",") > s.rfind("."):
            dec = ","
            thou = "."
        else:
            dec = "."
            thou = ","
        s = s.replace(thou, "")
        s = s.replace(dec, ".")
    elif "," in s:
        # single comma => decimal comma; multiple commas => thousand seps
        if s.count(",") > 1:
            s = s.replace(",", "")
        else:
            s = s.replace(".", "")
            s = s.replace(",", ".")
    else:
        # only dots
        if s.count(".") > 1:
            parts = s.split(".")
            if len(parts[-1]) == 3:
                # likely thousand separators only
                s = "".join(parts)
            else:
                s = "".join(parts[:-1]) + "." + parts[-1]

    try:
        val = float(s)
    except Exception:
        return None

    return -val if neg else val


def _coerce_amount_series(series: pd.Series) -> pd.Series:
    """Rask parsing av beløp.

    For store filer (millioner av rader) er `.apply(_parse_amount)` veldig treg.
    Vi gjør derfor en vektorisert parsing for de vanligste formatene og faller
    tilbake til `_parse_amount` kun for få "rare" verdier.

    Støtter typiske formater:
    - 1 234,50 / 1.234,50 / 1234,50
    - 1,234.50 / 1234.50
    - (34,50) / 100-  (negative)
    """

    if series is None:
        return pd.Series(dtype=float)

    if series.empty:
        return pd.to_numeric(series, errors="coerce")

    # Numeric dtypes -> direkte
    try:
        if pd.api.types.is_numeric_dtype(series.dtype):
            return pd.to_numeric(series, errors="coerce")
    except Exception:
        pass

    s = series.astype("string").fillna("")
    s = s.str.replace(" ", " ", regex=False).str.strip()

    # Marker negative: (..), trailing -, leading -
    neg_paren = s.str.startswith("(") & s.str.endswith(")")
    s = s.mask(neg_paren, s.str.slice(1, -1))

    trailing_minus = s.str.endswith("-", na=False)
    leading_minus = s.str.startswith("-", na=False)
    neg = neg_paren | trailing_minus | leading_minus

    # Fjern minus-tegn før normalisering
    try:
        s = s.str.removesuffix("-").str.removeprefix("-")
    except Exception:  # pragma: no cover (eldre pandas)
        s = s.str.replace(r"-$", "", regex=True)
        s = s.str.replace(r"^-", "", regex=True)

    # Fjern alt bortsett fra 0-9 , . -
    s = s.str.replace(_amount_keep_re, "", regex=True)

    # Tomt / "nan" tokens -> NA
    s = s.replace({"": pd.NA, "nan": pd.NA, "none": pd.NA, "null": pd.NA})

    # Normaliser separatorer:
    # - Hvis både . og , finnes: siste separator er desimal
    has_comma = s.str.contains(",", regex=False, na=False)
    has_dot = s.str.contains(".", regex=False, na=False)

    both = has_comma & has_dot
    if both.any():
        sb = s.where(both)
        last_comma = sb.str.rfind(",")
        last_dot = sb.str.rfind(".")
        comma_decimal = last_comma > last_dot

        # comma as decimal => remove dots, replace comma with dot
        idx = both & comma_decimal
        if idx.any():
            s.loc[idx] = s.loc[idx].str.replace(".", "", regex=False).str.replace(",", ".", regex=False)

        # dot as decimal => remove commas
        idx = both & (~comma_decimal)
        if idx.any():
            s.loc[idx] = s.loc[idx].str.replace(",", "", regex=False)

    # Kun komma
    only_comma = has_comma & (~has_dot)
    if only_comma.any():
        comma_count = s.str.count(",").fillna(0)

        thousand = only_comma & s.str.match(r"^-?\d{1,3}(,\d{3})+$", na=False)
        if thousand.any():
            s.loc[thousand] = s.loc[thousand].str.replace(",", "", regex=False)

        single = only_comma & (~thousand) & (comma_count == 1)
        if single.any():
            s.loc[single] = s.loc[single].str.replace(".", "", regex=False).str.replace(",", ".", regex=False)

    # Kun punktum
    only_dot = has_dot & (~has_comma)
    if only_dot.any():
        dot_count = s.str.count(r"\.").fillna(0)

        thousand = only_dot & s.str.match(r"^-?\d{1,3}(\.\d{3})+$", na=False)
        if thousand.any():
            s.loc[thousand] = s.loc[thousand].str.replace(".", "", regex=False)

        # Flere punktum som ikke er rene tusenskiller
        multi = only_dot & (dot_count > 1) & (~thousand)
        if multi.any():
            parts = s.loc[multi].str.split(".", regex=False)
            last_len = parts.str[-1].str.len()
            thou_only = last_len == 3

            if thou_only.any():
                idx = parts.index[thou_only]
                s.loc[idx] = s.loc[idx].str.replace(".", "", regex=False)

            dec_idx = parts.index[~thou_only]
            if len(dec_idx) > 0:
                def _join_last(lst):
                    if not lst or len(lst) == 1:
                        return "".join(lst)
                    return "".join(lst[:-1]) + "." + lst[-1]

                s.loc[dec_idx] = parts.loc[dec_idx].apply(_join_last)

    out = pd.to_numeric(s, errors="coerce")
    out = out.where(~neg, -out)

    # Rare fallback: hvis få "rare" rader fortsatt ikke lot seg parse, bruk treg parser
    try:
        rare_mask = out.isna() & s.notna()
        if rare_mask.sum() > 0 and rare_mask.sum() <= 5000:
            out.loc[rare_mask] = series.loc[rare_mask].apply(_parse_amount)
            out = pd.to_numeric(out, errors="coerce")
    except Exception:
        pass

    return out


_EXCEL_DATE_ORIGIN = "1899-12-30"

def _coerce_date_series(series: pd.Series) -> pd.Series:
    """Coerce date values to datetime.

    Handles common Norwegian/ISO date strings and Excel serial dates
    (days since 1899-12-30).
    """
    if series.empty:
        return pd.to_datetime(series, errors="coerce")

    # Excel serial dates often arrive as strings (we read Excel with dtype=str).
    num = pd.to_numeric(series, errors="coerce")
    serial_mask = num.notna() & num.between(20000, 80000)

    parsed = pd.to_datetime(series, errors="coerce", format="mixed", dayfirst=True)

    if serial_mask.any():
        serial_dates = pd.to_datetime(num, unit="D", origin=_EXCEL_DATE_ORIGIN, errors="coerce")
        parsed = parsed.where(~serial_mask, serial_dates)

    return parsed


_leading_zero_re = re.compile(r"^0\d+")


def _coerce_int_like_series(series: pd.Series) -> pd.Series:
    """Convert to Int64 if every non-null value is numeric and not leading-zero coded."""
    null_mask = _nullish_mask(series)

    # If any non-null values look like codes with leading zeros, keep as-is
    try:
        s = series.where(~null_mask).astype(str).str.strip()
        if s.dropna().apply(lambda x: bool(_leading_zero_re.match(x))).any():
            return series
    except Exception:
        pass

    num = pd.to_numeric(series.where(~null_mask), errors="coerce")

    # Only convert if all non-null values convert
    if num.where(~null_mask).notna().all():
        # Int-like?
        if ((num.dropna() % 1) == 0).all():
            return num.round(0).astype("Int64")
    return series


def build_from_file(
    source: Union[str, Path, pd.DataFrame],
    mapping: Optional[Dict[str, str]] = None,
    *,
    sheet_name: Optional[Union[str, int]] = None,
    header_row: int = 1,
    include_lowercase_aliases: bool = True,
    # Nyere GUI-kode (klient/versjonering + SAF-T) sender inn dette flagget.
    # Vi aksepterer det for kompatibilitet; vi bygger likt uansett når vi
    # allerede har et DataFrame.
    is_saft: bool = False,
) -> pd.DataFrame:
    """Build a canonical dataset DataFrame.

    Parameters
    ----------
    source:
        Either a file path to .xlsx/.xls/.csv/.txt, or an already loaded
        pandas DataFrame.
    mapping:
        Dict mapping canonical field name -> source column header.
    sheet_name:
        Excel sheet name/index (only for Excel files).
    header_row:
        1-indexed header row (as shown to the user).
    include_lowercase_aliases:
        If True, adds lowercase alias columns (e.g. "konto") mirroring the
        canonical ones (e.g. "Konto"). This is mainly for backward
        compatibility in tests.

    Returns
    -------
    pd.DataFrame
        DataFrame with canonical column names (and optionally lowercase
        aliases).
    """
    if isinstance(source, pd.DataFrame):
        # Do not mutate caller DF
        df = source.copy()
    else:
        path = Path(source)

        # Normalise options
        opts = _ReadOptions(sheet_name=sheet_name, header_row=_coerce_header_row(header_row))

        usecols: Optional[list[str]] = list(mapping.values()) if mapping else None

        # First try direct read with requested sheet/header
        try:
            df = _read_with_usecols(path, usecols=usecols, opts=opts)
        except Exception:
            # Only guess sheet/header if the user did not explicitly set sheet_name
            if usecols and path.suffix.lower() in {".xlsx", ".xls"} and sheet_name is None:
                guessed = _guess_excel_sheet_and_header(path, expected_cols=usecols)
                if guessed is not None:
                    df = _read_with_usecols(path, usecols=usecols, opts=guessed)
                else:
                    raise
            else:
                raise

    # Clean/normalize column labels (inkl. blanke/"Unnamed" -> kolX)
    df.columns = make_safe_unique_column_names(list(df.columns), placeholder_prefix="kol")

    if mapping:
        missing_src = [src for src in mapping.values() if src not in df.columns]
        if missing_src:
            raise ValueError(
                "Følgende kolonner finnes ikke i datasettet (etter header-normalisering): "
                + ", ".join(sorted(set(missing_src)))
            )

        # Rename source columns -> canonical
        renamed = {src: canon for canon, src in mapping.items() if src in df.columns}
        df = df.rename(columns=renamed)

        # Keep only the mapped canonical fields in mapping order
        fields = list(mapping.keys())
        df = df[fields].copy()

        # Type coercions
        if "Dato" in df.columns:
            df["Dato"] = _coerce_date_series(df["Dato"])

        if "Beløp" in df.columns:
            df["Beløp"] = _coerce_amount_series(df["Beløp"]).fillna(0.0)

        for int_field in ("Kundenr", "Leverandørnr"):
            if int_field in df.columns:
                df[int_field] = _coerce_int_like_series(df[int_field])

        if include_lowercase_aliases:
            for canon in fields:
                lc = canon.lower()
                if lc not in df.columns:
                    df[lc] = df[canon]

    return df


def _read_with_usecols(path: Path, *, usecols: Optional[Iterable[str]], opts: _ReadOptions) -> pd.DataFrame:
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xlsm"):
        # Pandas (via openpyxl) can take *forever* on some Excel exports because
        # the sheet "used range" (max_row) is polluted by formatting all the way
        # down to row 1,048,576. That makes openpyxl iterate a million rows of
        # blanks. We instead stream the sheet with openpyxl in read_only mode and
        # stop after a long run of empty rows.

        from openpyxl import load_workbook

        HEADER_MAX_COLS = 500
        LOOKAHEAD_ROWS = 25

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            # Resolve worksheet
            ws = None
            if isinstance(opts.sheet_name, int):
                idx = int(opts.sheet_name)
                ws = wb.worksheets[idx] if 0 <= idx < len(wb.worksheets) else wb.active
            elif isinstance(opts.sheet_name, str) and opts.sheet_name.strip():
                ws = wb[opts.sheet_name]
            else:
                ws = wb.active

            header_excel_row = int(opts.header_row) + 1  # openpyxl is 1-indexed

            # Les header + litt lookahead for å finne faktisk bredde.
            # Dette gjør at vi ikke mister en trailing blank header-kolonne hvis
            # det finnes data i kolonnen.
            sample_rows: list[list[object]] = []
            for row in ws.iter_rows(
                min_row=header_excel_row,
                max_row=header_excel_row + LOOKAHEAD_ROWS,
                min_col=1,
                max_col=HEADER_MAX_COLS,
                values_only=True,
            ):
                sample_rows.append(list(row))

            header_list = sample_rows[0] if sample_rows else []
            last_nonempty = -1
            for r in sample_rows:
                for i, v in enumerate(r):
                    if v is None:
                        continue
                    if str(v).strip() == "":
                        continue
                    last_nonempty = max(last_nonempty, i)

            if last_nonempty >= 0:
                header_list = header_list[: last_nonempty + 1]
            else:
                header_list = []

            cols = make_safe_unique_column_names(header_list, placeholder_prefix="kol")

            wanted: Optional[list[str]] = None
            if usecols:
                wanted = [str(c).strip() for c in usecols if str(c).strip()]

            if wanted:
                idx_map = {c: i for i, c in enumerate(cols)}
                keep_idx = [idx_map[c] for c in wanted if c in idx_map]
                keep_cols = [cols[i] for i in keep_idx]

                # If none of the requested columns exist on the selected header
                # row, treat this as a hard failure. Returning an empty
                # DataFrame hides the real problem and prevents build_from_file
                # from auto-guessing the correct sheet/header.
                if not keep_idx:
                    raise ValueError(
                        "Ingen av de ønskede kolonnene ble funnet på valgt header-rad. "
                        "Prøv å velge riktig ark/header-rad, eller la importen gjette automatisk."
                    )
            else:
                keep_idx = list(range(len(cols)))
                keep_cols = cols

            max_col = (max(keep_idx) + 1) if keep_idx else len(cols)

            rows_out: list[list[object]] = []
            blank_run = 0
            # High threshold to avoid cutting off legitimate data with occasional
            # blank lines, but still bail out quickly on polluted used ranges.
            BLANK_RUN_STOP = 500

            for row in ws.iter_rows(
                min_row=header_excel_row + 1,
                min_col=1,
                max_col=max_col,
                values_only=True,
            ):
                selected: list[object] = []
                all_blank = True
                for i in keep_idx:
                    v = row[i] if i < len(row) else None
                    if v is None:
                        selected.append(None)
                        continue
                    s = str(v).strip()
                    if s == "":
                        selected.append(None)
                        continue
                    all_blank = False
                    selected.append(s)

                if all_blank:
                    blank_run += 1
                    if blank_run >= BLANK_RUN_STOP:
                        break
                    continue
                blank_run = 0
                rows_out.append(selected)

            return pd.DataFrame(rows_out, columns=keep_cols)
        finally:
            try:
                wb.close()
            except Exception:
                # Some openpyxl versions don't expose close() on read_only wbs.
                pass

    if suffix == ".xls":
        # Legacy binary Excel files still go through pandas.
        return pd.read_excel(
            path,
            sheet_name=opts.sheet_name or 0,
            header=opts.header_row,
            usecols=list(usecols) if usecols else None,
            dtype=str,
        )


    # CSV / text
    def _sniff_csv_sep_fast() -> Optional[str]:
        """Rask separator-sjekk for store CSV-filer.

        Leser et lite utdrag og velger den separatoren som ser mest sannsynlig ut.
        Gir None hvis vi ikke kan avgjøre (da brukes pandas' egen sniffing).
        """
        try:
            # Les noen linjer forbi header-raden (hvis valgt)
            max_lines = 50
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                lines = []
                for _ in range((opts.header_row or 0) + max_lines):
                    ln = f.readline()
                    if not ln:
                        break
                    if ln.strip() == "":
                        continue
                    lines.append(ln)
            if not lines:
                return None

            # Bruk de siste linjene (nærmest header/data) for telling
            sample = lines[-10:]
            candidates = [",", ";", "\t", "|"]
            counts = {}
            for sep in candidates:
                counts[sep] = sum(ln.count(sep) for ln in sample)

            best = max(counts, key=counts.get)
            # Må ha litt signal for å velge
            if counts[best] <= 0:
                return None
            return best
        except Exception:
            return None

    def _read_csv_fast(*, sep: Optional[str], usecols_idx: Optional[list[int]] = None, nrows: Optional[int] = None) -> pd.DataFrame:
        # engine='c' er mye raskere enn 'python' for store filer når sep er kjent
        engine = "c" if sep is not None else "python"
        return pd.read_csv(
            path,
            sep=sep,
            engine=engine,
            header=opts.header_row,
            usecols=usecols_idx,
            dtype=str,
            encoding_errors="replace",
            memory_map=True,
            keep_default_na=False,
            na_filter=False,
            nrows=nrows,
        )

    sniffed_sep = _sniff_csv_sep_fast()

    if usecols is None:
        # Read everything
        return _read_csv_fast(sep=sniffed_sep)

    # Try common delimiters first (more stable than pandas' sniffer with preamble rows)
    seps = [sniffed_sep] if sniffed_sep else []
    seps += [",", ";", "\t", "|"]
    # dedupe while preserving order
    seen = set()
    seps = [s for s in seps if (s is not None and not (s in seen or seen.add(s)))]

    wanted = [str(c).strip() for c in usecols if str(c).strip()]

    def _read_csv_with_sep(sep: Optional[str]) -> pd.DataFrame:
        # Les kun header for å finne rå kolonnenavn (inkl. "Unnamed: x").
        head = _read_csv_fast(sep=sep, nrows=0)
        raw_cols = list(head.columns)
        safe_cols = make_safe_unique_column_names(raw_cols, placeholder_prefix="kol")
        idx_map = {c: i for i, c in enumerate(safe_cols)}
        keep_idx = [idx_map[c] for c in wanted if c in idx_map]
        if not keep_idx:
            raise ValueError(
                "Ingen av de ønskede kolonnene ble funnet på valgt header-rad. "
                "Prøv å velge riktig header-rad, eller la importen gjette automatisk."
            )

        keep_idx_sorted = sorted(set(keep_idx))
        df = _read_csv_fast(sep=sep, usecols_idx=keep_idx_sorted)
        df.columns = [safe_cols[i] for i in keep_idx_sorted]
        return df

    last_err: Optional[Exception] = None

    for sep in seps:
        try:
            return _read_csv_with_sep(sep)
        except Exception as e:
            last_err = e

    # Fallback to pandas sniffer
    try:
        return _read_csv_with_sep(None)
    except Exception:
        if last_err is not None:
            raise last_err
        raise
