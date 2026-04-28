"""Excel-arbeidspapir for SB versjonsdiff.

Genererer arbeidsbok med fire ark:
  - Oppsummering: totaler og nøkkeltall
  - Nye konti: konti som kun finnes i ny versjon
  - Fjernede konti: konti som kun finnes i gammel versjon
  - Endrede saldoer: konti med endret IB eller UB
"""

from __future__ import annotations

from datetime import datetime
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.shared.workpapers.forside import build_forside_sheet


_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
_HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
_ADDED_FILL = PatternFill("solid", fgColor="E2EFDA")     # Grønn — nye
_REMOVED_FILL = PatternFill("solid", fgColor="FCE4EC")   # Rød — fjernede
_CHANGED_FILL = PatternFill("solid", fgColor="FFF2CC")   # Gul — endrede
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_AMOUNT_FMT = '#,##0.00;[Red]-#,##0.00'


def build_sb_diff_workpaper(
    diff_result,
    *,
    client: str | None = None,
    year: str | int | None = None,
    version_a_label: str = "Forrige",
    version_b_label: str = "Gjeldende",
) -> Workbook:
    """Bygg komplett SB versjonsdiff arbeidspapir."""
    wb = Workbook()

    _build_summary_sheet(
        wb, diff_result.summary,
        client=client, year=year,
        version_a_label=version_a_label,
        version_b_label=version_b_label,
    )
    _build_konti_sheet(wb, "Nye konti", diff_result.added, _ADDED_FILL,
                        cols=("konto", "kontonavn", "ib", "ub"))
    _build_konti_sheet(wb, "Fjernede konti", diff_result.removed, _REMOVED_FILL,
                        cols=("konto", "kontonavn", "ib", "ub"))
    _build_changed_sheet(wb, diff_result.changed,
                          version_a_label=version_a_label,
                          version_b_label=version_b_label)

    build_forside_sheet(wb, workpaper_navn="SB versjonsdiff")

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    return wb


def _build_summary_sheet(
    wb: Workbook,
    summary: Dict[str, object],
    *,
    client: str | None,
    year: str | int | None,
    version_a_label: str,
    version_b_label: str,
) -> None:
    ws = wb.create_sheet("Oppsummering")
    title_parts = ["SB Versjonsdiff"]
    if client:
        title_parts.append(str(client))
    if year not in {None, ""}:
        title_parts.append(str(year))
    ws["A1"] = "  ·  ".join(title_parts)
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A1"].fill = _TITLE_FILL
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 26

    ws["A3"] = "Sammenligning"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "Versjon A:"
    ws["B4"] = version_a_label
    ws["A5"] = "Versjon B:"
    ws["B5"] = version_b_label
    ws["A6"] = "Generert:"
    ws["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    rows = [
        ("Konti i versjon A", summary.get("konti_a_total", 0)),
        ("Konti i versjon B", summary.get("konti_b_total", 0)),
        ("Sum UB versjon A", summary.get("sum_ub_a", 0.0)),
        ("Sum UB versjon B", summary.get("sum_ub_b", 0.0)),
        ("", ""),
        ("Nye konti", summary.get("nye_konti", 0)),
        ("Fjernede konti", summary.get("fjernede_konti", 0)),
        ("Endrede saldoer", summary.get("endrede_konti", 0)),
        ("Uendrede konti", summary.get("uendrede_konti", 0)),
    ]
    start_row = 8
    for i, (label, val) in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=bool(label))
        c = ws.cell(row=r, column=2, value=val)
        if isinstance(val, float):
            c.number_format = _AMOUNT_FMT

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22


def _build_konti_sheet(
    wb: Workbook, name: str, df: pd.DataFrame, header_fill: PatternFill,
    *, cols: tuple[str, ...],
) -> None:
    ws = wb.create_sheet(name)
    headers = {"konto": "Konto", "kontonavn": "Kontonavn", "ib": "IB", "ub": "UB"}
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=headers.get(c, c))
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = _BORDER

    if df is None or df.empty:
        ws.cell(row=2, column=1, value="(ingen)").font = Font(italic=True, color="888888")
    else:
        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            for c_idx, c in enumerate(cols, 1):
                v = row.get(c)
                cell = ws.cell(row=r_idx, column=c_idx, value=v)
                cell.border = _BORDER
                if c in {"ib", "ub"}:
                    cell.number_format = _AMOUNT_FMT
                    cell.alignment = Alignment(horizontal="right")

    widths = {"konto": 14, "kontonavn": 36, "ib": 16, "ub": 16}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 16)


def _build_changed_sheet(
    wb: Workbook, df: pd.DataFrame,
    *, version_a_label: str, version_b_label: str,
) -> None:
    ws = wb.create_sheet("Endrede saldoer")
    headers = [
        ("konto", "Konto", 14),
        ("kontonavn", "Kontonavn", 36),
        ("ib_a", f"IB {version_a_label}", 16),
        ("ib_b", f"IB {version_b_label}", 16),
        ("diff_ib", "Diff IB", 14),
        ("ub_a", f"UB {version_a_label}", 16),
        ("ub_b", f"UB {version_b_label}", 16),
        ("diff_ub", "Diff UB", 14),
    ]
    for i, (_, label, _w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = Font(bold=True)
        cell.fill = _CHANGED_FILL
        cell.border = _BORDER

    if df is None or df.empty:
        ws.cell(row=2, column=1, value="(ingen endringer)").font = Font(italic=True, color="888888")
    else:
        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            for c_idx, (col_key, _label, _w) in enumerate(headers, 1):
                v = row.get(col_key)
                cell = ws.cell(row=r_idx, column=c_idx, value=v)
                cell.border = _BORDER
                if col_key in {"ib_a", "ib_b", "diff_ib", "ub_a", "ub_b", "diff_ub"}:
                    cell.number_format = _AMOUNT_FMT
                    cell.alignment = Alignment(horizontal="right")

    for i, (_, _label, w) in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
