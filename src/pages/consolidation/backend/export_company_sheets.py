"""Company, elimination and associate export sheets."""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from .models import AssociateCase, CompanyTB, EliminationJournal

_HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_AMOUNT_FMT = "#,##0.00;[Red]-#,##0.00"


def _excel_col(idx: int) -> str:
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(val)
        return 0.0 if pd.isna(f) else f
    except (ValueError, TypeError):
        return 0.0


_KIND_LABELS = {
    "manual": "Manuell",
    "from_suggestion": "Forslag",
    "template": "Template",
    "equity_method": "EK-metode",
}


def build_elimineringer(
    wb: Workbook,
    eliminations: list[EliminationJournal],
    company_names: dict[str, str] | None = None,
) -> None:
    ws = wb.create_sheet("Elimineringer")

    if not eliminations:
        ws["A1"] = "Ingen elimineringer registrert."
        return

    name_map = company_names or {}
    headers = ["Bilag", "Type", "Regnr", "Selskap", "Beloep", "Beskrivelse"]
    row = 1

    for journal in eliminations:
        label = journal.display_label
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, size=12)
        kind_label = _KIND_LABELS.get(journal.kind, journal.kind)
        ws.cell(row=row, column=2, value=kind_label)
        balanced_text = "Balansert" if journal.is_balanced else f"UBALANSE ({journal.net:.2f})"
        ws.cell(row=row, column=4, value=balanced_text)
        row += 1

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.border = _BORDER
        row += 1

        for line in journal.lines:
            ws.cell(row=row, column=1, value=label).border = _BORDER
            ws.cell(row=row, column=2, value=kind_label).border = _BORDER
            ws.cell(row=row, column=3, value=line.regnr).border = _BORDER
            company_display = name_map.get(line.company_id, line.company_id[:16])
            ws.cell(row=row, column=4, value=company_display).border = _BORDER
            c = ws.cell(row=row, column=5, value=line.amount)
            c.number_format = _AMOUNT_FMT
            c.border = _BORDER
            ws.cell(row=row, column=6, value=line.description).border = _BORDER
            row += 1

        row += 1

    for col, w in {"A": 20, "B": 12, "C": 10, "D": 20, "E": 16, "F": 30}.items():
        ws.column_dimensions[col].width = w


def build_company_sheets(
    wb: Workbook,
    companies: list[CompanyTB],
    mapped_tbs: dict[str, pd.DataFrame],
    *,
    regnr_to_name: dict[int, str] | None = None,
    hide_zero: bool = False,
) -> None:
    _col_headers = {
        "konto": "Konto",
        "kontonavn": "Kontonavn",
        "regnr": "Regnr",
        "rl_navn": "Regnskapslinje",
        "ib": "IB",
        "netto": "Bevegelse",
        "ub": "UB",
    }
    _col_order = ["konto", "kontonavn", "regnr", "rl_navn", "ib", "netto", "ub"]
    _amount_cols = {"ib", "netto", "ub"}

    for company in companies:
        tb = mapped_tbs.get(company.company_id)
        if tb is None or tb.empty:
            continue

        prefix = "Grunnlag" if getattr(company, "is_line_basis", False) else "TB"
        sheet_name = f"{prefix} - {company.name}"[:31]
        ws = wb.create_sheet(sheet_name)

        if getattr(company, "is_line_basis", False):
            line_cols = [
                ("regnr", "Regnr"),
                ("regnskapslinje", "Regnskapslinje"),
                ("source_regnskapslinje", "Kildelinje"),
                ("ub", "UB"),
                ("source_page", "Side"),
                ("confidence", "Score"),
                ("review_status", "Status"),
            ]
            show_cols = [col for col, _label in line_cols if col in tb.columns]
            for col_idx, col in enumerate(show_cols, start=1):
                label = next(label for key, label in line_cols if key == col)
                cell = ws.cell(row=1, column=col_idx, value=label)
                cell.font = Font(bold=True)
                cell.fill = _HEADER_FILL
                cell.border = _BORDER

            row_idx = 2
            for _, row in tb.iterrows():
                if hide_zero and abs(_safe_float(row.get("ub"))) < 0.005:
                    continue
                for col_idx, col in enumerate(show_cols, start=1):
                    val = row.get(col)
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = _BORDER
                    if col in {"ub", "confidence"}:
                        cell.number_format = _AMOUNT_FMT if col == "ub" else "0.000"
                row_idx += 1

            ws.freeze_panes = "A2"
            widths = {
                "regnr": 10,
                "regnskapslinje": 30,
                "source_regnskapslinje": 30,
                "ub": 16,
                "source_page": 8,
                "confidence": 10,
                "review_status": 12,
            }
            for i, col in enumerate(show_cols, start=1):
                ws.column_dimensions[_excel_col(i)].width = widths.get(col, 16)
            continue

        has_regnr = "regnr" in tb.columns
        show_cols = [c for c in _col_order if c in tb.columns or c == "rl_navn"]
        if not has_regnr:
            show_cols = [c for c in show_cols if c not in ("regnr", "rl_navn")]

        for col_idx, col in enumerate(show_cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=_col_headers.get(col, col))
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.border = _BORDER

        row_idx = 2
        for _, row in tb.iterrows():
            if hide_zero:
                data_vals = [float(row.get(c, 0) or 0) for c in _amount_cols if c in tb.columns]
                if all(abs(v) < 0.005 for v in data_vals):
                    continue
            for col_idx, col in enumerate(show_cols, start=1):
                if col == "rl_navn":
                    regnr_raw = row.get("regnr")
                    try:
                        rn = int(regnr_raw) if pd.notna(regnr_raw) else None
                    except (ValueError, TypeError):
                        rn = None
                    val = (regnr_to_name or {}).get(rn, "") if rn is not None else ""
                else:
                    val = row.get(col)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = _BORDER
                if col in _amount_cols:
                    cell.number_format = _AMOUNT_FMT
            row_idx += 1

        ws.freeze_panes = "A2"
        for i, col in enumerate(show_cols, start=1):
            if col == "kontonavn":
                ws.column_dimensions[_excel_col(i)].width = 35
            elif col == "rl_navn":
                ws.column_dimensions[_excel_col(i)].width = 30
            elif col in _amount_cols:
                ws.column_dimensions[_excel_col(i)].width = 16
            else:
                ws.column_dimensions[_excel_col(i)].width = 12


def build_associate_sheets(
    wb: Workbook,
    associate_cases: list[AssociateCase],
    eliminations: list[EliminationJournal],
    *,
    company_names: dict[str, str],
    regnr_to_name: dict[int, str],
) -> None:
    if not associate_cases:
        return

    for case in associate_cases:
        sheet_name = f"EK - {case.name}"[:31] if case.name else f"EK - {case.case_id[:8]}"
        ws = wb.create_sheet(sheet_name)
        row = 1

        def write_kv(label: str, value: object, *, bold: bool = False) -> None:
            nonlocal row
            key_cell = ws.cell(row=row, column=1, value=label)
            key_cell.font = Font(bold=True)
            val_cell = ws.cell(row=row, column=2, value=value)
            if bold:
                val_cell.font = Font(bold=True)
            row += 1

        ownership = _safe_float(case.ownership_pct)
        opening = _safe_float(case.opening_carrying_amount)
        share_of_result = _safe_float(case.share_of_result)
        other_equity = _safe_float(case.share_of_other_equity)
        dividends = _safe_float(case.dividends)
        impairment = _safe_float(case.impairment)
        excess_value = _safe_float(case.excess_value_amortization)
        manual_total = sum(_safe_float(adj.amount) for adj in case.manual_adjustment_rows)
        closing = opening + share_of_result + other_equity - dividends - impairment - excess_value + manual_total

        write_kv("Tilknyttet selskap", case.name, bold=True)
        write_kv("Investor", company_names.get(case.investor_company_id, case.investor_company_id))
        write_kv("Eierandel %", ownership)
        write_kv("Status", case.status)
        write_kv("Kilde", case.source_mode)
        write_kv("Anskaffelsesdato", case.acquisition_date)
        write_kv("Bilag", case.journal_id)
        row += 1

        ws.cell(row=row, column=1, value="Arbeidspapir").font = Font(bold=True, size=12)
        row += 1
        rows = [
            ("Inngående bokført verdi", opening),
            ("Andel resultat", share_of_result),
            ("Andre EK-bevegelser", other_equity),
            ("Utbytte", -dividends),
            ("Nedskrivning", -impairment),
            ("Merverdi/amortisering", -excess_value),
            ("Manuelle justeringer", manual_total),
            ("Utgående bokført verdi", closing),
        ]
        for label, value in rows:
            write_kv(label, value, bold=label == "Utgående bokført verdi")

        if case.manual_adjustment_rows:
            row += 1
            ws.cell(row=row, column=1, value="Manuelle justeringer").font = Font(bold=True)
            row += 1
            headers = ["Label", "Beløp", "Motpost regnr", "Motpost", "Beskrivelse"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = _HEADER_FILL
                cell.border = _BORDER
            row += 1
            for adj in case.manual_adjustment_rows:
                values = [
                    adj.label,
                    _safe_float(adj.amount),
                    int(adj.offset_regnr or 0),
                    regnr_to_name.get(int(adj.offset_regnr or 0), ""),
                    adj.description,
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = _BORDER
                    if col_idx == 2:
                        cell.number_format = _AMOUNT_FMT
                row += 1

        journal = next((item for item in eliminations if item.journal_id == case.journal_id), None)
        if journal is not None:
            row += 1
            ws.cell(row=row, column=1, value="Generert EK-føring").font = Font(bold=True, size=12)
            row += 1
            headers = ["Regnr", "Regnskapslinje", "Beløp", "Beskrivelse"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = _HEADER_FILL
                cell.border = _BORDER
            row += 1
            for line in journal.lines:
                values = [
                    int(line.regnr or 0),
                    regnr_to_name.get(int(line.regnr or 0), ""),
                    _safe_float(line.amount),
                    line.description,
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = _BORDER
                    if col_idx == 3:
                        cell.number_format = _AMOUNT_FMT
                row += 1

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 26
        ws.column_dimensions["E"].width = 34
