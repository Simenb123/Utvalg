"""Main workbook sheets for consolidation export."""

from __future__ import annotations

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .control_rows import append_control_rows
from .models import CompanyTB, EliminationJournal

_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
_HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
_SUM_FILL = PatternFill("solid", fgColor="F3F6F9")
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_AMOUNT_FMT = "#,##0.00;[Red]-#,##0.00"


def _excel_col(idx: int) -> str:
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(val)
        return 0.0 if pd.isna(f) else f
    except (ValueError, TypeError):
        return 0.0


def build_konsernoppstilling(
    wb: Workbook,
    result_df: pd.DataFrame,
    *,
    client: str | None = None,
    year: str | None = None,
    companies: list[CompanyTB] | None = None,
    parent_company_id: str = "",
    hide_zero: bool = False,
) -> None:
    augmented = append_control_rows(result_df)
    if augmented is not None:
        result_df = augmented
    ws = wb.active
    ws.title = "Konsernoppstilling"

    title_parts = ["Konsernoppstilling"]
    if client:
        title_parts.append(str(client))
    if year:
        title_parts.append(str(year))
    title = " - ".join(title_parts)

    meta_cols = {"regnr", "regnskapslinje", "sumpost", "formel"}
    data_cols = [c for c in result_df.columns if c not in meta_cols]
    company_names = {c.name for c in (companies or []) if getattr(c, "name", None)}
    parent_name = next(
        (c.name for c in (companies or []) if c.company_id == parent_company_id and c.name),
        "",
    )
    aggregate_order = ["Mor", "Doetre", "sum_foer_elim", "eliminering", "konsolidert"]
    company_cols = [c for c in data_cols if c in company_names]
    other_cols = [c for c in data_cols if c not in company_names and c not in aggregate_order]

    ordered_company_cols: list[str] = []
    if parent_name and parent_name in company_cols:
        ordered_company_cols.append(parent_name)
    ordered_company_cols.extend(
        sorted((c for c in company_cols if c != parent_name), key=lambda name: str(name).lower())
    )
    ordered_data_cols = ordered_company_cols + other_cols + [c for c in aggregate_order if c in data_cols]
    all_headers = ["Nr", "Regnskapslinje"] + ordered_data_cols
    total_cols = len(all_headers)
    last_col_letter = _excel_col(total_cols)

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws["A1"].fill = _TITLE_FILL

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    header_row = 4
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

    data_row = header_row + 1
    for _, row in result_df.iterrows():
        regnr = int(row.get("regnr", 0))
        is_sum = bool(row.get("sumpost", False))

        if hide_zero and not is_sum:
            data_vals = [_safe_float(row.get(dc)) for dc in ordered_data_cols]
            if all(abs(v) < 0.005 for v in data_vals):
                continue

        values = [regnr, str(row.get("regnskapslinje", "") or "")]
        for dc in ordered_data_cols:
            values.append(_safe_float(row.get(dc)))

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.border = _BORDER
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
            if col_idx >= 3:
                cell.number_format = _AMOUNT_FMT
            if is_sum:
                cell.font = Font(bold=True)
                cell.fill = _SUM_FILL

        data_row += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_col_letter}{max(data_row - 1, 4)}"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    for i in range(3, total_cols + 1):
        ws.column_dimensions[_excel_col(i)].width = 16


def build_konsolidert_sb(
    wb: Workbook,
    result_df: pd.DataFrame,
    *,
    companies: list[CompanyTB],
    parent_company_id: str = "",
    eliminations: list[EliminationJournal],
) -> None:
    if result_df is None or result_df.empty:
        return

    ws = wb.create_sheet("Konsolidert SB")

    company_names = [c.name for c in companies]
    headers = ["Regnr", "Regnskapslinje"] + company_names + ["Sum foer elim", "Eliminering", "Konsolidert"]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER

    amt_start = 3

    for row_idx, (_, row) in enumerate(result_df.iterrows(), start=2):
        regnr = int(row.get("regnr", 0))
        is_sum = bool(row.get("sumpost", False))
        vals: list[object] = [regnr, row.get("regnskapslinje", "")]

        for cname in company_names:
            vals.append(_safe_float(row.get(cname)))

        vals.append(_safe_float(row.get("sum_foer_elim")))
        vals.append(_safe_float(row.get("eliminering")))
        vals.append(_safe_float(row.get("konsolidert")))

        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            if col_idx >= amt_start:
                cell.number_format = _AMOUNT_FMT
            if is_sum:
                cell.fill = _SUM_FILL
                cell.font = Font(bold=True)

    ws.freeze_panes = "C2"
    col_widths = [8, 30] + [16] * len(company_names) + [16, 16, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[_excel_col(i)].width = w
