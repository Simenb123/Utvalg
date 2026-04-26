"""dataset_pane_io.py

Små, testbare I/O-hjelpere for DatasetPane.

Formål
------
DatasetPane trenger lettvekts-operasjoner for:
  - liste arknavn (Excel)
  - lese en liten sample til forhåndsvisning
  - lese kun header-raden
  - best-effort autodeteksjon av header-rad (for mest mulig automatikk)

Vi bruker openpyxl i read_only streaming-mode og *alltid* begrenser rader/kolonner
for å unngå at "forurenset used-range" i Excel gjør at innlesing oppleves som heng.

For CSV leser vi en liten sample med Python sin csv-modul i stedet for pandas.
Det gjør at vi tåler "metadata-rader" med færre kolonner før selve header-raden
(typisk SAF-T/rapporter).

Denne modulen har ingen Tk-avhengighet og kan enhetstestes.
"""

from __future__ import annotations

from pathlib import Path

import csv

from column_names import make_safe_unique_column_names
from excel_sheet_guess import guess_best_excel_sheet

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
_CSV_DELIMS = [";", ",", "\t", "|"]


def is_excel_path(path: Path) -> bool:
    return path.suffix.lower() in EXCEL_SUFFIXES


def is_csv_path(path: Path) -> bool:
    return path.suffix.lower() == ".csv"


def list_excel_sheets(path: Path) -> list[str]:
    """Returner arknavn fra en Excel-fil (best effort)."""

    import openpyxl

    if not path.exists() or not path.is_file():
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _is_nonempty_cell(v: object) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() != ""
    return True


def read_excel_rows(
    path: Path,
    sheet_name: str | None,
    *,
    start_row: int = 1,
    nrows: int = 20,
    max_cols: int = 50,
    # Bakoverkompatibilitet: noen kall bruker max_rows i stedet for nrows
    max_rows: int | None = None,
) -> list[list[object]]:
    """Les et lite rektangel fra et Excel-ark.

    Leser *kun* (nrows x max_cols) fra gitt start_row.
    """

    import openpyxl

    if not path.exists() or not path.is_file():
        raise FileNotFoundError(str(path))

    if max_rows is not None:
        max_rows_i = max(1, int(max_rows))
        # Hvis begge er satt, må de samsvare for å unngå stille feil.
        if nrows != 20 and nrows != max_rows_i:
            raise ValueError("Både nrows og max_rows er angitt, men de er ulike")
        nrows = max_rows_i

    start_row = max(1, int(start_row))
    nrows = max(1, int(nrows))
    max_cols = max(1, int(max_cols))
    end_row = start_row + nrows - 1

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        rows: list[list[object]] = []
        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=1,
            max_col=max_cols,
            values_only=True,
        ):
            rows.append(list(row))

        # Trim trailing empty columns across the sample
        last_col = 0
        for r in rows:
            for idx, val in enumerate(r, start=1):
                if _is_nonempty_cell(val):
                    last_col = max(last_col, idx)
        if last_col > 0:
            rows = [r[:last_col] for r in rows]
        else:
            rows = [r[:0] for r in rows]

        return rows
    finally:
        try:
            wb.close()
        except Exception:
            pass


def read_excel_header(
    path: Path,
    sheet_name: str | None,
    *,
    header_row: int,
    max_cols: int = 500,
) -> list[str]:
    """Les kun header-raden (1-indeksert) fra et Excel-ark."""

    # NB: Vi leser også noen rader etter header for å kunne beholde
    # trailing blanke header-celler *hvis det finnes data i kolonnen*.
    rows = read_excel_rows(path, sheet_name, start_row=header_row, nrows=5, max_cols=max_cols)
    if not rows:
        return []

    # read_excel_rows trimmer trailing tomme kolonner basert på *hele* sample.
    # Det betyr at lengden på radene allerede representerer en best-effort
    # "faktisk" bredde.
    header_raw = list(rows[0])
    width = max((len(r) for r in rows), default=len(header_raw))
    if len(header_raw) < width:
        header_raw = header_raw + [None] * (width - len(header_raw))
    else:
        header_raw = header_raw[:width]

    return make_safe_unique_column_names(header_raw, placeholder_prefix="kol")


def _read_csv_sample_text(path: Path, *, max_chars: int = 16384) -> str:
    # Best-effort dekoding. Vi bryr oss mest om å få tak i separator og en sample.
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        return f.read(max_chars)


def _detect_csv_delimiter(sample: str) -> str:
    counts = {d: sample.count(d) for d in _CSV_DELIMS}
    # typisk norske filer: ';'
    return max(counts.items(), key=lambda kv: kv[1])[0] if counts else ";"


def _detect_csv_dialect(path: Path) -> csv.Dialect:
    sample = _read_csv_sample_text(path)
    sniffer = csv.Sniffer()
    try:
        return sniffer.sniff(sample, delimiters=_CSV_DELIMS)
    except Exception:
        delim = _detect_csv_delimiter(sample)

        class _FallbackDialect(csv.Dialect):
            delimiter = delim
            quotechar = '"'
            doublequote = True
            skipinitialspace = False
            lineterminator = "\n"
            quoting = csv.QUOTE_MINIMAL

        return _FallbackDialect()


def _trim_trailing_empty_columns(rows: list[list[str]]) -> list[list[str]]:
    last_col = 0
    for r in rows:
        for idx, val in enumerate(r, start=1):
            if str(val).strip() != "":
                last_col = max(last_col, idx)
    if last_col <= 0:
        return [r[:0] for r in rows]
    return [r[:last_col] for r in rows]


def read_csv_rows(
    path: Path,
    *,
    start_row: int = 1,
    nrows: int = 20,
    max_cols: int = 50,
    # Bakoverkompatibilitet: noen kall bruker max_rows i stedet for nrows
    max_rows: int | None = None,
) -> list[list[object]]:
    """Les de første nrows radene fra CSV uten header (best effort).

    - Bruker csv-modulen for å tåle varierende kolonneantall per rad.
    - Trimmer til max_cols.
    - Trimmer trailing tomme kolonner på tvers av sample.
    """

    if not path.exists() or not path.is_file():
        raise FileNotFoundError(str(path))

    if max_rows is not None:
        max_rows_i = max(1, int(max_rows))
        if nrows != 20 and nrows != max_rows_i:
            raise ValueError("Både nrows og max_rows er angitt, men de er ulike")
        nrows = max_rows_i

    start_row = max(1, int(start_row))
    nrows = max(1, int(nrows))
    max_cols = max(1, int(max_cols))

    dialect = _detect_csv_dialect(path)

    rows: list[list[str]] = []
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.reader(f, dialect)
        for idx, row in enumerate(reader, start=1):
            if idx < start_row:
                continue
            if len(rows) >= nrows:
                break
            r = ["" if v is None else str(v) for v in row]
            if len(r) > max_cols:
                r = r[:max_cols]
            rows.append(r)

    return _trim_trailing_empty_columns(rows)


def read_csv_header(path: Path, *, header_row: int) -> list[str]:
    """Les header fra CSV der header_row er 1-indeksert."""

    if not path.exists() or not path.is_file():
        raise FileNotFoundError(str(path))

    header_row = max(1, int(header_row))

    dialect = _detect_csv_dialect(path)

    # Vi leser header-raden + litt "lookahead" for å finne faktisk bredde.
    # Dette gjør at vi ikke mister en trailing blank header-kolonne hvis det
    # finnes data i kolonnen på de neste radene.

    header_values: list[str] | None = None
    tail_rows: list[list[str]] = []

    LOOKAHEAD = 25
    MAX_COLS = 500

    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.reader(f, dialect)
        for idx, r in enumerate(reader, start=1):
            if idx < header_row:
                continue

            row = ["" if v is None else str(v) for v in r]
            if len(row) > MAX_COLS:
                row = row[:MAX_COLS]

            if idx == header_row:
                header_values = [v.strip() for v in row]
            else:
                # data-rader: behold råstreng (vi strip'er ikke aggressivt)
                tail_rows.append(row)

            if idx >= header_row + LOOKAHEAD:
                break

    if header_values is None:
        return []

    # Finn siste kolonne som har innhold i header eller i lookahead.
    last_col = 0
    for r in [header_values] + tail_rows:
        for j, v in enumerate(r, start=1):
            if str(v).strip() != "":
                last_col = max(last_col, j)

    if last_col <= 0:
        return []

    if len(header_values) < last_col:
        header_values = header_values + [""] * (last_col - len(header_values))
    header_values = header_values[:last_col]

    return make_safe_unique_column_names(header_values, placeholder_prefix="kol")


def read_data_sample(
    path: Path,
    sheet_name: str | None,
    *,
    header_row: int,
    nrows: int = 60,
    expected_width: int | None = None,
    max_cols: int = 200,
) -> list[list[object]]:
    """Les en liten sample av data-rader *etter* header-raden.

    Brukes for mer intelligent mapping (innholdsbasert).

    Parametre
    ---------
    header_row:
        1-indeksert header-rad.
    expected_width:
        Forventet antall kolonner (typisk len(headers)). Hvis satt,
        padder/trunker vi alle rader til denne bredden.
    """

    header_row = max(1, int(header_row))
    nrows = max(1, int(nrows))
    max_cols = max(1, int(max_cols))

    if expected_width is not None:
        try:
            max_cols = max(max_cols, int(expected_width))
        except Exception:
            pass

    if is_excel_path(path):
        rows = read_excel_rows(path, sheet_name, start_row=header_row + 1, nrows=nrows, max_cols=max_cols)
    elif is_csv_path(path):
        # read_csv_rows kan ikke starte midt i filen; vi leser header_row+nrows
        # og tar ut data-radene.
        raw = read_csv_rows(path, nrows=header_row + nrows, max_cols=max_cols)
        rows = raw[header_row : header_row + nrows] if header_row <= len(raw) else []
    else:
        raise ValueError(f"Ukjent filtype: {path.suffix}")

    if expected_width is None:
        expected_width = max((len(r) for r in rows), default=0)

    width = max(0, int(expected_width or 0))
    if width <= 0:
        return []

    out: list[list[object]] = []
    for r in rows:
        rr = list(r)
        if len(rr) < width:
            # CSV gir strenger; Excel kan gi None -> bruk None som padding
            pad_val: object = "" if is_csv_path(path) else None
            rr = rr + [pad_val] * (width - len(rr))
        else:
            rr = rr[:width]
        out.append(rr)

    return out


def auto_detect_header_and_headers(
    path: Path,
    sheet_name: str | None,
    *,
    fallback_header_row: int = 1,
    scan_rows: int = 80,
    scan_max_cols: int = 120,
    header_max_cols: int = 500,
) -> tuple[int, list[str]]:
    """Best-effort: autodetekter header-rad og returner (header_row, headers).

    Brukes for å gjøre DatasetPane mer "automatisk": Ved filvalg/arkvalg kan vi
    lese en liten sample, gjette header-rad, og så lese header på den raden.

    Detektering kaster aldri (detect_header_row returnerer None ved usikkerhet).
    I så fall faller vi tilbake til fallback_header_row.
    """

    from header_detection import detect_header_row

    fallback_header_row = max(1, int(fallback_header_row))
    scan_rows = max(1, int(scan_rows))
    scan_max_cols = max(1, int(scan_max_cols))
    header_max_cols = max(1, int(header_max_cols))

    if is_excel_path(path):
        rows = read_excel_rows(path, sheet_name, start_row=1, nrows=scan_rows, max_cols=scan_max_cols)
        idx0 = detect_header_row(rows)
        header_row = int(idx0 + 1) if idx0 is not None else fallback_header_row
        headers = read_excel_header(path, sheet_name, header_row=header_row, max_cols=header_max_cols)
        return header_row, headers

    if is_csv_path(path):
        rows = read_csv_rows(path, nrows=scan_rows, max_cols=scan_max_cols)
        idx0 = detect_header_row(rows)
        header_row = int(idx0 + 1) if idx0 is not None else fallback_header_row
        headers = read_csv_header(path, header_row=header_row)
        return header_row, headers

    raise ValueError(f"Ukjent filtype: {path.suffix}")
