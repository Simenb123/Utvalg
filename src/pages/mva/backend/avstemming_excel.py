"""Excel-eksport for MVA-avstemming.

Genererer en formatert Excel-fil med:
- Ark 1: MVA-avstemming (termin-sammenligning HB vs Skatteetaten)
- Ark 2: MVA per kode (detaljert pivot med avgift OG grunnlag per termin)
- Ark 3: MVA-kontroller (K1 salg vs grunnlag, K2/K3 detaljer)
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Optional

import pandas as pd

log = logging.getLogger(__name__)


def write_mva_avstemming_excel(
    path: str | Path,
    *,
    mva_pivot: pd.DataFrame,
    reconciliation: Optional[pd.DataFrame] = None,
    kontroller: Any = None,
    client: str = "",
    year: str | int = "",
    skatteetaten: Any = None,
) -> Path:
    """Skriv MVA-avstemming til Excel.

    Args:
        path: Utfil.
        mva_pivot: Detaljert MVA-pivot (kode × termin, med G_T*-kolonner).
        reconciliation: Avstemmings-DataFrame (fra build_reconciliation).
        kontroller: MvaKontrollResult (fra build_mva_kontroller).
        client: Klientnavn.
        year: Regnskapsår.
        skatteetaten: SkatteetatenData (for header-info).

    Returns:
        Path til lagret fil.
    """
    out = Path(path)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # Ark 1: Avstemming
        if reconciliation is not None and not reconciliation.empty:
            reconciliation.to_excel(writer, sheet_name="MVA-avstemming", index=False)

        # Ark 2: MVA per kode (avgift + grunnlag)
        if mva_pivot is not None and not mva_pivot.empty:
            export_cols = ["MVA-kode", "Beskrivelse"]
            for t in range(1, 7):
                col = f"T{t}"
                if col in mva_pivot.columns:
                    export_cols.append(col)
            if "Sum" in mva_pivot.columns:
                export_cols.append("Sum")
            # Grunnlag-kolonner
            for t in range(1, 7):
                col = f"G_T{t}"
                if col in mva_pivot.columns:
                    export_cols.append(col)
            if "G_Sum" in mva_pivot.columns:
                export_cols.append("G_Sum")

            pivot_export = mva_pivot[[c for c in export_cols if c in mva_pivot.columns]].copy()

            # Rename grunnlag-kolonner for lesbarhet
            rename_map = {}
            for t in range(1, 7):
                rename_map[f"G_T{t}"] = f"Grunnlag T{t}"
            rename_map["G_Sum"] = "Grunnlag Sum"
            pivot_export.rename(columns=rename_map, inplace=True)

            pivot_export.to_excel(writer, sheet_name="MVA per kode", index=False)

        # Ark 3: Kontroller
        if kontroller is not None:
            _write_kontroller(writer, kontroller)

    # Formatering med openpyxl
    try:
        _polish_workbook(out, client=client, year=year, skatteetaten=skatteetaten)
    except Exception:
        log.exception("Formatering av Excel feilet (filen er likevel lagret)")

    return out


def _write_kontroller(writer: pd.ExcelWriter, kontroller: Any) -> None:
    """Skriv kontroller til Excel-ark."""
    # K-oppsummering
    if kontroller.summary:
        df_summary = pd.DataFrame(kontroller.summary)
        df_summary.to_excel(writer, sheet_name="Kontroller", index=False, startrow=0)

    # K1: Salg vs grunnlag
    if not kontroller.salg_vs_grunnlag.empty:
        start_row = len(kontroller.summary) + 3 if kontroller.summary else 0
        kontroller.salg_vs_grunnlag.to_excel(
            writer, sheet_name="Kontroller", index=False, startrow=start_row
        )

    # K2: Salgskontoer uten MVA (eget ark om mange)
    if not kontroller.salg_uten_mva.empty:
        kontroller.salg_uten_mva.head(500).to_excel(
            writer, sheet_name="K2 Salg uten MVA", index=False
        )

    # K3: Andre kontoer med utg. MVA (eget ark om mange)
    if not kontroller.andre_med_utg_mva.empty:
        kontroller.andre_med_utg_mva.head(500).to_excel(
            writer, sheet_name="K3 Andre med utg MVA", index=False
        )


def _polish_workbook(
    path: Path,
    *,
    client: str = "",
    year: str | int = "",
    skatteetaten: Any = None,
) -> None:
    """Formater Excel-filen med openpyxl."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, numbers
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)

    for ws in wb.worksheets:
        # Auto-filter
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions

        # Formater tallkolonner
        for col_idx in range(1, (ws.max_column or 0) + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header and header not in ("Termin", "MVA-kode", "Beskrivelse", "Kontroll", "Status", "Kommentar"):
                for row_idx in range(2, (ws.max_row or 0) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

        # Header-format
        header_fill = PatternFill(start_color="2C4A6E", end_color="2C4A6E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, (ws.max_column or 0) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Kolonnebredder
        for col_idx in range(1, (ws.max_column or 0) + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, min((ws.max_row or 0) + 1, 50)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)

        # Marker Sum-rad (siste rad) med bold
        if ws.max_row and ws.max_row > 1:
            last_row = ws.max_row
            for col_idx in range(1, (ws.max_column or 0) + 1):
                cell = ws.cell(row=last_row, column=col_idx)
                cell.font = Font(bold=True)

        # Marker avvik i avstemming-arket
        if ws.title == "MVA-avstemming":
            _highlight_diff_column(ws)

        # Marker avvik/merk i kontroller-arket
        if ws.title == "Kontroller":
            _highlight_status_column(ws)

    wb.save(path)


def _highlight_diff_column(ws) -> None:
    """Marker Differanse-kolonne med rødt/grønt."""
    from openpyxl.styles import Font

    diff_col = None
    for col_idx in range(1, (ws.max_column or 0) + 1):
        if ws.cell(row=1, column=col_idx).value == "Differanse":
            diff_col = col_idx
            break

    if not diff_col:
        return

    red_font = Font(color="C0392B", bold=True)
    green_font = Font(color="27AE60")
    for row_idx in range(2, (ws.max_row or 0) + 1):
        cell = ws.cell(row=row_idx, column=diff_col)
        val = cell.value
        if isinstance(val, (int, float)):
            if abs(val) > 1.0:
                cell.font = red_font
            else:
                cell.font = green_font


def _highlight_status_column(ws) -> None:
    """Marker Status-kolonne med farger."""
    from openpyxl.styles import Font

    status_col = None
    for col_idx in range(1, (ws.max_column or 0) + 1):
        if ws.cell(row=1, column=col_idx).value == "Status":
            status_col = col_idx
            break

    if not status_col:
        return

    for row_idx in range(2, (ws.max_row or 0) + 1):
        cell = ws.cell(row=row_idx, column=status_col)
        val = str(cell.value or "").strip()
        if val == "AVVIK":
            cell.font = Font(color="C0392B", bold=True)
        elif val == "MERK":
            cell.font = Font(color="E67E22", bold=True)
        elif val == "OK":
            cell.font = Font(color="27AE60", bold=True)
