"""Eksporterer scoping-resultat til en formatert Excel-fil.

Struktur:
- **Oversikt**: nøkkeltall, aggregeringskontroll, fordeling
- **Alle**: alle regnskapslinjer med klassifisering og beslutning
- **Inn**: kun linjer scopet inn (revisjonshandlinger trengs)
- **Ut**: kun linjer scopet ut (med begrunnelse)

Hver fane har samme kolonneoppsett som GUI-visningen i Scoping-fanen.
"""

from __future__ import annotations

import logging
from pathlib import Path
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .engine import ScopingResult

log = logging.getLogger(__name__)

# Farger — matcher GUI-fargene i Scoping-fanen
_FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
_FILL_VESENTLIG = PatternFill("solid", fgColor="FDE8E8")
_FILL_MODERAT = PatternFill("solid", fgColor="FEF3C7")
_FILL_IKKE = PatternFill("solid", fgColor="D1FAE5")
_FILL_MANUELL = PatternFill("solid", fgColor="DBEAFE")
_FILL_SUMMARY = PatternFill("solid", fgColor="F0F0F0")
_FILL_OK = PatternFill("solid", fgColor="D1FAE5")
_FILL_WARN = PatternFill("solid", fgColor="FDE8E8")
_FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
_FILL_INFO = PatternFill("solid", fgColor="EDF3FA")

_FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
_FONT_BOLD = Font(bold=True, size=10)
_FONT_NORMAL = Font(size=10)
_FONT_SUMMARY = Font(size=10, color="999999")
_FONT_TITLE = Font(bold=True, size=14, color="FFFFFF")
_FONT_SECTION = Font(bold=True, size=11, color="1F4E79")
_FONT_BIG_NUMBER = Font(bold=True, size=14, color="1F4E79")

_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

_FILL_MAP = {
    "vesentlig": _FILL_VESENTLIG,
    "moderat": _FILL_MODERAT,
    "ikke_vesentlig": _FILL_IKKE,
    "manuell": _FILL_MANUELL,
}


def _fmt_amount(value: float) -> str:
    if abs(value) < 0.5:
        return "0"
    return f"{value:,.0f}".replace(",", " ")


def _class_label(c: str) -> str:
    return {
        "vesentlig": "Vesentlig",
        "moderat": "Moderat",
        "ikke_vesentlig": "Ikke vesentlig",
        "manuell": "Manuell",
    }.get(c, "")


def _scope_label(s: str) -> str:
    return {"inn": "Inn", "ut": "Ut"}.get(s, "")


def _build_headers(year_int: int | None, has_prior: bool) -> list[tuple[str, int]]:
    """Returner kolonneoverskrifter + bredder, matcher GUI-rekkefølge."""
    ub_label = f"UB {year_int}" if year_int is not None else "UB"
    ub_fjor_label = f"UB {year_int - 1}" if year_int is not None else "UB i fjor"
    headers: list[tuple[str, int]] = [
        ("Regnr", 8),
        ("Regnskapslinje", 30),
        ("Type", 6),
        (ub_label, 15),
    ]
    if has_prior:
        headers.extend(
            [
                (ub_fjor_label, 15),
                ("Endring", 15),
                ("Endring %", 10),
            ]
        )
    headers.extend(
        [
            ("% av PM", 10),
            ("Klassifisering", 16),
            ("Scoping", 10),
            ("Revisjonshandling", 30),
            ("Begrunnelse", 35),
            ("Handl.", 8),
        ]
    )
    return headers


def _write_header_row(ws, row: int, headers: list[tuple[str, int]]) -> int:
    """Skriv kolonneoverskrifter og sett bredder. Returnerer neste rad."""
    for col_idx, (hdr, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    return row + 1


def _write_line_row(
    ws,
    row: int,
    line: Any,
    *,
    has_prior: bool,
    amount_cols: set[int],
    pct_cols: set[int],
) -> int:
    """Skriv én linje (regnskapslinje) i tabellen. Returnerer neste rad."""
    if line.is_summary:
        fill = _FILL_SUMMARY
        font = _FONT_SUMMARY
    else:
        fill = _FILL_MAP.get(line.classification, PatternFill())
        font = _FONT_NORMAL

    values: list[object] = [
        int(line.regnr) if line.regnr.isdigit() else line.regnr,
        line.regnskapslinje,
        line.line_type,
        line.amount,
    ]
    if has_prior:
        values.extend(
            [
                line.amount_prior,
                line.change_amount,
                f"{line.change_pct:+.1f}%"
                if line.change_pct is not None and not line.is_summary
                else "",
            ]
        )
    values.extend(
        [
            "" if line.is_summary else f"{line.pct_of_pm:.0f}%",
            _class_label(line.classification),
            _scope_label(line.scoping),
            line.audit_action if not line.is_summary else "",
            line.rationale,
            line.action_count if line.action_count else "",
        ]
    )

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = font
        cell.fill = fill
        cell.border = _THIN_BORDER
        if col_idx in amount_cols:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx in pct_cols:
            cell.alignment = Alignment(horizontal="right")
    return row + 1


def _write_lines_sheet(
    wb: openpyxl.Workbook,
    sheet_title: str,
    lines: list,
    *,
    year_int: int | None,
    has_prior: bool,
    description: str = "",
) -> None:
    """Bygg én fane med tabell over et utvalg regnskapslinjer."""
    ws = wb.create_sheet(title=sheet_title)

    # Tittel
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = sheet_title.upper()
    c.font = _FONT_TITLE
    c.fill = _FILL_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    row = 2
    if description:
        ws.cell(row=row, column=1, value=description).font = _FONT_NORMAL
        row += 1
    ws.cell(row=row, column=1, value=f"Antall linjer: {len(lines)}").font = _FONT_BOLD
    row += 2

    headers = _build_headers(year_int, has_prior)
    row = _write_header_row(ws, row, headers)

    amount_cols = {4}
    if has_prior:
        amount_cols.update({5, 6})
    pct_cols = {7} if has_prior else {5}

    for line in lines:
        row = _write_line_row(
            ws, row, line,
            has_prior=has_prior,
            amount_cols=amount_cols,
            pct_cols=pct_cols,
        )

    # Frys topptekst
    ws.freeze_panes = ws[f"A{4 + (1 if description else 0)}"]


def _write_oversikt_sheet(
    wb: openpyxl.Workbook,
    result: ScopingResult,
    *,
    client_name: str,
    year: str,
) -> None:
    """Bygg Oversikt-fanen — nøkkeltall + aggregeringskontroll."""
    ws = wb.active
    ws.title = "Oversikt"

    # Tittel
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "SCOPING — OVERSIKT"
    c.font = _FONT_TITLE
    c.fill = _FILL_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    row = 3

    # Klient-info
    ws.cell(row=row, column=1, value="Klient").font = _FONT_SECTION
    row += 1
    info = [
        ("Navn:", client_name or "—"),
        ("År:", year or "—"),
        ("Generert:", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, val in info:
        ws.cell(row=row, column=1, value=label).font = _FONT_BOLD
        ws.cell(row=row, column=2, value=val).font = _FONT_NORMAL
        row += 1

    row += 1

    # Vesentlighetsterskler
    ws.cell(row=row, column=1, value="Vesentlighetsterskler").font = _FONT_SECTION
    row += 1
    thresholds = [
        ("Overall materiality (OM):", _fmt_amount(result.om)),
        ("Performance materiality (PM):", _fmt_amount(result.pm)),
        ("Clearly trivial (SUM):", _fmt_amount(result.sum_threshold)),
    ]
    for label, val in thresholds:
        ws.cell(row=row, column=1, value=label).font = _FONT_BOLD
        ws.cell(row=row, column=2, value=val).font = _FONT_NORMAL
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        row += 1

    row += 1

    # Aggregeringsrisiko
    non_sum = [l for l in result.lines if not l.is_summary]
    scoped_out = sum(abs(l.amount) for l in non_sum if l.scoping == "ut")
    ok = scoped_out < result.om if result.om > 0 else True
    total_saldo = sum(abs(l.amount) for l in non_sum)
    pct_om = round(scoped_out / result.om * 100, 1) if result.om > 0 else 0
    pct_total = round(scoped_out / total_saldo * 100, 1) if total_saldo > 0 else 0

    ws.cell(row=row, column=1, value="Aggregeringskontroll").font = _FONT_SECTION
    row += 1
    agg_rows = [
        ("Sum scopet ut (abs.):", _fmt_amount(scoped_out)),
        ("Sum scopet ut som % av OM:", f"{pct_om}%"),
        ("Sum scopet ut som % av total saldo:", f"{pct_total}%"),
    ]
    for label, val in agg_rows:
        ws.cell(row=row, column=1, value=label).font = _FONT_NORMAL
        ws.cell(row=row, column=2, value=val).font = _FONT_BOLD
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        row += 1
    row += 1
    cell = ws.cell(row=row, column=1, value="Vurdering:")
    cell.font = _FONT_BOLD
    verdict = ws.cell(row=row, column=2, value="OK — under OM" if ok else "ADVARSEL — overskrider OM!")
    verdict.font = _FONT_BOLD
    verdict.fill = _FILL_OK if ok else _FILL_WARN
    row += 2

    # Klassifiserings-fordeling
    ws.cell(row=row, column=1, value="Fordeling — klassifisering").font = _FONT_SECTION
    row += 1
    vesentlige = sum(1 for l in non_sum if l.classification == "vesentlig")
    moderate = sum(1 for l in non_sum if l.classification == "moderat")
    ikke_ves = sum(1 for l in non_sum if l.classification == "ikke_vesentlig")
    manuelle = sum(1 for l in non_sum if l.classification == "manuell")

    klass_rows = [
        ("Vesentlige:", vesentlige, _FILL_VESENTLIG),
        ("Moderate:", moderate, _FILL_MODERAT),
        ("Ikke vesentlige:", ikke_ves, _FILL_IKKE),
        ("Manuelle:", manuelle, _FILL_MANUELL),
        ("Totalt linjer:", len(non_sum), _FILL_INFO),
    ]
    for label, val, fill in klass_rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = _FONT_NORMAL
        c1.fill = fill
        c2 = ws.cell(row=row, column=2, value=val)
        c2.font = _FONT_BOLD
        c2.fill = fill
        c2.alignment = Alignment(horizontal="right")
        row += 1
    row += 1

    # Scoping-fordeling
    ws.cell(row=row, column=1, value="Fordeling — scoping").font = _FONT_SECTION
    row += 1
    scoped_inn = sum(1 for l in non_sum if l.scoping == "inn")
    scoped_ut_n = sum(1 for l in non_sum if l.scoping == "ut")
    ikke_besluttet = sum(1 for l in non_sum if not l.scoping)

    scope_rows = [
        ("Scopet inn:", scoped_inn),
        ("Scopet ut:", scoped_ut_n),
        ("Ikke besluttet:", ikke_besluttet),
    ]
    for label, val in scope_rows:
        ws.cell(row=row, column=1, value=label).font = _FONT_NORMAL
        c2 = ws.cell(row=row, column=2, value=val)
        c2.font = _FONT_BOLD
        c2.alignment = Alignment(horizontal="right")
        row += 1
    row += 1

    # Beløps-fordeling
    ws.cell(row=row, column=1, value="Beløp — scoping").font = _FONT_SECTION
    row += 1
    sum_inn = sum(abs(l.amount) for l in non_sum if l.scoping == "inn")
    sum_ut = sum(abs(l.amount) for l in non_sum if l.scoping == "ut")
    sum_uavklart = sum(abs(l.amount) for l in non_sum if not l.scoping)

    sum_rows = [
        ("Sum scopet inn (abs.):", _fmt_amount(sum_inn)),
        ("Sum scopet ut (abs.):", _fmt_amount(sum_ut)),
        ("Sum ikke besluttet (abs.):", _fmt_amount(sum_uavklart)),
        ("Sum total (abs.):", _fmt_amount(total_saldo)),
    ]
    for label, val in sum_rows:
        ws.cell(row=row, column=1, value=label).font = _FONT_NORMAL
        c2 = ws.cell(row=row, column=2, value=val)
        c2.font = _FONT_BOLD
        c2.alignment = Alignment(horizontal="right")
        row += 1

    # Kolonnebredder
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22


def export_scoping(
    result: ScopingResult,
    path: str | Path,
    *,
    client_name: str = "",
    year: str = "",
) -> Path:
    """Eksporter scoping-resultat til Excel-arbeidspapir.

    Genererer 4 faner:
      - Oversikt: nøkkeltall, aggregeringskontroll, fordeling
      - Alle: alle regnskapslinjer (matcher GUI-visning)
      - Inn: kun scopet-inn-linjer
      - Ut: kun scopet-ut-linjer
    """
    path = Path(path)
    wb = openpyxl.Workbook()

    try:
        year_int = int(year) if year else None
    except (TypeError, ValueError):
        year_int = None
    has_prior = any(line.amount_prior is not None for line in result.lines)

    # Fane 1: Oversikt (bruker default workbook.active)
    _write_oversikt_sheet(wb, result, client_name=client_name, year=year)

    # Fane 2: Alle linjer (inkl. summary-linjer for visuell hierarki)
    _write_lines_sheet(
        wb, "Alle", result.lines,
        year_int=year_int, has_prior=has_prior,
        description="Alle regnskapslinjer med klassifisering og beslutning. "
                    "Sumposter vises i grå.",
    )

    # Fane 3: Scopet inn (kun ikke-sum-linjer)
    inn_lines = [l for l in result.lines if not l.is_summary and l.scoping == "inn"]
    _write_lines_sheet(
        wb, "Inn", inn_lines,
        year_int=year_int, has_prior=has_prior,
        description="Linjer scopet INN — disse trenger revisjonshandlinger.",
    )

    # Fane 4: Scopet ut (kun ikke-sum-linjer)
    ut_lines = [l for l in result.lines if not l.is_summary and l.scoping == "ut"]
    _write_lines_sheet(
        wb, "Ut", ut_lines,
        year_int=year_int, has_prior=has_prior,
        description="Linjer scopet UT — med begrunnelse for utelatelse.",
    )

    wb.save(str(path))
    return path
