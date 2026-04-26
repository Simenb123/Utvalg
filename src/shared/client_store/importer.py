from __future__ import annotations

import re
from dataclasses import dataclass
import threading
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

import pandas as pd

from . import store as client_store

_PROGRESS_CB = Callable[[int, int, str], None]


def _clean_client_no(val: Any) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    digits = "".join(ch for ch in s if ch.isdigit())
    return digits or None


def _read_client_names_xlsx(file_path: Path) -> List[str]:
    """Les klientnavn fra Excel (XLSX) raskt.

    Vi bruker openpyxl i read_only-modus for å unngå at store/format-tunge
    arbeidsbøker blir trege ved import.

    Støtter Visena-format (Firma + Knr/Klientnr) og lager "<Knr> <Firma>".
    """

    try:
        from openpyxl import load_workbook
    except Exception:
        # Fallback til pandas dersom openpyxl ikke er tilgjengelig
        df = pd.read_excel(file_path)
        if df is None or df.empty:
            return []
        return [str(v).strip() for v in df[df.columns[0]].dropna().tolist() if str(v).strip()]

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)

        # Finn header-rad (typisk første ikke-tomme rad)
        header = None
        for _ in range(0, 20):
            header = next(rows, None)
            if header is None:
                return []
            # minst én verdi -> header
            if any(v is not None and str(v).strip() for v in header):
                break

        if header is None:
            return []

        headers = [str(h).strip() if h is not None else "" for h in header]
        headers_l = [h.lower() for h in headers]

        # Prøv Visena-format: Firma + Knr
        col_firma_idx = None
        col_knr_idx = None
        for i, h in enumerate(headers_l):
            if h == "firma":
                col_firma_idx = i
            if h in {"knr", "klientnr", "klientnr."}:
                col_knr_idx = i

        # Fallback: første kolonne som heter klient/kunde
        col_client_idx = None
        if col_firma_idx is None:
            for i, h in enumerate(headers_l):
                if h in {"klient", "client", "kundenavn", "kunde", "firma"}:
                    col_client_idx = i
                    break

        out: List[str] = []
        for row in rows:
            if row is None:
                continue

            if col_firma_idx is not None and col_knr_idx is not None:
                name = str(row[col_firma_idx]).strip() if col_firma_idx < len(row) and row[col_firma_idx] is not None else ""
                knr = _clean_client_no(row[col_knr_idx]) if col_knr_idx < len(row) else None
                if not name:
                    continue
                out.append(f"{knr} {name}" if knr else name)
                continue

            idx = col_client_idx if col_client_idx is not None else 0
            if idx < len(row) and row[idx] is not None:
                s = str(row[idx]).strip()
                if s:
                    out.append(s)

        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _read_client_names_csv(file_path: Path) -> List[str]:
    txt = file_path.read_text(encoding="utf-8", errors="ignore")
    lines = [ln.strip() for ln in txt.splitlines() if ln.strip()]
    if not lines:
        return []
    # Prøv å detecte header med klientkolonne
    header = [h.strip() for h in re.split(r"[;,\t]", lines[0])]
    # hvis header ser ut som en enkel liste (ingen separator)
    if len(header) == 1:
        # typisk 1-kolonne CSV med header: "Klient"
        if header[0].lower() in {"klient", "client", "kundenavn", "kunde", "firma"}:
            return lines[1:]
        return lines
    # finn kolonne som heter klient
    idx = None
    for i, h in enumerate(header):
        if h.lower() in {"klient", "client", "kundenavn", "kunde", "firma"}:
            idx = i
            break
    if idx is None:
        # fallback: første kolonne
        idx = 0
    out: List[str] = []
    for ln in lines[1:]:
        parts = [p.strip() for p in re.split(r"[;,\t]", ln)]
        if idx < len(parts) and parts[idx]:
            out.append(parts[idx])
    return out


def _read_client_names_txt(file_path: Path) -> List[str]:
    return [ln.strip() for ln in file_path.read_text(encoding="utf-8", errors="ignore").splitlines() if ln.strip()]


def read_client_names_from_file(file_path: Path) -> List[str]:
    """Leser en klientliste fra fil og returnerer unike display-strenger.

    - XLSX: støtter Visena-format (Firma + Knr) og genererer "<Knr> <Firma>".
    - CSV/TXT: leser linjer/kolonne.

    Merk: Denne funksjonen dedupliserer på *display-strengen* (case/whitespace),
    ikke på klientnummer. Importplanen tar derimot hensyn til klientnummer hvis det finnes.
    """

    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(file_path)

    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        raw = _read_client_names_xlsx(file_path)
    elif suffix in {".csv"}:
        raw = _read_client_names_csv(file_path)
    else:
        raw = _read_client_names_txt(file_path)

    seen: set[str] = set()
    out: List[str] = []
    for v in raw:
        s = (v or "").strip()
        if not s:
            continue
        key = " ".join(s.lower().split())
        if key in seen:
            continue
        seen.add(key)
        out.append(s)
    return out


_CLIENT_NO_PREFIX_RE = re.compile(r"^\s*(\d{2,12})\s+")


def _extract_client_no(display_name: str) -> Optional[str]:
    """Prøv å hente klientnummer fra en display-streng som starter med tall."""

    m = _CLIENT_NO_PREFIX_RE.match(display_name or "")
    if not m:
        return None
    return m.group(1)


def _norm_name(s: str) -> str:
    return " ".join((s or "").strip().lower().split())


def _key_for_display(display_name: str) -> str:
    no = _extract_client_no(display_name)
    if no:
        return f"no:{no}"
    return f"name:{_norm_name(display_name)}"


@dataclass(frozen=True)
class ImportPlan:
    """Plan for import uten å gjøre endringer på disk."""

    found: int
    # Antall unike nøkler (klientnr hvis finnes, ellers normalisert navn)
    unique_keys: int
    new_clients: List[str]
    existing_clients: List[str]
    # (eksisterende_display, ny_display_fra_fil)
    rename_candidates: List[Tuple[str, str]]
    # display-strenger fra fil som kolliderer på key (samme klientnr/navn)
    duplicates_in_file: List[str]


def plan_import_clients(file_path: Path, *, existing_clients: Optional[Sequence[str]] = None) -> ImportPlan:
    """Lager en importplan ved å sammenligne filen mot eksisterende klienter.

    Matching-regler:
    - Hvis display-strengen starter med klientnummer ("1234 ..."), brukes klientnr som primærnøkkel.
      Da unngår vi duplikater hvis navn endrer seg, men klientnr er likt.
    - Ellers brukes normalisert navn som nøkkel.
    """

    file_names = read_client_names_from_file(file_path)
    found = len(file_names)

    if existing_clients is None:
        existing_clients = client_store.list_clients()

    existing_by_key: Dict[str, str] = {}
    for disp in existing_clients:
        key = _key_for_display(disp)
        # Hvis det finnes duplikater i eksisterende data, behold første (trygg).
        existing_by_key.setdefault(key, disp)

    file_by_key: Dict[str, str] = {}
    duplicates_in_file: List[str] = []
    for disp in file_names:
        key = _key_for_display(disp)
        if key in file_by_key and _norm_name(file_by_key[key]) != _norm_name(disp):
            duplicates_in_file.append(disp)
            continue
        file_by_key.setdefault(key, disp)

    new_clients: List[str] = []
    existing_hits: List[str] = []
    rename_candidates: List[Tuple[str, str]] = []
    for key, disp in file_by_key.items():
        if key not in existing_by_key:
            new_clients.append(disp)
            continue

        existing_hits.append(disp)
        existing_disp = existing_by_key[key]
        if key.startswith("no:") and _norm_name(existing_disp) != _norm_name(disp):
            rename_candidates.append((existing_disp, disp))

    return ImportPlan(
        found=found,
        unique_keys=len(file_by_key),
        new_clients=new_clients,
        existing_clients=existing_hits,
        rename_candidates=rename_candidates,
        duplicates_in_file=duplicates_in_file,
    )


def import_clients_from_file(
    file_path: Path,
    *,
    progress_cb: Optional[_PROGRESS_CB] = None,
    update_names: bool = False,
    plan: Optional[ImportPlan] = None,
    cancel_event: Optional[threading.Event] = None,
) -> Dict[str, Any]:
    """Importer klienter fra fil.

    Standard (trygt):
    - Oppretter kun nye klienter (ingen sletting).
    - Ved re-import brukes klientnummer (hvis tilgjengelig) for å unngå duplikater.

    Valgfritt:
    - update_names=True oppdaterer display-navn for eksisterende klienter som matcher på klientnr.
      Dette flytter ikke mapper og sletter ikke data – det oppdaterer kun meta.json.
    """

    if plan is None:
        plan = plan_import_clients(file_path)

    created = 0
    renamed = 0
    cancelled = False

    tasks_total = len(plan.new_clients) + (len(plan.rename_candidates) if update_names else 0)
    tasks_done = 0

    def _is_cancelled() -> bool:
        try:
            return bool(cancel_event and cancel_event.is_set())
        except Exception:
            return False

    # Opprett nye
    for disp in plan.new_clients:
        if _is_cancelled():
            cancelled = True
            break

        client_store.ensure_client(disp, persist_index=False)
        created += 1
        tasks_done += 1
        if progress_cb:
            progress_cb(tasks_done, tasks_total, disp)

    # Oppdater navn (ikke destruktivt)
    if update_names and not cancelled:
        for old_disp, new_disp in plan.rename_candidates:
            if _is_cancelled():
                cancelled = True
                break

            if client_store.update_client_display_name(old_disp, new_disp):
                renamed += 1
            tasks_done += 1
            if progress_cb:
                progress_cb(tasks_done, tasks_total, f"(oppdaterer navn) {old_disp} -> {new_disp}")

    # Oppdater hurtigindeks for raskere oppstart (best effort)
    try:
        client_store.persist_clients_index()
    except Exception:
        pass

    return {
        "found": plan.found,
        "unique_keys": plan.unique_keys,
        "created": created,
        "renamed": renamed,
        "existing": len(plan.existing_clients),
        "skipped_existing": len(plan.existing_clients),
        "duplicates_in_file": len(plan.duplicates_in_file),
        "cancelled": cancelled,
    }
