"""Arbeidspapir: Klientinfo, roller & eierskap.

Samler BRREG-data (enhet + roller) og aksjonærregister-data (fra
`ar_store.get_client_ownership_overview`) i én Excel-arbeidsbok med fem ark:

1. Oversikt      — org.info, status, MVA, næring
2. Roller        — DL, styreleder, styremedlemmer
3. Aksjonærer    — eierskap i klienten (fra ownership.db)
4. Kryssreferanse — aksjonærer som også har rolle (matching på navn + år)
5. Eide selskaper — selskaper klienten eier andeler i

Builder-funksjonen tar ferdig-hentede data (ingen nettverkskall her) slik
at modulen er enkel å teste og kan kjøres både on-line og fra cache.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Callable, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.pages.ar.backend.ownership_chain import walk_indirect_chain
from .forside import build_forside_sheet


# ---------------------------------------------------------------------------
# Styling

_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
_HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_BAD_FILL = PatternFill("solid", fgColor="FCE4EC")
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_PCT_FMT = '0.00"%"'


# ---------------------------------------------------------------------------
# Navne-match

_WS_RE = re.compile(r"\s+")
_PUNCT_RE = re.compile(r"[.,\-'`]")


def normalize_person_name(name: object) -> str:
    """Normaliserer et personnavn for robust matching.

    - Unicode-strip (NFKD) + case fold
    - Fjerner tegnsetting, kollapser mellomrom
    - Sorterer navne-delene så "Hansen, Ola" matcher "Ola Hansen"
    """
    if not name:
        return ""
    text = unicodedata.normalize("NFKD", str(name))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.casefold()
    text = _PUNCT_RE.sub(" ", text)
    text = _WS_RE.sub(" ", text).strip()
    if not text:
        return ""
    parts = sorted(p for p in text.split(" ") if p)
    return " ".join(parts)


def _birth_year(raw: object) -> str:
    """Henter fire-sifret år fra en dato-streng. Tom hvis ikke mulig."""
    if not raw:
        return ""
    s = str(raw).strip()
    m = re.search(r"(19|20)\d{2}", s)
    return m.group(0) if m else ""


@dataclass
class CrossMatch:
    shareholder_name: str
    shareholder_orgnr: str
    shareholder_kind: str  # "person" / "enhet" / ""
    ownership_pct: float  # direkte for direct, effektiv (sub×holding/100) for indirect
    roles: list[str]
    match_confidence: str  # "Høy" (navn + fødselsår) / "Mulig" (kun navn)
    notat: str = ""
    match_type: str = "direct"  # "direct" eller "indirect"
    via: str = ""  # holdingselskap-navn for indirekte match


def owner_birth_year(owner: dict[str, Any]) -> str:
    """Hent fødselsår fra aksjonærrad hvis eieren er en person.

    Skatteetatens CSV legger "Fødselsår/orgnr" i ett felt. For personer er
    dette enten 4-sifret år eller 11-sifret fødselsnummer (6 første sifre
    er DDMMYY). For selskaper er det et 9-sifret orgnr.
    """
    kind = str(owner.get("shareholder_kind") or "").lower()
    if kind not in {"person", "unknown"}:
        return ""
    raw = str(owner.get("shareholder_orgnr") or "")
    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) == 4:
        return digits
    if len(digits) == 11:
        yy = digits[4:6]
        try:
            year_int = int(yy)
        except ValueError:
            return ""
        # Fødselsnummer-konvensjonen: individnummer 000-499 ⇒ 1900-tallet,
        # 500-999 ⇒ 2000-tallet (med noen nyanser for 1854-1899 som er sjeldne).
        individ = digits[6:9]
        try:
            ind_int = int(individ)
        except ValueError:
            ind_int = 0
        century = 1900 if ind_int < 500 else 2000
        return str(century + year_int)
    return ""


def _match_role_confidence(
    role_entries: list[dict[str, str]], owner_year: str
) -> tuple[str, str]:
    """Return (confidence, warning-notat) for a navnematch mot rolleinnehavere.

    Konfidensverdier: "Bekreftet" (fødselsår matcher) / "Navn-match" (kun navn).
    """
    role_years = {_birth_year(r.get("fodselsdato")) for r in role_entries}
    role_years.discard("")
    if owner_year and owner_year in role_years:
        return "Bekreftet", ""
    if owner_year and role_years and owner_year not in role_years:
        notat = (
            f"Obs: fødselsår {owner_year} hos aksjonær matcher ikke "
            f"rolleinnehavers {sorted(role_years)[0]} — kan være ulike personer "
            "med samme navn."
        )
        return "Navn-match", notat
    return "Navn-match", ""


def _fmt_pct(value: float) -> str:
    """Norsk tallformat: 33,33 %."""
    return f"{value:.2f}".replace(".", ",") + " %"


def _fmt_roles_natural(roles: list[str]) -> str:
    """Formatér en rolleliste til naturlig norsk tekst.

    ["Daglig leder"]                                 → "Daglig leder"
    ["Daglig leder", "Styrets leder"]                → "Daglig leder og styrets leder"
    ["Daglig leder", "Styrets leder", "Styremedlem"] → "Daglig leder, styrets leder og styremedlem"
    """
    cleaned = [r.strip() for r in roles if r and r.strip()]
    if not cleaned:
        return "Rolleinnehaver"
    if len(cleaned) == 1:
        return cleaned[0]
    head = cleaned[0]
    tail = [r[0].lower() + r[1:] if r else r for r in cleaned[1:]]
    if len(tail) == 1:
        return f"{head} og {tail[0]}"
    return f"{head}, " + ", ".join(tail[:-1]) + f" og {tail[-1]}"


def build_cross_matches(
    owners: Iterable[dict[str, Any]],
    roller: Iterable[dict[str, str]],
    *,
    indirect_owners_fn: Optional[Callable[[str], list[dict[str, Any]]]] = None,
    max_indirect_depth: int = 1,
) -> list[CrossMatch]:
    """Finn aksjonærer som også har rolle i selskapet.

    Matcher på normalisert navn. Hvis begge sider har fødselsår, brukes
    det som tilleggssjekk (`"Bekreftet"`). Ellers `"Navn-match"`.

    Når `indirect_owners_fn` er gitt, sjekker vi også **indirekte eierskap**:
    for hver selskaps-aksjonær slås dens eiere opp og matches mot roller.
    `max_indirect_depth` styrer hvor mange ledd opp vi rekurserer (default 1).
    Effektiv eierandel = produkt av alle eierandeler i kjeden.
    """
    owners_list = list(owners or [])
    roller_list = list(roller or [])

    role_index: dict[str, list[dict[str, str]]] = {}
    for role in roller_list:
        key = normalize_person_name(role.get("navn"))
        if not key:
            continue
        role_index.setdefault(key, []).append(role)

    matches: list[CrossMatch] = []

    # Direkte match: aksjonær i klienten som også er rolleinnehaver.
    for owner in owners_list:
        name = owner.get("shareholder_name") or ""
        key = normalize_person_name(name)
        if not key or key not in role_index:
            continue

        role_entries = role_index[key]
        roles = [r.get("rolle", "") for r in role_entries if r.get("rolle")]
        confidence, warn = _match_role_confidence(role_entries, owner_birth_year(owner))

        pct = float(owner.get("ownership_pct") or 0.0)
        notat = f"Direkte aksjonær i klienten med {_fmt_pct(pct)} eierandel."
        if warn:
            notat = f"{notat} {warn}"

        matches.append(
            CrossMatch(
                shareholder_name=str(name),
                shareholder_orgnr=str(owner.get("shareholder_orgnr") or ""),
                shareholder_kind=str(owner.get("shareholder_kind") or ""),
                ownership_pct=pct,
                roles=roles,
                match_confidence=confidence,
                notat=notat,
                match_type="direct",
                via="",
            )
        )

    # Indirekte match: BFS nedover (fra klient) gjennom selskaps-aksjonærer.
    if indirect_owners_fn is not None and max_indirect_depth >= 1:
        chain_nodes, _breaks = walk_indirect_chain(
            owners_list, indirect_owners_fn, max_indirect_depth,
        )

        for node in chain_nodes:
            for sub in node["sub_owners"]:
                sub_name = str(sub.get("shareholder_name") or "")
                sub_key = normalize_person_name(sub_name)
                if not sub_key or sub_key not in role_index:
                    continue

                sub_pct = float(sub.get("ownership_pct") or 0.0)
                role_entries = role_index[sub_key]
                roles = [r.get("rolle", "") for r in role_entries if r.get("rolle")]
                confidence, warn = _match_role_confidence(
                    role_entries, owner_birth_year(sub)
                )

                # Effektiv pct = produkt av alle pct i kjeden × sub_pct / 100^n
                effective = sub_pct
                for _, pct_step in node["chain"]:
                    effective = effective * pct_step / 100.0

                immediate_holding = node["chain"][-1][0]
                chain_str = " → ".join(
                    f"{nm} ({_fmt_pct(p)})" for nm, p in node["chain"]
                )
                notat = (
                    f"Eier {_fmt_pct(sub_pct)} av {immediate_holding}, "
                    f"som eier {_fmt_pct(node['chain'][0][1])} av klienten. "
                    f"Effektiv indirekte eierandel: {_fmt_pct(effective)}."
                )
                if len(node["chain"]) > 1:
                    notat = f"{notat} Eierkjede: klient ← {chain_str}."
                if warn:
                    notat = f"{notat} {warn}"

                matches.append(
                    CrossMatch(
                        shareholder_name=sub_name,
                        shareholder_orgnr=str(sub.get("shareholder_orgnr") or ""),
                        shareholder_kind=str(sub.get("shareholder_kind") or ""),
                        ownership_pct=effective,
                        roles=roles,
                        match_confidence=confidence,
                        notat=notat,
                        match_type="indirect",
                        via=immediate_holding,
                    )
                )

    # Direct først, så indirect. Innenfor hver: synkende pct.
    matches.sort(
        key=lambda m: (
            m.match_type != "direct",
            -m.ownership_pct,
            m.shareholder_name.casefold(),
        )
    )
    return matches


def build_conclusion_text(matches: list[CrossMatch]) -> str:
    """Naturlig-språk oppsummering for forsiden.

    Returnerer én setning per match, klar til å limes inn i en
    "Konklusjon"-blokk. Returnerer en tom-state-setning når det ikke er
    noen match.
    """
    if not matches:
        return (
            "Ingen av aksjonærene har overlappende roller (daglig leder, "
            "styreverv, revisor eller regnskapsfører). Rolleinnehavere og "
            "aksjonærer er separate kretser."
        )

    lines: list[str] = []
    for m in matches:
        role_text = _fmt_roles_natural(m.roles)
        pct = _fmt_pct(m.ownership_pct)
        if m.match_type == "direct":
            lines.append(
                f"• {role_text} {m.shareholder_name} er direkte aksjonær i "
                f"klienten med {pct} eierandel."
            )
        else:
            lines.append(
                f"• {role_text} {m.shareholder_name} er indirekte aksjonær "
                f"via sitt eierskap i {m.via} — effektiv eierandel {pct}."
            )
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Workbook

def build_klientinfo_workpaper(
    *,
    client: str,
    year: str,
    client_orgnr: str = "",
    enhet: Optional[dict[str, Any]] = None,
    roller: Optional[list[dict[str, str]]] = None,
    owners: Optional[list[dict[str, Any]]] = None,
    owned_companies: Optional[list[dict[str, Any]]] = None,
    owners_year_used: str = "",
    indirect_owners_fn: Optional[Callable[[str], list[dict[str, Any]]]] = None,
) -> Workbook:
    """Bygg Excel-arbeidsbok med klientinfo, roller og eierskap."""
    wb = Workbook()
    enhet = enhet or {}
    roller = roller or []
    owners = owners or []
    owned_companies = owned_companies or []

    _build_oversikt_sheet(
        wb, client=client, year=year, client_orgnr=client_orgnr, enhet=enhet,
    )
    _build_roller_sheet(wb, roller=roller, client=client, year=year)
    _build_aksjonaerer_sheet(
        wb, owners=owners, client=client, year=year, owners_year_used=owners_year_used,
    )
    matches = build_cross_matches(
        owners, roller,
        indirect_owners_fn=indirect_owners_fn,
        max_indirect_depth=5,
    )
    _build_kryssreferanse_sheet(wb, matches=matches, client=client, year=year)
    _build_eide_sheet(wb, owned_companies=owned_companies, client=client, year=year)

    conclusion = build_conclusion_text(matches)
    build_forside_sheet(
        wb,
        workpaper_navn="Klientinfo, roller & eierskap",
        extra_blocks=[("Konklusjon — roller og eierskap", conclusion)],
    )

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    return wb


# ---------------------------------------------------------------------------
# Ark 1: Oversikt

def _build_oversikt_sheet(
    wb: Workbook,
    *,
    client: str,
    year: str,
    client_orgnr: str,
    enhet: dict[str, Any],
) -> None:
    ws = wb.create_sheet("Oversikt")

    title = "Klientinfo"
    if client:
        title += f" — {client}"
    if year:
        title += f" {year}"

    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = _TITLE_FILL

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    row = 4
    orgform = _lookup_desc(enhet.get("organisasjonsform"))
    naering = _lookup_desc(enhet.get("naeringskode1"))
    naering_full = str(enhet.get("naeringsnavn") or naering)
    mva = enhet.get("registrertIMvaregisteret")
    navn = str(enhet.get("navn") or client)

    items: list[tuple[str, Any, bool]] = [
        ("Navn", navn, False),
        ("Org.nr", client_orgnr, False),
        ("Organisasjonsform", orgform, False),
        ("Næring", naering_full, False),
        ("MVA-registrert", "Ja" if mva else "Nei", False),
        ("Stiftelsesdato", _normalize_date(enhet.get("stiftelsesdato")), False),
    ]

    status_items = _status_flags(enhet)
    for label, value, red in status_items:
        items.append((label, value, red))

    for label, value, red in items:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(bold=True)
        c2 = ws.cell(row=row, column=2, value=value if value not in (None, "") else "–")
        if red:
            c2.font = Font(bold=True, color="C62828")
            c2.fill = _BAD_FILL
        row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60
    ws.sheet_properties.tabColor = "4472C4"


def _lookup_desc(raw: Any) -> str:
    if not raw:
        return ""
    if isinstance(raw, dict):
        return str(raw.get("beskrivelse") or raw.get("kode") or "")
    return str(raw)


def _normalize_date(raw: Any) -> str:
    if not raw:
        return ""
    s = str(raw)
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}.{m.group(2)}.{m.group(1)}"
    return s


def _status_flags(enhet: dict[str, Any]) -> list[tuple[str, str, bool]]:
    """Returnerer statusrader — rød hvis konkurs/avvikling/slettet."""
    flags: list[tuple[str, str, bool]] = []
    if enhet.get("konkurs"):
        flags.append(("Status", "Konkurs", True))
    elif enhet.get("underTvangsavvikling"):
        flags.append(("Status", "Under tvangsavvikling", True))
    elif enhet.get("underAvvikling"):
        flags.append(("Status", "Under avvikling", True))
    elif enhet.get("slettedato"):
        flags.append(("Status", f"Slettet {_normalize_date(enhet.get('slettedato'))}", True))
    else:
        flags.append(("Status", "Aktiv", False))
    return flags


# ---------------------------------------------------------------------------
# Ark 2: Roller

def _build_roller_sheet(
    wb: Workbook,
    *,
    roller: list[dict[str, str]],
    client: str,
    year: str,
) -> None:
    ws = wb.create_sheet("Roller")

    title = "Roller (fra BRREG)"
    if client:
        title += f" — {client}"

    headers = ["Rolle", "Navn", "Fødselsår"]
    _write_title_and_header(ws, title, headers, span=len(headers))

    if not roller:
        ws.cell(row=5, column=1, value="Ingen roller registrert i Enhetsregisteret.")
        ws["A5"].font = Font(italic=True, color="666666")
    else:
        for idx, rolle in enumerate(roller, start=5):
            ws.cell(row=idx, column=1, value=rolle.get("rolle", "")).border = _BORDER
            ws.cell(row=idx, column=2, value=rolle.get("navn", "")).border = _BORDER
            ws.cell(row=idx, column=3, value=_birth_year(rolle.get("fodselsdato"))).border = _BORDER

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "4472C4"


# ---------------------------------------------------------------------------
# Ark 3: Aksjonærer

def _build_aksjonaerer_sheet(
    wb: Workbook,
    *,
    owners: list[dict[str, Any]],
    client: str,
    year: str,
    owners_year_used: str,
) -> None:
    ws = wb.create_sheet("Aksjonærer")

    title = "Aksjonærer (eiere)"
    if client:
        title += f" — {client}"
    if year:
        title += f" {year}"

    headers = ["Aksjonær", "Orgnr / Fødselsår", "Type", "Fødselsår", "Antall aksjer", "Eierandel (%)"]
    _write_title_and_header(ws, title, headers, span=len(headers))

    if not owners:
        ws.cell(row=5, column=1, value="Ingen aksjonærer registrert.")
        ws["A5"].font = Font(italic=True, color="666666")
    else:
        for idx, owner in enumerate(owners, start=5):
            ws.cell(row=idx, column=1, value=owner.get("shareholder_name", "")).border = _BORDER
            ws.cell(row=idx, column=2, value=owner.get("shareholder_orgnr", "")).border = _BORDER
            ws.cell(row=idx, column=3, value=owner.get("shareholder_kind", "")).border = _BORDER
            ws.cell(row=idx, column=4, value=owner_birth_year(owner)).border = _BORDER
            c_shares = ws.cell(row=idx, column=5, value=int(owner.get("shares") or 0))
            c_shares.border = _BORDER
            c_shares.number_format = "#,##0"
            pct = float(owner.get("ownership_pct") or 0.0)
            c_pct = ws.cell(row=idx, column=6, value=pct)
            c_pct.border = _BORDER
            c_pct.number_format = _PCT_FMT
            if pct >= 50.0:
                c_pct.font = Font(bold=True)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "70AD47"


# ---------------------------------------------------------------------------
# Ark 4: Kryssreferanse

def _build_kryssreferanse_sheet(
    wb: Workbook,
    *,
    matches: list[CrossMatch],
    client: str,
    year: str,
) -> None:
    ws = wb.create_sheet("Kryssreferanse")

    title = "Aksjonærer som også har rolle"
    if client:
        title += f" — {client}"

    headers = ["Navn", "Type", "Via", "Eierandel (%)", "Roller", "Konfidens", "Merknad"]
    _write_title_and_header(ws, title, headers, span=len(headers))

    if not matches:
        ws.cell(row=5, column=1, value="Ingen aksjonærer ble matchet mot roller.")
        ws["A5"].font = Font(italic=True, color="006100")
    else:
        for idx, m in enumerate(matches, start=5):
            ws.cell(row=idx, column=1, value=m.shareholder_name).border = _BORDER
            type_label = "Direkte" if m.match_type == "direct" else "Indirekte"
            c_type = ws.cell(row=idx, column=2, value=type_label)
            c_type.border = _BORDER
            if m.match_type == "indirect":
                c_type.fill = _WARN_FILL
            ws.cell(row=idx, column=3, value=m.via).border = _BORDER
            c_pct = ws.cell(row=idx, column=4, value=m.ownership_pct)
            c_pct.border = _BORDER
            c_pct.number_format = _PCT_FMT
            ws.cell(row=idx, column=5, value=", ".join(m.roles)).border = _BORDER
            c_conf = ws.cell(row=idx, column=6, value=m.match_confidence)
            c_conf.border = _BORDER
            if m.match_confidence == "Navn-match":
                c_conf.fill = _WARN_FILL
            ws.cell(row=idx, column=7, value=m.notat).border = _BORDER

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 70
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "FF0000" if matches else "70AD47"


# ---------------------------------------------------------------------------
# Ark 5: Eide selskaper

def _build_eide_sheet(
    wb: Workbook,
    *,
    owned_companies: list[dict[str, Any]],
    client: str,
    year: str,
) -> None:
    ws = wb.create_sheet("Eide selskaper")

    title = "Selskaper klienten eier andeler i"
    if client:
        title += f" — {client}"
    if year:
        title += f" {year}"

    headers = ["Selskap", "Orgnr", "Eierandel (%)", "Relasjon", "Intern klient", "Kilde"]
    _write_title_and_header(ws, title, headers, span=len(headers))

    if not owned_companies:
        ws.cell(row=5, column=1, value="Klienten har ingen registrerte eierandeler.")
        ws["A5"].font = Font(italic=True, color="666666")
    else:
        for idx, row in enumerate(owned_companies, start=5):
            ws.cell(row=idx, column=1, value=row.get("company_name", "")).border = _BORDER
            ws.cell(row=idx, column=2, value=row.get("company_orgnr", "")).border = _BORDER
            pct = float(row.get("ownership_pct") or 0.0)
            c_pct = ws.cell(row=idx, column=3, value=pct)
            c_pct.border = _BORDER
            c_pct.number_format = _PCT_FMT
            ws.cell(row=idx, column=4, value=row.get("relation_type", "")).border = _BORDER
            ws.cell(row=idx, column=5, value=row.get("matched_client") or "").border = _BORDER
            ws.cell(row=idx, column=6, value=row.get("source", "")).border = _BORDER

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 18
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "4472C4"


# ---------------------------------------------------------------------------
# Felles

def _write_title_and_header(ws, title: str, headers: list[str], *, span: int) -> None:
    last_col_letter = chr(ord("A") + span - 1)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = _TITLE_FILL

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")
