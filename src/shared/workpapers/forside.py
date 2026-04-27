"""Felles "Beskrivelse"-fane for innebygde Excel-arbeidspapirer.

Funksjonen `build_forside_sheet` leser aktiv `ActionContext` (hvis satt)
og legger en forside som første fane. Uten kontekst gjør funksjonen
ingenting — kallere trenger ikke vite om eksporten ble startet fra en
handling eller ikke.
"""

from __future__ import annotations

from datetime import datetime
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import action_context

_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
_LABEL_FILL = PatternFill("solid", fgColor="F3F6F9")


def build_forside_sheet(
    wb: Workbook,
    *,
    workpaper_navn: str = "",
    extra_blocks: Optional[Iterable[tuple[str, str]]] = None,
) -> bool:
    """Legg "Beskrivelse" som første fane hvis det finnes aktiv ActionContext.

    `extra_blocks` er en liste av `(label, tekst)`-par som legges til etter
    Handlingsbeskrivelse og Revisors kommentar — f.eks. en auto-generert
    konklusjon fra arbeidspapiret.

    Returnerer `True` hvis fanen ble lagt til. Gjør ingenting (og returnerer
    `False`) hvis ingen handling-kontekst er satt.
    """
    ctx = action_context.current()
    if ctx is None:
        return False

    ws = wb.create_sheet("Beskrivelse", 0)

    title = workpaper_navn or ctx.workpaper_navn or "Arbeidspapir"
    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].fill = _TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generert {_now_str()}"
    ws["A2"].font = Font(italic=True, color="666666")

    row = 4
    meta = [
        ("Handling", ctx.handling_navn),
        ("Type", ctx.handling_type),
        ("Område", ctx.omraade),
        ("Regnr", ctx.regnr),
        ("Klient", ctx.client),
        ("År", ctx.year),
        ("Kjørt av", ctx.kjort_av),
    ]
    for label, value in meta:
        if not value:
            continue
        _write_label(ws, row, label, value)
        row += 1

    row += 1
    _write_block(ws, row, "Handlingsbeskrivelse", ctx.beskrivelse or "(ingen beskrivelse)")
    row += 2 + _block_row_span(ctx.beskrivelse)

    if ctx.kommentar:
        _write_block(ws, row, "Revisors kommentar", ctx.kommentar)
        row += 2 + _block_row_span(ctx.kommentar)

    for label, text in (extra_blocks or []):
        if not text:
            continue
        _write_block(ws, row, label, text)
        row += 2 + _block_row_span(text)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    ws.sheet_properties.tabColor = "70AD47"
    return True


def _write_label(ws, row: int, label: str, value: str) -> None:
    c1 = ws.cell(row=row, column=1, value=label)
    c1.font = Font(bold=True)
    c1.fill = _LABEL_FILL
    c1.alignment = Alignment(vertical="top")
    c2 = ws.cell(row=row, column=2, value=value)
    c2.alignment = Alignment(vertical="top", wrap_text=True)


def _write_block(ws, row: int, label: str, text: str) -> None:
    header = ws.cell(row=row, column=1, value=label)
    header.font = Font(bold=True)
    header.fill = _LABEL_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    body = ws.cell(row=row + 1, column=1, value=text)
    body.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=4)
    ws.row_dimensions[row + 1].height = max(48, 14 * _estimated_lines(text))


def _estimated_lines(text: str) -> int:
    if not text:
        return 1
    lines = text.count("\n") + 1
    # ~90 tegn pr. linje ved full bredde
    for segment in text.split("\n"):
        lines += max(0, len(segment) // 90)
    return min(30, max(1, lines))


def _block_row_span(text: str) -> int:
    return 1


def _now_str() -> str:
    return datetime.now().strftime("%d.%m.%Y %H:%M")
