"""document_control_export.py — Eksport av bilagskontroll-rapport.

Genererer:
  - Excel (.xlsx) med oppsummering + detaljark med HB vs PDF-verdier
  - PDF-klar HTML-rapport som kan skrives ut fra browser
"""
from __future__ import annotations

import os
import sys
import tempfile
import webbrowser
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Data structure ────────────────────────────────────────────────────
# Each bilag dict should have:
#   bilag_nr, supplier_name, status, hb_fields, pdf_fields, avvik, notes
# where hb_fields/pdf_fields are {field_key: value} dicts.

FIELD_DEFS: list[tuple[str, str]] = [
    ("supplier_name",   "Leverandør"),
    ("supplier_orgnr",  "Org.nr."),
    ("invoice_number",  "Fakturanr."),
    ("invoice_date",    "Fakturadato"),
    ("due_date",        "Forfallsdato"),
    ("subtotal_amount", "Beløp ekskl. mva"),
    ("vat_amount",      "MVA"),
    ("total_amount",    "Total"),
    ("currency",        "Valuta"),
]

STATUS_LABELS = {
    "ok": "OK",
    "avvik": "Avvik",
    "ikke_funnet": "Ikke funnet",
    "feil": "Feil",
}

# ── Excel export ──────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_OK_FILL     = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
_AVVIK_FILL  = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
_IKKE_FILL   = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def export_to_excel(
    *,
    client: str,
    year: str,
    bilag_data: list[dict[str, Any]],
    output_path: Path | None = None,
) -> Path:
    """Export document control results to Excel."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Oppsummering ─────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Oppsummering"

    n_ok = sum(1 for b in bilag_data if b["status"] == "ok")
    n_avvik = sum(1 for b in bilag_data if b["status"] == "avvik")
    n_ikke = sum(1 for b in bilag_data if b["status"] == "ikke_funnet")
    n_total = len(bilag_data)

    info_rows = [
        ("Bilagskontroll — dokumentkontroll", ""),
        ("", ""),
        ("Klient", client),
        ("Regnskapsår", year),
        ("Rapportdato", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("", ""),
        ("Antall bilag kontrollert", n_total),
        ("OK", n_ok),
        ("Avvik", n_avvik),
        ("Ikke funnet", n_ikke),
    ]
    for r, (label, value) in enumerate(info_rows, 1):
        ws_sum.cell(row=r, column=1, value=label).font = Font(
            name="Calibri", size=11, bold=(r == 1 or label in ("OK", "Avvik", "Ikke funnet"))
        )
        ws_sum.cell(row=r, column=2, value=value).font = Font(name="Calibri", size=11)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 40

    # ── Bilagsliste i oppsummering ────────────────────────────────────
    list_start = len(info_rows) + 2
    list_headers = ["Bilagsnr.", "Leverandør", "Total (HB)", "Total (PDF)", "Status", "Avvik"]
    for c, h in enumerate(list_headers, 1):
        cell = ws_sum.cell(row=list_start, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER

    for i, b in enumerate(bilag_data):
        r = list_start + 1 + i
        status = b["status"]
        row_fill = _OK_FILL if status == "ok" else _AVVIK_FILL if status == "avvik" else _IKKE_FILL

        vals = [
            b["bilag_nr"],
            b.get("hb_fields", {}).get("supplier_name", "") or b.get("pdf_fields", {}).get("supplier_name", ""),
            b.get("hb_fields", {}).get("total_amount", ""),
            b.get("pdf_fields", {}).get("total_amount", ""),
            STATUS_LABELS.get(status, status),
            "; ".join(b.get("avvik", [])) if b.get("avvik") else "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws_sum.cell(row=r, column=c, value=v)
            cell.fill = row_fill
            cell.border = _THIN_BORDER
            cell.font = Font(name="Calibri", size=10)

    for c in range(1, len(list_headers) + 1):
        ws_sum.column_dimensions[get_column_letter(c)].width = [12, 30, 15, 15, 12, 50][c - 1]

    # ── Sheet 2: Detaljer — HB vs PDF per felt ───────────────────────
    ws_det = wb.create_sheet("Detaljer")

    det_headers = ["Bilagsnr.", "Status"]
    for _, label in FIELD_DEFS:
        det_headers.append(f"HB: {label}")
        det_headers.append(f"PDF: {label}")
        det_headers.append(f"Match")
    det_headers.append("Avvik")
    det_headers.append("Notater")

    for c, h in enumerate(det_headers, 1):
        cell = ws_det.cell(row=1, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)

    for i, b in enumerate(bilag_data):
        r = 2 + i
        status = b["status"]
        row_fill = _OK_FILL if status == "ok" else _AVVIK_FILL if status == "avvik" else _IKKE_FILL
        hb = b.get("hb_fields", {})
        pdf = b.get("pdf_fields", {})

        col = 1
        ws_det.cell(row=r, column=col, value=b["bilag_nr"]).fill = row_fill
        ws_det.cell(row=r, column=col).border = _THIN_BORDER
        col += 1
        ws_det.cell(row=r, column=col, value=STATUS_LABELS.get(status, status)).fill = row_fill
        ws_det.cell(row=r, column=col).border = _THIN_BORDER
        col += 1

        for key, _ in FIELD_DEFS:
            hb_val = hb.get(key, "")
            pdf_val = pdf.get(key, "")
            match = "✓" if _values_match(key, hb_val, pdf_val) else ("—" if not hb_val else "✗")

            for v in (hb_val, pdf_val, match):
                cell = ws_det.cell(row=r, column=col, value=v)
                cell.fill = row_fill
                cell.border = _THIN_BORDER
                cell.font = Font(name="Calibri", size=10)
                col += 1

        avvik_str = "; ".join(b.get("avvik", []))
        ws_det.cell(row=r, column=col, value=avvik_str).fill = row_fill
        ws_det.cell(row=r, column=col).border = _THIN_BORDER
        col += 1
        ws_det.cell(row=r, column=col, value=b.get("notes", "")).fill = row_fill
        ws_det.cell(row=r, column=col).border = _THIN_BORDER

    # Auto-width for detail sheet
    for c in range(1, len(det_headers) + 1):
        ws_det.column_dimensions[get_column_letter(c)].width = 14

    # Save
    if output_path is None:
        output_path = Path(tempfile.gettempdir()) / f"Bilagskontroll_{client}_{year}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def _values_match(key: str, hb: str, pdf: str) -> bool:
    if not hb or not pdf:
        return False
    hb_n = hb.strip().lower().replace("\xa0", " ")
    pdf_n = pdf.strip().lower().replace("\xa0", " ")
    if hb_n == pdf_n:
        return True
    # Numeric comparison for amounts
    if key in ("subtotal_amount", "vat_amount", "total_amount"):
        try:
            a = float(hb_n.replace(" ", "").replace(",", "."))
            b = float(pdf_n.replace(" ", "").replace(",", "."))
            return abs(a - b) < 0.01
        except ValueError:
            pass
    return False


# ── HTML/PDF export ───────────────────────────────────────────────────

def export_to_html(
    *,
    client: str,
    year: str,
    bilag_data: list[dict[str, Any]],
    output_path: Path | None = None,
) -> Path:
    """Export document control results to a printable HTML report."""
    n_ok = sum(1 for b in bilag_data if b["status"] == "ok")
    n_avvik = sum(1 for b in bilag_data if b["status"] == "avvik")
    n_ikke = sum(1 for b in bilag_data if b["status"] == "ikke_funnet")

    rows_html = []
    for b in bilag_data:
        status = b["status"]
        status_cls = "ok" if status == "ok" else "avvik" if status == "avvik" else "ikke"
        hb = b.get("hb_fields", {})
        pdf = b.get("pdf_fields", {})
        avvik = b.get("avvik", [])

        # Build field comparison rows
        field_rows = []
        for key, label in FIELD_DEFS:
            hb_val = escape(str(hb.get(key, "") or ""))
            pdf_val = escape(str(pdf.get(key, "") or ""))
            matched = _values_match(key, hb.get(key, ""), pdf.get(key, ""))
            match_icon = "✓" if matched else ("—" if not hb_val else "✗")
            match_cls = "match-ok" if matched else ("match-na" if not hb_val else "match-fail")
            field_rows.append(
                f'<tr><td class="field-label">{escape(label)}</td>'
                f'<td>{hb_val}</td>'
                f'<td>{pdf_val}</td>'
                f'<td class="{match_cls}">{match_icon}</td></tr>'
            )

        avvik_html = "<br>".join(escape(a) for a in avvik) if avvik else "Ingen avvik"
        notes_html = escape(b.get("notes", "") or "")

        rows_html.append(f"""
        <div class="bilag-card {status_cls}">
            <div class="bilag-header">
                <span class="bilag-nr">Bilag {escape(str(b['bilag_nr']))}</span>
                <span class="bilag-supplier">{escape(hb.get('supplier_name', '') or pdf.get('supplier_name', ''))}</span>
                <span class="bilag-status status-{status_cls}">{escape(STATUS_LABELS.get(status, status))}</span>
            </div>
            <table class="field-table">
                <thead><tr><th>Felt</th><th>Regnskap (HB)</th><th>PDF (innlest)</th><th></th></tr></thead>
                <tbody>{''.join(field_rows)}</tbody>
            </table>
            <div class="avvik-section"><strong>Avvik:</strong> {avvik_html}</div>
            {"<div class='notes-section'><strong>Notater:</strong> " + notes_html + "</div>" if notes_html else ""}
        </div>
        """)

    html = f"""<!DOCTYPE html>
<html lang="no">
<head>
<meta charset="utf-8">
<title>Bilagskontroll — {escape(client)} {escape(year)}</title>
<style>
    @page {{ margin: 15mm; }}
    body {{ font-family: 'Segoe UI', Calibri, Arial, sans-serif; font-size: 11px;
            color: #222; max-width: 210mm; margin: 0 auto; padding: 15mm; }}
    h1 {{ font-size: 18px; margin-bottom: 4px; }}
    .meta {{ color: #555; margin-bottom: 12px; }}
    .summary {{ display: flex; gap: 20px; margin-bottom: 20px; padding: 10px;
                background: #f5f5f5; border-radius: 4px; }}
    .summary-item {{ text-align: center; }}
    .summary-item .count {{ font-size: 22px; font-weight: bold; }}
    .summary-item .label {{ font-size: 10px; color: #666; }}
    .count-ok {{ color: #2e7d32; }}
    .count-avvik {{ color: #e65100; }}
    .count-ikke {{ color: #c62828; }}
    .bilag-card {{ border: 1px solid #ddd; border-radius: 4px; margin-bottom: 12px;
                   padding: 8px 12px; page-break-inside: avoid; }}
    .bilag-card.ok {{ border-left: 4px solid #4caf50; }}
    .bilag-card.avvik {{ border-left: 4px solid #ff9800; }}
    .bilag-card.ikke {{ border-left: 4px solid #f44336; }}
    .bilag-header {{ display: flex; justify-content: space-between; align-items: center;
                     margin-bottom: 6px; font-weight: bold; }}
    .bilag-nr {{ font-size: 13px; }}
    .bilag-status {{ padding: 2px 8px; border-radius: 3px; font-size: 10px; }}
    .status-ok {{ background: #e8f5e9; color: #2e7d32; }}
    .status-avvik {{ background: #fff3e0; color: #e65100; }}
    .status-ikke {{ background: #ffebee; color: #c62828; }}
    .field-table {{ width: 100%; border-collapse: collapse; margin-bottom: 6px; }}
    .field-table th {{ text-align: left; font-size: 10px; color: #666;
                       border-bottom: 1px solid #ddd; padding: 2px 4px; }}
    .field-table td {{ padding: 2px 4px; font-size: 11px; border-bottom: 1px solid #f0f0f0; }}
    .field-label {{ font-weight: 500; color: #444; width: 120px; }}
    .match-ok {{ color: #2e7d32; font-weight: bold; }}
    .match-fail {{ color: #c62828; font-weight: bold; }}
    .match-na {{ color: #999; }}
    .avvik-section {{ font-size: 10px; color: #e65100; margin-top: 4px; }}
    .notes-section {{ font-size: 10px; color: #555; margin-top: 2px; }}
    @media print {{
        body {{ padding: 0; }}
        .bilag-card {{ break-inside: avoid; }}
    }}
</style>
</head>
<body>
    <h1>Bilagskontroll — dokumentkontroll</h1>
    <div class="meta">
        Klient: <strong>{escape(client)}</strong> &nbsp;|&nbsp;
        Regnskapsår: <strong>{escape(year)}</strong> &nbsp;|&nbsp;
        Rapportdato: <strong>{datetime.now().strftime("%d.%m.%Y %H:%M")}</strong>
    </div>
    <div class="summary">
        <div class="summary-item">
            <div class="count">{len(bilag_data)}</div>
            <div class="label">Kontrollert</div>
        </div>
        <div class="summary-item">
            <div class="count count-ok">{n_ok}</div>
            <div class="label">OK</div>
        </div>
        <div class="summary-item">
            <div class="count count-avvik">{n_avvik}</div>
            <div class="label">Avvik</div>
        </div>
        <div class="summary-item">
            <div class="count count-ikke">{n_ikke}</div>
            <div class="label">Ikke funnet</div>
        </div>
    </div>
    {''.join(rows_html)}
    <div class="meta" style="margin-top: 20px; text-align: center; font-size: 9px;">
        Generert av Utvalg — bilagskontroll
    </div>
</body>
</html>"""

    if output_path is None:
        output_path = Path(tempfile.gettempdir()) / f"Bilagskontroll_{client}_{year}.html"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    return output_path


# ── Convenience: open file ────────────────────────────────────────────

def open_file(path: Path) -> None:
    """Open a file with the system default application."""
    try:
        if sys.platform.startswith("win"):
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            os.system(f"open '{path}'")
        else:
            os.system(f"xdg-open '{path}'")
    except Exception:
        pass
