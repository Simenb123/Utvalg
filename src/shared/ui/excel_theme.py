"""Vaak-theme helpers for openpyxl exports.

Central definitions of ``PatternFill``, ``Font``, ``Border`` and named
styles so every Excel workbook produced by Utvalg-1 shares the same
visual identity with the Tkinter GUI. All color values come from
``vaak_tokens``.

Typical use::

    from openpyxl import Workbook
    from src.shared.ui.excel_theme import (
        FILL_TITLE, FONT_TITLE, register_vaak_styles,
    )

    wb = Workbook()
    register_vaak_styles(wb)
    ws = wb.active
    ws["A1"].style = "vaak_title"

The ``register_vaak_styles`` helper is idempotent and safe to call from
every module that touches a workbook.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.workbook import Workbook

from . import tokens as vt

# ---------------------------------------------------------------------------
# Fills
# ---------------------------------------------------------------------------
FILL_TITLE = PatternFill("solid", fgColor=vt.BG_SAND)
FILL_HEADER = PatternFill("solid", fgColor=vt.BG_SAND_SOFT)
FILL_SUBHEADER = PatternFill("solid", fgColor=vt.BORDER)
FILL_SUMLINE = PatternFill("solid", fgColor=vt.SAGE_WASH)
FILL_SUMLINE_MAJOR = PatternFill("solid", fgColor=vt.FOREST)
FILL_ZEBRA = PatternFill("solid", fgColor=vt.BG_ZEBRA)
FILL_TOTAL = PatternFill("solid", fgColor=vt.FOREST)
FILL_NEUTRAL = PatternFill("solid", fgColor=vt.BG_NEUTRAL)
FILL_POS_SOFT = PatternFill("solid", fgColor=vt.POS_SOFT)
FILL_NEG_SOFT = PatternFill("solid", fgColor=vt.NEG_SOFT)
FILL_WARN_SOFT = PatternFill("solid", fgColor=vt.WARN_SOFT)

# ---------------------------------------------------------------------------
# Fonts
# ---------------------------------------------------------------------------
_BODY = vt.FONT_FAMILY_BODY

FONT_TITLE = Font(name=_BODY, size=16, bold=True, color=vt.TEXT_PRIMARY)
FONT_H1 = Font(name=_BODY, size=13, bold=True, color=vt.TEXT_PRIMARY)
FONT_H2 = Font(name=_BODY, size=11, bold=True, color=vt.TEXT_PRIMARY)
FONT_HEADER = Font(name=_BODY, size=10, bold=True, color=vt.TEXT_PRIMARY)
FONT_SUBHEADER = Font(name=_BODY, size=10, bold=True, color=vt.TEXT_PRIMARY)
FONT_BODY = Font(name=_BODY, size=10, color=vt.TEXT_PRIMARY)
FONT_MUTED = Font(name=_BODY, size=9, color=vt.TEXT_MUTED)
FONT_TOTAL = Font(name=_BODY, size=11, bold=True, color=vt.TEXT_ON_FOREST)
FONT_SUM = Font(name=_BODY, size=10, bold=True, color=vt.TEXT_PRIMARY)
FONT_SUM_MAJOR = Font(name=_BODY, size=11, bold=True, color=vt.TEXT_ON_FOREST)
FONT_POS = Font(name=_BODY, size=10, color=vt.POS_TEXT)
FONT_NEG = Font(name=_BODY, size=10, color=vt.NEG_TEXT)
FONT_WARN = Font(name=_BODY, size=10, color=vt.WARN_TEXT)

# ---------------------------------------------------------------------------
# Borders
# ---------------------------------------------------------------------------
_BORDER_SIDE = Side(style="thin", color=vt.BORDER)
_BORDER_SOFT_SIDE = Side(style="thin", color=vt.BORDER_SOFT)
_BORDER_MEDIUM = Side(style="medium", color=vt.FOREST)

BORDER_CELL = Border(
    left=_BORDER_SOFT_SIDE,
    right=_BORDER_SOFT_SIDE,
    top=_BORDER_SOFT_SIDE,
    bottom=_BORDER_SOFT_SIDE,
)
BORDER_HEADER = Border(
    left=_BORDER_SIDE,
    right=_BORDER_SIDE,
    top=_BORDER_SIDE,
    bottom=_BORDER_SIDE,
)
BORDER_THICK_BOTTOM = Border(bottom=_BORDER_MEDIUM)
BORDER_TOP_ACCENT = Border(top=_BORDER_MEDIUM)

# ---------------------------------------------------------------------------
# Alignment
# ---------------------------------------------------------------------------
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_TITLE = Alignment(horizontal="left", vertical="center", wrap_text=False)


# ---------------------------------------------------------------------------
# Named styles
# ---------------------------------------------------------------------------
_NAMED_STYLES: dict[str, dict] = {
    "vaak_title": dict(font=FONT_TITLE, fill=FILL_TITLE, alignment=ALIGN_TITLE),
    "vaak_h1": dict(font=FONT_H1, fill=FILL_HEADER, alignment=ALIGN_LEFT),
    "vaak_h2": dict(font=FONT_H2, alignment=ALIGN_LEFT),
    "vaak_header": dict(
        font=FONT_HEADER,
        fill=FILL_HEADER,
        alignment=ALIGN_HEADER,
        border=BORDER_HEADER,
    ),
    "vaak_subheader": dict(
        font=FONT_SUBHEADER, fill=FILL_SUBHEADER, alignment=ALIGN_LEFT
    ),
    "vaak_body": dict(font=FONT_BODY, alignment=ALIGN_LEFT, border=BORDER_CELL),
    "vaak_muted": dict(font=FONT_MUTED, alignment=ALIGN_LEFT),
    "vaak_sum": dict(font=FONT_SUM, fill=FILL_SUMLINE, alignment=ALIGN_LEFT),
    "vaak_sum_major": dict(
        font=FONT_SUM_MAJOR, fill=FILL_SUMLINE_MAJOR, alignment=ALIGN_LEFT
    ),
    "vaak_total": dict(
        font=FONT_TOTAL,
        fill=FILL_TOTAL,
        alignment=ALIGN_LEFT,
        border=BORDER_TOP_ACCENT,
    ),
    "vaak_pos": dict(font=FONT_POS, alignment=ALIGN_RIGHT),
    "vaak_neg": dict(font=FONT_NEG, alignment=ALIGN_RIGHT),
    "vaak_zebra": dict(font=FONT_BODY, fill=FILL_ZEBRA, alignment=ALIGN_LEFT),
}


def register_vaak_styles(wb: Workbook) -> None:
    """Idempotently add all Vaak named styles to ``wb``.

    Safe to call from any export module — skips entries already present.
    """
    existing = set(wb.named_styles)
    for name, attrs in _NAMED_STYLES.items():
        if name in existing:
            continue
        style = NamedStyle(name=name)
        for key, value in attrs.items():
            setattr(style, key, value)
        wb.add_named_style(style)


__all__ = [
    "FILL_TITLE",
    "FILL_HEADER",
    "FILL_SUBHEADER",
    "FILL_SUMLINE",
    "FILL_SUMLINE_MAJOR",
    "FILL_ZEBRA",
    "FILL_TOTAL",
    "FILL_NEUTRAL",
    "FILL_POS_SOFT",
    "FILL_NEG_SOFT",
    "FILL_WARN_SOFT",
    "FONT_TITLE",
    "FONT_H1",
    "FONT_H2",
    "FONT_HEADER",
    "FONT_SUBHEADER",
    "FONT_BODY",
    "FONT_MUTED",
    "FONT_TOTAL",
    "FONT_SUM",
    "FONT_SUM_MAJOR",
    "FONT_POS",
    "FONT_NEG",
    "FONT_WARN",
    "BORDER_CELL",
    "BORDER_HEADER",
    "BORDER_THICK_BOTTOM",
    "BORDER_TOP_ACCENT",
    "ALIGN_LEFT",
    "ALIGN_CENTER",
    "ALIGN_RIGHT",
    "ALIGN_HEADER",
    "ALIGN_TITLE",
    "register_vaak_styles",
]
