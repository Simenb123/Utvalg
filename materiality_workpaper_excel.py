from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from materiality_engine import BENCHMARK_KEYS, BENCHMARK_LABELS, calculate_materiality, get_default_percentages


_TITLE_FILL = PatternFill("solid", fgColor="2F6D62")
_SECTION_FILL = PatternFill("solid", fgColor="E5F1EE")
_ALT_FILL = PatternFill("solid", fgColor="F7FAF8")
_TEXT_COLOR = "1F2430"
_MUTED_COLOR = "667085"


def export_materiality_workpaper(
    target: str | Path,
    *,
    client: str,
    year: str,
    active_materiality: Mapping[str, Any] | None,
    selection_threshold_label: str,
    state_updated_at: str = "",
    crm_client_number: str = "",
    crm_lookup: object | None = None,
    calculation_payload: Mapping[str, Any] | None = None,
    benchmark_amounts: Mapping[str, float] | None = None,
) -> Path:
    out = Path(target)
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Oversikt"

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row, 1, "Vesentlighetsarbeidspapir")
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = _TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24

    row += 1
    subtitle = f"{client or 'Ukjent klient'} | {year or '-'}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row, 1, subtitle)
    cell.font = Font(italic=True, color=_MUTED_COLOR)

    row += 2
    row = _write_meta_section(
        ws,
        row,
        "Dokumentasjon",
        [
            ("Klient", client),
            ("År", year),
            ("Eksportert", datetime.now().strftime("%d.%m.%Y %H:%M")),
            ("Brukes i Utvalg", selection_threshold_label),
            ("Sist lagret i appen", state_updated_at or "-"),
        ],
    )

    row += 1
    row = _write_meta_section(
        ws,
        row,
        "Aktiv verdi",
        [
            ("Kilde", _display_source_label(active_materiality)),
            ("OM", _mapping_value(active_materiality, "overall_materiality")),
            ("PM", _mapping_value(active_materiality, "performance_materiality")),
            ("ClearlyTriv", _mapping_value(active_materiality, "clearly_trivial")),
        ],
    )

    row += 1
    row = _write_meta_section(
        ws,
        row,
        "CRMSystem",
        [
            ("Valgt klientnr", crm_client_number or "-"),
            ("Match", _crm_match_text(crm_lookup)),
            ("Oppdragsår", _crm_record_value(crm_lookup, "engagement_year")),
            ("Materiality", _crm_record_value(crm_lookup, "materiality")),
            ("Arbeidsvesentlighet (PM)", _crm_record_value(crm_lookup, "pmateriality")),
            ("ClearlyTriv", _crm_record_value(crm_lookup, "clearly_triv")),
            ("Kilde oppdatert", _crm_record_value(crm_lookup, "source_updated_at", default="-")),
            ("Sist synket til CRM", _crm_record_value(crm_lookup, "last_synced_at_utc", default="-")),
            ("CRM-DB", _crm_lookup_value(crm_lookup, "db_path", default="-")),
        ],
    )

    row += 1
    row = _write_meta_section(
        ws,
        row,
        "Lokal beregning",
        [
            ("Benchmark", _benchmark_label(calculation_payload)),
            ("Benchmarkgrunnlag", _mapping_value(calculation_payload, "benchmark_amount")),
            ("Referanse %", _reference_pct_text(calculation_payload)),
            ("Referanse beløp", _reference_amount_text(calculation_payload)),
            ("Total vesentlighet", _mapping_value(calculation_payload, "overall_materiality")),
            ("PM % av total", _mapping_value(calculation_payload, "pm_pct")),
            ("PM", _mapping_value(calculation_payload, "performance_materiality")),
            ("Grense ubet feil % av PM", _mapping_value(calculation_payload, "trivial_pct")),
            ("ClearlyTriv", _mapping_value(calculation_payload, "clearly_trivial")),
        ],
    )

    ref_ws = wb.create_sheet("Referanseverdier")
    _write_reference_sheet(ref_ws, benchmark_amounts or {}, calculation_payload)

    _format_sheet(ws)
    _format_sheet(ref_ws)
    wb.save(out)
    return out


def _write_meta_section(ws, row: int, title: str, items: list[tuple[str, Any]]) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    title_cell = ws.cell(row, 1, title)
    title_cell.font = Font(bold=True, color=_TEXT_COLOR)
    title_cell.fill = _SECTION_FILL
    title_cell.alignment = Alignment(horizontal="left")

    row += 1
    for idx, (label, value) in enumerate(items):
        ws.cell(row, 1, label).font = Font(bold=False, color=_MUTED_COLOR)
        val_cell = ws.cell(row, 2, _normalize_cell_value(value))
        if idx % 2 == 0:
            ws.cell(row, 1).fill = _ALT_FILL
            val_cell.fill = _ALT_FILL
        if isinstance(val_cell.value, (int, float)):
            val_cell.number_format = "#,##0.0" if _is_percent_label(label) else "#,##0"
            val_cell.alignment = Alignment(horizontal="right")
        else:
            val_cell.alignment = Alignment(horizontal="left")
        row += 1
    return row


def _write_reference_sheet(ws, benchmark_amounts: Mapping[str, float], calculation_payload: Mapping[str, Any] | None) -> None:
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "Referanseverdier og benchmarkgrunnlag"
    title.font = Font(bold=True, color="FFFFFF", size=14)
    title.fill = _TITLE_FILL

    headers = ["Benchmark", "Grunnlag", "Fra %", "Til %", "Fra beløp", "Til beløp", "Valgt"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color=_TEXT_COLOR)
        cell.fill = _SECTION_FILL
        cell.alignment = Alignment(horizontal="center")

    selected_key = str((calculation_payload or {}).get("benchmark_key") or "").strip()
    row = 4
    for key in BENCHMARK_KEYS:
        amount = abs(float(benchmark_amounts.get(key) or 0.0))
        om_pct, pm_pct, trivial_pct = get_default_percentages(key)
        calc = calculate_materiality(key, amount, om_pct=om_pct, pm_pct=pm_pct, trivial_pct=trivial_pct)

        values = [
            BENCHMARK_LABELS.get(key, key),
            amount,
            calc.reference_pct_low / 100.0,
            calc.reference_pct_high / 100.0,
            calc.reference_amount_low,
            calc.reference_amount_high,
            "Ja" if key == selected_key else "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            if row % 2 == 0:
                cell.fill = _ALT_FILL
            if col in {2, 5, 6} and isinstance(value, (int, float)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col in {3, 4} and isinstance(value, (int, float)):
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
        row += 1


def _format_sheet(ws) -> None:
    thin = Side(style="thin", color="D7D1C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell.border = border
                if cell.alignment is None or not cell.alignment.horizontal:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    widths = [28, 22, 22, 14, 14, 14, 12]
    for idx in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths[idx - 1] if idx <= len(widths) else 18

    ws.freeze_panes = "A3"


def _normalize_cell_value(value: Any) -> Any:
    if value in (None, ""):
        return "-"
    return value


def _is_percent_label(label: str) -> bool:
    return "%" in label


def _display_source_label(payload: Mapping[str, Any] | None) -> str:
    if not isinstance(payload, Mapping):
        return "-"
    source = str(payload.get("source") or "").strip()
    if source == "crmsystem":
        return "CRMSystem"
    if source == "local_calculation":
        return "Lokal beregning"
    return source or "-"


def _mapping_value(payload: Mapping[str, Any] | None, key: str, default: Any = "-") -> Any:
    if not isinstance(payload, Mapping):
        return default
    value = payload.get(key, default)
    return default if value in (None, "") else value


def _benchmark_label(payload: Mapping[str, Any] | None) -> str:
    key = str((payload or {}).get("benchmark_key") or "").strip()
    return BENCHMARK_LABELS.get(key, key or "-")


def _reference_pct_text(payload: Mapping[str, Any] | None) -> str:
    if not isinstance(payload, Mapping):
        return "-"
    low = payload.get("reference_pct_low")
    high = payload.get("reference_pct_high")
    try:
        return f"{float(low):.1f}% - {float(high):.1f}%"
    except Exception:
        return "-"


def _reference_amount_text(payload: Mapping[str, Any] | None) -> str:
    if not isinstance(payload, Mapping):
        return "-"
    low = payload.get("reference_amount_low")
    high = payload.get("reference_amount_high")
    try:
        return f"{float(low):,.0f} - {float(high):,.0f}".replace(",", " ")
    except Exception:
        return "-"


def _crm_lookup_value(crm_lookup: object | None, key: str, default: Any = "-") -> Any:
    if crm_lookup is None:
        return default
    value = getattr(crm_lookup, key, default)
    return default if value in (None, "") else str(value)


def _crm_record_value(crm_lookup: object | None, key: str, default: Any = "-") -> Any:
    record = getattr(crm_lookup, "record", None) if crm_lookup is not None else None
    if record is None:
        return default
    value = getattr(record, key, default)
    return default if value in (None, "") else value


def _crm_match_text(crm_lookup: object | None) -> str:
    if crm_lookup is None:
        return "-"
    record = getattr(crm_lookup, "record", None)
    if record is None:
        return "-"
    matched = getattr(crm_lookup, "matched_client_number", None)
    current = getattr(record, "client_number", None)
    if matched and current and matched != current:
        return f"{matched} -> {current}"
    return str(current or "-")
