from __future__ import annotations

"""Fane: "Valgte kontoer (populasjon)" i motpost-arbeidspapiret.

Hvorfor egen fane?
- Oversikt skal være en ren "dashboard" uten brede tabeller.
- Tabellen med valgte kontoer har ofte mange kolonner og lange kontonavn.

Inneholder:
- Kort key/value-blokk (retning, scope, lenke tilbake)
- Tabell med valgte kontoer (Kredit/Debet/Netto + avstemming)

NB: Dette er presentasjonslogikk. Beregninger skjer primært i :mod:`motpost.konto_core`.
"""

from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook

from .common import _set_cell, _write_df_table, _write_kv_sheet


SHEET_NAME_VALGTE_KONTOER = "Valgte kontoer (populasjon)"


def _build_valgte_kontoer_df(
    *,
    data: Any,
    df_scope: pd.DataFrame,
    selected_accounts: list[str],
    konto_navn_map: Mapping[str, str],
) -> pd.DataFrame:
    """Bygg en "print-DF" for tabellen på fanen "Valgte kontoer".

    Input:
      - data.df_selected (pivot fra core)
      - df_scope (alle linjer i scope)

    Output:
      - DF med kolonner (best effort):
        Konto, Kontonavn, Kredit, Debet, Netto, Sum valgte kontoer, Andel av valgt, Antall bilag
    """

    df_selected = getattr(data, "df_selected", pd.DataFrame())
    if df_selected is None or df_selected.empty:
        return pd.DataFrame()

    df_sel_print = df_selected.copy()

    # Standardiser kolonnenavn
    if "Sum" in df_sel_print.columns and "Sum valgte kontoer" not in df_sel_print.columns:
        df_sel_print = df_sel_print.rename(columns={"Sum": "Sum valgte kontoer"})

    if "Kontonavn" not in df_sel_print.columns:
        try:
            df_sel_print["Kontonavn"] = df_sel_print.get("Konto", "").map(
                lambda k: konto_navn_map.get(str(k), "")
            )
        except Exception:
            df_sel_print["Kontonavn"] = ""

    # Andel (0-1) basert på absolutt beløp
    total_abs = 0.0
    try:
        total_abs = float(df_sel_print["Sum valgte kontoer"].astype(float).abs().sum())
    except Exception:
        total_abs = 0.0

    if total_abs:
        try:
            df_sel_print["Andel av valgt"] = (
                df_sel_print["Sum valgte kontoer"].astype(float).abs() / total_abs
            )
        except Exception:
            df_sel_print["Andel av valgt"] = 0.0
    else:
        df_sel_print["Andel av valgt"] = 0.0

    # Kredit/Debet/Netto per valgt konto (i scope) – avstemming.
    # Beløp følger hovedbok-konvensjon: kreditlinjer er negative beløp, debetlinjer positive.
    try:
        if (
            df_scope is not None
            and not df_scope.empty
            and selected_accounts
            and "Beløp" in df_scope.columns
        ):
            konto_col_scope = "Konto_str" if "Konto_str" in df_scope.columns else "Konto"
            belop_num = pd.to_numeric(df_scope["Beløp"], errors="coerce").fillna(0.0)
            konto_s = df_scope[konto_col_scope].astype(str)

            sel_set = {str(a).strip() for a in selected_accounts if str(a).strip()}
            sel_mask = konto_s.isin(sel_set)

            if sel_mask.any():
                konto_sel = konto_s[sel_mask]
                belop_sel = belop_num[sel_mask]

                kredit_by = belop_sel.where(belop_sel < 0, 0.0).groupby(konto_sel).sum().to_dict()
                debet_by = belop_sel.where(belop_sel > 0, 0.0).groupby(konto_sel).sum().to_dict()
                netto_by = belop_sel.groupby(konto_sel).sum().to_dict()
            else:
                kredit_by = {}
                debet_by = {}
                netto_by = {}

            konto_key = df_sel_print["Konto"].astype(str)
            df_sel_print["Kredit"] = konto_key.map(kredit_by).fillna(0.0).astype(float)
            df_sel_print["Debet"] = konto_key.map(debet_by).fillna(0.0).astype(float)
            df_sel_print["Netto"] = konto_key.map(netto_by).fillna(0.0).astype(float)
        else:
            df_sel_print["Kredit"] = 0.0
            df_sel_print["Debet"] = 0.0
            df_sel_print["Netto"] = 0.0
    except Exception:
        df_sel_print["Kredit"] = 0.0
        df_sel_print["Debet"] = 0.0
        df_sel_print["Netto"] = 0.0

    # Kolonneordre
    keep = [
        c
        for c in [
            "Konto",
            "Kontonavn",
            "Kredit",
            "Debet",
            "Netto",
            "Sum valgte kontoer",
            "Andel av valgt",
            "Antall bilag",
        ]
        if c in df_sel_print.columns
    ]
    df_sel_print = df_sel_print[keep]

    # Sorter etter absolutt beløp
    if "Sum valgte kontoer" in df_sel_print.columns:
        df_sel_print = df_sel_print.sort_values(
            by="Sum valgte kontoer",
            key=lambda s: s.astype(float).abs(),
            ascending=False,
        )

    return df_sel_print


def write_valgte_kontoer_sheet(
    wb: Workbook,
    *,
    data: Any,
    df_scope: pd.DataFrame,
    selected_accounts: list[str],
    direction: str,
    konto_navn_map: Mapping[str, str],
) -> None:
    """Opprett og skriv fanen "Valgte kontoer (populasjon)"."""

    sheet_name = SHEET_NAME_VALGTE_KONTOER

    # Opprett ark (eller hent eksisterende)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    selected_sum = float(getattr(data, "selected_sum", 0.0) or 0.0)
    bilag_count = int(getattr(data, "bilag_count", 0) or 0)

    kv = [
        ("Til oversikt", "=HYPERLINK(\"#'Oversikt'!A1\",\"Oversikt\")"),
        ("Retning (valgte kontoer)", direction),
        ("Antall valgte kontoer", int(len(selected_accounts))),
        ("Antall bilag i scope", bilag_count),
        ("Sum valgte kontoer (retning)", selected_sum),
    ]

    next_row = _write_kv_sheet(
        ws,
        sheet_name,
        kv,
        key_col_width=22,
        value_col_width=32,
        apply_column_widths=True,
    )

    # Format for sum-valgte
    try:
        for r in range(1, ws.max_row + 1):
            k = ws.cell(row=r, column=1).value
            if k == "Sum valgte kontoer (retning)":
                ws.cell(row=r, column=2).number_format = "#,##0.00"
    except Exception:
        pass

    df_sel_print = _build_valgte_kontoer_df(
        data=data,
        df_scope=df_scope,
        selected_accounts=selected_accounts,
        konto_navn_map=konto_navn_map,
    )

    if df_sel_print is None or df_sel_print.empty:
        _set_cell(ws, next_row, 1, "(ingen valgte kontoer i scope)")
        return

    _write_df_table(
        ws,
        df_sel_print,
        f"Valgte kontoer ({direction})",
        start_row=next_row,
        start_col=1,
        add_summary_row=True,
        max_col_width=32,
    )
