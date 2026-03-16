from __future__ import annotations

"""Kombinasjoner-fanen i motpost-arbeidspapiret.

Denne fanen er "hovedtabellen":
- kombinasjon (motkontoer i samme bilag)
- summer på valgte kontoer (retning)
- netto valgt retning (kun bilag med overvekt i valgt retning)
- status (Forventet/Outlier/Umerket) + kommentar fra GUI

NB: Flere tester forventer at tabellen heter `TKombinasjoner`.
"""

from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from ..combo_workflow import (
    build_combo_totals_df,
    combo_display_name,
    compute_selected_net_sum_by_combo,
    normalize_combo_status,
    status_label,
    status_sort_key,
)

from .common import _write_df_table


def build_and_write_kombinasjoner_sheet(
    wb: Workbook,
    *,
    df_scope: pd.DataFrame,
    selected_accounts: list[str],
    direction: str,
    konto_navn_map: Mapping[str, str],
    status_map: Mapping[str, str],
    comment_map: Mapping[str, str],
    net_col: str,
) -> pd.DataFrame:
    ws_combo = wb.create_sheet("Kombinasjoner")

    df_combos = build_combo_totals_df(
        df_scope,
        selected_accounts,
        selected_direction=direction,
    )
python
    if not df_combos.empty:
        df_combos = df_combos.copy()

        # Netto-kolonne for valgte kontoer.
        # Når retning=Kredit/Debet ønsker vi "overvekt" i valgt retning:
        # - Kredit: bilag med netto debet på valgte kontoer bidrar 0
        # - Debet:  bilag med netto kredit på valgte kontoer bidrar 0
        try:
            net_map = compute_selected_net_sum_by_combo(
                df_scope,
                selected_accounts,
                selected_direction=direction,
            )
            net_values = df_combos["Kombinasjon"].astype(str).map(
                lambda c: float(net_map.get(str(c).strip(), 0.0) or 0.0)
            )

            if net_col not in df_combos.columns:
                try:
                    insert_at = list(df_combos.columns).index("Sum valgte kontoer") + 1
                except ValueError:
                    insert_at = len(df_combos.columns)
                df_combos.insert(insert_at, net_col, net_values)
        except Exception:
            # Best effort: net-kolonnen er kun et hjelpetall.
            pass

        # Legg til kontonavn for kombinasjonen (lesbarhet i revisjonsdokumentasjon)
        try:
            uniq = df_combos["Kombinasjon"].astype(str).fillna("").unique().tolist()
            combo_name_map = {c: combo_display_name(c, konto_navn_map) for c in uniq if str(c).strip()}
            df_combos.insert(
                list(df_combos.columns).index("Kombinasjon") + 1,
                "Kombinasjon (navn)",
                df_combos["Kombinasjon"].astype(str).map(lambda c: combo_name_map.get(str(c).strip(), "")),
            )
        except Exception:
            pass

        df_combos["Status"] = df_combos["Kombinasjon"].map(
            lambda c: status_label(status_map.get(str(c), ""), neutral_label="Umerket")
        )

        # Kommentar per kombinasjon (fra GUI) – tas med i dokumentasjonen
        try:
            df_combos.insert(
                list(df_combos.columns).index("Status") + 1,
                "Kommentar",
                df_combos["Kombinasjon"].astype(str).map(lambda c: comment_map.get(str(c).strip(), "")),
            )
        except Exception:
            df_combos["Kommentar"] = df_combos["Kombinasjon"].astype(str).map(
                lambda c: comment_map.get(str(c).strip(), "")
            )

        # Sorter så Outlier kommer først (stabil)
        df_combos["_status_order"] = df_combos["Kombinasjon"].map(
            lambda c: status_sort_key(status_map.get(str(c), ""))
        )
        df_combos = (
            df_combos.sort_values(by=["_status_order", "Antall bilag", "Kombinasjon #"], ascending=[True, False, True])
            .drop(columns=["_status_order"])
        )

    _write_df_table(ws_combo, df_combos, "Kombinasjoner")

    # Frysbokser: behold tittel + header synlig ved scrolling
    try:
        ws_combo.freeze_panes = "A4"
    except Exception:
        pass

    # Fargekoding av rader i kombinasjonstabellen (best effort)
    try:
        if df_combos is not None and not df_combos.empty and "Status" in df_combos.columns:
            status_col = list(df_combos.columns).index("Status") + 1
            start_row = 4  # data starter etter tittel(1) + blank(2) + header(3)
            end_row = start_row + len(df_combos) - 1
            fill_expected = PatternFill("solid", fgColor="C6EFCE")
            fill_outlier = PatternFill("solid", fgColor="FFF2CC")
            for r in range(start_row, end_row + 1):
                v = ws_combo.cell(row=r, column=status_col).value
                if v == "Forventet":
                    fill = fill_expected
                elif v == "Outlier":
                    fill = fill_outlier
                else:
                    continue
                for c in range(1, len(df_combos.columns) + 1):
                    ws_combo.cell(row=r, column=c).fill = fill
    except Exception:
        pass

    return df_combos
