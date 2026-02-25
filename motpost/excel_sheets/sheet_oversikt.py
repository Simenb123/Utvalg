"""Bygger "Oversikt"-arket i motpostanalyse-eksporten.

Oversikt er et presentasjons-/arbeidspapir-ark:
- Kompakt parameterblokk (start på rad 3)
- Statusoppsummering + outlier-indeks
- Arbeidspapirfelt (arbeidsvesentlighet + konklusjon)

Finpuss:
- Beløp vises med tusenskille og uten desimaler på Oversikt
- Handling/konklusjon stopper i kolonne H
- Outlier-indeks starter i kolonne A
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .common import (
    DEFAULT_INT_FORMAT,
    HEADER_FONT,
    THIN_BORDER,
    TITLE_FONT,
    hide_gridlines,
    set_column_widths,
    _write_df_table,
)

_FILL_SECTION = PatternFill(patternType="solid", fgColor="FFE2EFDA")  # lys grønn
_FILL_INPUT = PatternFill(patternType="solid", fgColor="FFFFF2CC")  # lys gul


@dataclass
class OversiktLayout:
    next_row: int


def _amount_cell(ws: Worksheet, row: int, col: int, value: float | int | None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = DEFAULT_INT_FORMAT


def _is_money_like_header(header: str) -> bool:
    h = str(header or "").strip().lower()
    if not h:
        return False
    if h.startswith("antall"):
        return False
    if "andel" in h or "%" in h:
        return False
    return any(k in h for k in ("sum", "netto", "populasjon", "beløp"))


def _force_int_format_on_table(
    ws: Worksheet,
    *,
    first_data_row: int,
    last_row: int,
    start_col: int,
    headers: list[str],
) -> None:
    money_cols = [i for i, h in enumerate(headers) if _is_money_like_header(h)]
    for idx in money_cols:
        col = start_col + idx
        for r in range(first_data_row, last_row + 1):
            ws.cell(row=r, column=col).number_format = DEFAULT_INT_FORMAT




def _ensure_wrap_text(ws: Worksheet, *, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    """Sikre wrap_text=True på et område uten å endre annen alignment."""

    for rr in range(start_row, end_row + 1):
        for cc in range(start_col, end_col + 1):
            cell = ws.cell(row=rr, column=cc)
            try:
                cell.alignment = cell.alignment.copy(wrap_text=True)
            except Exception:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=True,
                )


def write_oversikt_sheet(
    ws: Worksheet,
    *,
    data: Any,
    direction: str,
    selected_accounts: list[str],
    selected_sum: float,
    population_net: float,
    sum_label: str,
    population_label: str,
    net_label: str,
    df_status: pd.DataFrame,
    df_outlier_index: pd.DataFrame,
    outlier_sheet_name: str,
) -> OversiktLayout:
    hide_gridlines(ws)

    # Outlier-indeksen kan starte i kolonne A, derfor er A ikke bare "marg".
    set_column_widths(
        ws,
        {
            "A": 6.0,
            "B": 35.0,
            "C": 55.0,
            "D": 20.0,
            "E": 16.0,
            "F": 14.0,
            "G": 14.0,
            "H": 40.0,
            "I": 4.0,
            "J": 4.0,
        },
    )

    ws["A1"].value = "Motpostanalyse - Oversikt"
    ws["A1"].font = TITLE_FONT

    # Standard frys etter rad 1 (ingen kolonnefrys)
    ws.freeze_panes = "A2"

    # Hurtignavigasjon
    ws["E3"].value = '=HYPERLINK("#\'Data\'!A1","Gå til Data")'
    ws["E3"].font = Font(color="0000FF", underline="single")
    ws["E4"].value = f'=HYPERLINK("#\'{outlier_sheet_name}\'!A1","Gå til Outliers")'
    ws["E4"].font = Font(color="0000FF", underline="single")

    _write_handling_box(ws, start_row=6, start_col=5)

    # Parametre (kompakt) – start på rad 3
    generated = datetime.now().strftime("%d.%m.%Y %H:%M")

    bilag_count = getattr(data, "bilag_count", None)
    row_count = getattr(data, "row_count", None)
    try:
        df_scope = getattr(data, "df_scope", None)
        if df_scope is not None:
            if bilag_count is None and "Bilag_str" in df_scope.columns:
                bilag_count = int(df_scope["Bilag_str"].nunique())
            if row_count is None:
                row_count = int(len(df_scope))
    except Exception:
        pass

    param_rows: list[tuple[str, Any]] = [
        ("Generert", generated),
        ("Retning (valgte kontoer)", direction),
        ("Antall valgte kontoer", int(len(selected_accounts))),
        ("Antall bilag", int(bilag_count or 0)),
        ("Antall rader", int(row_count or 0)),
        (population_label, float(selected_sum)),
    ]

    r = 3
    for k, v in param_rows:
        ck = ws.cell(row=r, column=2, value=k)
        ck.font = HEADER_FONT
        ck.alignment = Alignment(vertical="top")

        cv = ws.cell(row=r, column=3, value=v)
        cv.alignment = Alignment(vertical="top")
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            cv.number_format = DEFAULT_INT_FORMAT
            cv.alignment = Alignment(horizontal="right", vertical="top")
        # Linjemarkeringer (tynne rutenettlinjer) rundt parameterblokken
        ck.border = THIN_BORDER
        cv.border = THIN_BORDER

        r += 1

    # Statusoppsummering
    start_status = r + 2
    res_status = _write_df_table(
        ws,
        df_status,
        "Oversikt forventet / ikke forventet",
        start_row=start_status,
        start_col=2,
        add_summary_row=True,
        auto_width=False,
    )
    _force_int_format_on_table(
        ws,
        first_data_row=res_status.first_data_row,
        last_row=res_status.last_row,
        start_col=res_status.start_col,
        headers=[str(c) for c in (df_status.columns.tolist() if df_status is not None else [])],
    )

    next_row = res_status.last_row + 2

    # Outlier-indeks (kun outliers) – start i kolonne A
    if df_outlier_index is not None and not df_outlier_index.empty:
        res_out = _write_df_table(
            ws,
            df_outlier_index,
            "Outliers (ikke forventet) – dokumenter i egne faner",
            start_row=next_row,
            start_col=1,
            add_summary_row=False,
            auto_width=False,
        )
        _force_int_format_on_table(
            ws,
            first_data_row=res_out.first_data_row,
            last_row=res_out.last_row,
            start_col=res_out.start_col,
            headers=[str(c) for c in (df_outlier_index.columns.tolist() if df_outlier_index is not None else [])],
        )
        # Wrap text i outlier-indeksen slik at kolonner kan være smalere
        _ensure_wrap_text(
            ws,
            start_row=res_out.header_row,
            end_row=res_out.last_row,
            start_col=res_out.start_col,
            end_col=res_out.last_col,
        )

        next_row = res_out.last_row + 2

    rest_population = _extract_rest_population(df_status)
    return _write_materiality_and_conclusion(ws, start_row=next_row, rest_population=rest_population)


def _write_handling_box(ws: Worksheet, *, start_row: int, start_col: int) -> None:
    """Handling-boks på høyre side. Stopper i kolonne H."""

    end_col = 8  # H

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    header = ws.cell(row=start_row, column=start_col, value="Handling")
    header.font = HEADER_FONT
    header.fill = _FILL_SECTION
    header.alignment = Alignment(vertical="center")

    lines = [
        "1. Opparbeid en forståelse av uforventede kombinasjoner og dokumenter vurderingen i fanene #x.",
        "2. Vurder om det er relevant å detaljteste på bilagsnivå",
        "3. Oppsummer og konkluder",
    ]

    for i, text in enumerate(lines, start=1):
        r = start_row + i
        ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=end_col)
        cell = ws.cell(row=r, column=start_col, value=text)
        cell.alignment = Alignment(vertical="top", wrap_text=False)


def _extract_rest_population(df_status: pd.DataFrame) -> float:
    try:
        if df_status is None or df_status.empty:
            return 0.0
        if "Status" not in df_status.columns:
            return 0.0
        m = df_status[df_status["Status"] == "ikke vesentlig (ikke markert)"]
        if m.empty:
            return 0.0
        return float(m.iloc[0].get("Sum valgte kontoer") or 0.0)
    except Exception:
        return 0.0


def _write_materiality_and_conclusion(
    ws: Worksheet,
    *,
    start_row: int,
    rest_population: float,
) -> OversiktLayout:
    r = start_row

    ws.cell(row=r, column=2, value="Arbeidsvesentlighetsgrense").font = HEADER_FONT

    thr = ws.cell(row=r, column=3)
    thr.value = None
    thr.number_format = DEFAULT_INT_FORMAT
    thr.fill = _FILL_INPUT
    thr.border = THIN_BORDER
    thr.alignment = Alignment(horizontal="right")

    hint = ws.cell(row=r, column=4, value="\u2190 legg inn beløp")
    hint.font = Font(italic=True, color="666666")
    hint.alignment = Alignment(vertical="top")

    try:
        dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
        dv.errorTitle = "Ugyldig verdi"
        dv.error = "Skriv inn et beløp (>= 0) eller la feltet stå tomt."
        dv.promptTitle = "Arbeidsvesentlighetsgrense"
        dv.prompt = "Legg inn tolererbar feil (beløp)."
        ws.add_data_validation(dv)
        dv.add(thr)
    except Exception:
        pass

    r += 1
    ws.cell(row=r, column=2, value="Restpopulasjon").font = HEADER_FONT
    _amount_cell(ws, r, 3, float(rest_population))
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")

    r += 1
    ws.cell(row=r, column=2, value="differanse").font = HEADER_FONT
    thr_addr = ws.cell(row=start_row, column=3).coordinate
    rest_addr = ws.cell(row=start_row + 1, column=3).coordinate
    diff = ws.cell(row=r, column=3)
    diff.value = f'=IF({thr_addr}="","",{thr_addr}-ABS({rest_addr}))'
    diff.number_format = DEFAULT_INT_FORMAT
    diff.border = THIN_BORDER
    diff.alignment = Alignment(horizontal="right")

    r += 2
    header = ws.cell(row=r, column=2, value="Konklusjon")
    header.font = HEADER_FONT
    header.fill = _FILL_SECTION

    box_top = r + 1
    box_bottom = box_top + 6

    # Konklusjonsfeltet stopper i kolonne H (8)
    ws.merge_cells(start_row=box_top, start_column=2, end_row=box_bottom, end_column=8)
    cell = ws.cell(row=box_top, column=2, value="Sum restpopulasjon er under grensen for tolererbar feil. Ytterligere handlinger vurderes ikke hensiktsmessig.")
    cell.font = Font(italic=True, color="666666")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    fill_box = PatternFill(patternType="solid", fgColor="FFFFFFFF")
    for rr in range(box_top, box_bottom + 1):
        for cc in range(2, 9):
            c = ws.cell(row=rr, column=cc)
            c.border = THIN_BORDER
            c.fill = fill_box
        ws.row_dimensions[rr].height = 22

    return OversiktLayout(next_row=box_bottom + 2)


def add_navigation_links(ws_over: Worksheet) -> None:
    """Bakoverkompatibilitet: delegert "Innhold"-liste."""

    try:
        from .oversikt_nav import add_navigation_links_impl

        add_navigation_links_impl(ws_over)
    except Exception:
        return
