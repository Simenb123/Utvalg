"""Legacy/kompatibilitet: navigasjonslenker på Oversikt.

I nyere mal brukes normalt kun faste "Gå til"-lenker.
Denne implementasjonen beholdes likevel for bakoverkompatibilitet.
"""

from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


def add_navigation_links_impl(ws_over: Worksheet) -> None:
    """Skriv en enkel "Innhold"-liste i kolonne J (best-effort)."""

    wb = ws_over.parent

    try:
        link_items: list[tuple[str, str]] = []

        # Legacy/kompatibilitet: "Valgte kontoer (populasjon)"
        try:
            from .sheet_valgte_kontoer import SHEET_NAME_VALGTE_KONTOER

            if (
                SHEET_NAME_VALGTE_KONTOER in wb.sheetnames
                and wb[SHEET_NAME_VALGTE_KONTOER].sheet_state != "hidden"
            ):
                link_items.append((SHEET_NAME_VALGTE_KONTOER, SHEET_NAME_VALGTE_KONTOER))
        except Exception:
            pass

        # Foretrukket rekkefølge i arbeidsflyt
        for sheet_name in (
            "Data",
            "Kombinasjoner",
            "Outlier-kombinasjoner",
            "Outlier - alle transaksjoner",
            "Outlier – Full bilagsutskrift",
            "Outlier - Full bilagsutskrift",
        ):
            if sheet_name in wb.sheetnames and wb[sheet_name].sheet_state != "hidden":
                if all(existing_sheet != sheet_name for _lbl, existing_sheet in link_items):
                    link_items.append((sheet_name, sheet_name))

        if not link_items:
            return

        nav_row = 1
        nav_col = 10  # J
        nav_col_letter = "J"

        ws_over.cell(row=nav_row, column=nav_col, value="Innhold").font = Font(bold=True)

        # Ikke krymp bredde hvis arket allerede har satt den bredere.
        dim = ws_over.column_dimensions[nav_col_letter]
        current_w = dim.width
        target_w = 22
        if current_w is None or float(current_w) < target_w:
            dim.width = target_w

        for i, (label, sheet_name) in enumerate(link_items, start=1):
            cell = ws_over.cell(row=nav_row + i, column=nav_col, value=label)
            cell.hyperlink = f"#'{sheet_name}'!A1"
            cell.style = "Hyperlink"
    except Exception:
        return
