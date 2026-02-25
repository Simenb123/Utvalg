"""Outlier-ark og outlier-detaljfaner (#n).

Dette er "arbeidspapir"-versjonen av eksporten:
- Outlier - alle transaksjoner: full bilagsutskrift for outlier-bilag.
- #<n>: én fane per outlier-kombinasjon, med:
  - nøkkelfelt (A)
  - bilagsoppsummering (B)
  - handling + resultat (C)

Faner for outliers skal alltid vises (outlier-transaksjonsarket alltid, selv om det ikke er outliers).
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Alignment

from ..combinations import build_bilag_to_motkonto_combo
from ..utils import _bilag_str, _konto_str
from .common import (
    DEFAULT_DATE_FORMAT,
    DEFAULT_INT_FORMAT,
    DEFAULT_MONEY_FORMAT,
    HEADER_FONT,
    TAB_OUTLIER_YELLOW,
    THIN_BORDER,
    TITLE_FONT,
    hide_gridlines,
    set_column_widths,
    set_tab_color,
    _write_df_table,
)


OUTLIER_LINES_COLUMNS = [
    "Status",
    "Kombinasjon",
    "Kombinasjon (navn)",
    "Bilag",
    "Dato",
    "Tekst",
    "Konto",
    "Kontonavn",
    "Debet",
    "Kredit",
    "Beløp",
]

# Excel begrensning: maks 1 048 576 rader totalt per ark.
# Outlier-tabellen bruker rad 1 (tittel), rad 2 (notat), rad 3 (header), og data starter på rad 4.
EXCEL_MAX_ROWS = 1_048_576
OUTLIER_HEADER_ROWS = 3
OUTLIER_MAX_DATA_ROWS = EXCEL_MAX_ROWS - OUTLIER_HEADER_ROWS



@dataclass
class OutlierFrames:
    df_lines: pd.DataFrame
    combo_first_row: dict[str, int]
    bilag_first_row: dict[tuple[str, str], int]
    bilag_to_combo: dict[str, str]
    excluded_blank_motpost_count: int = 0

    # Hvor mange transaksjonslinjer som *ville* blitt eksportert i outlier-arket
    # (brukes til notat og UI-beslutning).
    outlier_lines_count: int = 0

    # Om bilagslinjer faktisk er inkludert i eksporten.
    transactions_included: bool = True

    # Hvis bilagslinjer er utelatt automatisk (f.eks. pga Excel-radgrense), legg inn forklaring.
    transactions_omitted_reason: str = ""


def build_outlier_frames(
    df_scope: pd.DataFrame,
    *,
    selected_accounts: list[str],
    outlier_combos: list[str],
    combo_name_map: dict[str, str],
    include_transactions: bool = True,
    max_data_rows: int = OUTLIER_MAX_DATA_ROWS,
) -> OutlierFrames:
    """Bygg DataFrame for outlier-transaksjoner + mapper for hyperlinking.

    `include_transactions=False` brukes for å lage en "lett" eksport som ikke
    tar med full bilagsutskrift (og dermed unngår enorme Excel-filer).

    `max_data_rows` beskytter mot Excel sin radbegrensning.
    """

    empty = pd.DataFrame(columns=OUTLIER_LINES_COLUMNS)

    if df_scope is None or df_scope.empty or not outlier_combos:
        return OutlierFrames(
            df_lines=empty,
            combo_first_row={},
            bilag_first_row={},
            bilag_to_combo={},
            excluded_blank_motpost_count=0,
            outlier_lines_count=0,
            transactions_included=bool(include_transactions),
        )

    # Sørg for bilag/konto som str (bruk eksisterende *_str hvis mulig)
    bilag_s = df_scope["Bilag_str"].astype(str) if "Bilag_str" in df_scope.columns else df_scope.get("Bilag", pd.Series([], dtype=object)).map(_bilag_str)
    konto_s = df_scope["Konto_str"].astype(str) if "Konto_str" in df_scope.columns else df_scope.get("Konto", pd.Series([], dtype=object)).map(_konto_str)

    # Bygg bilag -> kombinasjon for hele scope (vektorisert)
    bilag_to_combo = build_bilag_to_motkonto_combo(df_scope, selected_accounts)

    outlier_combo_set = {str(c) for c in (outlier_combos or []) if str(c).strip()}
    out_bilag = {b for b, combo in bilag_to_combo.items() if combo in outlier_combo_set}

    # Ekskluder blanke bilagsnummer fra full bilagsutskrift (gir svært store grupper ved import-feil)
    blank = {b for b in out_bilag if not str(b).strip()}
    excluded_blank = int(len(blank))
    out_bilag = {b for b in out_bilag if str(b).strip()}

    if not out_bilag:
        return OutlierFrames(
            df_lines=empty,
            combo_first_row={},
            bilag_first_row={},
            bilag_to_combo=bilag_to_combo,
            excluded_blank_motpost_count=excluded_blank,
            outlier_lines_count=0,
            transactions_included=bool(include_transactions),
        )

    # Antall linjer som ville bli skrevet i outlier-arket (uten å kopiere alle rader)
    outlier_lines_count = int(bilag_s.isin(out_bilag).sum())

    # Beskyttelse: Excel radgrense
    if include_transactions and outlier_lines_count > int(max_data_rows):
        return OutlierFrames(
            df_lines=empty,
            combo_first_row={},
            bilag_first_row={},
            bilag_to_combo=bilag_to_combo,
            excluded_blank_motpost_count=excluded_blank,
            outlier_lines_count=outlier_lines_count,
            transactions_included=False,
            transactions_omitted_reason=(
                f"For mange rader for Excel (beregnet {outlier_lines_count} > maks {int(max_data_rows)}). "
                "Eksporter uten bilagslinjer, eller filtrer scope mer."
            ),
        )

    if not include_transactions or outlier_lines_count == 0:
        return OutlierFrames(
            df_lines=empty,
            combo_first_row={},
            bilag_first_row={},
            bilag_to_combo=bilag_to_combo,
            excluded_blank_motpost_count=excluded_blank,
            outlier_lines_count=outlier_lines_count,
            transactions_included=False,
        )

    # Full bilagsutskrift: filtrer scope og bygg print-DF
    mask = bilag_s.isin(out_bilag)
    df_out = df_scope.loc[mask].copy()

    # Sikkerhetskolonner (ikke endre original df_scope)
    if "Bilag_str" not in df_out.columns:
        df_out["Bilag_str"] = bilag_s.loc[mask].astype(str).values
    if "Konto_str" not in df_out.columns:
        df_out["Konto_str"] = konto_s.loc[mask].astype(str).values

    # Dato-kolonne
    date_src = "Dato_dt" if "Dato_dt" in df_out.columns else ("Dato" if "Dato" in df_out.columns else None)

    if "Beløp_num" in df_out.columns and pd.api.types.is_numeric_dtype(df_out["Beløp_num"]):
        amount = df_out["Beløp_num"].astype(float)
    else:
        amount = pd.to_numeric(df_out.get("Beløp", 0), errors="coerce").fillna(0.0).astype(float)

    bilag_key = df_out["Bilag_str"].astype(str)

    df_print = pd.DataFrame(
        {
            "Status": "Outlier",
            "Kombinasjon": bilag_key.map(bilag_to_combo),
            "Kombinasjon (navn)": bilag_key.map(bilag_to_combo).map(combo_name_map),
            "Bilag": bilag_key,
            "Dato": df_out[date_src] if date_src else None,
            "Tekst": df_out.get("Tekst", ""),
            "Konto": df_out.get("Konto_str", df_out.get("Konto", "")),
            "Kontonavn": df_out.get("Kontonavn", ""),
            "Debet": amount.where(amount > 0, 0.0),
            "Kredit": (-amount).where(amount < 0, 0.0),
            "Beløp": amount,
        }
    )

    df_print = df_print[OUTLIER_LINES_COLUMNS]

    # Sortér for stabilitet (og penere eksport)
    df_print = df_print.sort_values(
        by=["Kombinasjon", "Bilag", "Dato", "Konto"],
        kind="mergesort",
    ).reset_index(drop=True)

    # Mapper: første rad i outlier-arket for hver kombinasjon/bilag
    combo_first: dict[str, int] = {}
    bilag_first: dict[tuple[str, str], int] = {}

    # Outlier-arket har header på rad 3, data starter rad 4
    base_row = 4
    for idx, row in df_print.iterrows():
        excel_row = base_row + int(idx)
        combo = str(row.get("Kombinasjon") or "")
        bilag = str(row.get("Bilag") or "")
        if combo and combo not in combo_first:
            combo_first[combo] = excel_row
        key = (combo, bilag)
        if combo and bilag and key not in bilag_first:
            bilag_first[key] = excel_row

    return OutlierFrames(
        df_lines=df_print,
        combo_first_row=combo_first,
        bilag_first_row=bilag_first,
        bilag_to_combo=bilag_to_combo,
        excluded_blank_motpost_count=excluded_blank,
        outlier_lines_count=int(len(df_print)),
        transactions_included=True,
    )
def write_outlier_transactions_sheet(
    wb: Workbook,
    *,
    frames: OutlierFrames,
    sheet_name: str,
) -> Worksheet:
    """Skriv "Outlier - alle transaksjoner"."""

    ws = wb.create_sheet(sheet_name)
    hide_gridlines(ws)
    set_tab_color(ws, TAB_OUTLIER_YELLOW)

    df = frames.df_lines
    if df is None or df.empty:
        df = pd.DataFrame(columns=OUTLIER_LINES_COLUMNS)

    _write_df_table(
        ws,
        df,
        "Outlier – Full bilagsutskrift",
        start_row=1,
        start_col=1,
        add_summary_row=False,
        max_col_width=45,
    )

    # Note / dokumentasjon
    note_parts: list[str] = []

    if not frames.transactions_included:
        if frames.outlier_lines_count:
            n = f"{int(frames.outlier_lines_count):,}".replace(",", " ")
            note_parts.append(f"Bilagslinjer er utelatt i eksport. (Beregnet antall linjer i full bilagsutskrift: {n}.)")
        else:
            note_parts.append("Bilagslinjer er utelatt i eksport.")
        if frames.transactions_omitted_reason:
            note_parts.append(str(frames.transactions_omitted_reason))
    elif frames.df_lines is None or frames.df_lines.empty:
        note_parts.append("Ingen outliers er markert i kombinasjonslisten.")

    if frames.excluded_blank_motpost_count:
        note_parts.append(
            f"NB: {int(frames.excluded_blank_motpost_count)} bilagsgrupper uten bilagsnummer er utelatt fra full bilagsutskrift."
        )

    if note_parts:
        ws["A2"].value = " ".join(note_parts)
        ws["A2"].alignment = Alignment(wrap_text=True)
    # Freeze på første datarad
    ws.freeze_panes = "A4"

    # Kolonnebredder (inspirert av eksempelmal)
    set_column_widths(
        ws,
        {
            "A": 14,
            "B": 22,
            "C": 52,
            "D": 12,
            "E": 11,
            "F": 45,
            "G": 10,
            "H": 35,
            "I": 14,
            "J": 14,
            "K": 14,
        },
    )

    # Datoformat
    # Header er rad 3, Dato-kolonnen er E
    for r in range(4, 4 + len(df)):
        ws.cell(row=r, column=5).number_format = DEFAULT_DATE_FORMAT

    return ws


def _first_nonempty_text(values: pd.Series) -> str:
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return ""


def _direction_mask(series: pd.Series, direction: str) -> pd.Series:
    d = str(direction or "").strip().lower()
    if d.startswith("k"):
        return series < 0
    if d.startswith("d"):
        return series > 0
    return pd.Series([True] * len(series), index=series.index)


def _build_bilag_summary_df(
    *,
    df_scope: pd.DataFrame,
    selected_accounts: list[str],
    direction: str,
    bilags_for_combo: list[str],
    bilag_first_row_map: dict[tuple[str, str], int],
    outlier_sheet_name: str,
    combo: str,
    include_outlier_transactions: bool,
    sum_label: str,
) -> pd.DataFrame:
    """Bilagsoppsummering per kombinasjon."""

    if not bilags_for_combo:
        return pd.DataFrame(columns=["Bilag", "Dato", "Tekst", sum_label, "Netto valgte kontoer", "Bilagslinjer"])

    sel_set = {str(_konto_str(x)) for x in selected_accounts}
    df_sel = df_scope[df_scope.get("Konto_str", "").astype(str).isin(sel_set)].copy()

    if df_sel.empty:
        return pd.DataFrame(columns=["Bilag", "Dato", "Tekst", sum_label, "Netto valgte kontoer", "Bilagslinjer"])

    amount = pd.to_numeric(df_sel.get("Beløp", 0), errors="coerce").fillna(0.0)
    dir_mask = _direction_mask(amount, direction)

    bilag_key = df_sel.get("Bilag_str")
    if bilag_key is None:
        bilag_key = df_sel.get("Bilag")

    bilag_key = bilag_key.astype(str)

    sum_dir = amount.where(dir_mask).groupby(bilag_key).sum()
    net_all = amount.groupby(bilag_key).sum()

    # Meta
    date_src = "Dato_dt" if "Dato_dt" in df_scope.columns else ("Dato" if "Dato" in df_scope.columns else None)
    meta_cols = ["Bilag_str"]
    if date_src:
        meta_cols.append(date_src)
    if "Tekst" in df_scope.columns:
        meta_cols.append("Tekst")

    meta = df_scope[meta_cols].copy()
    meta["Bilag_str"] = meta["Bilag_str"].astype(str)

    if date_src:
        meta_dates = meta.groupby("Bilag_str")[date_src].min()
    else:
        meta_dates = pd.Series(index=bilags_for_combo, dtype=object)

    if "Tekst" in meta.columns:
        meta_text = meta.groupby("Bilag_str")["Tekst"].apply(_first_nonempty_text)
    else:
        meta_text = pd.Series(index=bilags_for_combo, dtype=str)

    rows: list[dict[str, Any]] = []
    for bilag in bilags_for_combo:
        bilag_s = str(bilag)
        if include_outlier_transactions:
            rownum = bilag_first_row_map.get((combo, bilag_s), 1)
            link = f'=HYPERLINK("#\'{outlier_sheet_name}\'!A{rownum}","Gå til bilagslinjer")'
        else:
            link = "Utelatt i eksport"

        rows.append(
            {
                "Bilag": bilag_s,
                "Dato": meta_dates.get(bilag_s, None) if hasattr(meta_dates, "get") else None,
                "Tekst": meta_text.get(bilag_s, "") if hasattr(meta_text, "get") else "",
                sum_label: float(sum_dir.get(bilag_s, 0.0)) if hasattr(sum_dir, "get") else 0.0,
                "Netto valgte kontoer": float(net_all.get(bilag_s, 0.0)) if hasattr(net_all, "get") else 0.0,
                "Bilagslinjer": link,
            }
        )

    df_sum = pd.DataFrame(rows)

    # Sortér på bilag som nummer dersom mulig
    def _to_int(x: str) -> int:
        try:
            return int(float(x))
        except Exception:
            return 0

    df_sum["_bilag_sort"] = df_sum["Bilag"].map(_to_int)
    df_sum = df_sum.sort_values(by=["_bilag_sort"]).drop(columns=["_bilag_sort"]).reset_index(drop=True)

    return df_sum


def _write_text_box(
    ws: Worksheet,
    *,
    title_row: int,
    title: str,
    body: str,
    start_col: int = 1,
    end_col: int = 10,
    height_rows: int = 4,
) -> int:
    """Skriv en enkel tekstboks med tittel + merge."""

    ws.cell(row=title_row, column=start_col, value=title).font = HEADER_FONT

    top = title_row + 1
    bottom = top + height_rows - 1
    ws.merge_cells(start_row=top, start_column=start_col, end_row=bottom, end_column=end_col)
    cell = ws.cell(row=top, column=start_col, value=body)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for rr in range(top, bottom + 1):
        for cc in range(start_col, end_col + 1):
            ws.cell(row=rr, column=cc).border = THIN_BORDER

    return bottom + 1


def write_outlier_detail_sheets(
    wb: Workbook,
    *,
    df_kombinasjoner: pd.DataFrame,
    frames: OutlierFrames,
    df_scope: pd.DataFrame,
    selected_accounts: list[str],
    direction: str,
    sum_label: str,
    net_label: str,
    outlier_sheet_name: str,
    include_outlier_transactions: bool = True,
) -> list[str]:
    """Opprett #<n>-faner for outlier-kombinasjoner.

    Returnerer liste med arkfanenavn som ble opprettet.
    """

    if df_kombinasjoner is None or df_kombinasjoner.empty:
        return []

    required_cols = {"#", "Kombinasjon", "Kombinasjon (navn)", "Antall bilag", "Status"}
    missing = required_cols - set(df_kombinasjoner.columns)
    if missing:
        raise ValueError(f"df_kombinasjoner mangler kolonner: {sorted(missing)}")

    # Kun outliers
    df_out = df_kombinasjoner[df_kombinasjoner["Status"].astype(str).str.lower().isin({"ikke forventet", "outlier"})].copy()
    if df_out.empty:
        return []

    # Stabil sortering
    df_out = df_out.sort_values(by=["#"], kind="mergesort")

    created: list[str] = []

    for _, row in df_out.iterrows():
        combo_num = int(row["#"])
        combo = str(row["Kombinasjon"])
        combo_name = str(row.get("Kombinasjon (navn)", ""))
        ant_bilag = int(row.get("Antall bilag", 0) or 0)
        kommentar = str(row.get("Kommentar", "") or "")

        sum_val = float(row.get(sum_label, 0.0) or 0.0) if sum_label in row else 0.0
        net_val = float(row.get(net_label, 0.0) or 0.0) if net_label in row else 0.0

        sheet_title = f"#{combo_num}"
        ws = wb.create_sheet(sheet_title)
        hide_gridlines(ws)
        set_tab_color(ws, TAB_OUTLIER_YELLOW)

        # Kolonnebredder (inspirert av eksempelmal)
        set_column_widths(
            ws,
            {
                "A": 33,
                "B": 55,
                "C": 46,
                "D": 14,
                "E": 18,
                "F": 18,
                "G": 18,
                "H": 18,
                "I": 4,
                "J": 4,
            },
        )

        # Tittel
        ws["A1"].value = f"Kombinasjon #{combo_num}"
        ws["A1"].font = TITLE_FONT

        # Nøkkelfelt (A)
        ws["A3"].value = "Til oversikt"
        ws["A3"].font = HEADER_FONT
        ws["B3"].value = '=HYPERLINK("#\'Oversikt\'!A1","Gå til oversikt")'

        ws["A4"].value = "Status"
        ws["A4"].font = HEADER_FONT
        ws["B4"].value = "Outlier"

        ws["A5"].value = "Kombinasjon"
        ws["A5"].font = HEADER_FONT
        ws["B5"].value = combo

        ws["A6"].value = "Kombinasjon (navn)"
        ws["A6"].font = HEADER_FONT
        ws["B6"].value = combo_name

        ws["A7"].value = sum_label
        ws["A7"].font = HEADER_FONT
        ws["B7"].value = sum_val
        ws["B7"].number_format = DEFAULT_MONEY_FORMAT

        ws["A8"].value = net_label
        ws["A8"].font = HEADER_FONT
        ws["B8"].value = net_val
        ws["B8"].number_format = DEFAULT_MONEY_FORMAT

        ws["A9"].value = "Antall bilag"
        ws["A9"].font = HEADER_FONT
        ws["B9"].value = ant_bilag
        ws["B9"].number_format = DEFAULT_INT_FORMAT

        ws["A10"].value = "Kommentar (fra GUI)"
        ws["A10"].font = HEADER_FONT
        ws["B10"].value = kommentar
        ws["B10"].alignment = Alignment(wrap_text=True, vertical="top")

        ws["A11"].value = "Bilagslinjer (outlier)"
        ws["A11"].font = HEADER_FONT
        if include_outlier_transactions and getattr(frames, "transactions_included", True):
            first_row = frames.combo_first_row.get(combo, 1)
            ws["B11"].value = f'=HYPERLINK("#\'{outlier_sheet_name}\'!A{first_row}","Gå til bilagslinjer")'
        else:
            ws["B11"].value = "Bilagslinjer er utelatt i eksport"
            ws["B11"].alignment = Alignment(wrap_text=True)

        # Handling + resultat (skal ikke havne nederst på arket)
        next_row = 13

        next_row = _write_text_box(
            ws,
            title_row=next_row,
            title="Handling",
            body="1. Opparbeid en forståelse av kombinasjonen og dokumenter.\n2. Vurder om det er relevant å detaljteste på bilagsnivå",
            start_col=1,
            end_col=8,
            height_rows=4,
        )
        next_row = _write_text_box(
            ws,
            title_row=next_row + 1,
            title="Resultat",
            body="",
            start_col=1,
            end_col=8,
            height_rows=6,
        )

        bilag_start_row = next_row + 1

        # Bilagsoppsummering (B)
        bilags_for_combo = [b for b, c in frames.bilag_to_combo.items() if c == combo and str(b).strip()]
        df_bilag = _build_bilag_summary_df(
            df_scope=df_scope,
            selected_accounts=selected_accounts,
            direction=direction,
            bilags_for_combo=bilags_for_combo,
            bilag_first_row_map=frames.bilag_first_row,
            outlier_sheet_name=outlier_sheet_name,
            combo=combo,
            include_outlier_transactions=include_outlier_transactions and getattr(frames, "transactions_included", True),
            sum_label=sum_label,
        )

        res_bilag = _write_df_table(
            ws,
            df_bilag,
            "Bilagsoppsummering (B)",
            start_row=bilag_start_row,
            start_col=1,
            add_summary_row=True,
            max_col_width=45,
        )

        # Standard frys etter rad 1 (ingen frosne kolonner).
        # Bruker ønsker ikke å fryse langt ned på detalj-/kombinasjonsfaner.
        ws.freeze_panes = "A2"

        # Datoformat i bilagstabell (kolonne B)
        # Header er på res_bilag.header_row
        date_col_idx = None
        if df_bilag is not None and not df_bilag.empty and "Dato" in df_bilag.columns:
            date_col_idx = 1 + list(df_bilag.columns).index("Dato")
            for rr in range(res_bilag.first_data_row, res_bilag.last_data_row + 1):
                ws.cell(row=rr, column=date_col_idx).number_format = DEFAULT_DATE_FORMAT

        created.append(sheet_title)

    return created
