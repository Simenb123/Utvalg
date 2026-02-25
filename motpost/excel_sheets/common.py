"""Felles Excel-hjelpere for motpostanalyse.

Mål:
- Enkle, robuste helper-funksjoner for å skrive pandas DataFrames til openpyxl-ark.
- Konsistent tabell-styling (Excel Table) + summeringsrader.
- Fornuftige kolonnebredder uten at hyperlink-formler gjør kolonner enorme.

NB: Denne modulen brukes både av eksisterende tester og av eksport-layouten.
Endringer bør derfor være bakoverkompatible.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
import re
from typing import Any, Iterable, Mapping

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.workbook import Workbook


DEFAULT_MONEY_FORMAT = "#,##0.00"
DEFAULT_INT_FORMAT = "#,##0"
# Oversikt/arbeidspapir viser ofte summer uten øre.
NO_DECIMAL_MONEY_FORMAT = "#,##0"
DEFAULT_PERCENT_FORMAT = "0.00%"
DEFAULT_DATE_FORMAT = "dd.mm.yyyy"


def _to_argb(color: str | None) -> str | None:
    """Normaliser farger til Excel ARGB (8 hex).

    Viktig: Hvis vi bruker 6-sifret RGB direkte kan openpyxl ende opp med å
    lagre fargen med alpha=00 (transparent) i XML. Excel kan da gi varsel om at
    arbeidsboken inneholder korrupt innhold og forsøke å reparere.
    """

    if not color:
        return None
    c = str(color).replace("#", "").upper()
    if len(c) == 6:
        return "FF" + c
    if len(c) == 8:
        return c
    raise ValueError(f"Ugyldig fargekode: {color!r}. Forventet RGB(6) eller ARGB(8).")


# Enkle farger som matcher eksempelmalen
TAB_OUTLIER_YELLOW = "FFFFEB9C"  # lys gul (mal)
TAB_EXPECTED_GREEN = "FFC6EFCE"  # lys grønn

FILL_OUTLIER = PatternFill(patternType="solid", fgColor=_to_argb(TAB_OUTLIER_YELLOW))
FILL_EXPECTED = PatternFill(patternType="solid", fgColor=_to_argb(TAB_EXPECTED_GREEN))


TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True)

THIN_SIDE = Side(style="thin", color=_to_argb("DDDDDD"))
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def set_tab_color(ws: Worksheet, rgb: str | None) -> None:
    """Setter farge på arkfanen (Excel tab color)."""

    if not rgb:
        ws.sheet_properties.tabColor = None
        return
    ws.sheet_properties.tabColor = Color(rgb=_to_argb(rgb))


def hide_gridlines(ws: Worksheet) -> None:
    """Skjul gridlines for mer "arbeidspapir"-preg."""

    ws.sheet_view.showGridLines = False


def set_column_widths(
    ws: Worksheet,
    widths: Mapping[str, float],
    *,
    do_not_shrink: bool = False,
) -> None:
    """Setter kolonnebredder.

    widths: mapping {"A": 3.4, "B": 33, ...}

    do_not_shrink=True gjør at vi ikke krymper en kolonne som allerede er bredere.
    """

    for col_letter, width in widths.items():
        if not col_letter:
            continue
        col_letter = str(col_letter).upper()
        dim = ws.column_dimensions[col_letter]
        existing = dim.width
        if do_not_shrink and existing is not None and existing > float(width):
            continue
        dim.width = float(width)


def _looks_like_cell_ref(name: str) -> bool:
    """Returner True hvis navnet ligner på en A1-celle-referanse.

    Excel tabellnavn (displayName) kan bli "reparert" hvis navnet ligner på en
    cellereferanse (f.eks. TT2, A1, XFD1048576). Dette gir ofte:
      - fjernede formler i sheet.xml
      - reparerte tabeller i table*.xml

    Vi er konservative og tar høyde for at Excel-kolonner i moderne versjoner
    har opptil 3 bokstaver (A..XFD).
    """

    return bool(re.match(r"^[A-Za-z]{1,3}\d+$", name))


def _safe_table_name(ws_title: str) -> str:
    """Lager et Excel-table name som er kompatibelt.

    Krav (Excel):
    - Kun [A-Za-z0-9_]
    - Starter med bokstav eller underscore
    - Må ikke ligne på en cellereferanse (f.eks. "TT2")
    """

    # Fjern alt som ikke er [A-Za-z0-9_]
    base = re.sub(r"[^A-Za-z0-9_]", "", ws_title or "")
    base = base.strip("_")

    if not base:
        base = "Table"

    # Ikke start med tall (Excel-krav). Bruk prefix som samtidig hindrer at
    # resultatet blir noe som kan ligne på en cellereferanse.
    if not re.match(r"^[A-Za-z_]", base):
        base = f"Sheet_{base}"

    name = f"T{base}"

    # Unngå tabellnavn som ser ut som en celle (A1, TT2, ...)
    if _looks_like_cell_ref(name):
        name = f"{name}_tbl"

    return name


def _collect_table_names(wb: Workbook) -> set[str]:
    names: set[str] = set()
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            if getattr(tbl, "displayName", None):
                names.add(tbl.displayName)
    return names


def _make_unique_table_name(ws: Worksheet, base_name: str) -> str:
    # Excel krever at displayName er unikt i hele arbeidsboken (ikke bare per ark).
    existing = _collect_table_names(ws.parent)

    if base_name not in existing:
        return base_name

    i = 2
    while True:
        candidate = f"{base_name}_{i}"
        if candidate not in existing:
            return candidate
        i += 1


def _apply_table_style(ws: Worksheet, ref: str) -> Table:
    base_name = _safe_table_name(ws.title)
    table_name = _make_unique_table_name(ws, base_name)

    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    return table


def _style_title_cell(cell) -> None:
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_header_row(ws: Worksheet, header_row: int, start_col: int, end_col: int) -> None:
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _is_money_header(header: str) -> bool:
    h = header.strip().lower()
    if not h:
        return False

    # Eksplisitte
    if h in {
        "beløp",
        "motbeløp",
        "netto",
        "debet",
        "kredit",
        "kontrollsum",
        "sum",
        "differanse",
    }:
        return True

    # Mønstre
    if "sum valgte kontoer" in h:
        return True
    if "netto" in h and "andel" not in h:
        return True
    if "populasjon" in h:
        return True
    if "motposter" in h:
        return True
    if "beløp" in h and "andel" not in h:
        return True

    return False


def _is_percent_header(header: str) -> bool:
    h = header.strip().lower()
    if not h:
        return False

    if h.startswith("%") or h.endswith("%"):
        return True

    if "andel" in h:
        # f.eks. "Andel %", "Andel av total", "% andel bilag"
        return True

    return False


def _format_number_by_header(cell, header: str) -> None:
    h = str(header or "").strip()

    if h.lower().startswith("dato"):
        cell.number_format = DEFAULT_DATE_FORMAT
        return

    if h.lower().startswith("antall"):
        cell.number_format = DEFAULT_INT_FORMAT
        return

    if _is_money_header(h):
        cell.number_format = DEFAULT_MONEY_FORMAT
        return

    if _is_percent_header(h):
        cell.number_format = DEFAULT_PERCENT_FORMAT
        return


def _display_len(value: Any) -> int:
    """Best-effort 'display length' for auto-width.

    Viktig: Hyperlink-formler kan være ekstremt lange (inkl. sheet refs).
    Vi ønsker bredde basert på synlig label, ikke formelen.
    """

    if value is None:
        return 0

    # Excel-formler (streng)
    if isinstance(value, str) and value.startswith("="):
        # HYPERLINK("...","Label") -> bruk lengden på Label hvis mulig
        m = re.search(r"HYPERLINK\([^,]+,\s*\"([^\"]*)\"\)", value, flags=re.IGNORECASE)
        if m:
            return len(m.group(1))
        # fallback: ikke bruk hele formelen
        return min(len(value), 12)

    return len(str(value))


def _auto_fit_column_widths(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    *,
    max_col_width: int = 45,
    header_cap: int = 20,
    do_not_shrink: bool = True,
) -> None:
    """Auto-fit kolonnebredder basert på innhold, med cap.

    do_not_shrink=True gjør at vi aldri reduserer en kolonnebredde som allerede
    er større (nyttig når vi skriver flere tabeller på samme ark).
    """

    for c in range(start_col, end_col + 1):
        col_letter = get_column_letter(c)
        current = ws.column_dimensions[col_letter].width

        max_len = 0
        for r in range(start_row, end_row + 1):
            cell_len = _display_len(ws.cell(row=r, column=c).value)
            if r == start_row:
                cell_len = min(cell_len, header_cap)
            max_len = max(max_len, cell_len)

        # Litt padding
        width = min(max(max_len + 2, 8), max_col_width)

        if do_not_shrink and current is not None and current > width:
            continue
        ws.column_dimensions[col_letter].width = float(width)


@dataclass
class TableWriteResult:
    start_row: int
    start_col: int
    header_row: int
    first_data_row: int
    last_data_row: int
    last_row: int
    last_col: int


def _write_df_table(
    ws: Worksheet,
    df: pd.DataFrame,
    title: str,
    *,
    start_row: int = 1,
    start_col: int = 1,
    add_summary_row: bool = True,
    max_col_width: int = 45,
    auto_width: bool = True,
    do_not_shrink_columns: bool = True,
) -> TableWriteResult:
    """Skriv en DataFrame som en Excel Table med tittel og (valgfri) summeringsrad.

    Layout:
        start_row:    title
        start_row+1:  (blank)
        start_row+2:  header
        start_row+3.. data
        (optional)    summary row (utenfor selve tabellen)

    Returnerer koordinater som er nyttige for videre layout.
    """

    df = df.copy() if df is not None else pd.DataFrame()

    # Title
    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    _style_title_cell(title_cell)

    # Blank row (ikke skriv noe)

    headers = [str(c) for c in df.columns.tolist()]
    header_row = start_row + 2
    first_data_row = header_row + 1

    # Header
    for j, h in enumerate(headers, start=start_col):
        ws.cell(row=header_row, column=j, value=h)

    last_col = start_col + max(len(headers) - 1, 0)
    if headers:
        _style_header_row(ws, header_row, start_col, last_col)

    # Data rows
    for i, (_, row) in enumerate(df.iterrows(), start=first_data_row):
        for j, h in enumerate(headers, start=start_col):
            v = row.get(h)
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            _format_number_by_header(cell, h)

    # Hvis df er tom: ikke forsøk å lage en Excel Table (Excel reparerer ofte slike filer).
    if len(df) == 0 or not headers:
        last_data_row = first_data_row - 1
        last_row = last_data_row
        if auto_width and headers:
            _auto_fit_column_widths(
                ws,
                start_row=header_row,
                end_row=header_row,
                start_col=start_col,
                end_col=last_col,
                max_col_width=max_col_width,
                do_not_shrink=do_not_shrink_columns,
            )
        return TableWriteResult(
            start_row=start_row,
            start_col=start_col,
            header_row=header_row,
            first_data_row=first_data_row,
            last_data_row=first_data_row - 1,
            last_row=last_row,
            last_col=last_col,
        )

    last_data_row = first_data_row + len(df) - 1
    last_row = last_data_row

    # Apply Excel Table style over header+data (ikke summary row)
    ref = f"{get_column_letter(start_col)}{header_row}:{get_column_letter(last_col)}{last_data_row}"
    table = _apply_table_style(ws, ref)
    table_name = table.displayName

    # Summary row under tabellen
    if add_summary_row:
        sum_row = last_data_row + 1
        ws.cell(row=sum_row, column=start_col, value="SUM").font = HEADER_FONT

        for j, h in enumerate(headers, start=start_col):
            if j == start_col:
                continue
            cell = ws.cell(row=sum_row, column=j)

            if _is_money_header(h) or _is_percent_header(h) or h.lower().startswith("antall"):
                cell.value = f"=SUBTOTAL(109,{table_name}[{h}])"
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top")
                _format_number_by_header(cell, h)

        last_row = sum_row

    # Auto width
    if auto_width:
        _auto_fit_column_widths(
            ws,
            start_row=header_row,
            end_row=last_row,
            start_col=start_col,
            end_col=last_col,
            max_col_width=max_col_width,
            do_not_shrink=do_not_shrink_columns,
        )

    return TableWriteResult(
        start_row=start_row,
        start_col=start_col,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        last_row=last_row,
        last_col=last_col,
    )


def _write_kv_sheet(
    ws: Worksheet,
    title: str,
    kv: Iterable[tuple[str, Any]],
    *,
    start_row: int = 1,
    start_col: int = 1,
    key_col_width: float = 32,
    value_col_width: float = 55,
    apply_column_widths: bool = True,
) -> int:
    """Skriver en enkel nøkkel/verdi-blokk.

    Returnerer neste ledige rad (1-indeksert) etter blokken.

    start_col gjør det mulig å ha venstremarg (f.eks. start_col=2 for å bruke kolonne B/C).
    """

    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    _style_title_cell(title_cell)

    key_col_letter = get_column_letter(start_col)
    val_col_letter = get_column_letter(start_col + 1)

    if apply_column_widths:
        ws.column_dimensions[key_col_letter].width = float(key_col_width)
        ws.column_dimensions[val_col_letter].width = float(value_col_width)

    r = start_row + 2
    for k, v in kv:
        ck = ws.cell(row=r, column=start_col, value=str(k))
        ck.font = HEADER_FONT
        ck.alignment = Alignment(vertical="top", wrap_text=True)

        cv = ws.cell(row=r, column=start_col + 1, value=v)
        cv.alignment = Alignment(vertical="top", wrap_text=True)

        # Best-effort format
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            cv.number_format = DEFAULT_MONEY_FORMAT
        if isinstance(v, datetime):
            cv.number_format = "dd.mm.yyyy hh:mm"

        r += 1

    return r


def _set_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    *,
    bold: bool = False,
    fill: str | None = None,
    number_format: str | None = None,
    alignment: Alignment | None = None,
    border: Border | None = THIN_BORDER,
) -> Any:
    """Bakoverkompatibel helper for eldre arkbyggere.

    Noen eldre sheet-moduler/tester importerer fortsatt `_set_cell`. Vi beholder den
    for å unngå ImportError, selv om den nye eksport-layouten ikke er avhengig av den.
    """

    cell = ws.cell(row=row, column=col, value=value)
    if bold:
        cell.font = HEADER_FONT
    if fill:
        cell.fill = PatternFill(patternType="solid", fgColor=_to_argb(fill))
    if number_format:
        cell.number_format = number_format
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def _write_kv_block(
    ws: Worksheet,
    kv: Iterable[tuple[str, Any]],
    *,
    start_row: int,
    start_col: int = 1,
) -> int:
    """Skriver en KV-blokk uten tittel. Returnerer neste rad."""

    r = start_row
    for k, v in kv:
        ck = ws.cell(row=r, column=start_col, value=str(k))
        ck.font = HEADER_FONT
        ck.alignment = Alignment(vertical="top", wrap_text=True)

        cv = ws.cell(row=r, column=start_col + 1, value=v)
        cv.alignment = Alignment(vertical="top", wrap_text=True)

        if isinstance(v, (int, float)) and not isinstance(v, bool):
            cv.number_format = DEFAULT_MONEY_FORMAT
        if isinstance(v, datetime):
            cv.number_format = "dd.mm.yyyy hh:mm"

        r += 1
    return r
