"""page_statistikk_excel.py — Excel-eksport for Statistikk-fanen.

Utskilt fra page_statistikk.py. Krever ingen tkinter-avhengighet og kan
kalles direkte fra tester eller batch-jobber. Compute-funksjonene
(``_compute_kontoer`` osv.) importeres lazy fra ``page_statistikk`` for å
unngå sirkulær import mellom UI-laget og eksport-laget.
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd

from page_statistikk_compute import _AMT_FMT

log = logging.getLogger(__name__)


def _compute_motpost_rl(grp_mp: pd.DataFrame, page: object) -> pd.DataFrame:
    """Aggregér konto-motpost på regnskapslinje-nivå (samme logikk som GUI-toggle)."""
    empty = pd.DataFrame(columns=["Regnr", "Regnskapslinje", "Beløp", "Andel", "AntallBilag"])
    if grp_mp is None or grp_mp.empty or page is None:
        return empty
    try:
        from regnskapslinje_mapping_service import context_from_page, resolve_accounts_to_rl
        ctx = context_from_page(page)
        mapping = resolve_accounts_to_rl(grp_mp["Konto"].astype(str).tolist(), context=ctx)
    except Exception as exc:
        log.warning("_compute_motpost_rl: %s", exc)
        return empty

    df = grp_mp.copy()
    df["Konto"] = df["Konto"].astype(str)
    mapping = mapping.rename(columns={"konto": "Konto"})[["Konto", "regnr", "regnskapslinje"]]
    merged = df.merge(mapping, on="Konto", how="left")

    def _fmt_regnr(v: object) -> str:
        try:
            if v is None or pd.isna(v):
                return ""
            return str(int(v))
        except Exception:
            return ""

    merged["Regnr"] = merged["regnr"].map(_fmt_regnr)
    merged["Regnskapslinje"] = merged["regnskapslinje"].fillna("").astype(str)
    merged.loc[merged["Regnr"] == "", "Regnskapslinje"] = "— umappet —"
    merged["_b"] = pd.to_numeric(merged["Beløp"], errors="coerce").fillna(0.0)
    merged["_n"] = pd.to_numeric(merged["AntallBilag"], errors="coerce").fillna(0).astype(int)

    agg = (
        merged.groupby(["Regnr", "Regnskapslinje"], sort=False, dropna=False)
        .agg(Beløp=("_b", "sum"), AntallBilag=("_n", "sum"))
        .reset_index()
    )
    agg["_abs"] = agg["Beløp"].abs()
    total = agg["_abs"].sum()
    agg["Andel"] = (agg["_abs"] / total * 100).round(1) if total > 0 else 0.0
    agg = agg.sort_values("_abs", ascending=False).drop(columns=["_abs"])
    return agg[["Regnr", "Regnskapslinje", "Beløp", "Andel", "AntallBilag"]].reset_index(drop=True)


def _compute_kombinasjoner_export(df_all: pd.DataFrame, df_rl: pd.DataFrame) -> pd.DataFrame:
    """Bygg kombinasjonstabell for eksport (samme engine som GUI-fanen)."""
    empty_cols = [
        "Kombinasjon #", "Kombinasjon", "Kombinasjon (navn)",
        "Antall bilag", "Sum valgte kontoer", "% andel bilag", "Outlier",
    ]
    if (
        df_all is None or df_all.empty
        or df_rl is None or df_rl.empty
        or "Konto" not in df_rl.columns
    ):
        return pd.DataFrame(columns=empty_cols)
    rl_kontoer = set(df_rl["Konto"].dropna().astype(str).unique())
    if not rl_kontoer:
        return pd.DataFrame(columns=empty_cols)
    try:
        from motpost.konto_core import build_motpost_data
        from motpost.combinations import build_motkonto_combinations
        mp = build_motpost_data(df_all, rl_kontoer, selected_direction="Alle")
        return build_motkonto_combinations(mp.df_scope, rl_kontoer)
    except Exception as exc:
        log.warning("_compute_kombinasjoner_export: %s", exc)
        return pd.DataFrame(columns=empty_cols)


def write_workbook(
    path: str, *, regnr: int, rl_name: str,
    df_rl: pd.DataFrame, df_all: pd.DataFrame,
    page: object, client: str = "", year: str = "",
    konto_set: set[str] | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from page_statistikk_compute import (
        _compute_bilag,
        _compute_kontoer,
        _compute_maned_pivot,
        _compute_motpost,
        _compute_mva,
        _safe_float,
        _safe_int,
    )

    TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
    HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
    SUM_FILL = PatternFill("solid", fgColor="D6E2EF")
    AVVIK_FILL = PatternFill("solid", fgColor="FCE4EC")
    OK_FILL = PatternFill("solid", fgColor="E8F5E9")
    THIN = Side(style="thin", color="D9D9D9")
    B = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    NEG_FONT = Font(color="C62828")
    ts = (f"  |  {client}" if client else "") + (f"  {year}" if year else "")

    def _title(ws: object, title: str, n: int) -> None:
        last = get_column_letter(n)
        ws.merge_cells(f"A1:{last}1")  # type: ignore[union-attr]
        ws["A1"] = title  # type: ignore[index]
        ws["A1"].font = Font(size=13, bold=True)  # type: ignore[index]
        ws["A1"].fill = TITLE_FILL  # type: ignore[index]
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")  # type: ignore[index]
        ws.row_dimensions[1].height = 22  # type: ignore[union-attr]
        ws.merge_cells(f"A2:{last}2")  # type: ignore[union-attr]
        ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"  # type: ignore[index]
        ws["A2"].font = Font(italic=True, color="666666", size=9)  # type: ignore[index]

    def _header(ws: object, row: int, cols: list[str]) -> None:
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=col)  # type: ignore[union-attr]
            c.font = Font(bold=True, size=10)
            c.fill = HEADER_FILL
            c.border = B
            c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 18  # type: ignore[union-attr]

    def _amt(ws: object, r: int, c: int, v: object, neg: bool = False) -> None:
        cell = ws.cell(row=r, column=c, value=_safe_float(v))  # type: ignore[union-attr]
        cell.border = B
        cell.number_format = _AMT_FMT
        cell.alignment = Alignment(horizontal="right")
        if neg:
            cell.font = NEG_FONT

    def _sum_row(
        ws: object,
        row: int,
        n_cols: int,
        *,
        label: str = "Sum",
        label_col: int = 1,
        amounts: dict[int, float] | None = None,
        ints: dict[int, int] | None = None,
    ) -> None:
        """Fyller en summeringslinje med uthevet SUM_FILL-bakgrunn."""
        amounts = amounts or {}
        ints = ints or {}
        for ci in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=ci)  # type: ignore[union-attr]
            cell.fill = SUM_FILL
            cell.border = B
        lbl = ws.cell(row=row, column=label_col, value=label)  # type: ignore[union-attr]
        lbl.font = Font(bold=True)
        lbl.alignment = Alignment(horizontal="left")
        for ci, v in amounts.items():
            c = ws.cell(row=row, column=ci, value=_safe_float(v))  # type: ignore[union-attr]
            c.number_format = _AMT_FMT
            c.font = Font(bold=True, color="C62828" if _safe_float(v) < 0 else "000000")
            c.alignment = Alignment(horizontal="right")
        for ci, v in ints.items():
            c = ws.cell(row=row, column=ci, value=_safe_int(v))  # type: ignore[union-attr]
            c.number_format = "#,##0"
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="right")

    wb = Workbook()

    # Ark 1: Sammendrag
    ws1 = wb.active
    ws1.title = "Sammendrag"
    _title(ws1, f"Statistikk – {regnr} {rl_name}{ts}", 6)
    _header(ws1, 4, ["UB", "UB i fjor", "Endring (kr)", "Endring %", "Antall bilag", ""])
    pivot_df = getattr(page, "_pivot_df_last", None)
    kpi = None
    if pivot_df is not None and not pivot_df.empty:
        kpi = next((r for _, r in pivot_df.iterrows() if _safe_int(r.get("regnr", -1)) == regnr), None)
    if kpi is not None:
        for i, (col, fmt) in enumerate([
            ("UB", _AMT_FMT), ("UB_fjor", _AMT_FMT), ("Endring", _AMT_FMT),
            ("Endring_pct", '0.0"%"'), ("Antall", "#,##0"),
        ], 1):
            v = _safe_int(kpi.get(col)) if fmt == "#,##0" else _safe_float(kpi.get(col))
            c = ws1.cell(5, i, v)
            c.border = B
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right")
    ws1.cell(7, 1).value = "Kontoer"
    ws1.cell(7, 1).font = Font(bold=True, size=11)
    _header(ws1, 8, ["Konto", "Kontonavn", "IB", "Bevegelse", "UB", "Antall"])
    grp_k, _ib_label = _compute_kontoer(df_rl, page, konto_set=konto_set)
    dr = 9
    sum_ib = sum_bev = sum_ub = 0.0
    sum_ant = 0
    has_ib = False
    has_ub = False
    for _, row in grp_k.iterrows():
        ws1.cell(dr, 1, str(row["Konto"])).border = B
        ws1.cell(dr, 2, str(row.get("Kontonavn", "") or "")).border = B
        for ci, cn in [(3, "IB"), (4, "Bevegelse"), (5, "UB")]:
            raw = row.get(cn)
            if raw is not None and str(raw) not in ("", "nan"):
                val = _safe_float(raw)
                _amt(ws1, dr, ci, raw, val < 0)
                if cn == "IB":
                    sum_ib += val
                    has_ib = True
                elif cn == "Bevegelse":
                    sum_bev += val
                elif cn == "UB":
                    sum_ub += val
                    has_ub = True
            else:
                ws1.cell(dr, ci, None).border = B
        ws1.cell(dr, 6, _safe_int(row["Antall"])).border = B
        sum_ant += _safe_int(row["Antall"])
        dr += 1
    if not grp_k.empty:
        amounts: dict[int, float] = {4: sum_bev}
        if has_ib:
            amounts[3] = sum_ib
        if has_ub:
            amounts[5] = sum_ub
        _sum_row(ws1, dr, 6, amounts=amounts, ints={6: sum_ant})
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 35
    for l in ["C", "D", "E"]:
        ws1.column_dimensions[l].width = 18
    ws1.column_dimensions["F"].width = 10
    ws1.freeze_panes = "A5"

    # Ark 2: Månedspivot
    ws2 = wb.create_sheet("Månedspivot")
    months, pivot = _compute_maned_pivot(df_rl)
    nc = 2 + len(months) + 1
    _title(ws2, f"Månedspivot – {regnr} {rl_name}{ts}", nc)
    _header(ws2, 4, ["Konto", "Kontonavn"] + months + ["Sum"])
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 30
    dr = 5
    tot_m = {m: 0.0 for m in months}
    gt = 0.0
    for _, row in pivot.iterrows():
        ws2.cell(dr, 1, str(row["Konto"])).border = B
        ws2.cell(dr, 2, str(row.get("Kontonavn", "") or "")).border = B
        for j, m in enumerate(months, 3):
            v = _safe_float(row.get(m, 0))
            tot_m[m] = tot_m.get(m, 0.0) + v
            if v != 0.0:
                _amt(ws2, dr, j, v, v < 0)
            else:
                ws2.cell(dr, j, None).border = B
        s = _safe_float(row.get("Sum", 0))
        gt += s
        _amt(ws2, dr, nc, s, s < 0)
        dr += 1
    # Sum-rad
    ws2.cell(dr, 1, "Sum").font = Font(bold=True)
    ws2.cell(dr, 1).fill = SUM_FILL
    ws2.cell(dr, 1).border = B
    ws2.cell(dr, 2).fill = SUM_FILL
    ws2.cell(dr, 2).border = B
    for j, m in enumerate(months, 3):
        c = ws2.cell(dr, j, tot_m[m])
        c.font = Font(bold=True, color="C62828" if tot_m[m] < 0 else "000000")
        c.fill = SUM_FILL
        c.border = B
        c.number_format = _AMT_FMT
        c.alignment = Alignment(horizontal="right")
    c_gt = ws2.cell(dr, nc, gt)
    c_gt.font = Font(bold=True, color="C62828" if gt < 0 else "000000")
    c_gt.fill = SUM_FILL
    c_gt.border = B
    c_gt.number_format = _AMT_FMT
    c_gt.alignment = Alignment(horizontal="right")
    for j in range(3, nc + 1):
        ws2.column_dimensions[get_column_letter(j)].width = 14
    ws2.freeze_panes = "C5"

    # Ark 3: Bilag-analyse
    ws3 = wb.create_sheet("Bilag-analyse")
    _title(ws3, f"Bilag-analyse – {regnr} {rl_name}{ts}", 6)
    _header(ws3, 4, ["Bilag", "Dato", "Tekst", "Sum beløp", "Antall poster", "Kontoer"])
    grp_b = _compute_bilag(df_rl)
    dr = 5
    sum_b = 0.0
    sum_p = 0
    for _, row in grp_b.iterrows():
        ws3.cell(dr, 1, str(row.get("Bilag", ""))).border = B
        ws3.cell(dr, 2, str(row.get("Dato", ""))).border = B
        ws3.cell(dr, 3, str(row.get("Tekst", ""))).border = B
        v = _safe_float(row["Sum beløp"])
        _amt(ws3, dr, 4, v, v < 0)
        ws3.cell(dr, 5, _safe_int(row["Antall poster"])).border = B
        ws3.cell(dr, 6, str(row.get("Kontoer", ""))).border = B
        sum_b += v
        sum_p += _safe_int(row["Antall poster"])
        dr += 1
    if not grp_b.empty:
        _sum_row(ws3, dr, 6, label=f"Sum ({len(grp_b)} bilag)", amounts={4: sum_b}, ints={5: sum_p})
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 45
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 25
    ws3.freeze_panes = "A5"

    # Ark 4: MVA-analyse
    ws4 = wb.create_sheet("MVA-analyse")
    _title(ws4, f"MVA-analyse – {regnr} {rl_name}{ts}", 7)
    _header(ws4, 4, ["MVA-kode", "Antall", "Grunnlag", "MVA-beløp", "Sats %", "Effektiv %", "Status"])
    mva_result = _compute_mva(df_rl, df_all)
    grp_mva = mva_result["rows"]
    dr = 5
    sum_ant = 0
    sum_grunn = 0.0
    sum_mva = 0.0
    for _, row in grp_mva.iterrows():
        status = str(row.get("Status", ""))
        fill = OK_FILL if "\u2713" in status else (AVVIK_FILL if "\u26a0" in status else None)
        ws4.cell(dr, 1, str(row.get("MVA-kode", ""))).border = B
        ws4.cell(dr, 2, _safe_int(row["Antall"])).border = B
        _amt(ws4, dr, 3, row["Grunnlag"])
        _amt(ws4, dr, 4, row["MVA-beløp"])
        ws4.cell(dr, 5, round(_safe_float(row.get("Sats %")), 1)).border = B
        ws4.cell(dr, 6, round(_safe_float(row.get("Effektiv %")), 1)).border = B
        ws4.cell(dr, 7, status).border = B
        if fill:
            for ci in range(1, 8):
                ws4.cell(dr, ci).fill = fill
        sum_ant += _safe_int(row["Antall"])
        sum_grunn += _safe_float(row["Grunnlag"])
        sum_mva += _safe_float(row["MVA-beløp"])
        dr += 1
    if not grp_mva.empty:
        _sum_row(ws4, dr, 7, amounts={3: sum_grunn, 4: sum_mva}, ints={2: sum_ant})
    for col_letter, w in zip(["A", "B", "C", "D", "E", "F", "G"], [10, 8, 18, 15, 8, 10, 30]):
        ws4.column_dimensions[col_letter].width = w
    ws4.freeze_panes = "A5"

    # Ark 5: Motpostfordeling (konto-nivå + RL-aggregert)
    ws5 = wb.create_sheet("Motpostfordeling")
    _title(ws5, f"Motpostfordeling – {regnr} {rl_name}{ts}", 5)
    _header(ws5, 4, ["Konto", "Kontonavn", "Beløp", "Andel %", "Antall bilag"])
    grp_mp = _compute_motpost(df_all, df_rl)
    dr = 5
    sum_bel = 0.0
    sum_andel = 0.0
    sum_bilag = 0
    for _, row in grp_mp.iterrows():
        ws5.cell(dr, 1, str(row.get("Konto", ""))).border = B
        ws5.cell(dr, 2, str(row.get("Kontonavn", ""))).border = B
        v = _safe_float(row["Beløp"])
        _amt(ws5, dr, 3, v, v < 0)
        ws5.cell(dr, 4, round(float(row["Andel"]), 1)).border = B
        ws5.cell(dr, 5, _safe_int(row["AntallBilag"])).border = B
        sum_bel += v
        sum_andel += float(row["Andel"])
        sum_bilag += _safe_int(row["AntallBilag"])
        dr += 1
    if not grp_mp.empty:
        _sum_row(
            ws5, dr, 5,
            amounts={3: sum_bel},
            ints={5: sum_bilag},
        )
        ws5.cell(dr, 4, round(sum_andel, 1)).font = Font(bold=True)
        ws5.cell(dr, 4).alignment = Alignment(horizontal="right")
        dr += 1

    # --- RL-aggregert del ---
    dr += 1
    ws5.cell(dr, 1, "Motpost pr regnskapslinje").font = Font(bold=True, size=11)
    dr += 1
    _header(ws5, dr, ["Nr", "Regnskapslinje", "Beløp", "Andel %", "Antall bilag"])
    dr += 1
    grp_rl = _compute_motpost_rl(grp_mp, page)
    sum_bel_rl = 0.0
    sum_andel_rl = 0.0
    sum_bilag_rl = 0
    for _, row in grp_rl.iterrows():
        ws5.cell(dr, 1, str(row.get("Regnr", "") or "")).border = B
        ws5.cell(dr, 2, str(row.get("Regnskapslinje", "") or "")).border = B
        v = _safe_float(row["Beløp"])
        _amt(ws5, dr, 3, v, v < 0)
        ws5.cell(dr, 4, round(float(row["Andel"]), 1)).border = B
        ws5.cell(dr, 5, _safe_int(row["AntallBilag"])).border = B
        sum_bel_rl += v
        sum_andel_rl += float(row["Andel"])
        sum_bilag_rl += _safe_int(row["AntallBilag"])
        dr += 1
    if not grp_rl.empty:
        _sum_row(
            ws5, dr, 5,
            amounts={3: sum_bel_rl},
            ints={5: sum_bilag_rl},
        )
        ws5.cell(dr, 4, round(sum_andel_rl, 1)).font = Font(bold=True)
        ws5.cell(dr, 4).alignment = Alignment(horizontal="right")

    ws5.column_dimensions["A"].width = 10
    ws5.column_dimensions["B"].width = 40
    ws5.column_dimensions["C"].width = 18
    ws5.column_dimensions["D"].width = 10
    ws5.column_dimensions["E"].width = 14
    ws5.freeze_panes = "A5"

    # Ark 6: Kombinasjoner
    ws6 = wb.create_sheet("Kombinasjoner")
    _title(ws6, f"Kombinasjoner – {regnr} {rl_name}{ts}", 5)
    _header(ws6, 4, ["Nr", "Kombinasjon", "Antall bilag", "Sum valgte kontoer", "Andel %"])
    grp_kombo = _compute_kombinasjoner_export(df_all, df_rl)
    dr = 5
    sum_belk = 0.0
    sum_antk = 0
    for _, row in grp_kombo.iterrows():
        ws6.cell(dr, 1, _safe_int(row.get("Kombinasjon #"))).border = B
        combo_label = str(
            row.get("Kombinasjon (navn)") or row.get("Kombinasjon", "") or ""
        )
        ws6.cell(dr, 2, combo_label).border = B
        ws6.cell(dr, 3, _safe_int(row.get("Antall bilag"))).border = B
        v = _safe_float(row.get("Sum valgte kontoer"))
        _amt(ws6, dr, 4, v, v < 0)
        ws6.cell(dr, 5, round(_safe_float(row.get("% andel bilag")), 1)).border = B
        sum_belk += v
        sum_antk += _safe_int(row.get("Antall bilag"))
        dr += 1
    if not grp_kombo.empty:
        _sum_row(
            ws6, dr, 5,
            label=f"Sum ({len(grp_kombo)} kombinasjoner)",
            amounts={4: sum_belk},
            ints={3: sum_antk},
        )
    ws6.column_dimensions["A"].width = 6
    ws6.column_dimensions["B"].width = 60
    ws6.column_dimensions["C"].width = 14
    ws6.column_dimensions["D"].width = 20
    ws6.column_dimensions["E"].width = 10
    ws6.freeze_panes = "A5"

    out = Path(path)
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
