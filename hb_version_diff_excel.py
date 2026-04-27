"""Excel-arbeidspapir for HB versjonsdiff.

Genererer arbeidsbok med tre ark:
  - Oppsummering: totaler og nøkkeltall
  - Nye bilag: transaksjoner som kun finnes i ny versjon
  - Fjernede bilag: transaksjoner som kun finnes i gammel versjon
  - Endrede bilag: bilag med endret innhold (sum/linjer/beløp)
"""

from __future__ import annotations

from datetime import datetime
from typing import Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.shared.workpapers.forside import build_forside_sheet

_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
_HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
_ADDED_FILL = PatternFill("solid", fgColor="E2EFDA")     # Grønn - nye
_REMOVED_FILL = PatternFill("solid", fgColor="FCE4EC")   # Rød - fjernede
_CHANGED_FILL = PatternFill("solid", fgColor="FFF2CC")   # Gul - endrede
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_AMOUNT_FMT = '#,##0.00;[Red]-#,##0.00'


def build_hb_diff_workpaper(
    diff_result,
    *,
    client: str | None = None,
    year: str | int | None = None,
    version_a_label: str = "Forrige",
    version_b_label: str = "Gjeldende",
) -> Workbook:
    """Bygg komplett HB versjonsdiff arbeidspapir."""
    wb = Workbook()

    _build_summary_sheet(
        wb, diff_result.summary,
        client=client, year=year,
        version_a_label=version_a_label,
        version_b_label=version_b_label,
    )
    _build_added_sheet(wb, diff_result.added, client=client, year=year)
    _build_removed_sheet(wb, diff_result.removed, client=client, year=year)
    _build_changed_sheet(wb, diff_result.changed, client=client, year=year)

    build_forside_sheet(wb, workpaper_navn="HB versjonsdiff")

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    return wb


# ---------------------------------------------------------------------------
# Oppsummering
# ---------------------------------------------------------------------------

def _build_summary_sheet(
    wb: Workbook,
    summary: Dict[str, object],
    *,
    client: str | None = None,
    year: str | int | None = None,
    version_a_label: str = "Forrige",
    version_b_label: str = "Gjeldende",
) -> None:
    ws = wb.create_sheet("Oppsummering")

    title_parts = ["HB Versjonsdiff"]
    if client:
        title_parts.append(str(client))
    if year not in {None, ""}:
        title_parts.append(str(year))
    title = " — ".join(title_parts)

    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = _TITLE_FILL

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    ws.merge_cells("A3:D3")
    ws["A3"] = f"Sammenligning: {version_a_label} → {version_b_label}"
    ws["A3"].font = Font(italic=True)

    row = 5
    items = [
        (f"Bilag i {version_a_label}", summary.get("bilag_a_total", 0)),
        (f"Bilag i {version_b_label}", summary.get("bilag_b_total", 0)),
        ("", ""),
        ("Nye bilag", summary.get("nye_bilag", 0)),
        ("Fjernede bilag", summary.get("fjernede_bilag", 0)),
        ("Endrede bilag", summary.get("endrede_bilag", 0)),
        ("Uendrede bilag", summary.get("uendrede_bilag", 0)),
        ("", ""),
        ("Nye transaksjonslinjer", summary.get("nye_transaksjoner", 0)),
        ("Fjernede transaksjonslinjer", summary.get("fjernede_transaksjoner", 0)),
    ]
    for label, value in items:
        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.font = Font(bold=True)
        cell_val = ws.cell(row=row, column=2, value=value if value != "" else None)
        if isinstance(value, float):
            cell_val.number_format = _AMOUNT_FMT

        # Fargekode nøkkeltall
        if label == "Nye bilag" and isinstance(value, int) and value > 0:
            cell_val.fill = _ADDED_FILL
            cell_val.font = Font(bold=True, color="006100")
        elif label == "Fjernede bilag" and isinstance(value, int) and value > 0:
            cell_val.fill = _REMOVED_FILL
            cell_val.font = Font(bold=True, color="CC0000")
        elif label == "Endrede bilag" and isinstance(value, int) and value > 0:
            cell_val.fill = _CHANGED_FILL
            cell_val.font = Font(bold=True, color="7F6000")
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.sheet_properties.tabColor = "4472C4"


# ---------------------------------------------------------------------------
# Nye bilag
# ---------------------------------------------------------------------------

def _build_added_sheet(
    wb: Workbook,
    added_df: pd.DataFrame,
    *,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.create_sheet("Nye bilag")

    title = "Nye bilag (kun i ny versjon)"
    if client:
        title += f" — {client}"
    if year not in {None, ""}:
        title += f" {year}"

    _write_transaction_sheet(ws, title, added_df)
    ws.sheet_properties.tabColor = "70AD47"  # Grønn


# ---------------------------------------------------------------------------
# Fjernede bilag
# ---------------------------------------------------------------------------

def _build_removed_sheet(
    wb: Workbook,
    removed_df: pd.DataFrame,
    *,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.create_sheet("Fjernede bilag")

    title = "Fjernede bilag (kun i gammel versjon)"
    if client:
        title += f" — {client}"
    if year not in {None, ""}:
        title += f" {year}"

    _write_transaction_sheet(ws, title, removed_df)
    ws.sheet_properties.tabColor = "FF0000"  # Rød


# ---------------------------------------------------------------------------
# Endrede bilag
# ---------------------------------------------------------------------------

def _build_changed_sheet(
    wb: Workbook,
    changed_df: pd.DataFrame,
    *,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.create_sheet("Endrede bilag")

    title = "Endrede bilag (ulikt innhold)"
    if client:
        title += f" — {client}"
    if year not in {None, ""}:
        title += f" {year}"

    if changed_df.empty:
        _write_title_row(ws, title, span=5)
        ws.cell(row=5, column=1, value="Ingen endrede bilag funnet")
        ws["A5"].font = Font(color="006100")
        ws.sheet_properties.tabColor = "70AD47"
        return

    headers = ["Bilag", "Sum forrige", "Linjer forrige", "Sum gjeldende",
               "Linjer gjeldende", "Diff sum", "Diff linjer"]
    cols = ["bilag", "sum_a", "linjer_a", "sum_b", "linjer_b", "diff_sum", "diff_linjer"]
    amount_cols = {1, 3, 5}  # sum_a, sum_b, diff_sum

    _write_title_row(ws, title, span=len(headers))

    # Header
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, (_, row) in enumerate(changed_df.iterrows(), start=5):
        for col_idx, col_name in enumerate(cols):
            raw = row.get(col_name, "")
            if col_idx in amount_cols:
                value = float(raw) if raw is not None and raw == raw else 0.0
            elif col_idx in {2, 4, 6}:  # linjer - heltall
                value = int(raw) if raw is not None and raw == raw else 0
            else:
                value = str(raw) if raw is not None and raw == raw else ""
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
            cell.border = _BORDER
            if col_idx in amount_cols:
                cell.number_format = _AMOUNT_FMT
                cell.alignment = Alignment(horizontal="right")
            # Marker diff-kolonner
            if col_idx == 5 and isinstance(value, float) and abs(value) > 0.01:
                cell.fill = _CHANGED_FILL
            elif col_idx == 6 and isinstance(value, (int, float)) and value != 0:
                cell.fill = _CHANGED_FILL

    ws.column_dimensions["A"].width = 14
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 16
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "FFC000"  # Oransje


# ---------------------------------------------------------------------------
# Hjelpere
# ---------------------------------------------------------------------------

def _write_title_row(ws, title: str, *, span: int) -> None:
    last_col = chr(ord("A") + span - 1)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = _TITLE_FILL

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")


def _write_transaction_sheet(ws, title: str, df: pd.DataFrame) -> None:
    """Skriv et transaksjonsark med dynamiske kolonner fra DataFrame."""
    if df.empty:
        _write_title_row(ws, title, span=4)
        ws.cell(row=5, column=1, value="Ingen transaksjoner")
        ws["A5"].font = Font(color="666666")
        ws.freeze_panes = "A5"
        return

    cols = list(df.columns)
    _write_title_row(ws, title, span=len(cols))

    # Header
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, (_, row) in enumerate(df.iterrows(), start=5):
        for col_idx, col_name in enumerate(cols):
            raw = row[col_name]
            if isinstance(raw, float):
                value = raw if raw == raw else 0.0
            else:
                value = raw if raw is not None and raw == raw else ""
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
            cell.border = _BORDER
            if isinstance(raw, float):
                cell.number_format = _AMOUNT_FMT
                cell.alignment = Alignment(horizontal="right")

    # Autobredde (enkel heuristikk)
    for col_idx, col_name in enumerate(cols, start=1):
        letter = chr(ord("A") + col_idx - 1)
        ws.column_dimensions[letter].width = max(12, len(str(col_name)) + 4)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{chr(ord('A') + len(cols) - 1)}{max(5, 4 + len(df))}"
