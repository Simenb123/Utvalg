"""page_statistikk.py — Statistikk-fane for Utvalg."""
from __future__ import annotations

import logging
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd

log = logging.getLogger(__name__)

try:
    import tkinter as tk
    from tkinter import messagebox, ttk
    from tkinter.filedialog import asksaveasfilename
except Exception:  # pragma: no cover
    tk = None  # type: ignore
    ttk = None  # type: ignore
    messagebox = None  # type: ignore

_AMT_FMT = "#,##0.00;[Red]-#,##0.00"

# MVA-koder → normalsats (brukes i avstemming)
_MVA_SATS_MAP: dict[str, float] = {
    "1": 0.25, "11": 0.15, "12": 0.12, "13": 0.0,
    "3": 0.25, "31": 0.15, "32": 0.12, "33": 0.0,
}
# Kontoer for utgående MVA
_MVA_KONTO_FRA = 2700
_MVA_KONTO_TIL = 2799


# ---------------------------------------------------------------------------
# Hjelpere
# ---------------------------------------------------------------------------

def _safe_float(v: object) -> float:
    try:
        f = float(v)  # type: ignore[arg-type]
        return f if f == f else 0.0
    except Exception:
        return 0.0


def _safe_int(v: object) -> int:
    try:
        return int(float(v))  # type: ignore[arg-type]
    except Exception:
        return 0


def _fmt_amount(v: object) -> str:
    try:
        f = float(v)  # type: ignore[arg-type]
        if f != f:
            return "\u2013"
        return f"{f:,.0f}".replace(",", "\u202f")
    except Exception:
        return "\u2013"


def _fmt_pct(v: object) -> str:
    try:
        f = float(v)  # type: ignore[arg-type]
        if f != f:
            return "\u2013"
        return f"{f:.1f} %"
    except Exception:
        return "\u2013"


def _open_file(path: str) -> None:
    try:
        if sys.platform == "win32":
            os.startfile(path)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Treeview-sortering
# ---------------------------------------------------------------------------

_SORT_STATE: dict[tuple[int, str], bool] = {}  # (id(tree), col) → ascending


def _sort_col(tree: object, col: str) -> None:
    """Sorter treeview etter kolonne ved klikk på header. Sum-rader holdes nederst."""
    key = (id(tree), col)
    ascending = not _SORT_STATE.get(key, False)
    _SORT_STATE[key] = ascending

    items = list(tree.get_children(""))  # type: ignore[union-attr]
    pinned = [i for i in items if "sum" in (tree.item(i, "tags") or ())]  # type: ignore[union-attr]
    sortable = [i for i in items if i not in set(pinned)]

    def _key(iid: str) -> tuple:
        raw = str(tree.set(iid, col))  # type: ignore[union-attr]
        cleaned = raw.replace("\u202f", "").replace("\xa0", "").replace(" ", "").replace("%", "").replace("\u2013", "0")
        try:
            return (0, float(cleaned))
        except ValueError:
            return (1, raw.lower())

    sortable.sort(key=_key, reverse=not ascending)

    # Oppdater heading-tekst med pil
    for c in tree["columns"]:  # type: ignore[index]
        txt = str(tree.heading(c, "text")).rstrip(" ▲▼")  # type: ignore[union-attr]
        tree.heading(c, text=txt)  # type: ignore[union-attr]
    arrow = " \u25b2" if ascending else " \u25bc"
    cur_txt = str(tree.heading(col, "text")).rstrip(" ▲▼")  # type: ignore[union-attr]
    tree.heading(col, text=cur_txt + arrow)  # type: ignore[union-attr]

    for idx, iid in enumerate(sortable + pinned):
        tree.move(iid, "", idx)  # type: ignore[union-attr]


def _attach_sort(tree: object) -> None:
    """Kobler klikk-sortering til alle kolonner i ett treeview."""
    for col in tree["columns"]:  # type: ignore[index]
        tree.heading(col, command=lambda c=col, t=tree: _sort_col(t, c))  # type: ignore[union-attr]


# ---------------------------------------------------------------------------
# Databeregning
# ---------------------------------------------------------------------------

def _get_konto_ranges(page: object, regnr: int) -> list[tuple[int, int]]:
    intervals = getattr(page, "_rl_intervals", None)
    regnskapslinjer = getattr(page, "_rl_regnskapslinjer", None)
    if intervals is None or (hasattr(intervals, "empty") and intervals.empty):
        return []
    leaf_set: set[int] = {regnr}
    if regnskapslinjer is not None and not (hasattr(regnskapslinjer, "empty") and regnskapslinjer.empty):
        try:
            from regnskap_mapping import expand_regnskapslinje_selection, normalize_regnskapslinjer
            regn = normalize_regnskapslinjer(regnskapslinjer)
            if bool(regn.loc[regn["regnr"].astype(int) == regnr, "sumpost"].any()):
                expanded = expand_regnskapslinje_selection(
                    regnskapslinjer=regnskapslinjer, selected_regnr=[regnr]
                )
                if expanded:
                    leaf_set = set(expanded)
        except Exception as exc:
            log.warning("_get_konto_ranges: %s", exc)
    ranges: list[tuple[int, int]] = []
    try:
        for _, row in intervals.iterrows():
            if int(row["regnr"]) in leaf_set:
                ranges.append((int(row["fra"]), int(row["til"])))
    except Exception as exc:
        log.warning("_get_konto_ranges loop: %s", exc)
    return ranges


def _filter_df(df: pd.DataFrame, ranges: list[tuple[int, int]]) -> pd.DataFrame:
    if not ranges or df is None or df.empty or "Konto" not in df.columns:
        return pd.DataFrame(columns=(df.columns if df is not None else []))
    try:
        num = pd.to_numeric(df["Konto"], errors="coerce")
        mask = pd.Series(False, index=df.index)
        for fra, til in ranges:
            mask |= (num >= fra) & (num <= til)
        return df.loc[mask].copy()
    except Exception as exc:
        log.warning("_filter_df: %s", exc)
        return pd.DataFrame(columns=df.columns)


def _compute_kontoer(df_rl: pd.DataFrame, page: object) -> tuple[pd.DataFrame, str]:
    """Returnerer (grp_df, ib_col_label) der label er 'UB fjor' eller 'IB'."""
    if df_rl.empty or "Beløp" not in df_rl.columns:
        return pd.DataFrame(columns=["Konto", "Kontonavn", "IB", "Bevegelse", "UB", "Antall"]), "IB"
    df = df_rl.copy()
    df["_b"] = pd.to_numeric(df["Beløp"], errors="coerce").fillna(0)
    grp = (
        df.groupby("Konto", sort=False)
        .agg(Kontonavn=("Kontonavn", "first"), Bevegelse=("_b", "sum"), Antall=("_b", "count"))
        .reset_index()
        .sort_values("Konto")
    )

    ib_map: dict[str, float] = {}
    ub_map: dict[str, float] = {}
    ib_label = "IB"

    # Foretrekk UB fra fjorårets saldobalanse som IB-kolonne
    sb_prev = getattr(page, "_rl_sb_prev_df", None)
    if sb_prev is not None and not sb_prev.empty and "konto" in sb_prev.columns:
        for _, r in sb_prev.iterrows():
            k = str(r["konto"])
            ib_map[k] = _safe_float(r.get("ub"))   # fjorår UB = årets IB
        ib_label = "UB fjor"

    # Aktuelle SB for UB (inneværende år)
    try:
        sb = page._get_effective_sb_df()  # type: ignore[union-attr]
    except Exception:
        sb = getattr(page, "_rl_sb_df", None)
    if sb is not None and not sb.empty and "konto" in sb.columns:
        for _, r in sb.iterrows():
            k = str(r["konto"])
            if not ib_map:   # fall back til IB fra sb dersom ingen prev år
                ib_map[k] = _safe_float(r.get("ib"))
            ub_map[k] = _safe_float(r.get("ub"))
        if not ib_map:
            ib_label = "IB"

    grp["IB"] = grp["Konto"].astype(str).map(ib_map)
    grp["UB"] = grp["Konto"].astype(str).map(ub_map)
    return grp[["Konto", "Kontonavn", "IB", "Bevegelse", "UB", "Antall"]], ib_label


def _compute_extra_stats(df_rl: pd.DataFrame) -> dict:
    """Beregner ekstra analytiske nøkkeltall for visning under nøkkeltall-bandet."""
    out: dict = {}
    if df_rl.empty or "Beløp" not in df_rl.columns:
        return out

    beløp = pd.to_numeric(df_rl["Beløp"], errors="coerce").fillna(0)
    total_abs = beløp.abs().sum()

    # Konsentrasjon: topp 10 bilag som % av total
    if "Bilag" in df_rl.columns and total_abs > 0:
        bilag_abs = (
            df_rl.assign(_b=beløp).groupby("Bilag")["_b"].sum().abs().nlargest(10)
        )
        out["top10_pct"] = bilag_abs.sum() / total_abs * 100
        out["n_bilag"] = df_rl["Bilag"].nunique()

    # Unike kunder
    if "Kunder" in df_rl.columns:
        out["n_kunder"] = int(df_rl["Kunder"].dropna().nunique())

    # Månedlig spredning
    if "Dato" in df_rl.columns:
        try:
            dato = pd.to_datetime(df_rl["Dato"], dayfirst=True, errors="coerce")
            mnd_sum = (
                df_rl.assign(_b=beløp, _mnd=dato.dt.to_period("M"))
                .groupby("_mnd")["_b"].sum()
            )
            if len(mnd_sum) > 1:
                out["mnd_snitt"] = float(mnd_sum.mean())
                out["mnd_std"] = float(mnd_sum.std())
                out["mnd_max_name"] = str(mnd_sum.abs().idxmax())
                out["mnd_max_val"] = float(mnd_sum[mnd_sum.abs().idxmax()])
                # Anomali: måneder der abs(avvik fra snitt) > 1.5 × std
                if out["mnd_std"] > 0:
                    avvik = (mnd_sum - out["mnd_snitt"]).abs()
                    out["n_anomali_mnd"] = int((avvik > 1.5 * out["mnd_std"]).sum())
        except Exception:
            pass

    # Runde beløp-andel (beløp delbart med 1000)
    runde = (beløp.abs() % 1000 == 0) & (beløp.abs() >= 1000)
    out["runde_pct"] = float(runde.sum() / len(beløp) * 100) if len(beløp) > 0 else 0.0

    return out


def _compute_maned_pivot(df_rl: pd.DataFrame) -> tuple[list[str], pd.DataFrame]:
    if df_rl.empty or "Beløp" not in df_rl.columns:
        return [], pd.DataFrame()
    df = df_rl.copy()
    df["_b"] = pd.to_numeric(df["Beløp"], errors="coerce").fillna(0)
    if "Dato" in df.columns:
        try:
            df["_mnd"] = pd.to_datetime(df["Dato"], dayfirst=True, errors="coerce").dt.to_period("M").astype(str)
        except Exception:
            df["_mnd"] = "Ukjent"
    elif "Periode" in df.columns:
        df["_mnd"] = df["Periode"].astype(str)
    else:
        df["_mnd"] = "Ukjent"
    months = sorted(df["_mnd"].dropna().unique().tolist())
    konto_navn = df.groupby("Konto")["Kontonavn"].first()
    pivot = df.pivot_table(index="Konto", columns="_mnd", values="_b", aggfunc="sum", fill_value=0.0)
    pivot = pivot.reindex(columns=months, fill_value=0.0)
    pivot["Sum"] = pivot[months].sum(axis=1)
    pivot = pivot.reset_index()
    pivot.insert(1, "Kontonavn", pivot["Konto"].astype(str).map(konto_navn).fillna(""))
    return months, pivot.sort_values("Konto")


def _compute_bilag(df_rl: pd.DataFrame) -> pd.DataFrame:
    empty = pd.DataFrame(columns=["Bilag", "Dato", "Tekst", "Sum beløp", "Antall poster", "Kontoer"])
    if df_rl.empty or "Bilag" not in df_rl.columns or "Beløp" not in df_rl.columns:
        return empty
    df = df_rl.copy()
    df["_b"] = pd.to_numeric(df["Beløp"], errors="coerce").fillna(0)
    if "Dato" in df.columns:
        try:
            df["_dato"] = pd.to_datetime(df["Dato"], dayfirst=True, errors="coerce").dt.strftime("%d.%m.%Y")
        except Exception:
            df["_dato"] = df["Dato"].astype(str)
    else:
        df["_dato"] = ""
    tekst_col = "Tekst" if "Tekst" in df.columns else "_dato"

    def _ktoer(s: pd.Series) -> str:
        u = sorted(s.dropna().astype(str).unique())
        return ", ".join(u[:5]) + (" …" if len(u) > 5 else "")

    grp = df.groupby("Bilag", sort=False).agg(
        Dato=("_dato", "first"),
        Tekst=(tekst_col, "first"),
        SumBeløp=("_b", "sum"),
        Antall=("_b", "count"),
        Kontoer=("Konto", _ktoer),
    ).reset_index()
    grp["_abs"] = grp["SumBeløp"].abs()
    grp = grp.sort_values("_abs", ascending=False).drop(columns=["_abs"])
    return grp.rename(columns={"SumBeløp": "Sum beløp", "Antall": "Antall poster"})


def _compute_mva(df_rl: pd.DataFrame, df_all: Optional[pd.DataFrame] = None) -> dict:
    """
    MVA-analyse med to strategier for å finne faktisk MVA-beløp:
    1. Direkte: MVA-beløp-feltet på transaksjonslinjen
    2. Motpost: konto 2700-2799 med samme bilagsnummer

    Inkluderer rader uten MVA-kode som egen gruppe.
    Returnerer dict med 'rows' (DataFrame) og avstemmingstall.
    """
    _EMPTY_ROWS = pd.DataFrame(
        columns=["MVA-kode", "Antall", "Grunnlag", "MVA-beløp", "Sats %", "Effektiv %", "Status"]
    )
    _empty = {
        "rows": _EMPTY_ROWS, "total_bevegelse": 0.0,
        "total_med_kode": 0.0, "total_uten_kode": 0.0,
        "total_mva": 0.0, "total_forventet_mva": 0.0,
    }

    if df_rl.empty or "Beløp" not in df_rl.columns:
        return _empty

    df = df_rl.copy()
    df["_b"] = pd.to_numeric(df["Beløp"], errors="coerce").fillna(0)
    df["_mva_direkt"] = (
        pd.to_numeric(df["MVA-beløp"], errors="coerce").fillna(0)
        if "MVA-beløp" in df.columns else pd.Series(0.0, index=df.index)
    )

    total_bevegelse = float(df["_b"].sum())

    if "MVA-kode" not in df.columns:
        return {**_empty, "total_bevegelse": total_bevegelse, "total_uten_kode": total_bevegelse}

    # Splitt i med/uten kode
    has_kode = df["MVA-kode"].notna() & (df["MVA-kode"].astype(str).str.strip() != "")
    df_med = df[has_kode]
    df_uten = df[~has_kode]

    total_med_kode = float(df_med["_b"].sum())
    total_uten_kode = float(df_uten["_b"].sum())

    # Motpost-oppslag for MVA-kontoer (2700-2799)
    motpost_mva: dict[str, float] = {}
    if (
        df_all is not None and not df_all.empty
        and "Bilag" in df.columns and "Bilag" in df_all.columns
        and "Konto" in df_all.columns
    ):
        try:
            konto_num_all = pd.to_numeric(df_all["Konto"], errors="coerce")
            df_mva_k = df_all.loc[
                (konto_num_all >= _MVA_KONTO_FRA) & (konto_num_all <= _MVA_KONTO_TIL)
            ].copy()
            if not df_mva_k.empty:
                df_mva_k["_b"] = pd.to_numeric(df_mva_k["Beløp"], errors="coerce").fillna(0)
                for kode in df_med["MVA-kode"].unique():
                    ks = str(kode).strip()
                    bilag_k = set(
                        df_med[df_med["MVA-kode"].astype(str).str.strip() == ks]["Bilag"]
                        .dropna().astype(str).unique()
                    )
                    if bilag_k:
                        motpost_mva[ks] = float(
                            df_mva_k.loc[df_mva_k["Bilag"].astype(str).isin(bilag_k), "_b"].sum()
                        )
        except Exception as exc:
            log.warning("_compute_mva motpost-oppslag: %s", exc)

    rows: list[dict] = []
    total_faktisk_mva = 0.0
    total_forventet_mva = 0.0

    # --- Rader med MVA-kode ---
    if not df_med.empty:
        grp = df_med.groupby("MVA-kode", sort=False).agg(
            Antall=("_b", "count"),
            Grunnlag=("_b", "sum"),
            MVADirekt=("_mva_direkt", "sum"),
        ).reset_index()

        for _, row in grp.iterrows():
            kode = str(row["MVA-kode"]).strip()
            grunnlag = float(row["Grunnlag"])
            direkt = float(row["MVADirekt"])
            motpost = motpost_mva.get(kode, 0.0)

            actual_mva = direkt if abs(direkt) > 1 else (motpost if abs(motpost) > 1 else 0.0)
            eff_sats = abs(actual_mva / grunnlag) * 100 if abs(grunnlag) > 1 else 0.0
            forventet_sats = _MVA_SATS_MAP.get(kode)
            forventet_mva_kode = abs(grunnlag) * (forventet_sats or 0.0)

            total_faktisk_mva += actual_mva
            total_forventet_mva += forventet_mva_kode

            status = ""
            if forventet_sats is not None and abs(grunnlag) > 100:
                if abs(actual_mva) < 1:
                    if forventet_mva_kode > 100:
                        status = f"\u26a0 Ingen MVA funnet (forventet {_fmt_amount(forventet_mva_kode)})"
                else:
                    avvik = (
                        abs(forventet_mva_kode - abs(actual_mva)) / forventet_mva_kode * 100
                        if forventet_mva_kode > 0 else 0.0
                    )
                    status = "\u2713 OK" if avvik < 2.0 else f"\u26a0 Avvik {avvik:.1f}%"

            rows.append({
                "MVA-kode": kode,
                "Antall": _safe_int(row["Antall"]),
                "Grunnlag": grunnlag,
                "MVA-beløp": actual_mva,
                "Sats %": (forventet_sats or 0.0) * 100,
                "Effektiv %": eff_sats,
                "Status": status,
            })

    # --- Rader uten MVA-kode ---
    if not df_uten.empty:
        rows.append({
            "MVA-kode": "\u2013 Ingen kode",
            "Antall": len(df_uten),
            "Grunnlag": float(df_uten["_b"].sum()),
            "MVA-beløp": 0.0,
            "Sats %": 0.0,
            "Effektiv %": 0.0,
            "Status": "",
        })

    result_df = pd.DataFrame(rows)
    if not result_df.empty:
        # Sorter: rader med kode øverst (etter grunnlag), "Ingen kode" sist
        med_kode_df = result_df[~result_df["MVA-kode"].str.startswith("\u2013")].sort_values("Grunnlag")
        ingen_df = result_df[result_df["MVA-kode"].str.startswith("\u2013")]
        result_df = pd.concat([med_kode_df, ingen_df], ignore_index=True)

    return {
        "rows": result_df,
        "total_bevegelse": total_bevegelse,
        "total_med_kode": total_med_kode,
        "total_uten_kode": total_uten_kode,
        "total_mva": total_faktisk_mva,
        "total_forventet_mva": total_forventet_mva,
    }


def _compute_motpost(df_all: pd.DataFrame, df_rl: pd.DataFrame) -> pd.DataFrame:
    empty = pd.DataFrame(columns=["Konto", "Kontonavn", "Beløp", "Andel", "AntallBilag"])
    if df_rl.empty or df_all is None or df_all.empty:
        return empty
    if "Bilag" not in df_rl.columns or "Bilag" not in df_all.columns:
        return empty
    rl_bilag = set(df_rl["Bilag"].dropna().astype(str).unique())
    rl_kontoer = set(df_rl["Konto"].dropna().astype(str).unique()) if "Konto" in df_rl.columns else set()
    mask = df_all["Bilag"].astype(str).isin(rl_bilag)
    df_mp = df_all.loc[mask].copy()
    if "Konto" in df_mp.columns:
        df_mp = df_mp.loc[~df_mp["Konto"].astype(str).isin(rl_kontoer)]
    if df_mp.empty:
        return empty
    df_mp = df_mp.copy()
    df_mp["_b"] = pd.to_numeric(df_mp["Beløp"], errors="coerce").fillna(0)
    grp = (
        df_mp.groupby("Konto", sort=False)
        .agg(Kontonavn=("Kontonavn", "first"), Beløp=("_b", "sum"), AntallBilag=("Bilag", "nunique"))
        .reset_index()
    )
    grp["_abs"] = grp["Beløp"].abs()
    grp = grp.sort_values("_abs", ascending=False)
    total = grp["_abs"].sum()
    grp["Andel"] = (grp["_abs"] / total * 100).round(1) if total > 0 else 0.0
    return grp[["Konto", "Kontonavn", "Beløp", "Andel", "AntallBilag"]]


# ---------------------------------------------------------------------------
# Widget-hjelper
# ---------------------------------------------------------------------------

def _make_tree(
    parent: object,
    cols: tuple,
    widths: dict,
    *,
    text_cols: tuple = (),
    stretch_col: str | None = None,
    with_hscroll: bool = False,
) -> "ttk.Treeview":
    frame = ttk.Frame(parent)  # type: ignore[misc]
    frame.grid(row=0, column=0, sticky="nsew")
    frame.columnconfigure(0, weight=1)
    frame.rowconfigure(0, weight=1)

    tree = ttk.Treeview(frame, columns=cols, show="headings")  # type: ignore[misc]
    for col in cols:
        tree.column(col, width=widths.get(col, 100),
                    anchor="w" if col in text_cols else "e",
                    stretch=(col == stretch_col))
        tree.heading(col, text=col)

    vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)  # type: ignore[misc]
    tree.configure(yscrollcommand=vsb.set)
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")

    if with_hscroll:
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)  # type: ignore[misc]
        tree.configure(xscrollcommand=hsb.set)
        hsb.grid(row=1, column=0, sticky="ew")

    _attach_sort(tree)
    return tree


# ---------------------------------------------------------------------------
# StatistikkPage
# ---------------------------------------------------------------------------

class StatistikkPage(ttk.Frame):  # type: ignore[misc]

    def __init__(self, parent: object) -> None:
        super().__init__(parent)  # type: ignore[call-arg]
        self._analyse_page: Optional[object] = None
        self._current_regnr: Optional[int] = None
        self._current_rl_name: str = ""
        self._rl_options: list[tuple[int, str]] = []
        self._maned_frame: Optional[object] = None
        self._df_rl_last: Optional[pd.DataFrame] = None
        self._df_all_last: Optional[pd.DataFrame] = None
        self._mva_result_last: Optional[dict] = None
        self._motpost_data_last: Optional[pd.DataFrame] = None
        self._build_ui()

    # ------------------------------------------------------------------
    # Offentlig API

    def set_analyse_page(self, page: object) -> None:
        self._analyse_page = page

    def show_regnr(self, regnr: int) -> None:
        for r, name in self._rl_options:
            if r == regnr:
                self._var_rl.set(f"{r} \u2013 {name}")
                self._current_regnr = regnr
                self._current_rl_name = name
                self._refresh()
                return
        self._current_regnr = regnr
        self._current_rl_name = str(regnr)
        self._refresh()

    def refresh_from_session(self, session: object = None, **_kw: object) -> None:
        self._reload_rl_options()
        if self._current_regnr is not None:
            self._refresh()

    # ------------------------------------------------------------------
    # UI

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)

        # Topplinje
        top = ttk.Frame(self, padding=(8, 6, 8, 4))
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)
        ttk.Label(top, text="Regnskapslinje:").grid(row=0, column=0, sticky="w")
        self._var_rl = tk.StringVar()
        self._combo = ttk.Combobox(top, textvariable=self._var_rl, state="readonly", width=55)
        self._combo.grid(row=0, column=1, sticky="ew", padx=(8, 4))
        self._combo.bind("<<ComboboxSelected>>", self._on_combo_select)
        ttk.Button(top, text="Vis", command=self._refresh, width=8).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(top, text="Eksporter arbeidsdokument \u2192 Excel", command=self._export).grid(row=0, column=3)

        # Nøkkeltall + ekstra stats
        kpi_outer = ttk.LabelFrame(self, text="Nøkkeltall", padding=(8, 4, 8, 6))
        kpi_outer.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 4))
        kpi_outer.columnconfigure(0, weight=1)

        kpi_frame = ttk.Frame(kpi_outer)
        kpi_frame.grid(row=0, column=0, sticky="ew")
        self._kpi_vars: dict[str, tk.StringVar] = {}
        for i, (key, label) in enumerate([
            ("ub", "UB"), ("ub_fjor", "UB i fjor"), ("endring_kr", "Endring (kr)"),
            ("endring_pct", "Endring %"), ("antall", "Antall bilag"),
        ]):
            f = ttk.Frame(kpi_frame)
            f.grid(row=0, column=i, padx=(0 if i == 0 else 20, 0), sticky="w")
            ttk.Label(f, text=label, font=("", 8), foreground="#666666").pack(anchor="w")
            var = tk.StringVar(value="\u2013")
            self._kpi_vars[key] = var
            ttk.Label(f, textvariable=var, font=("", 12, "bold")).pack(anchor="w")

        # Separator
        ttk.Separator(kpi_outer, orient="horizontal").grid(row=1, column=0, sticky="ew", pady=(6, 4))

        # Ekstra analytiske nøkkeltall (rad 2)
        ext_frame = ttk.Frame(kpi_outer)
        ext_frame.grid(row=2, column=0, sticky="ew")
        self._ext_vars: dict[str, tk.StringVar] = {}
        for i, (key, label) in enumerate([
            ("top10", "Topp 10 bilag"),
            ("n_bilag", "Unike bilag"),
            ("n_kunder", "Unike kunder"),
            ("mnd_max", "Største måned"),
            ("anomali", "Anomale måneder"),
            ("runde", "Runde beløp"),
        ]):
            f = ttk.Frame(ext_frame)
            f.grid(row=0, column=i, padx=(0 if i == 0 else 16, 0), sticky="w")
            ttk.Label(f, text=label, font=("", 8), foreground="#888888").pack(anchor="w")
            var = tk.StringVar(value="\u2013")
            self._ext_vars[key] = var
            ttk.Label(f, textvariable=var, font=("", 10)).pack(anchor="w")

        # Notebook
        self._nb = ttk.Notebook(self)
        self._nb.grid(row=2, column=0, sticky="nsew", padx=8, pady=(0, 6))

        self._tab_kontoer = ttk.Frame(self._nb)
        self._tab_maned = ttk.Frame(self._nb)
        self._tab_bilag = ttk.Frame(self._nb)
        self._tab_mva = ttk.Frame(self._nb)
        self._tab_motpost = ttk.Frame(self._nb)

        self._nb.add(self._tab_kontoer, text="Kontoer")
        self._nb.add(self._tab_maned, text="Månedspivot")
        self._nb.add(self._tab_bilag, text="Bilag-analyse")
        self._nb.add(self._tab_mva, text="MVA-analyse")
        self._nb.add(self._tab_motpost, text="Motpostfordeling")

        for tab in (self._tab_kontoer, self._tab_maned, self._tab_bilag, self._tab_mva, self._tab_motpost):
            tab.columnconfigure(0, weight=1)
            tab.rowconfigure(0, weight=1)

        # --- Kontoer ---
        self._tree_kontoer = _make_tree(
            self._tab_kontoer,
            ("Konto", "Kontonavn", "IB", "Bevegelse", "UB", "Antall"),
            {"Konto": 80, "Kontonavn": 260, "IB": 140, "Bevegelse": 140, "UB": 140, "Antall": 80},
            text_cols=("Konto", "Kontonavn"), stretch_col="Kontonavn",
        )
        self._tree_kontoer.bind("<Double-Button-1>", self._on_kontoer_doubleclick)
        ttk.Label(
            self._tab_kontoer, text="Dobbeltklikk en rad for å se transaksjoner",
            foreground="#888888", font=("", 8),
        ).grid(row=1, column=0, sticky="w", padx=6, pady=(2, 2))

        # --- Månedspivot (dynamisk) ---
        self._maned_frame = self._tab_maned

        # --- Bilag-analyse ---
        self._tree_bilag = _make_tree(
            self._tab_bilag,
            ("Bilag", "Dato", "Tekst", "Sum beløp", "Antall poster", "Kontoer"),
            {"Bilag": 80, "Dato": 90, "Tekst": 280, "Sum beløp": 140, "Antall poster": 90, "Kontoer": 200},
            text_cols=("Bilag", "Dato", "Tekst", "Kontoer"), stretch_col="Tekst",
        )
        self._tree_bilag.bind("<Double-Button-1>", self._on_bilag_doubleclick)
        ttk.Label(
            self._tab_bilag, text="Dobbeltklikk et bilag for å se enkeltposteringene",
            foreground="#888888", font=("", 8),
        ).grid(row=1, column=0, sticky="w", padx=6, pady=(2, 2))

        # --- MVA-analyse ---
        self._tab_mva.rowconfigure(1, weight=0)
        self._tab_mva.rowconfigure(2, weight=0)
        self._tree_mva = _make_tree(
            self._tab_mva,
            ("MVA-kode", "Antall", "Grunnlag", "MVA-beløp", "Sats %", "Effektiv %", "Status"),
            {"MVA-kode": 100, "Antall": 70, "Grunnlag": 150, "MVA-beløp": 150,
             "Sats %": 70, "Effektiv %": 80, "Status": 240},
            text_cols=("MVA-kode", "Status"), stretch_col="Status",
        )
        self._tree_mva.tag_configure("ok", foreground="#2E7D32")
        self._tree_mva.tag_configure("avvik", foreground="#C62828")
        self._tree_mva.tag_configure("ingen", foreground="#888888")
        self._tree_mva.bind("<Double-Button-1>", self._on_mva_doubleclick)

        # Avstemmingspanel
        avsf = ttk.LabelFrame(self._tab_mva, text="Avstemming mot totale salgsinntekter", padding=(8, 4))
        avsf.grid(row=1, column=0, sticky="ew", padx=6, pady=(4, 2))
        avsf.columnconfigure(1, weight=1)
        avsf.columnconfigure(3, weight=1)
        self._mva_avs_vars: dict[str, tk.StringVar] = {}
        _avs_fields = [
            ("bev", "Total bevegelse (RL):", 0, 0),
            ("med", "  Herav med MVA-kode:", 1, 0),
            ("uten", "  Herav uten kode:", 2, 0),
            ("faktisk", "Total faktisk MVA:", 0, 2),
            ("forventet", "Forventet MVA:", 1, 2),
            ("avvik", "Avvik:", 2, 2),
        ]
        for key, label, row_i, col_i in _avs_fields:
            ttk.Label(avsf, text=label, foreground="#555555").grid(
                row=row_i, column=col_i, sticky="w", padx=(0 if col_i == 0 else 20, 4)
            )
            var = tk.StringVar(value="\u2013")
            self._mva_avs_vars[key] = var
            fg = "#2E7D32" if key == "avvik" else "#222222"
            ttk.Label(avsf, textvariable=var, foreground=fg, font=("TkFixedFont", 10)).grid(
                row=row_i, column=col_i + 1, sticky="e"
            )

        ttk.Label(
            self._tab_mva, text="Dobbeltklikk en kode for å se transaksjoner",
            foreground="#888888", font=("", 8),
        ).grid(row=2, column=0, sticky="w", padx=6, pady=(0, 2))

        # --- Motpostfordeling ---
        self._tree_motpost = _make_tree(
            self._tab_motpost,
            ("Konto", "Kontonavn", "Beløp", "Andel %", "Antall bilag"),
            {"Konto": 80, "Kontonavn": 260, "Beløp": 140, "Andel %": 80, "Antall bilag": 90},
            text_cols=("Konto", "Kontonavn"), stretch_col="Kontonavn",
        )
        self._tree_motpost.bind("<Double-Button-1>", self._on_motpost_doubleclick)
        motpost_bot = ttk.Frame(self._tab_motpost)
        motpost_bot.grid(row=1, column=0, sticky="ew", padx=6, pady=(2, 2))
        motpost_bot.columnconfigure(0, weight=1)
        ttk.Label(
            motpost_bot, text="Dobbeltklikk en konto for å se tilhørende bilag",
            foreground="#888888", font=("", 8),
        ).grid(row=0, column=0, sticky="w")
        ttk.Button(
            motpost_bot, text="\U0001f4ca  Vis flowchart", command=self._show_motpost_flowchart,
        ).grid(row=0, column=1, sticky="e")

        # Statuslinje
        self._status_var = tk.StringVar(value="Velg en regnskapslinje og trykk Vis")
        ttk.Label(self, textvariable=self._status_var, foreground="#555555").grid(
            row=3, column=0, sticky="w", padx=8, pady=(0, 4)
        )

    # ------------------------------------------------------------------
    # Dropdown

    def _reload_rl_options(self) -> None:
        page = self._analyse_page
        if page is None:
            return
        regnskapslinjer = getattr(page, "_rl_regnskapslinjer", None)
        if regnskapslinjer is None or (hasattr(regnskapslinjer, "empty") and regnskapslinjer.empty):
            self._rl_options = []
            try:
                self._combo["values"] = []
            except Exception:
                pass
            return
        try:
            from regnskap_mapping import normalize_regnskapslinjer
            regn = normalize_regnskapslinjer(regnskapslinjer)
            self._rl_options = [
                (int(r["regnr"]), str(r.get("regnskapslinje", "") or ""))
                for _, r in regn.iterrows()
            ]
            self._combo["values"] = [f"{r} \u2013 {n}" for r, n in self._rl_options]
        except Exception as exc:
            log.warning("_reload_rl_options: %s", exc)

    def _on_combo_select(self, event: object = None) -> None:
        val = self._var_rl.get()
        if not val:
            return
        try:
            parts = val.split("\u2013", 1)
            self._current_regnr = int(parts[0].strip())
            self._current_rl_name = parts[1].strip() if len(parts) > 1 else str(self._current_regnr)
        except (ValueError, IndexError):
            pass

    # ------------------------------------------------------------------
    # Refresh

    def _refresh(self) -> None:
        self._on_combo_select()
        if self._current_regnr is None:
            self._status_var.set("Velg en regnskapslinje")
            return
        page = self._analyse_page
        if page is None:
            self._status_var.set("Ikke koblet til Analyse-siden")
            return
        df_all = getattr(page, "_df_filtered", None)
        if df_all is None or (hasattr(df_all, "empty") and df_all.empty):
            self._status_var.set("Ingen transaksjonsdata lastet")
            return

        ranges = _get_konto_ranges(page, self._current_regnr)
        df_rl = _filter_df(df_all, ranges)

        # Lagre for drill-down
        self._df_rl_last = df_rl
        self._df_all_last = df_all

        self._update_kpi(page, self._current_regnr)
        kontoer_data, ib_label = _compute_kontoer(df_rl, page)
        self._populate_kontoer(kontoer_data, ib_label)
        self._rebuild_maned_pivot(df_rl)
        self._populate_bilag(_compute_bilag(df_rl))
        mva_result = _compute_mva(df_rl, df_all)
        self._mva_result_last = mva_result
        self._populate_mva(mva_result)
        motpost_data = _compute_motpost(df_all, df_rl)
        self._motpost_data_last = motpost_data
        self._populate_motpost(motpost_data)
        self._populate_extra_stats(_compute_extra_stats(df_rl))

        self._status_var.set(
            f"{self._current_regnr} \u2013 {self._current_rl_name}"
            f"  \u00b7  {len(df_rl):,} transaksjoner"
        )

    def _update_kpi(self, page: object, regnr: int) -> None:
        pivot_df = getattr(page, "_pivot_df_last", None)
        blank = "\u2013"
        if pivot_df is None or pivot_df.empty:
            for var in self._kpi_vars.values():
                var.set(blank)
            return
        row = next(
            (r for _, r in pivot_df.iterrows() if _safe_int(r.get("regnr", -1)) == regnr),
            None,
        )
        if row is None:
            for var in self._kpi_vars.values():
                var.set(blank)
            return
        self._kpi_vars["ub"].set(_fmt_amount(row.get("UB")))
        self._kpi_vars["ub_fjor"].set(_fmt_amount(row.get("UB_fjor")))
        self._kpi_vars["endring_kr"].set(_fmt_amount(row.get("Endring")))
        self._kpi_vars["endring_pct"].set(_fmt_pct(row.get("Endring_pct")))
        antall = row.get("Antall")
        self._kpi_vars["antall"].set(str(_safe_int(antall)) if antall is not None else blank)

    # ------------------------------------------------------------------
    # Populate

    def _populate_extra_stats(self, stats: dict) -> None:
        blank = "\u2013"
        top10 = stats.get("top10_pct")
        self._ext_vars["top10"].set(f"{top10:.0f} %" if top10 is not None else blank)
        n_b = stats.get("n_bilag")
        self._ext_vars["n_bilag"].set(f"{n_b:,}" if n_b is not None else blank)
        n_k = stats.get("n_kunder")
        self._ext_vars["n_kunder"].set(f"{n_k:,}" if n_k is not None else blank)
        maks = stats.get("mnd_max_name")
        maks_v = stats.get("mnd_max_val")
        self._ext_vars["mnd_max"].set(f"{maks}  {_fmt_amount(maks_v)}" if maks else blank)
        n_an = stats.get("n_anomali_mnd")
        self._ext_vars["anomali"].set(
            f"{n_an} mnd" + (" \u26a0" if n_an and n_an > 0 else "  \u2713")
            if n_an is not None else blank
        )
        runde = stats.get("runde_pct")
        self._ext_vars["runde"].set(f"{runde:.0f} %" if runde is not None else blank)

    def _populate_kontoer(self, grp: pd.DataFrame, ib_label: str = "IB") -> None:
        tree = self._tree_kontoer
        # Oppdater kolonneoverskrift dynamisk
        try:
            tree.heading("IB", text=ib_label)
        except Exception:
            pass
        tree.delete(*tree.get_children())
        if grp.empty:
            return

        sum_ib = sum_bev = sum_ub = sum_ant = 0.0
        has_sb = False

        for _, row in grp.iterrows():
            ib_v = row.get("IB")
            ub_v = row.get("UB")
            ib_ok = ib_v is not None and str(ib_v) not in ("", "nan")
            ub_ok = ub_v is not None and str(ub_v) not in ("", "nan")
            bev = _safe_float(row["Bevegelse"])
            ant = _safe_int(row["Antall"])
            if ib_ok:
                sum_ib += _safe_float(ib_v)
                has_sb = True
            if ub_ok:
                sum_ub += _safe_float(ub_v)
            sum_bev += bev
            sum_ant += ant
            tree.insert("", tk.END, values=(
                str(row["Konto"]),
                str(row.get("Kontonavn", "") or ""),
                _fmt_amount(ib_v) if ib_ok else "",
                _fmt_amount(bev),
                _fmt_amount(ub_v) if ub_ok else "",
                ant,
            ))

        # Totalsrad
        tree.insert("", tk.END, values=(
            "", "Sum",
            _fmt_amount(sum_ib) if has_sb else "",
            _fmt_amount(sum_bev),
            _fmt_amount(sum_ub) if has_sb else "",
            _safe_int(sum_ant),
        ), tags=("sum",))
        tree.tag_configure("sum", font=("", 10, "bold"))

    def _rebuild_maned_pivot(self, df_rl: pd.DataFrame) -> None:
        frame = self._maned_frame
        if frame is None:
            return
        for child in frame.winfo_children():
            try:
                child.destroy()
            except Exception:
                pass

        months, pivot = _compute_maned_pivot(df_rl)
        if pivot.empty:
            ttk.Label(frame, text="Ingen data").grid(row=0, column=0, padx=8, pady=8)
            return

        mnds = [m.replace("-", "\u2011") for m in months]
        all_cols = ("Konto", "Kontonavn") + tuple(mnds) + ("Sum",)

        inner = ttk.Frame(frame)
        inner.grid(row=0, column=0, sticky="nsew")
        inner.columnconfigure(0, weight=1)
        inner.rowconfigure(0, weight=1)

        tree = ttk.Treeview(inner, columns=all_cols, show="headings")  # type: ignore[misc]
        tree.column("Konto", width=70, anchor="w", stretch=False)
        tree.heading("Konto", text="Konto")
        tree.column("Kontonavn", width=220, anchor="w", stretch=True)
        tree.heading("Kontonavn", text="Kontonavn")
        for col, m in zip(mnds, months):
            try:
                from calendar import month_abbr
                yr, mn = m.split("-")
                short = f"{month_abbr[int(mn)]} {yr[2:]}"
            except Exception:
                short = col
            tree.column(col, width=90, anchor="e", stretch=False)
            tree.heading(col, text=short)
        tree.column("Sum", width=110, anchor="e", stretch=False)
        tree.heading("Sum", text="Sum")

        vsb = ttk.Scrollbar(inner, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(inner, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        _attach_sort(tree)

        # Negative verdier i rødt
        tree.tag_configure("neg", foreground="#C62828")

        totals: dict[str, float] = {c: 0.0 for c in mnds}
        grand_total = 0.0
        for _, row in pivot.iterrows():
            vals: list = [str(row["Konto"]), str(row.get("Kontonavn", "") or "")]
            row_neg = False
            for col, m in zip(mnds, months):
                v = _safe_float(row.get(m, 0))
                totals[col] = totals.get(col, 0.0) + v
                vals.append(_fmt_amount(v) if v != 0.0 else "")
                if v < 0:
                    row_neg = True
            s = _safe_float(row.get("Sum", 0))
            grand_total += s
            vals.append(_fmt_amount(s))
            tree.insert("", tk.END, values=tuple(vals), tags=("neg",) if row_neg else ())

        # Sum-rad
        sv = ["", "Sum"] + [_fmt_amount(totals[c]) for c in mnds] + [_fmt_amount(grand_total)]
        tree.insert("", tk.END, values=tuple(sv), tags=("sum",))
        tree.tag_configure("sum", font=("", 10, "bold"))

    def _populate_bilag(self, grp: pd.DataFrame) -> None:
        tree = self._tree_bilag
        tree.delete(*tree.get_children())
        for _, row in grp.iterrows():
            bel = _safe_float(row["Sum beløp"])
            tree.insert("", tk.END, values=(
                str(row.get("Bilag", "") or ""),
                str(row.get("Dato", "") or ""),
                str(row.get("Tekst", "") or ""),
                _fmt_amount(bel),
                _safe_int(row["Antall poster"]),
                str(row.get("Kontoer", "") or ""),
            ), tags=("neg",) if bel < 0 else ())
        self._tree_bilag.tag_configure("neg", foreground="#C62828")

    def _populate_mva(self, result: dict) -> None:
        tree = self._tree_mva
        tree.delete(*tree.get_children())

        # Oppdater avstemmingspanel
        bev = result.get("total_bevegelse", 0.0)
        med = result.get("total_med_kode", 0.0)
        uten = result.get("total_uten_kode", 0.0)
        faktisk = result.get("total_mva", 0.0)
        forventet = result.get("total_forventet_mva", 0.0)
        avvik_kr = abs(faktisk) - forventet

        self._mva_avs_vars["bev"].set(_fmt_amount(bev))
        self._mva_avs_vars["med"].set(_fmt_amount(med))
        self._mva_avs_vars["uten"].set(_fmt_amount(uten))
        self._mva_avs_vars["faktisk"].set(_fmt_amount(faktisk))
        self._mva_avs_vars["forventet"].set(_fmt_amount(-forventet if faktisk < 0 else forventet))
        avvik_txt = _fmt_amount(avvik_kr)
        if abs(avvik_kr) < 1 and forventet > 0:
            avvik_txt += "  \u2713 OK"
        elif forventet > 0:
            avvik_txt += "  \u26a0"
        self._mva_avs_vars["avvik"].set(avvik_txt)

        grp = result.get("rows", pd.DataFrame())
        if grp is None or grp.empty:
            return

        for _, row in grp.iterrows():
            kode = str(row.get("MVA-kode", ""))
            status = str(row.get("Status", ""))
            ingen = kode.startswith("\u2013")
            tag = "ingen" if ingen else ("ok" if "\u2713" in status else ("avvik" if "\u26a0" in status else ""))
            tree.insert("", tk.END, values=(
                kode,
                _safe_int(row["Antall"]),
                _fmt_amount(row["Grunnlag"]),
                _fmt_amount(row["MVA-beløp"]) if not ingen else "",
                _fmt_pct(row.get("Sats %")) if not ingen else "",
                _fmt_pct(row.get("Effektiv %")) if not ingen else "",
                status,
            ), tags=(tag,) if tag else ())

    def _populate_motpost(self, grp: pd.DataFrame) -> None:
        tree = self._tree_motpost
        tree.delete(*tree.get_children())
        for _, row in grp.iterrows():
            bel = _safe_float(row["Beløp"])
            tree.insert("", tk.END, values=(
                str(row.get("Konto", "") or ""),
                str(row.get("Kontonavn", "") or ""),
                _fmt_amount(bel),
                f"{float(row['Andel']):.1f}",
                _safe_int(row["AntallBilag"]),
            ), tags=("neg",) if bel < 0 else ())
        self._tree_motpost.tag_configure("neg", foreground="#C62828")

    # ------------------------------------------------------------------
    # Drill-down: dobbeltklikk i Bilag-analyse

    # ------------------------------------------------------------------
    # Drill-down helpers

    def _open_tx_popup(self, title: str, df: pd.DataFrame) -> None:
        """Generisk transaksjonspopup — viser df som en sortérbar tabell."""
        if df is None or df.empty:
            messagebox.showinfo("Ingen data", "Ingen transaksjoner funnet.", parent=self)
            return
        top = tk.Toplevel(self)
        top.title(title)
        top.geometry("980x440")
        top.transient(self)

        frame = ttk.Frame(top, padding=8)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)

        pref_cols = ("Dato", "Bilag", "Konto", "Kontonavn", "Tekst", "Beløp", "MVA-kode", "MVA-beløp")
        avail = [c for c in pref_cols if c in df.columns]
        widths = {"Dato": 90, "Bilag": 80, "Konto": 70, "Kontonavn": 180,
                  "Tekst": 270, "Beløp": 120, "MVA-kode": 70, "MVA-beløp": 110}
        text_cols_p = ("Dato", "Bilag", "Konto", "Kontonavn", "Tekst", "MVA-kode")

        tree = ttk.Treeview(frame, columns=avail, show="headings")
        for col in avail:
            tree.heading(col, text=col)
            tree.column(col, width=widths.get(col, 100),
                        anchor="w" if col in text_cols_p else "e")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        _attach_sort(tree)
        tree.tag_configure("neg", foreground="#C62828")

        sum_bel = 0.0
        for _, row in df.iterrows():
            dato_str = ""
            if "Dato" in row.index:
                try:
                    d = pd.to_datetime(row["Dato"], dayfirst=True, errors="coerce")
                    dato_str = d.strftime("%d.%m.%Y") if not pd.isna(d) else str(row["Dato"])
                except Exception:
                    dato_str = str(row.get("Dato", ""))
            bel = _safe_float(row.get("Beløp"))
            sum_bel += bel
            vals: list = []
            for col in avail:
                if col == "Dato":
                    vals.append(dato_str)
                elif col in ("Beløp", "MVA-beløp"):
                    vals.append(_fmt_amount(row.get(col)))
                else:
                    vals.append(str(row.get(col, "") or ""))
            tree.insert("", tk.END, values=tuple(vals), tags=("neg",) if bel < 0 else ())

        bot = ttk.Frame(frame)
        bot.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        bot.columnconfigure(0, weight=1)
        ttk.Label(
            bot, text=f"{len(df):,} poster  |  Sum beløp: {_fmt_amount(sum_bel)}",
            font=("", 10, "bold"),
        ).grid(row=0, column=0, sticky="w")
        ttk.Button(bot, text="Lukk", command=top.destroy).grid(row=0, column=1, sticky="e")

    def _on_kontoer_doubleclick(self, event: object = None) -> None:
        tree = self._tree_kontoer
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], "values")
        if not vals or not str(vals[0]).strip():
            return
        konto = str(vals[0]).strip()
        df_rl = self._df_rl_last
        if df_rl is None or df_rl.empty or "Konto" not in df_rl.columns:
            return
        df_detail = df_rl[df_rl["Konto"].astype(str) == konto]
        kontonavn = str(vals[1]) if len(vals) > 1 else konto
        self._open_tx_popup(f"Transaksjoner — konto {konto} {kontonavn}", df_detail)

    def _on_mva_doubleclick(self, event: object = None) -> None:
        tree = self._tree_mva
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], "values")
        if not vals:
            return
        kode = str(vals[0]).strip()
        df_rl = self._df_rl_last
        if df_rl is None or df_rl.empty or "MVA-kode" not in df_rl.columns:
            return
        if kode.startswith("\u2013"):
            # "Ingen kode" — vis rader uten MVA-kode
            mask = df_rl["MVA-kode"].isna() | (df_rl["MVA-kode"].astype(str).str.strip() == "")
            df_detail = df_rl[mask]
            title = "Transaksjoner uten MVA-kode"
        else:
            df_detail = df_rl[df_rl["MVA-kode"].astype(str).str.strip() == kode]
            title = f"Transaksjoner — MVA-kode {kode}"
        self._open_tx_popup(title, df_detail)

    def _on_motpost_doubleclick(self, event: object = None) -> None:
        tree = self._tree_motpost
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], "values")
        if not vals or not str(vals[0]).strip():
            return
        konto = str(vals[0]).strip()
        kontonavn = str(vals[1]) if len(vals) > 1 else konto
        df_rl = self._df_rl_last
        df_all = self._df_all_last
        if df_rl is None or df_all is None or "Bilag" not in df_rl.columns:
            return
        # Finn bilag fra RL, filtrer df_all til de bilagene + korrekt konto
        rl_bilag = set(df_rl["Bilag"].dropna().astype(str).unique())
        mask = df_all["Bilag"].astype(str).isin(rl_bilag) & (df_all["Konto"].astype(str) == konto)
        df_detail = df_all[mask]
        self._open_tx_popup(f"Motpostbilag — konto {konto} {kontonavn}", df_detail)

    def _show_motpost_flowchart(self) -> None:
        """Genererer D3 Sankey-diagram som HTML og åpner i nettleser."""
        import json
        import tempfile
        import webbrowser

        grp = self._motpost_data_last
        df_rl = self._df_rl_last
        if grp is None or grp.empty:
            messagebox.showinfo("Ingen data", "Vis statistikk for en regnskapslinje først.", parent=self)
            return

        grp_top = grp.head(15).reset_index(drop=True)

        # --- Bygg nodedata og lenker ---
        src_name = f"{self._current_regnr} {self._current_rl_name}"
        if df_rl is not None and not df_rl.empty and "Konto" in df_rl.columns:
            kontoer = sorted(df_rl["Konto"].dropna().astype(str).unique())
            kontoer_str = ", ".join(kontoer[:5]) + ("…" if len(kontoer) > 5 else "")
            src_name += f"  ({kontoer_str})"

        nodes = [{"name": src_name, "group": "source"}]
        links = []
        for _, row in grp_top.iterrows():
            konto = str(row.get("Konto", ""))
            navn = str(row.get("Kontonavn", ""))
            beløp = abs(float(row.get("Beløp", 0)))
            andel = float(row.get("Andel", 0))
            node_name = f"{konto}  {navn}"
            nodes.append({
                "name": node_name,
                "group": "target",
                "beløp": beløp,
                "andel": andel,
            })
            links.append({
                "source": 0,
                "target": len(nodes) - 1,
                "value": max(beløp, 1.0),
                "andel": andel,
                "beløp_fmt": _fmt_amount(float(row.get("Beløp", 0))),
            })

        nodes_json = json.dumps(nodes, ensure_ascii=False)
        links_json = json.dumps(links, ensure_ascii=False)
        tittel = f"Motpostfordeling — {self._current_regnr} {self._current_rl_name}"

        html = f"""<!DOCTYPE html>
<html lang="no">
<head>
<meta charset="utf-8">
<title>{tittel}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', system-ui, sans-serif; background: #f8f9fa; color: #1a1a2e; }}
  #header {{ padding: 18px 28px 10px; border-bottom: 1px solid #dee2e6; background: #fff; }}
  #header h1 {{ font-size: 16px; font-weight: 600; color: #1565C0; }}
  #header p {{ font-size: 12px; color: #6c757d; margin-top: 4px; }}
  #chart {{ padding: 20px 28px; }}
  svg {{ width: 100%; overflow: visible; }}
  .link {{ fill: none; stroke-opacity: 0.45; transition: stroke-opacity 0.2s; }}
  .link:hover {{ stroke-opacity: 0.75; cursor: pointer; }}
  .node rect {{ rx: 6; ry: 6; stroke-width: 1.5; }}
  .node text {{ font-size: 12px; }}
  .tooltip {{
    position: fixed; background: rgba(0,0,0,0.82); color: #fff;
    padding: 8px 12px; border-radius: 6px; font-size: 12px;
    pointer-events: none; display: none; white-space: nowrap;
    box-shadow: 0 4px 12px rgba(0,0,0,0.3);
  }}
</style>
</head>
<body>
<div id="header">
  <h1>{tittel}</h1>
  <p>Hover over strømmene for detaljer. Bredden er proporsjonal med beløpet.</p>
</div>
<div id="chart"><svg id="sankey"></svg></div>
<div class="tooltip" id="tip"></div>

<script src="https://cdn.jsdelivr.net/npm/d3@7/dist/d3.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/d3-sankey@0.12/dist/d3-sankey.min.js"></script>
<script>
const rawNodes = {nodes_json};
const rawLinks = {links_json};

const W = Math.max(window.innerWidth - 56, 600);
const nodeCount = rawNodes.length;
const H = Math.max(nodeCount * 52 + 80, 360);

const svg = d3.select("#sankey")
  .attr("viewBox", `0 0 ${{W}} ${{H}}`)
  .attr("height", H);

const sankey = d3.sankey()
  .nodeId(d => d.index)
  .nodeWidth(200)
  .nodePadding(18)
  .extent([[16, 16], [W - 16, H - 16]]);

const graph = sankey({{
  nodes: rawNodes.map((d, i) => ({{ ...d, index: i }})),
  links: rawLinks.map(d => ({{ ...d }})),
}});

const color = d3.scaleOrdinal()
  .domain(["source"])
  .range(["#1565C0"])
  .unknown("#2196F3");

// Links
const link = svg.append("g").attr("fill", "none")
  .selectAll("path")
  .data(graph.links)
  .join("path")
    .attr("class", "link")
    .attr("d", d3.sankeyLinkHorizontal())
    .attr("stroke", d => color(graph.nodes[d.target.index].group))
    .attr("stroke-width", d => Math.max(1, d.width));

// Nodes
const node = svg.append("g")
  .selectAll("g")
  .data(graph.nodes)
  .join("g")
    .attr("class", "node");

node.append("rect")
  .attr("x", d => d.x0)
  .attr("y", d => d.y0)
  .attr("width", d => d.x1 - d.x0)
  .attr("height", d => Math.max(1, d.y1 - d.y0))
  .attr("fill", d => d.group === "source" ? "#E3F2FD" : "#EEF2FF")
  .attr("stroke", d => d.group === "source" ? "#1565C0" : "#3F51B5");

node.append("text")
  .attr("x", d => d.x0 + (d.x1 - d.x0) / 2)
  .attr("y", d => (d.y0 + d.y1) / 2 - (d.group === "source" ? 6 : 7))
  .attr("text-anchor", "middle")
  .attr("dominant-baseline", "middle")
  .attr("font-weight", "600")
  .attr("fill", d => d.group === "source" ? "#0D47A1" : "#1a1a2e")
  .attr("font-size", d => d.group === "source" ? "12px" : "12px")
  .each(function(d) {{
    const el = d3.select(this);
    const parts = d.name.split("  ");
    if (parts.length > 1 && d.group === "target") {{
      el.append("tspan")
        .attr("x", d.x0 + (d.x1 - d.x0) / 2)
        .attr("dy", "0")
        .attr("font-weight", "700")
        .attr("fill", "#1565C0")
        .text(parts[0]);
      el.append("tspan")
        .attr("x", d.x0 + (d.x1 - d.x0) / 2)
        .attr("dy", "1.3em")
        .attr("font-weight", "400")
        .attr("fill", "#333")
        .text(parts.slice(1).join("  ").trim());
    }} else {{
      el.text(d.name);
    }}
  }});

// Andel-label under kontonavn for target-noder
node.filter(d => d.group === "target")
  .append("text")
  .attr("x", d => d.x0 + (d.x1 - d.x0) / 2)
  .attr("y", d => (d.y0 + d.y1) / 2 + 18)
  .attr("text-anchor", "middle")
  .attr("font-size", "11px")
  .attr("fill", "#555")
  .text(d => {{
    const lnk = graph.links.find(l => l.target.index === d.index);
    if (!lnk) return "";
    return `${{lnk.andel.toFixed(1)}}%  ${{lnk.beløp_fmt}}`;
  }});

// Tooltip
const tip = document.getElementById("tip");
link.on("mousemove", function(event, d) {{
  tip.style.display = "block";
  tip.style.left = (event.clientX + 14) + "px";
  tip.style.top = (event.clientY - 10) + "px";
  tip.innerHTML = `<b>${{d.target.name}}</b><br>Andel: ${{d.andel.toFixed(1)}}%<br>Beløp: ${{d.beløp_fmt}}`;
}}).on("mouseleave", () => tip.style.display = "none");
</script>
</body>
</html>"""

        tmp = tempfile.NamedTemporaryFile(
            mode="w", suffix=".html", delete=False, encoding="utf-8",
            prefix="utvalg_motpost_",
        )
        tmp.write(html)
        tmp.close()
        webbrowser.open(f"file:///{tmp.name.replace(os.sep, '/')}")

    def _on_bilag_doubleclick(self, event: object = None) -> None:
        tree = self._tree_bilag
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], "values")
        if not vals:
            return
        bilag_nr = str(vals[0]).strip()
        df_rl = self._df_rl_last
        if df_rl is None or df_rl.empty or "Bilag" not in df_rl.columns:
            return
        df_detail = df_rl[df_rl["Bilag"].astype(str) == bilag_nr]
        if df_detail.empty:
            return
        self._open_tx_popup(f"Bilag {bilag_nr}", df_detail)

    # ------------------------------------------------------------------
    # Eksport

    def _export(self) -> None:
        if self._current_regnr is None:
            messagebox.showwarning("Ingen valgt linje", "Velg en regnskapslinje og trykk Vis.", parent=self)
            return
        page = self._analyse_page
        if page is None:
            return
        df_all = getattr(page, "_df_filtered", None)
        if df_all is None or (hasattr(df_all, "empty") and df_all.empty):
            messagebox.showwarning("Ingen data", "Ingen transaksjonsdata lastet.", parent=self)
            return

        try:
            import session as _s
            client = getattr(_s, "client", None) or ""
            year = getattr(_s, "year", None) or ""
        except Exception:
            client, year = "", ""

        name_safe = self._current_rl_name.replace("/", "-").replace("\\", "-").replace(":", "")[:40]
        default_name = f"Statistikk_{self._current_regnr}_{name_safe}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        init_dir = str(Path.home())
        try:
            import client_store
            if client and year:
                init_dir = str(client_store.exports_dir(client, year=year))
        except Exception:
            pass

        path = asksaveasfilename(
            parent=self, title="Lagre arbeidsdokument", defaultextension=".xlsx",
            filetypes=[("Excel-arbeidsbok", "*.xlsx")],
            initialfile=default_name, initialdir=init_dir,
        )
        if not path:
            return
        try:
            ranges = _get_konto_ranges(page, self._current_regnr)
            df_rl = _filter_df(df_all, ranges)
            _write_workbook(
                path, regnr=self._current_regnr, rl_name=self._current_rl_name,
                df_rl=df_rl, df_all=df_all, page=page, client=client, year=year,
            )
            self._status_var.set(f"Eksportert: {Path(path).name}")
            _open_file(path)
        except Exception as exc:
            messagebox.showerror("Eksport feilet", str(exc), parent=self)
            log.exception("StatistikkPage: eksport feilet")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _write_workbook(
    path: str, *, regnr: int, rl_name: str,
    df_rl: pd.DataFrame, df_all: pd.DataFrame,
    page: object, client: str = "", year: str = "",
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
    HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
    SUM_FILL = PatternFill("solid", fgColor="D6E2EF")
    AVVIK_FILL = PatternFill("solid", fgColor="FCE4EC")
    OK_FILL = PatternFill("solid", fgColor="E8F5E9")
    THIN = Side(style="thin", color="D9D9D9")
    B = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    NEG_FONT = Font(color="C62828")
    ts = (f"  |  {client}" if client else "") + (f"  {year}" if year else "")

    def _title(ws: object, title: str, n: int) -> None:
        last = get_column_letter(n)
        ws.merge_cells(f"A1:{last}1")  # type: ignore[union-attr]
        ws["A1"] = title  # type: ignore[index]
        ws["A1"].font = Font(size=13, bold=True)  # type: ignore[index]
        ws["A1"].fill = TITLE_FILL  # type: ignore[index]
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")  # type: ignore[index]
        ws.row_dimensions[1].height = 22  # type: ignore[union-attr]
        ws.merge_cells(f"A2:{last}2")  # type: ignore[union-attr]
        ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"  # type: ignore[index]
        ws["A2"].font = Font(italic=True, color="666666", size=9)  # type: ignore[index]

    def _header(ws: object, row: int, cols: list[str]) -> None:
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=col)  # type: ignore[union-attr]
            c.font = Font(bold=True, size=10)
            c.fill = HEADER_FILL
            c.border = B
            c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 18  # type: ignore[union-attr]

    def _amt(ws: object, r: int, c: int, v: object, neg: bool = False) -> None:
        cell = ws.cell(row=r, column=c, value=_safe_float(v))  # type: ignore[union-attr]
        cell.border = B
        cell.number_format = _AMT_FMT
        cell.alignment = Alignment(horizontal="right")
        if neg:
            cell.font = NEG_FONT

    wb = Workbook()

    # Ark 1: Sammendrag
    ws1 = wb.active
    ws1.title = "Sammendrag"
    _title(ws1, f"Statistikk – {regnr} {rl_name}{ts}", 6)
    _header(ws1, 4, ["UB", "UB i fjor", "Endring (kr)", "Endring %", "Antall bilag", ""])
    pivot_df = getattr(page, "_pivot_df_last", None)
    kpi = None
    if pivot_df is not None and not pivot_df.empty:
        kpi = next((r for _, r in pivot_df.iterrows() if _safe_int(r.get("regnr", -1)) == regnr), None)
    if kpi is not None:
        for i, (col, fmt) in enumerate([
            ("UB", _AMT_FMT), ("UB_fjor", _AMT_FMT), ("Endring", _AMT_FMT),
            ("Endring_pct", '0.0"%"'), ("Antall", "#,##0"),
        ], 1):
            v = _safe_int(kpi.get(col)) if fmt == "#,##0" else _safe_float(kpi.get(col))
            c = ws1.cell(5, i, v)
            c.border = B
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right")
    ws1.cell(7, 1).value = "Kontoer"
    ws1.cell(7, 1).font = Font(bold=True, size=11)
    _header(ws1, 8, ["Konto", "Kontonavn", "IB", "Bevegelse", "UB", "Antall"])
    grp_k, _ib_label = _compute_kontoer(df_rl, page)
    dr = 9
    for _, row in grp_k.iterrows():
        ws1.cell(dr, 1, str(row["Konto"])).border = B
        ws1.cell(dr, 2, str(row.get("Kontonavn", "") or "")).border = B
        for ci, cn in [(3, "IB"), (4, "Bevegelse"), (5, "UB")]:
            raw = row.get(cn)
            if raw is not None and str(raw) not in ("", "nan"):
                _amt(ws1, dr, ci, raw, _safe_float(raw) < 0)
            else:
                ws1.cell(dr, ci, None).border = B
        ws1.cell(dr, 6, _safe_int(row["Antall"])).border = B
        dr += 1
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 35
    for l in ["C", "D", "E"]:
        ws1.column_dimensions[l].width = 18
    ws1.column_dimensions["F"].width = 10
    ws1.freeze_panes = "A5"

    # Ark 2: Månedspivot
    ws2 = wb.create_sheet("Månedspivot")
    months, pivot = _compute_maned_pivot(df_rl)
    nc = 2 + len(months) + 1
    _title(ws2, f"Månedspivot – {regnr} {rl_name}{ts}", nc)
    _header(ws2, 4, ["Konto", "Kontonavn"] + months + ["Sum"])
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 30
    dr = 5
    tot_m = {m: 0.0 for m in months}
    gt = 0.0
    for _, row in pivot.iterrows():
        ws2.cell(dr, 1, str(row["Konto"])).border = B
        ws2.cell(dr, 2, str(row.get("Kontonavn", "") or "")).border = B
        for j, m in enumerate(months, 3):
            v = _safe_float(row.get(m, 0))
            tot_m[m] = tot_m.get(m, 0.0) + v
            if v != 0.0:
                _amt(ws2, dr, j, v, v < 0)
            else:
                ws2.cell(dr, j, None).border = B
        s = _safe_float(row.get("Sum", 0))
        gt += s
        _amt(ws2, dr, nc, s, s < 0)
        dr += 1
    # Sum-rad
    ws2.cell(dr, 1, "Sum").font = Font(bold=True)
    ws2.cell(dr, 1).fill = SUM_FILL
    ws2.cell(dr, 1).border = B
    ws2.cell(dr, 2).fill = SUM_FILL
    ws2.cell(dr, 2).border = B
    for j, m in enumerate(months, 3):
        c = ws2.cell(dr, j, tot_m[m])
        c.font = Font(bold=True, color="C62828" if tot_m[m] < 0 else "000000")
        c.fill = SUM_FILL
        c.border = B
        c.number_format = _AMT_FMT
        c.alignment = Alignment(horizontal="right")
    c_gt = ws2.cell(dr, nc, gt)
    c_gt.font = Font(bold=True, color="C62828" if gt < 0 else "000000")
    c_gt.fill = SUM_FILL
    c_gt.border = B
    c_gt.number_format = _AMT_FMT
    c_gt.alignment = Alignment(horizontal="right")
    for j in range(3, nc + 1):
        ws2.column_dimensions[get_column_letter(j)].width = 14
    ws2.freeze_panes = "C5"

    # Ark 3: Bilag-analyse
    ws3 = wb.create_sheet("Bilag-analyse")
    _title(ws3, f"Bilag-analyse – {regnr} {rl_name}{ts}", 6)
    _header(ws3, 4, ["Bilag", "Dato", "Tekst", "Sum beløp", "Antall poster", "Kontoer"])
    grp_b = _compute_bilag(df_rl)
    dr = 5
    for _, row in grp_b.iterrows():
        ws3.cell(dr, 1, str(row.get("Bilag", ""))).border = B
        ws3.cell(dr, 2, str(row.get("Dato", ""))).border = B
        ws3.cell(dr, 3, str(row.get("Tekst", ""))).border = B
        v = _safe_float(row["Sum beløp"])
        _amt(ws3, dr, 4, v, v < 0)
        ws3.cell(dr, 5, _safe_int(row["Antall poster"])).border = B
        ws3.cell(dr, 6, str(row.get("Kontoer", ""))).border = B
        dr += 1
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 45
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 25
    ws3.freeze_panes = "A5"

    # Ark 4: MVA-analyse
    ws4 = wb.create_sheet("MVA-analyse")
    _title(ws4, f"MVA-analyse – {regnr} {rl_name}{ts}", 7)
    _header(ws4, 4, ["MVA-kode", "Antall", "Grunnlag", "MVA-beløp", "Sats %", "Effektiv %", "Status"])
    mva_result = _compute_mva(df_rl, df_all)
    grp_mva = mva_result["rows"]
    dr = 5
    for _, row in grp_mva.iterrows():
        status = str(row.get("Status", ""))
        fill = OK_FILL if "\u2713" in status else (AVVIK_FILL if "\u26a0" in status else None)
        ws4.cell(dr, 1, str(row.get("MVA-kode", ""))).border = B
        ws4.cell(dr, 2, _safe_int(row["Antall"])).border = B
        _amt(ws4, dr, 3, row["Grunnlag"])
        _amt(ws4, dr, 4, row["MVA-beløp"])
        ws4.cell(dr, 5, round(_safe_float(row.get("Sats %")), 1)).border = B
        ws4.cell(dr, 6, round(_safe_float(row.get("Effektiv %")), 1)).border = B
        ws4.cell(dr, 7, status).border = B
        if fill:
            for ci in range(1, 8):
                ws4.cell(dr, ci).fill = fill
        dr += 1
    for col_letter, w in zip(["A", "B", "C", "D", "E", "F", "G"], [10, 8, 18, 15, 8, 10, 30]):
        ws4.column_dimensions[col_letter].width = w
    ws4.freeze_panes = "A5"

    # Ark 5: Motpostfordeling
    ws5 = wb.create_sheet("Motpostfordeling")
    _title(ws5, f"Motpostfordeling – {regnr} {rl_name}{ts}", 5)
    _header(ws5, 4, ["Konto", "Kontonavn", "Beløp", "Andel %", "Antall bilag"])
    grp_mp = _compute_motpost(df_all, df_rl)
    dr = 5
    for _, row in grp_mp.iterrows():
        ws5.cell(dr, 1, str(row.get("Konto", ""))).border = B
        ws5.cell(dr, 2, str(row.get("Kontonavn", ""))).border = B
        v = _safe_float(row["Beløp"])
        _amt(ws5, dr, 3, v, v < 0)
        ws5.cell(dr, 4, round(float(row["Andel"]), 1)).border = B
        ws5.cell(dr, 5, _safe_int(row["AntallBilag"])).border = B
        dr += 1
    ws5.column_dimensions["A"].width = 10
    ws5.column_dimensions["B"].width = 35
    ws5.column_dimensions["C"].width = 18
    ws5.column_dimensions["D"].width = 10
    ws5.column_dimensions["E"].width = 14
    ws5.freeze_panes = "A5"

    out = Path(path)
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
