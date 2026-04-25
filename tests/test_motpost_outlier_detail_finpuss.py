import pandas as pd
import pytest
from openpyxl import Workbook

from src.audit_actions.motpost.konto_core import build_motpost_data
from src.audit_actions.motpost.konto_core import build_motpost_excel_workbook
from src.audit_actions.motpost.excel_sheets.sheet_outlier_full_bilag import write_outlier_detail_sheets


def test_outlier_detail_boxes_stop_at_col_h_and_prefilled_handling():
    df = pd.DataFrame(
        [
            {
                "Bilag": 1,
                "Dato": "2025-01-01",
                "Tekst": "Utgående faktura",
                "Konto": "3000",
                "Kontonavn": "Salg",
                "Beløp": -1000.0,
            },
            {
                "Bilag": 1,
                "Dato": "2025-01-01",
                "Tekst": "Utgående faktura",
                "Konto": "1500",
                "Kontonavn": "Kundefordringer",
                "Beløp": 1000.0,
            },
        ]
    )

    data = build_motpost_data(df, {"3000"}, selected_direction="Kredit")
    wb = build_motpost_excel_workbook(data, combo_status_map={"1500": "outlier"})

    # Detaljarket heter typisk "#2" i dette minimale testscenariet
    ws = wb["#2"]

    merged = {str(rng) for rng in ws.merged_cells.ranges}

    # Handling-boks (title_row=13, height_rows=4) -> A14:H17
    assert "A14:H17" in merged
    assert "A14:J17" not in merged

    # Resultat-boks (title_row=19, height_rows=6) -> A20:H25
    assert "A20:H25" in merged
    assert "A20:J25" not in merged

    # Handling er forhåndsutfylt
    assert ws["A14"].value.startswith("1. Opparbeid en forståelse av kombinasjonen")
    assert "\n2. Vurder om det er relevant å detaljteste på bilagsnivå" in ws["A14"].value

    # Arkbredde: I/J er ikke satt bredt
    assert float(ws.column_dimensions["I"].width) <= 6
    assert float(ws.column_dimensions["J"].width) <= 6


def test_write_outlier_detail_sheets_raises_on_missing_required_columns():
    wb = Workbook()
    df_bad = pd.DataFrame({"foo": [1]})

    with pytest.raises(ValueError):
        write_outlier_detail_sheets(
            wb,
            df_kombinasjoner=df_bad,
            frames=None,  # brukes ikke i denne feilstien
            df_scope=pd.DataFrame(),
            selected_accounts=[],
            direction="Kredit",
            sum_label="Sum valgte kontoer (Kredit)",
            net_label="Netto kredit (valgte kontoer)",
            outlier_sheet_name="Outlier - alle transaksjoner",
        )
