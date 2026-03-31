"""Tests for consolidation.export — Excel-arbeidsbok."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from consolidation.models import (
    CompanyTB,
    EliminationJournal,
    EliminationLine,
    RunResult,
)
from consolidation.export import build_consolidation_workbook, save_consolidation_workbook


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sample_result_df() -> pd.DataFrame:
    return pd.DataFrame({
        "regnr": [10, 11, 20],
        "regnskapslinje": ["Eiendeler", "Inntekter", "SUM"],
        "sumpost": [False, False, True],
        "formel": [None, None, "=10+11"],
        "Morselskap AS": [150.0, -300.0, -150.0],
        "Datter AS": [50.0, -100.0, -50.0],
        "sum_foer_elim": [200.0, -400.0, -200.0],
        "eliminering": [0.0, 50.0, 50.0],
        "konsolidert": [200.0, -350.0, -150.0],
    })


def _sample_companies() -> list[CompanyTB]:
    return [
        CompanyTB(company_id="a", name="Morselskap AS", row_count=2),
        CompanyTB(company_id="b", name="Datter AS", row_count=2),
    ]


def _sample_eliminations() -> list[EliminationJournal]:
    return [
        EliminationJournal(
            journal_id="e1", name="Internhandel",
            lines=[
                EliminationLine(regnr=11, company_id="a", amount=50.0, description="Salg"),
                EliminationLine(regnr=11, company_id="b", amount=-50.0, description="Kjop"),
            ],
        ),
    ]


def _sample_mapped_tbs() -> dict[str, pd.DataFrame]:
    return {
        "a": pd.DataFrame({
            "konto": ["1000", "3000"],
            "kontonavn": ["Bank", "Salg"],
            "regnr": [10, 11],
            "ib": [0.0, 0.0],
            "ub": [150.0, -300.0],
            "netto": [150.0, -300.0],
        }),
        "b": pd.DataFrame({
            "konto": ["1500", "3500"],
            "kontonavn": ["Varelager", "Tjeneste"],
            "regnr": [10, 11],
            "ib": [0.0, 0.0],
            "ub": [50.0, -100.0],
            "netto": [50.0, -100.0],
        }),
    }


def _sample_run_result() -> RunResult:
    return RunResult(
        run_id="r1",
        company_ids=["a", "b"],
        elimination_ids=["e1"],
        warnings=[],
        result_hash="abc123def456",
    )


def _build_sample_wb() -> "Workbook":
    return build_consolidation_workbook(
        _sample_result_df(),
        _sample_companies(),
        _sample_eliminations(),
        _sample_mapped_tbs(),
        _sample_run_result(),
        client="TestKonsern",
        year="2025",
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestWorkbookStructure:
    def test_correct_sheet_names(self):
        wb = _build_sample_wb()
        names = wb.sheetnames
        assert "Konsernoppstilling" in names
        assert "Elimineringer" in names
        assert "TB - Morselskap AS" in names
        assert "TB - Datter AS" in names
        assert "Kontrollark" in names

    def test_five_sheets(self):
        wb = _build_sample_wb()
        assert len(wb.sheetnames) == 5


class TestKonsernoppstilling:
    def test_title_contains_client(self):
        wb = _build_sample_wb()
        ws = wb["Konsernoppstilling"]
        assert "TestKonsern" in str(ws["A1"].value)
        assert "2025" in str(ws["A1"].value)

    def test_header_row(self):
        wb = _build_sample_wb()
        ws = wb["Konsernoppstilling"]
        # Headers are in row 4
        headers = [ws.cell(row=4, column=c).value for c in range(1, 10)]
        assert headers[0] == "Nr"
        assert headers[1] == "Regnskapslinje"
        assert "Morselskap AS" in headers
        assert "konsolidert" in headers

    def test_data_values(self):
        wb = _build_sample_wb()
        ws = wb["Konsernoppstilling"]
        # First data row (regnr 10, row 5)
        assert ws.cell(row=5, column=1).value == 10
        assert ws.cell(row=5, column=2).value == "Eiendeler"

    def test_sum_lines_bold(self):
        wb = _build_sample_wb()
        ws = wb["Konsernoppstilling"]
        # SUM-linje er regnr 20 -> rad 7 (row 4 header + 3 data rows)
        cell = ws.cell(row=7, column=2)
        assert cell.value == "SUM"
        assert cell.font.bold is True


class TestElimineringer:
    def test_journal_name_present(self):
        wb = _build_sample_wb()
        ws = wb["Elimineringer"]
        assert ws.cell(row=1, column=1).value == "Internhandel"

    def test_voucher_label_present_when_available(self):
        wb = build_consolidation_workbook(
            _sample_result_df(),
            _sample_companies(),
            [
                EliminationJournal(
                    journal_id="e1",
                    voucher_no=1,
                    name="Internhandel",
                    lines=[
                        EliminationLine(regnr=11, company_id="a", amount=50.0, description="Salg"),
                        EliminationLine(regnr=11, company_id="b", amount=-50.0, description="Kjop"),
                    ],
                ),
            ],
            _sample_mapped_tbs(),
            _sample_run_result(),
        )
        ws = wb["Elimineringer"]
        assert ws.cell(row=1, column=1).value == "Bilag 1"

    def test_no_eliminations(self):
        wb = build_consolidation_workbook(
            _sample_result_df(), _sample_companies(), [],
            _sample_mapped_tbs(), _sample_run_result(),
        )
        ws = wb["Elimineringer"]
        assert "Ingen" in str(ws["A1"].value)


class TestCompanySheets:
    def test_company_tb_data(self):
        wb = _build_sample_wb()
        ws = wb["TB - Morselskap AS"]
        # Header row
        assert ws.cell(row=1, column=1).value == "Konto"
        # First data row
        assert str(ws.cell(row=2, column=1).value) == "1000"


class TestKontrollark:
    def test_contains_hash(self):
        wb = _build_sample_wb()
        ws = wb["Kontrollark"]
        # Finn hash-raden
        found = False
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0] and "hash" in str(row[0]).lower():
                assert row[1] == "abc123def456"
                found = True
        assert found, "Hash not found in Kontrollark"

    def test_contains_client(self):
        wb = _build_sample_wb()
        ws = wb["Kontrollark"]
        assert ws.cell(row=1, column=2).value == "TestKonsern"


class TestSaveWorkbook:
    def test_save_creates_file(self, tmp_path):
        out = tmp_path / "test_export.xlsx"
        result = save_consolidation_workbook(
            out,
            result_df=_sample_result_df(),
            companies=_sample_companies(),
            eliminations=_sample_eliminations(),
            mapped_tbs=_sample_mapped_tbs(),
            run_result=_sample_run_result(),
            client="Test",
            year="2025",
        )
        assert Path(result).exists()

        # Reopen and verify
        wb = load_workbook(result)
        assert "Konsernoppstilling" in wb.sheetnames
