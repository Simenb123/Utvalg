"""Tester for workpaper_forside — felles "Beskrivelse"-fane."""

from __future__ import annotations

from openpyxl import Workbook

import action_context
import src.shared.workpapers.forside as workpaper_forside


def test_forside_skipped_when_no_context() -> None:
    action_context.clear()
    wb = Workbook()
    added = workpaper_forside.build_forside_sheet(wb, workpaper_navn="Test")
    assert added is False
    assert "Beskrivelse" not in wb.sheetnames


def test_forside_added_when_context_pushed() -> None:
    action_context.clear()
    ctx = action_context.ActionContext(
        action_key="L:abc",
        handling_navn="IB/UB-kontroll",
        handling_type="substansiv",
        omraade="Bank",
        regnr="80",
        beskrivelse="Kontroller at IB stemmer med UB fjor.",
        kommentar="Ingen avvik funnet.",
        kjort_av="simen",
        client="ACME",
        year="2025",
    )
    with action_context.push(ctx):
        wb = Workbook()
        added = workpaper_forside.build_forside_sheet(wb, workpaper_navn="IB/UB")
        assert added is True
        assert "Beskrivelse" in wb.sheetnames
        assert wb.sheetnames[0] == "Beskrivelse"
        ws = wb["Beskrivelse"]
        assert ws["A1"].value == "IB/UB"
        texts = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if cell.value
        ]
        assert any("IB/UB-kontroll" in str(t) for t in texts)
        assert any("Bank" in str(t) for t in texts)
        assert any("stemmer med UB fjor" in str(t) for t in texts)
        assert any("Ingen avvik funnet" in str(t) for t in texts)


def test_push_restores_previous_context() -> None:
    action_context.clear()
    outer = action_context.ActionContext(action_key="L:outer", handling_navn="Ytre")
    inner = action_context.ActionContext(action_key="L:inner", handling_navn="Indre")
    with action_context.push(outer):
        assert action_context.current() is outer
        with action_context.push(inner):
            assert action_context.current() is inner
        assert action_context.current() is outer
    assert action_context.current() is None


def test_push_sets_kjort_at_if_missing() -> None:
    action_context.clear()
    ctx = action_context.ActionContext(action_key="L:x", handling_navn="X")
    assert ctx.kjort_at == ""
    with action_context.push(ctx):
        assert ctx.kjort_at  # ISO-tid satt
