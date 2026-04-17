from __future__ import annotations

from decimal import Decimal

import pandas as pd
from openpyxl import load_workbook

from a07_feature.export import export_a07_workbook


def test_export_a07_workbook_creates_expected_sheets_and_formats_numbers(tmp_path) -> None:
    out_path = tmp_path / "a07_kontroll.xlsx"

    exported = export_a07_workbook(
        out_path,
        overview_df=pd.DataFrame(
            [
                {
                    "Kode": "fastloenn",
                    "Navn": "Fastloenn",
                    "Belop": Decimal("7530516.49"),
                    "Status": "Ikke mappet",
                    "Kontoer": "",
                }
            ]
        ),
        reconcile_df=pd.DataFrame(
            [
                {
                    "Kode": "fastloenn",
                    "Navn": "Fastloenn",
                    "A07_Belop": Decimal("7530516.49"),
                    "GL_Belop": Decimal("7073783.57"),
                    "Diff": Decimal("456732.92"),
                    "AntallKontoer": 1,
                    "Kontoer": "5000",
                    "WithinTolerance": False,
                }
            ]
        ),
        mapping_df=pd.DataFrame([{"Konto": "5000", "Navn": "Lonn", "Kode": "fastloenn"}]),
        suggestions_df=pd.DataFrame(
            [
                {
                    "Kode": "fastloenn",
                    "KodeNavn": "fastloenn",
                    "Basis": "Endring",
                    "A07_Belop": Decimal("7530516.49"),
                    "ForslagKontoer": "5000",
                    "GL_Sum": Decimal("7073783.57"),
                    "Diff": Decimal("456732.92"),
                    "Score": 0.651,
                    "WithinTolerance": False,
                }
            ]
        ),
        unmapped_df=pd.DataFrame([{"Konto": "1000", "Navn": "Forskning", "GL_Belop": Decimal("0.00"), "Kode": ""}]),
    )

    assert exported == out_path
    assert exported.exists()

    workbook = load_workbook(exported)
    assert workbook.sheetnames == ["Kontroll", "Avstemming", "Mapping", "Forslag", "Umappede"]
    assert workbook["Kontroll"]["C2"].value == 7530516.49
    assert workbook["Kontroll"]["C2"].number_format == "#,##0.00"
    assert workbook["Avstemming"]["H2"].value == "Nei"


def test_export_a07_workbook_includes_control_statement_sheet_when_present(tmp_path) -> None:
    out_path = tmp_path / "a07_kontroll_med_kontrolloppstilling.xlsx"

    exported = export_a07_workbook(
        out_path,
        overview_df=pd.DataFrame([{"Kode": "fastloenn", "Navn": "Fastloenn", "Belop": Decimal("100.00")}]),
        reconcile_df=pd.DataFrame(),
        mapping_df=pd.DataFrame(),
        control_statement_df=pd.DataFrame(
            [
                {
                    "Gruppe": "aga_pliktig_loenn",
                    "Navn": "AGA-pliktig lonn",
                    "IB": Decimal("1000.00"),
                    "Endring": Decimal("250.50"),
                    "UB": Decimal("1250.50"),
                    "A07": Decimal("240.00"),
                    "Diff": Decimal("1010.50"),
                    "Status": "Manuell",
                    "AntallKontoer": 3,
                    "Kontoer": "5000, 5010, 5400",
                    "Kilder": "manual, history",
                }
            ]
        ),
        suggestions_df=pd.DataFrame(),
        unmapped_df=pd.DataFrame(),
    )

    assert exported == out_path
    workbook = load_workbook(exported)
    assert workbook.sheetnames == ["Kontroll", "Avstemming", "Mapping", "Kontrolloppstilling"]
    ws = workbook["Kontrolloppstilling"]
    assert [cell.value for cell in ws[1]] == [
        "Gruppe",
        "Navn",
        "IB",
        "Endring",
        "UB",
        "A07",
        "Diff",
        "Status",
        "AntallKontoer",
        "Kontoer",
        "Kilder",
    ]
    assert ws["C2"].value == 1000.0
    assert ws["C2"].number_format == "#,##0.00"
    assert ws["F2"].value == 240.0
    assert ws["F2"].number_format == "#,##0.00"
    assert ws["G2"].value == 1010.5
    assert ws["G2"].number_format == "#,##0.00"
    assert ws["H2"].value == "Manuell"
    assert ws["I2"].value == 3
    assert ws["I2"].number_format == "0"
