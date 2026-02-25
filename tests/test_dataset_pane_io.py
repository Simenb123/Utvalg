from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

import dataset_pane_io as dio


def _write_xlsx_with_header_not_first(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hovedbok"
    ws.append(["Rapport", "dummy"])  # row 1
    ws.append(["Mer info", "dummy"])  # row 2
    ws.append(["Konto", "Bilag", "Beløp", "", ""])  # row 3 (header)
    ws.append([3000, "B1", 100, "", ""])  # row 4 (data)
    wb.save(path)


def _write_xlsx_with_blank_headers(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HB"
    ws.append(["Konto", "", None, "Beløp"])
    ws.append([3000, "B1", "2026-01-01", "100"])  # data som viser at kol2/kol3 finnes
    wb.save(path)


def _write_xlsx_with_trailing_blank_header_but_data(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HB"
    ws.append(["A", ""])  # blank header i siste kolonne
    ws.append([1, 2])
    wb.save(path)


def test_read_excel_header_trims_trailing_empty_columns(tmp_path: Path) -> None:
    p = tmp_path / "hb.xlsx"
    _write_xlsx_with_header_not_first(p)

    headers = dio.read_excel_header(p, "Hovedbok", header_row=3, max_cols=20)
    assert headers == ["Konto", "Bilag", "Beløp"]


def test_read_excel_header_replaces_blank_headers_with_kolX(tmp_path: Path) -> None:
    p = tmp_path / "blank_headers.xlsx"
    _write_xlsx_with_blank_headers(p)

    headers = dio.read_excel_header(p, "HB", header_row=1, max_cols=20)
    assert headers == ["Konto", "kol2", "kol3", "Beløp"]


def test_read_excel_header_keeps_trailing_blank_header_if_data_exists(tmp_path: Path) -> None:
    p = tmp_path / "trailing_blank.xlsx"
    _write_xlsx_with_trailing_blank_header_but_data(p)

    headers = dio.read_excel_header(p, "HB", header_row=1, max_cols=20)
    assert headers == ["A", "kol2"]


def test_read_excel_rows_trims_trailing_empty_columns(tmp_path: Path) -> None:
    p = tmp_path / "hb.xlsx"
    _write_xlsx_with_header_not_first(p)

    rows = dio.read_excel_rows(p, "Hovedbok", start_row=3, nrows=2, max_cols=20)
    assert len(rows) == 2
    assert rows[0] == ["Konto", "Bilag", "Beløp"]
    assert rows[1][:3] == [3000, "B1", 100]
    assert len(rows[0]) == 3
    assert len(rows[1]) == 3


def test_read_csv_rows_does_not_raise_when_max_cols_exceeds_actual(tmp_path: Path) -> None:
    p = tmp_path / "hb.csv"
    p.write_text("A;B\n1;2\n3;4\n", encoding="utf-8")

    rows = dio.read_csv_rows(p, nrows=10, max_cols=80)
    assert rows[0] == ["A", "B"]
    assert rows[1] == ["1", "2"]
    assert all(len(r) == 2 for r in rows)


def test_read_csv_header_from_non_first_row(tmp_path: Path) -> None:
    p = tmp_path / "hb.csv"
    p.write_text(
        "Rapport;dummy\n"
        "Mer info;dummy\n"
        "Konto;Bilag;Beløp;Ekstra\n"
        "3000;B1;100;X\n",
        encoding="utf-8",
    )

    headers = dio.read_csv_header(p, header_row=3)
    assert headers == ["Konto", "Bilag", "Beløp", "Ekstra"]


def test_read_csv_header_replaces_blank_headers_with_kolX(tmp_path: Path) -> None:
    p = tmp_path / "blank_headers.csv"
    # blank header i kolonne 2
    p.write_text(
        "Konto;;Beløp\n"
        "3000;B1;100\n",
        encoding="utf-8",
    )

    headers = dio.read_csv_header(p, header_row=1)
    assert headers == ["Konto", "kol2", "Beløp"]


def test_read_csv_header_keeps_trailing_blank_header_if_data_exists(tmp_path: Path) -> None:
    p = tmp_path / "trailing_blank.csv"
    p.write_text(
        "A;\n"
        "1;2\n",
        encoding="utf-8",
    )

    headers = dio.read_csv_header(p, header_row=1)
    assert headers == ["A", "kol2"]


def test_auto_detect_header_and_headers_excel(tmp_path: Path) -> None:
    p = tmp_path / "hb.xlsx"
    _write_xlsx_with_header_not_first(p)

    header_row, headers = dio.auto_detect_header_and_headers(p, "Hovedbok")
    assert header_row == 3
    assert headers == ["Konto", "Bilag", "Beløp"]


def test_guess_best_excel_sheet_prefers_sheet_with_ledger_headers(tmp_path: Path) -> None:
    p = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Forside"
    ws1.append(["Dette er en forside"])
    ws1.append(["Generert av systemet"])

    ws2 = wb.create_sheet("Bilagsjournal")
    ws2.append(["Rapport"])  # tittel
    ws2.append(["Kontonr", "Bilagsnr", "Beløp"])  # header
    ws2.append([3000, 1001, 10])

    wb.save(p)

    best = dio.guess_best_excel_sheet(p)
    assert best == "Bilagsjournal"


def test_auto_detect_header_and_headers_csv(tmp_path: Path) -> None:
    p = tmp_path / "hb.csv"
    p.write_text(
        "Rapport;dummy\n"
        "Mer info;dummy\n"
        "Konto;Bilag;Beløp;Ekstra\n"
        "3000;B1;100;X\n",
        encoding="utf-8",
    )

    header_row, headers = dio.auto_detect_header_and_headers(p, None)
    assert header_row == 3
    assert headers == ["Konto", "Bilag", "Beløp", "Ekstra"]


def test_auto_detect_falls_back_when_uncertain(tmp_path: Path) -> None:
    # Kun tall og bare 2 kolonner -> detect_header_row blir usikker (score=0) og returnerer None
    p = tmp_path / "numbers.csv"
    p.write_text("1;2\n3;4\n5;6\n", encoding="utf-8")

    header_row, headers = dio.auto_detect_header_and_headers(p, None, fallback_header_row=1)
    assert header_row == 1
    assert headers == ["1", "2"]


def test_read_excel_header_missing_file_raises(tmp_path: Path) -> None:
    p = tmp_path / "missing.xlsx"
    with pytest.raises(FileNotFoundError):
        _ = dio.read_excel_header(p, None, header_row=1)
