"""Tester for mva_avstemming og page_analyse_mva."""

from __future__ import annotations

import pandas as pd
import pytest


# ---------------------------------------------------------------------------
# page_analyse_mva — month_to_termin
# ---------------------------------------------------------------------------

class TestMonthToTermin:
    def test_jan_feb(self):
        from page_analyse_mva import month_to_termin
        assert month_to_termin(1) == 1
        assert month_to_termin(2) == 1

    def test_mar_apr(self):
        from page_analyse_mva import month_to_termin
        assert month_to_termin(3) == 2
        assert month_to_termin(4) == 2

    def test_all_months(self):
        from page_analyse_mva import month_to_termin
        expected = {1: 1, 2: 1, 3: 2, 4: 2, 5: 3, 6: 3,
                    7: 4, 8: 4, 9: 5, 10: 5, 11: 6, 12: 6}
        for month, termin in expected.items():
            assert month_to_termin(month) == termin, f"month={month}"


# ---------------------------------------------------------------------------
# page_analyse_mva — build_mva_pivot
# ---------------------------------------------------------------------------

class TestBuildMvaPivot:
    def _make_df(self, rows):
        return pd.DataFrame(rows, columns=["MVA-kode", "MVA-beløp", "Dato", "Konto", "Beløp"])

    def test_empty_df(self):
        from page_analyse_mva import build_mva_pivot
        result = build_mva_pivot(pd.DataFrame())
        assert result.empty or len(result) == 0

    def test_no_mva_code_col(self):
        from page_analyse_mva import build_mva_pivot
        df = pd.DataFrame({"Konto": ["1000"], "Beløp": [100], "Dato": ["2025-01-15"]})
        result = build_mva_pivot(df)
        assert result.empty or len(result) == 0

    def test_basic_pivot(self):
        from page_analyse_mva import build_mva_pivot
        df = self._make_df([
            ["1", 250.0, "2025-01-15", "3000", 1000.0],
            ["1", 125.0, "2025-03-10", "3000", 500.0],
            ["11", -250.0, "2025-01-20", "4300", -1000.0],
        ])
        result = build_mva_pivot(df)

        # Sjekk at vi har kode 1 og 11 + 3 summeringsrader
        codes = result["MVA-kode"].tolist()
        assert "1" in codes
        assert "11" in codes

        # Sjekk at summeringsrader finnes
        descriptions = result["Beskrivelse"].tolist()
        assert any("utgående" in d.lower() for d in descriptions)
        assert any("inngående" in d.lower() for d in descriptions)
        assert any("netto" in d.lower() for d in descriptions)

    def test_termin_grouping(self):
        from page_analyse_mva import build_mva_pivot
        df = self._make_df([
            ["1", 100.0, "2025-01-15", "3000", 400.0],   # T1
            ["1", 200.0, "2025-02-15", "3000", 800.0],   # T1
            ["1", 300.0, "2025-05-10", "3000", 1200.0],  # T3
        ])
        result = build_mva_pivot(df)

        code_1 = result[result["MVA-kode"] == "1"].iloc[0]
        assert code_1["T1"] == 300.0  # 100 + 200
        assert code_1["T3"] == 300.0
        assert code_1["T2"] == 0.0

    def test_sum_column(self):
        from page_analyse_mva import build_mva_pivot
        df = self._make_df([
            ["1", 100.0, "2025-01-15", "3000", 400.0],
            ["1", 200.0, "2025-05-10", "3000", 800.0],
        ])
        result = build_mva_pivot(df)
        code_1 = result[result["MVA-kode"] == "1"].iloc[0]
        assert code_1["Sum"] == 300.0

    def test_empty_mva_codes_filtered(self):
        from page_analyse_mva import build_mva_pivot
        df = self._make_df([
            ["", 0.0, "2025-01-15", "3000", 1000.0],
            ["1", 250.0, "2025-01-15", "3000", 1000.0],
        ])
        result = build_mva_pivot(df)
        codes = [c for c in result["MVA-kode"].tolist() if c]
        assert "" not in codes
        assert "1" in codes


# ---------------------------------------------------------------------------
# mva_avstemming — month_to_termin (same logic, separate module)
# ---------------------------------------------------------------------------

class TestMvaAvstemmingMonthToTermin:
    def test_all_months(self):
        from mva_avstemming import month_to_termin
        for m in range(1, 13):
            t = month_to_termin(m)
            assert 1 <= t <= 6


# ---------------------------------------------------------------------------
# mva_avstemming — build_reconciliation
# ---------------------------------------------------------------------------

class TestBuildReconciliation:
    def test_basic_reconciliation(self):
        from mva_avstemming import SkatteetatenData, build_reconciliation

        # MVA-pivot med utgående (negativ i HB = kredit) og inngående (positiv = debet)
        mva_pivot = pd.DataFrame([
            {"MVA-kode": "1", "direction": "utgående", "T1": -1000.0, "T2": -2000.0,
             "T3": 0, "T4": 0, "T5": 0, "T6": 0, "Sum": -3000.0},
            {"MVA-kode": "11", "direction": "inngående", "T1": 400.0, "T2": 800.0,
             "T3": 0, "T4": 0, "T5": 0, "T6": 0, "Sum": 1200.0},
        ])

        skatt = SkatteetatenData(
            mva_per_termin={1: 600.0, 2: 1200.0},
        )

        result = build_reconciliation(mva_pivot, skatt)

        assert len(result) == 7  # 6 terminer + 1 sum-rad

        # T1: HB Utgående = |−1000| = 1000, HB Inngående = |400| = 400
        # HB Netto = 1000 - 400 = 600, Innrapportert = 600, Diff = 0
        t1 = result[result["Termin"] == "T1"].iloc[0]
        assert t1["HB Utgående"] == 1000.0
        assert t1["HB Inngående"] == 400.0
        assert t1["HB Netto"] == 600.0
        assert t1["Innrapportert"] == 600.0
        assert t1["Differanse"] == 0.0

    def test_sum_row(self):
        from mva_avstemming import SkatteetatenData, build_reconciliation

        mva_pivot = pd.DataFrame([
            {"MVA-kode": "1", "direction": "utgående", "T1": -100.0, "T2": -200.0,
             "T3": 0, "T4": 0, "T5": 0, "T6": 0, "Sum": -300.0},
        ])

        skatt = SkatteetatenData(mva_per_termin={1: 100.0, 2: 200.0})

        result = build_reconciliation(mva_pivot, skatt)
        sum_row = result[result["Termin"] == "Sum"].iloc[0]
        assert sum_row["HB Netto"] == 300.0
        assert sum_row["Innrapportert"] == 300.0
        assert sum_row["Differanse"] == 0.0

    def test_empty_pivot(self):
        from mva_avstemming import SkatteetatenData, build_reconciliation

        result = build_reconciliation(pd.DataFrame(), SkatteetatenData())
        assert len(result) == 7


# ---------------------------------------------------------------------------
# mva_avstemming — parse_skatteetaten_kontoutskrift
# ---------------------------------------------------------------------------

class TestParseSkatteetaten:
    def test_parse_real_file(self, tmp_path):
        """Test med en minimal mock-Excel-fil."""
        import openpyxl

        wb = openpyxl.Workbook()

        # Kontoutskrift gjelder
        ws1 = wb.active
        ws1.title = "Kontoutskrift gjelder"
        ws1.append(["Kontoutskrift"])
        ws1.append([None, "123456789", "Test AS"])
        ws1.append([None, "Periode:", "01.01.2025 - 31.12.2025"])

        # Krav
        ws2 = wb.create_sheet("Krav")
        ws2.append([
            "Forfallsdato", "Kravbeskrivelse", "Kravgruppe",
            "År", "Periode", "Periode fra-til", "Opprinnelig beløp",
        ])
        ws2.append([None, "Mva-melding", "Merverdiavgift", 2025, 1, "01.01.2025 - 28.02.2025", 50000.0])
        ws2.append([None, "Mva-melding", "Merverdiavgift", 2025, 2, "01.03.2025 - 30.04.2025", 75000.0])
        ws2.append([None, "Arbeidsgiveravgift", "Arbeidsgiveravgift", 2025, 1, "", 30000.0])
        ws2.append([None, "Forskuddstrekk", "Forskuddstrekk", 2025, 1, "", 80000.0])
        # Annet år — skal filtreres bort ved year=2025
        ws2.append([None, "Mva-melding", "Merverdiavgift", 2024, 6, "01.11.2024 - 31.12.2024", 99999.0])

        path = tmp_path / "kontoutskrift_test.xlsx"
        wb.save(path)

        from mva_avstemming import parse_skatteetaten_kontoutskrift

        result = parse_skatteetaten_kontoutskrift(path, year=2025)

        assert result.org_nr == "123456789"
        assert result.company == "Test AS"
        assert result.mva_per_termin[1] == 50000.0
        assert result.mva_per_termin[2] == 75000.0
        assert 6 not in result.mva_per_termin  # 2024 filtrert bort
        assert result.aga_per_termin[1] == 30000.0
        assert result.forskuddstrekk_per_termin[1] == 80000.0

    def test_parse_all_years(self, tmp_path):
        """Test uten årsfilter."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Kontoutskrift gjelder"
        ws1.append(["Kontoutskrift"])

        ws2 = wb.create_sheet("Krav")
        ws2.append(["Forfallsdato", "Kravbeskrivelse", "Kravgruppe", "År", "Periode", "Periode fra-til", "Opprinnelig beløp"])
        ws2.append([None, "Mva-melding", "Merverdiavgift", 2024, 6, "", 10000.0])
        ws2.append([None, "Mva-melding", "Merverdiavgift", 2025, 1, "", 20000.0])

        path = tmp_path / "kontoutskrift_all.xlsx"
        wb.save(path)

        from mva_avstemming import parse_skatteetaten_kontoutskrift
        result = parse_skatteetaten_kontoutskrift(path)  # Ingen årsfilter

        assert result.mva_per_termin.get(6, 0) == 10000.0
        assert result.mva_per_termin.get(1, 0) == 20000.0


# ---------------------------------------------------------------------------
# build_mva_pivot — grunnlag-kolonner
# ---------------------------------------------------------------------------

class TestBuildMvaPivotGrunnlag:
    def _make_df(self, rows):
        return pd.DataFrame(rows, columns=["MVA-kode", "MVA-beløp", "Dato", "Konto", "Beløp"])

    def test_grunnlag_columns_exist(self):
        from page_analyse_mva import build_mva_pivot
        df = self._make_df([
            ["1", 250.0, "2025-01-15", "3000", 1000.0],
        ])
        result = build_mva_pivot(df)
        assert "G_T1" in result.columns
        assert "G_Sum" in result.columns

    def test_grunnlag_values(self):
        from page_analyse_mva import build_mva_pivot
        df = self._make_df([
            ["1", 250.0, "2025-01-15", "3000", 1000.0],
            ["1", 125.0, "2025-03-10", "3000", 500.0],
        ])
        result = build_mva_pivot(df)
        code_1 = result[result["MVA-kode"] == "1"].iloc[0]
        assert code_1["G_T1"] == 1000.0
        assert code_1["G_T2"] == 500.0
        assert code_1["G_Sum"] == 1500.0


# ---------------------------------------------------------------------------
# mva_avstemming — build_mva_kontroller
# ---------------------------------------------------------------------------

class TestBuildMvaKontroller:
    def _make_df(self, rows):
        return pd.DataFrame(rows, columns=["Konto", "Kontonavn", "Dato", "Beløp", "MVA-kode"])

    def test_empty_df(self):
        from mva_avstemming import build_mva_kontroller
        result = build_mva_kontroller(pd.DataFrame())
        assert result.salg_vs_grunnlag.empty
        assert result.salg_uten_mva.empty
        assert result.andre_med_utg_mva.empty

    def test_salg_vs_grunnlag_match(self):
        """Salgskontoer med MVA-kode 1 — grunnlag bør matche salgsinntekter."""
        from mva_avstemming import build_mva_kontroller
        df = self._make_df([
            ["3000", "Salgsinntekter", "2025-01-15", -1000.0, "1"],
            ["3000", "Salgsinntekter", "2025-01-20", -2000.0, "1"],
        ])
        result = build_mva_kontroller(df)
        k1 = result.salg_vs_grunnlag
        t1 = k1[k1["Termin"] == "T1"].iloc[0]
        # Salg = -3000, grunnlag med utg. MVA = -3000 → diff = 0
        assert abs(t1["Differanse"]) < 0.01

    def test_salg_uten_mva(self):
        """Transaksjoner på 3xxx uten MVA-kode flagges."""
        from mva_avstemming import build_mva_kontroller
        df = self._make_df([
            ["3000", "Salgsinntekter", "2025-01-15", -1000.0, "1"],
            ["3100", "Andre salg", "2025-02-10", -500.0, ""],
        ])
        result = build_mva_kontroller(df)
        assert not result.salg_uten_mva.empty
        assert len(result.salg_uten_mva) == 1

    def test_andre_med_utg_mva(self):
        """Transaksjoner utenfor 3xxx med utgående MVA-kode flagges."""
        from mva_avstemming import build_mva_kontroller
        df = self._make_df([
            ["3000", "Salgsinntekter", "2025-01-15", -1000.0, "1"],
            ["6700", "Tap på fordringer", "2025-03-10", -200.0, "1"],
        ])
        result = build_mva_kontroller(df)
        assert not result.andre_med_utg_mva.empty
        assert len(result.andre_med_utg_mva) == 1

    def test_summary_ok(self):
        """Alle kontroller OK når alt matcher."""
        from mva_avstemming import build_mva_kontroller
        df = self._make_df([
            ["3000", "Salgsinntekter", "2025-01-15", -1000.0, "1"],
        ])
        result = build_mva_kontroller(df)
        statuses = [s["Status"] for s in result.summary]
        assert "AVVIK" not in statuses
        assert result.summary[1]["Status"] == "OK"  # K2: ingen salg uten MVA
        assert result.summary[2]["Status"] == "OK"  # K3: ingen andre med utg. MVA
