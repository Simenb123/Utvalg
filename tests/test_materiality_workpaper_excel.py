from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from src.pages.materiality.backend.workpaper_excel import export_materiality_workpaper


def _sheet_kv_map(ws) -> dict[str, object]:
    out: dict[str, object] = {}
    for label, value, *_rest in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if label not in (None, "") and value not in (None, ""):
            out.setdefault(str(label), value)
    return out


def test_export_materiality_workpaper_creates_expected_content(tmp_path: Path) -> None:
    crm_lookup = SimpleNamespace(
        db_path=Path("C:/crm.sqlite"),
        matched_client_number="7162",
        record=SimpleNamespace(
            client_number="7162",
            client_name="Spor Arkitekter AS",
            engagement_year=2025,
            materiality=250000,
            pmateriality=175000,
            clearly_triv=17500,
            source_updated_at="2026-04-05T12:00:00Z",
            last_synced_at_utc="2026-04-06T03:19:44Z",
        ),
    )
    calculation_payload = {
        "benchmark_key": "gross_profit",
        "benchmark_amount": 1609438,
        "reference_pct_low": 1.5,
        "reference_pct_high": 3.0,
        "reference_amount_low": 24142,
        "reference_amount_high": 48283,
        "overall_materiality": 120708,
        "pm_pct": 75.0,
        "performance_materiality": 90531,
        "trivial_pct": 10.0,
        "clearly_trivial": 9053,
    }
    active_materiality = {
        "source": "crmsystem",
        "overall_materiality": 250000,
        "performance_materiality": 175000,
        "clearly_trivial": 17500,
    }

    exported = export_materiality_workpaper(
        tmp_path / "vesentlighet",
        client="Spor Arkitekter AS",
        year="2025",
        active_materiality=active_materiality,
        selection_threshold_label="Arbeidsvesentlighet (PM)",
        state_updated_at="2026-04-06T03:35:10Z",
        crm_client_number="7162",
        crm_lookup=crm_lookup,
        calculation_payload=calculation_payload,
        benchmark_amounts={
            "revenue": 6000000,
            "gross_profit": 1609438,
            "profit_before_tax": 1609438,
            "total_assets": 4200000,
            "equity": 1800000,
        },
    )

    assert exported.suffix == ".xlsx"
    assert exported.exists()

    wb = load_workbook(exported, data_only=True)
    assert wb.sheetnames == ["Oversikt", "Referanseverdier"]

    overview = wb["Oversikt"]
    assert overview["A1"].value == "Vesentlighetsarbeidspapir"

    values = _sheet_kv_map(overview)
    assert values["Klient"] == "Spor Arkitekter AS"
    assert values["Brukes i Utvalg"] == "Arbeidsvesentlighet (PM)"
    assert values["Kilde"] == "CRMSystem"
    assert values["OM"] == 250000
    assert values["PM"] == 175000
    assert values["Valgt klientnr"] == "7162"
    assert values["Benchmark"] == "Bruttofortjeneste"
    assert values["Total vesentlighet"] == 120708

    ref_ws = wb["Referanseverdier"]
    row_by_name = {
        str(ref_ws.cell(row, 1).value): row
        for row in range(4, ref_ws.max_row + 1)
        if ref_ws.cell(row, 1).value
    }
    gross_profit_row = row_by_name["Bruttofortjeneste"]
    assert ref_ws.cell(gross_profit_row, 7).value == "Ja"
