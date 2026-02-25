# -*- coding: utf-8 -*-
from __future__ import annotations

import datetime as dt

import pandas as pd
import pytest
from openpyxl import Workbook

from dataset_build_fast import build_from_file
from excel_importer import infer_excel_sheet_and_headers


def _excel_serial(date_: dt.date) -> int:
    # Samme origin som vi bruker i dataset_build_fast._to_date_no: 1899-12-30
    origin = dt.date(1899, 12, 30)
    return (date_ - origin).days


def _write_workbook_header_not_first(path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"

    # Topprader som ofte finnes i eksport (tittel/metadata)
    ws["A1"] = "Hovedbokrapport"
    ws["A2"] = "Generert fra ERP"
    ws["A3"] = "Periode: 01.01.2025-31.01.2025"

    # Header på rad 4 (0-basert index 3)
    headers = ["Konto", "Kontonavn", "Bilag", "Beløp", "Bilagsdato", "Tekst"]
    ws.append(headers)

    # Data
    ws.append([3000, "Salg", 1001, "(34,50)", _excel_serial(dt.date(2025, 1, 15)), "Korrigering"])
    ws.append([3010, "Service", 1002, "100-", _excel_serial(dt.date(2025, 1, 16)), "Justering"])
    ws.append([3020, "Annet", 1003, "1 234,50", _excel_serial(dt.date(2025, 1, 17)), "Test tusenskiller"])

    wb.save(path)


def _write_workbook_multi_sheet(path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Forside"
    ws1["A1"] = "Dette er en forside"
    ws1["A2"] = "Ikke transaksjoner"

    ws2 = wb.create_sheet("Hovedbok")
    ws2.append(["Konto", "Kontonavn", "Bilag", "Beløp"])
    ws2.append([1500, "Kundefordringer", 2001, 10])
    ws2.append([1500, "Kundefordringer", 2002, 20])

    wb.save(path)


def test_infer_excel_sheet_and_headers_detects_header_row(tmp_path):
    p = tmp_path / "hb.xlsx"
    _write_workbook_header_not_first(p)

    sheet, header_row, cols = infer_excel_sheet_and_headers(str(p))

    assert sheet == "Rapport"
    assert header_row == 3  # 0-basert
    assert cols[:4] == ["Konto", "Kontonavn", "Bilag", "Beløp"]
    assert "Bilagsdato" in cols


def test_build_from_file_excel_header_not_first_row_parses_amount_and_dates(tmp_path):
    p = tmp_path / "hb.xlsx"
    _write_workbook_header_not_first(p)

    mapping = {
        "Konto": "Konto",
        "Kontonavn": "Kontonavn",
        "Bilag": "Bilag",
        "Beløp": "Beløp",
        "Dato": "Bilagsdato",
        "Tekst": "Tekst",
    }

    df = build_from_file(str(p), mapping=mapping)

    assert list(df.columns)[:6] == ["Konto", "Kontonavn", "Bilag", "Beløp", "Dato", "Tekst"]
    assert len(df) == 3

    # Beløp: parentes og trailing minus skal gi negative tall
    assert df["Beløp"].tolist() == pytest.approx([-34.5, -100.0, 1234.5], rel=1e-9)

    # Excel-serienummer skal bli dato
    assert pd.to_datetime(df.loc[df.index[0], "Dato"]).date() == dt.date(2025, 1, 15)

    # Konto/Bilag skal være tekst uten 'nan'
    assert df["Konto"].dtype.name in {"string", "object"}
    assert df["Konto"].astype(str).str.lower().str.contains("nan").sum() == 0
    assert df["Bilag"].astype(str).str.lower().str.contains("nan").sum() == 0


def test_build_from_file_selects_best_sheet_when_multiple_sheets(tmp_path):
    p = tmp_path / "multi.xlsx"
    _write_workbook_multi_sheet(p)

    mapping = {
        "Konto": "Konto",
        "Kontonavn": "Kontonavn",
        "Bilag": "Bilag",
        "Beløp": "Beløp",
    }

    df = build_from_file(str(p), mapping=mapping)

    assert len(df) == 2
    assert df["Beløp"].sum() == 30
    assert set(df["Bilag"].astype(str)) == {"2001", "2002"}
