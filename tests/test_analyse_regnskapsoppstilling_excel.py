from __future__ import annotations

import openpyxl
import pandas as pd


def test_build_regnskapsoppstilling_workbook_creates_sheets() -> None:
    from analyse_regnskapsoppstilling_excel import build_regnskapsoppstilling_workbook

    rl_df = pd.DataFrame(
        {
            "regnr": [10, 99],
            "regnskapslinje": ["Eiendeler", "SUM"],
            "IB": [100.0, 100.0],
            "Endring": [50.0, 50.0],
            "UB": [150.0, 150.0],
            "Antall": [3, 3],
        }
    )
    regn = pd.DataFrame(
        {
            "nr": [10, 99],
            "regnskapslinje": ["Eiendeler", "SUM"],
            "sumpost": ["nei", "ja"],
            "Formel": ["", "=10"],
        }
    )

    wb = build_regnskapsoppstilling_workbook(
        rl_df,
        regnskapslinjer=regn,
        client="Nbs Regnskap AS",
        year="2025",
    )

    assert "Regnskapsoppstilling" in wb.sheetnames
    assert "Beregningsgrunnlag" in wb.sheetnames
    ws = wb["Regnskapsoppstilling"]
    assert "Regnskapsoppstilling" in ws["A1"].value
    # Klient/år flyttet til undertittel-striperad 2
    assert "Nbs Regnskap AS" in (ws["A2"].value or "")
    # Standardkolonner (uten IB og Bevegelse)
    assert ws["A4"].value == "Nr"
    assert ws["B4"].value == "Regnskapslinje"
    assert ws["C4"].value == "I år"
    assert ws["D4"].value == "Antall poster"
    # Rad 5 er seksjonsoverskriften RESULTATREGNSKAP; regnr 10 og 99 følger på rad 6/7
    assert ws["A5"].value == "RESULTATREGNSKAP"
    assert ws["A6"].value == 10
    assert ws["A7"].value == 99
    # Sumpost (regnr 99) should be bold
    assert ws["B7"].font.bold is True


def test_columns_with_fjor_data() -> None:
    """Når UB_fjor finnes, skal Endring vs i fjor (kr) og (%) vises."""
    from analyse_regnskapsoppstilling_excel import build_regnskapsoppstilling_workbook

    rl_df = pd.DataFrame(
        {
            "regnr": [10, 20],
            "regnskapslinje": ["Salgsinntekt", "Varekostnad"],
            "IB": [0.0, 0.0],
            "Endring": [1000.0, 500.0],
            "UB": [1000.0, 500.0],
            "UB_fjor": [800.0, 600.0],
            "Antall": [10, 5],
        }
    )

    wb = build_regnskapsoppstilling_workbook(rl_df, client="Test", year="2025")
    ws = wb["Regnskapsoppstilling"]

    # Finn header-kolonner
    headers = [ws.cell(row=4, column=c).value for c in range(1, 10)]
    assert "I fjor" in headers
    assert "Endring" in headers
    assert "Endring (%)" in headers

    # Sjekk beregnet endring for regnr 10 (rad 6 pga. seksjonsoverskrift på rad 5):
    # UB=1000, UB_fjor=800 → endring=200
    endr_kr_col = headers.index("Endring") + 1
    assert ws.cell(row=6, column=endr_kr_col).value == 200.0


def test_nokkeltall_under_regnskapsoppstilling() -> None:
    """Nøkkeltall skal plasseres under regnskapsoppstillingen på samme ark."""
    from analyse_regnskapsoppstilling_excel import build_regnskapsoppstilling_workbook

    rl_df = pd.DataFrame(
        {
            "regnr": [10, 19, 20, 40, 70, 79, 80, 280, 660, 665, 715, 810, 820],
            "regnskapslinje": [
                "Salgsinntekt", "Sum driftsinntekter", "Varekostnad",
                "Lønnskostnad", "Annen driftskostnad", "Sum driftskostnader",
                "Driftsresultat", "Årsresultat",
                "Sum omløpsmidler", "Sum eiendeler",
                "Sum egenkapital", "Sum kortsiktig gjeld", "Sum gjeld",
            ],
            "UB": [
                -1000000, -1000000, 600000,
                200000, 100000, 900000,
                -100000, -80000,
                -500000, -800000,
                -300000, -400000, -500000,
            ],
            "Antall": [100, 100, 50, 80, 40, 170, 170, 170, 200, 200, 10, 150, 150],
        }
    )

    wb = build_regnskapsoppstilling_workbook(rl_df, client="Test", year="2025")
    ws = wb["Regnskapsoppstilling"]

    # Nøkkeltall skal IKKE ha eget ark
    assert "Nøkkeltall" not in wb.sheetnames

    # Finn nøkkeltall-seksjonen på hovedarket (søk etter "Nøkkeltall"-tittelen)
    nk_title_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Nøkkeltall":
            nk_title_row = r
            break
    assert nk_title_row is not None, "Nøkkeltall-seksjon ikke funnet på arket"

    # Skal være minst 3 rader under siste datarad (rad 4 header + 13 data = rad 17, +3 = 20)
    assert nk_title_row >= 20

    # Sjekk at kategorier finnes (kategorier ligger på kol 2)
    labels = []
    for r in range(nk_title_row, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val:
            labels.append(val)
    assert "Lønnsomhet" in labels
    assert "Likviditet" in labels
    assert "Soliditet" in labels


def test_nokkeltall_with_fjor() -> None:
    """Nøkkeltall skal vise fjorårstall og endring når UB_fjor finnes."""
    from analyse_regnskapsoppstilling_excel import build_regnskapsoppstilling_workbook

    rl_df = pd.DataFrame(
        {
            "regnr": [10, 19, 80, 280, 660, 665, 715, 810, 820],
            "regnskapslinje": [
                "Salgsinntekt", "Sum driftsinntekter",
                "Driftsresultat", "Årsresultat",
                "Sum omløpsmidler", "Sum eiendeler",
                "Sum egenkapital", "Sum kortsiktig gjeld", "Sum gjeld",
            ],
            "UB": [-1000000, -1000000, -100000, -80000,
                   -500000, -800000, -300000, -400000, -500000],
            "UB_fjor": [-900000, -900000, -80000, -60000,
                        -450000, -700000, -280000, -350000, -420000],
            "Antall": [100, 100, 100, 100, 100, 100, 10, 100, 100],
        }
    )

    wb = build_regnskapsoppstilling_workbook(rl_df, client="Test", year="2025")
    ws = wb["Regnskapsoppstilling"]

    # Finn nøkkeltall-header-raden
    nk_header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Nøkkeltall":
            nk_header_row = r + 1  # header er raden etter tittel
            break
    assert nk_header_row is not None

    # Sjekk nøkkeltall-header (layout: kol1 tom, kol2 Nøkkeltall, kol3 I år,
    # kol4 I fjor, kol5 Endring, kol6 Endring (%))
    assert ws.cell(row=nk_header_row, column=2).value == "Nøkkeltall"
    assert ws.cell(row=nk_header_row, column=3).value == "I år"
    assert ws.cell(row=nk_header_row, column=4).value == "I fjor"
    assert ws.cell(row=nk_header_row, column=5).value == "Endring"

    # Sjekk at "I fjor"-kolonnen (kol 4) har verdier (ikke tom)
    has_fjor_values = False
    for r in range(nk_header_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=4).value is not None:
            has_fjor_values = True
            break
    assert has_fjor_values, "I fjor-kolonnen har ingen verdier"


def test_beregningsgrunnlag_sheet() -> None:
    """Beregningsgrunnlag-arket skal inneholde formelreferanser."""
    from analyse_regnskapsoppstilling_excel import build_regnskapsoppstilling_workbook

    rl_df = pd.DataFrame(
        {
            "regnr": [10],
            "regnskapslinje": ["Salgsinntekt"],
            "UB": [-1000000],
            "Antall": [100],
        }
    )

    wb = build_regnskapsoppstilling_workbook(rl_df, client="Test", year="2025")
    assert "Beregningsgrunnlag" in wb.sheetnames
    ws = wb["Beregningsgrunnlag"]

    labels = []
    for row in range(5, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            labels.append(val)

    assert "Lønnsomhet" in labels
    assert "Bruttofortjeneste" in labels
    assert "Likviditetsgrad 1" in labels
    assert "Egenkapitalandel" in labels


def test_save_regnskapsoppstilling_workbook_writes_file(tmp_path) -> None:
    from analyse_regnskapsoppstilling_excel import save_regnskapsoppstilling_workbook

    rl_df = pd.DataFrame(
        {
            "regnr": [10],
            "regnskapslinje": ["Eiendeler"],
            "IB": [100.0],
            "Endring": [25.0],
            "UB": [125.0],
            "Antall": [2],
        }
    )

    out = save_regnskapsoppstilling_workbook(tmp_path / "rl_export", rl_df=rl_df, client="Test", year="2025")

    wb = openpyxl.load_workbook(out)
    assert "Regnskapsoppstilling" in wb.sheetnames
    assert "Kontospesifisert" in wb.sheetnames
    assert "Beregningsgrunnlag" in wb.sheetnames
    # A5 er seksjonsoverskrift; regnr 10 ligger på rad 6
    assert wb["Regnskapsoppstilling"]["A5"].value == "RESULTATREGNSKAP"
    assert wb["Regnskapsoppstilling"]["A6"].value == 10


def test_section_headers_emitted_at_boundaries() -> None:
    """Seksjonsoverskrifter skal skrives ved overgang Resultat → Eiendeler → EK."""
    from analyse_regnskapsoppstilling_excel import build_regnskapsoppstilling_workbook

    rl_df = pd.DataFrame(
        {
            "regnr":          [10,   280,             500,             715,               820],
            "regnskapslinje": ["Salg", "Årsresultat", "Anleggsmidler", "Sum egenkapital", "Sum gjeld"],
            "UB":             [1000,  400,             2000,            1500,              1500],
            "Antall":         [5,     0,               3,               0,                 0],
        }
    )

    wb = build_regnskapsoppstilling_workbook(rl_df, client="Test", year="2025")
    ws = wb["Regnskapsoppstilling"]

    # Samle alle A-kolonneverdier som er strenger (= seksjonsoverskrifter, ikke regnr)
    section_labels = [
        str(ws.cell(row=r, column=1).value)
        for r in range(5, ws.max_row + 1)
        if isinstance(ws.cell(row=r, column=1).value, str)
        and ws.cell(row=r, column=1).value
        in {"RESULTATREGNSKAP", "EIENDELER", "EGENKAPITAL OG GJELD"}
    ]
    assert section_labels == ["RESULTATREGNSKAP", "EIENDELER", "EGENKAPITAL OG GJELD"]


def test_kontospesifisert_sheet_lists_accounts_under_rl() -> None:
    """Kontospesifisert-arket skal vise SB-kontoer under de RL-linjene de mapper til."""
    from analyse_regnskapsoppstilling_excel import build_regnskapsoppstilling_workbook

    rl_df = pd.DataFrame(
        {
            "regnr":          [10,      80,               280,             500],
            "regnskapslinje": ["Salg",  "Driftsresultat", "Årsresultat",    "Anleggsmidler"],
            "UB":             [1000.0,  500.0,            400.0,           2000.0],
            "Antall":         [5,       0,                0,               3],
        }
    )
    regnskapslinjer = pd.DataFrame(
        {
            "regnr":          [10,      80,               280,             500],
            "regnskapslinje": ["Salg",  "Driftsresultat", "Årsresultat",   "Anleggsmidler"],
            "sumpost":        [False,   True,             True,            False],
            "fortegn":        ["+",     "+",              "+",             "+"],
        }
    )
    intervals = pd.DataFrame(
        {"regnr": [10, 500], "fra": ["3000", "1200"], "til": ["3099", "1299"]}
    )
    hb = pd.DataFrame(
        {
            "Konto":     ["3000",       "3010",        "1200"],
            "Kontonavn": ["Salg Norge", "Salg Utland", "Maskiner"],
            "Beløp":     [600.0,        400.0,         2000.0],
        }
    )
    sb = pd.DataFrame(
        {
            "konto":     ["3000",       "3010",        "1200"],
            "kontonavn": ["Salg Norge", "Salg Utland", "Maskiner"],
            "ib":        [0.0,          0.0,           1800.0],
            "ub":        [600.0,        400.0,         2000.0],
        }
    )

    wb = build_regnskapsoppstilling_workbook(
        rl_df,
        regnskapslinjer=regnskapslinjer,
        transactions_df=hb,
        sb_df=sb,
        intervals=intervals,
        client="Test AS",
        year="2025",
    )
    assert "Kontospesifisert" in wb.sheetnames
    ws = wb["Kontospesifisert"]

    # Samle alle verdier i kolonne A+B for å verifisere rekkefølgen
    rows = [(ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value)
            for r in range(5, ws.max_row + 1)]

    # Leaf-RL 10 "Salg" skal ha konto 3000 og 3010 under seg
    salg_idx = next(i for i, (a, b) in enumerate(rows) if a == 10 and b == "Salg")
    assert rows[salg_idx + 1] == ("3000", "Salg Norge")
    assert rows[salg_idx + 2] == ("3010", "Salg Utland")

    # Sum-poster (80, 280) skal IKKE ha konto-rader under seg
    driftsres_idx = next(i for i, (a, b) in enumerate(rows) if a == 80)
    # Neste rad etter Driftsres. skal være regnr 280 (Årsresultat), ikke en konto
    assert rows[driftsres_idx + 1][0] == 280

    # Leaf-RL 500 "Anleggsmidler" skal ha konto 1200 under seg
    anlegg_idx = next(i for i, (a, b) in enumerate(rows) if a == 500)
    assert rows[anlegg_idx + 1] == ("1200", "Maskiner")


def test_kontospesifisert_sheet_without_sb_data_is_empty_of_account_rows() -> None:
    """Uten sb_df/intervals skal Kontospesifisert ha RL-linjer, men ingen konto-rader."""
    from analyse_regnskapsoppstilling_excel import build_regnskapsoppstilling_workbook

    rl_df = pd.DataFrame(
        {
            "regnr":          [10,     500],
            "regnskapslinje": ["Salg", "Anleggsmidler"],
            "UB":             [1000.0, 2000.0],
            "Antall":         [5,      3],
        }
    )

    wb = build_regnskapsoppstilling_workbook(rl_df, client="Test", year="2025")
    assert "Kontospesifisert" in wb.sheetnames
    ws = wb["Kontospesifisert"]

    # Det skal IKKE finnes rader med konto-format (streng i A-kolonnen som ikke er seksjonsoverskrift)
    section_labels = {"RESULTATREGNSKAP", "EIENDELER", "EGENKAPITAL OG GJELD"}
    for r in range(5, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val:
            assert val in section_labels, f"Uventet streng på A{r}: {val!r}"
