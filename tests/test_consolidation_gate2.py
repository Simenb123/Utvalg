"""Gate 2 integration test — full consolidation round-trip.

Verifies the Gate 2 criterion:
  At least 2 companies imported, mapped, eliminated, run, exported
  in one client/year project, and state reopenable without loss.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from consolidation.models import (
    CompanyTB,
    ConsolidationProject,
    EliminationJournal,
    EliminationLine,
    MappingConfig,
    RunResult,
)
from consolidation.storage import (
    delete_project,
    load_company_tb,
    load_project,
    save_company_tb,
    save_project,
)
from consolidation.engine import run_consolidation
from consolidation.export import save_consolidation_workbook
from consolidation.mapping import map_company_tb


# ---------------------------------------------------------------------------
# Synthetic test data
# ---------------------------------------------------------------------------

def _intervals_df() -> pd.DataFrame:
    return pd.DataFrame({
        "fra": [1000, 3000],
        "til": [1999, 3999],
        "regnr": [10, 11],
    })


def _regnskapslinjer_df() -> pd.DataFrame:
    return pd.DataFrame({
        "regnr": [10, 11, 20],
        "regnskapslinje": ["Eiendeler", "Inntekter", "SUM"],
        "sumpost": [False, False, True],
        "formel": [None, None, "=10+11"],
        "sumnivaa": [None, None, None],
        "delsumnr": [None, None, None],
        "sumnr": [None, None, None],
        "sumnr2": [None, None, None],
        "sluttsumnr": [None, None, None],
    })


def _mor_tb() -> pd.DataFrame:
    return pd.DataFrame({
        "konto": ["1000", "1500", "3000", "3400"],
        "kontonavn": ["Bank", "Kundefordringer", "Salgsinntekt", "Interco salg"],
        "ib": [0.0, 0.0, 0.0, 0.0],
        "ub": [200.0, 80.0, -500.0, -150.0],
        "netto": [200.0, 80.0, -500.0, -150.0],
    })


def _datter_tb() -> pd.DataFrame:
    return pd.DataFrame({
        "konto": ["1100", "3100", "3200"],
        "kontonavn": ["Kasse", "Tjenestesalg", "Varesalg"],
        "ib": [0.0, 0.0, 0.0],
        "ub": [60.0, -200.0, -100.0],
        "netto": [60.0, -200.0, -100.0],
    })


@pytest.fixture
def _mock_config(monkeypatch):
    import regnskap_config
    monkeypatch.setattr(
        regnskap_config, "load_kontoplan_mapping",
        lambda **kw: _intervals_df(),
    )
    monkeypatch.setattr(
        regnskap_config, "load_regnskapslinjer",
        lambda **kw: _regnskapslinjer_df(),
    )


@pytest.fixture
def _mock_storage(monkeypatch, tmp_path):
    """Redirect client_store.years_dir to tmp_path."""
    import client_store
    monkeypatch.setattr(
        client_store, "years_dir",
        lambda client, year: tmp_path / client / "years" / year,
    )
    return tmp_path


# ---------------------------------------------------------------------------
# Gate 2: End-to-end
# ---------------------------------------------------------------------------

class TestGate2EndToEnd:
    """Full round-trip: create -> import -> map -> eliminate -> run -> export -> reopen."""

    def test_full_consolidation_roundtrip(self, _mock_config, _mock_storage, tmp_path):
        # ── 1. Create project ──
        project = ConsolidationProject(
            client="IntegTestKonsern",
            year="2025",
        )

        # ── 2. Import 2 companies ──
        mor = CompanyTB(company_id="mor", name="Morselskap AS", row_count=4, source_type="excel")
        datter = CompanyTB(company_id="dat", name="Datter AS", row_count=3, source_type="excel")
        project.companies = [mor, datter]

        mor_df = _mor_tb()
        dat_df = _datter_tb()

        save_company_tb("IntegTestKonsern", "2025", "mor", mor_df)
        save_company_tb("IntegTestKonsern", "2025", "dat", dat_df)

        # ── 3. Add elimination journal ──
        journal = EliminationJournal(
            journal_id="elim1",
            name="Internhandel",
            lines=[
                EliminationLine(regnr=11, company_id="mor", amount=150.0, description="Elim interco salg"),
                EliminationLine(regnr=11, company_id="dat", amount=-150.0, description="Elim interco kjop"),
            ],
        )
        project.eliminations = [journal]

        # Verify journal balances
        assert journal.is_balanced
        assert abs(journal.net) < 0.005

        # ── 4. Save project state ──
        save_project(project)

        # ── 5. Run consolidation ──
        tbs = {"mor": mor_df, "dat": dat_df}
        result_df, run_result = run_consolidation(project, tbs)

        # Verify result structure
        assert "Morselskap AS" in result_df.columns
        assert "Datter AS" in result_df.columns
        assert "sum_foer_elim" in result_df.columns
        assert "eliminering" in result_df.columns
        assert "konsolidert" in result_df.columns

        # Verify leaf values for regnr 10 (Eiendeler)
        leaf = result_df[~result_df["sumpost"]]
        r10 = leaf[leaf["regnr"] == 10].iloc[0]
        # Mor: 1000(200) + 1500(80) = 280, Datter: 1100(60) = 60
        assert r10["Morselskap AS"] == pytest.approx(280.0)
        assert r10["Datter AS"] == pytest.approx(60.0)
        assert r10["sum_foer_elim"] == pytest.approx(340.0)
        assert r10["eliminering"] == pytest.approx(0.0)
        assert r10["konsolidert"] == pytest.approx(340.0)

        # Verify leaf values for regnr 11 (Inntekter)
        r11 = leaf[leaf["regnr"] == 11].iloc[0]
        # Mor: 3000(-500) + 3400(-150) = -650, Datter: 3100(-200) + 3200(-100) = -300
        assert r11["Morselskap AS"] == pytest.approx(-650.0)
        assert r11["Datter AS"] == pytest.approx(-300.0)
        assert r11["sum_foer_elim"] == pytest.approx(-950.0)
        # Eliminering: +150 + (-150) = 0 on regnr 11
        assert r11["eliminering"] == pytest.approx(0.0)
        assert r11["konsolidert"] == pytest.approx(-950.0)

        # Verify sum line (formel: =10+11)
        sum_row = result_df[result_df["regnr"] == 20].iloc[0]
        assert sum_row["konsolidert"] == pytest.approx(340.0 + (-950.0))

        # Verify RunResult
        assert len(run_result.company_ids) == 2
        assert run_result.result_hash != ""
        assert len(run_result.warnings) == 0

        # ── 6. Export to Excel ──
        out_path = tmp_path / "export" / "consolidation.xlsx"
        mapped_tbs = {}
        for cid, df in tbs.items():
            mapped_df, _ = map_company_tb(df)
            mapped_tbs[cid] = mapped_df

        saved = save_consolidation_workbook(
            out_path,
            result_df=result_df,
            companies=project.companies,
            eliminations=project.eliminations,
            mapped_tbs=mapped_tbs,
            run_result=run_result,
            client=project.client,
            year=project.year,
        )
        assert Path(saved).exists()

        # Verify Excel content
        wb = load_workbook(saved)
        assert "Konsernoppstilling" in wb.sheetnames
        assert "Elimineringer" in wb.sheetnames
        assert "TB - Morselskap AS" in wb.sheetnames
        assert "TB - Datter AS" in wb.sheetnames
        assert "Kontrollark" in wb.sheetnames

        # Verify title cell
        ws = wb["Konsernoppstilling"]
        assert "IntegTestKonsern" in str(ws["A1"].value)
        assert "2025" in str(ws["A1"].value)

        # ── 7. Reopen state — verify persistence ──
        project.runs.append(run_result)
        save_project(project)

        reloaded = load_project("IntegTestKonsern", "2025")
        assert reloaded is not None
        assert reloaded.project_id == project.project_id
        assert reloaded.client == "IntegTestKonsern"
        assert reloaded.year == "2025"

        # Companies survived
        assert len(reloaded.companies) == 2
        mor_reloaded = reloaded.find_company("mor")
        assert mor_reloaded is not None
        assert mor_reloaded.name == "Morselskap AS"

        dat_reloaded = reloaded.find_company("dat")
        assert dat_reloaded is not None
        assert dat_reloaded.name == "Datter AS"

        # Eliminations survived
        assert len(reloaded.eliminations) == 1
        j_reloaded = reloaded.find_journal("elim1")
        assert j_reloaded is not None
        assert j_reloaded.name == "Internhandel"
        assert len(j_reloaded.lines) == 2
        assert j_reloaded.is_balanced

        # RunResult survived
        assert len(reloaded.runs) == 1
        assert reloaded.runs[0].result_hash == run_result.result_hash
        assert reloaded.runs[0].run_id == run_result.run_id

        # Company TBs survived on disk
        mor_tb_reloaded = load_company_tb("IntegTestKonsern", "2025", "mor")
        assert mor_tb_reloaded is not None
        assert len(mor_tb_reloaded) == 4
        assert list(mor_tb_reloaded["konto"]) == ["1000", "1500", "3000", "3400"]

        dat_tb_reloaded = load_company_tb("IntegTestKonsern", "2025", "dat")
        assert dat_tb_reloaded is not None
        assert len(dat_tb_reloaded) == 3

        # ── 8. Re-run with reloaded state produces same hash ──
        tbs_reloaded = {
            "mor": mor_tb_reloaded,
            "dat": dat_tb_reloaded,
        }
        _, run2 = run_consolidation(reloaded, tbs_reloaded)
        assert run2.result_hash == run_result.result_hash

    def test_delete_and_recreate(self, _mock_config, _mock_storage):
        """Delete project and verify clean recreation."""
        project = ConsolidationProject(
            client="DeleteTest",
            year="2025",
            companies=[CompanyTB(company_id="x", name="X AS", row_count=1)],
        )
        save_project(project)
        assert load_project("DeleteTest", "2025") is not None

        deleted = delete_project("DeleteTest", "2025")
        assert deleted is True
        assert load_project("DeleteTest", "2025") is None

        # Recreate
        project2 = ConsolidationProject(client="DeleteTest", year="2025")
        save_project(project2)
        reloaded = load_project("DeleteTest", "2025")
        assert reloaded is not None
        assert reloaded.project_id == project2.project_id

    def test_unbalanced_elimination_still_runs(self, _mock_config, _mock_storage):
        """Unbalanced elimination produces a warning but completes."""
        project = ConsolidationProject(
            client="UbalTest",
            year="2025",
            companies=[CompanyTB(company_id="a", name="Selskap A", row_count=2)],
            eliminations=[
                EliminationJournal(
                    journal_id="e1",
                    name="Ubalansert",
                    lines=[
                        EliminationLine(regnr=10, company_id="a", amount=100.0),
                    ],
                ),
            ],
        )
        tb = pd.DataFrame({
            "konto": ["1000", "3000"],
            "kontonavn": ["Bank", "Salg"],
            "ib": [0.0, 0.0],
            "ub": [50.0, -100.0],
            "netto": [50.0, -100.0],
        })

        result_df, run_result = run_consolidation(project, {"a": tb})

        assert any("ikke balansert" in w for w in run_result.warnings)
        # Eliminering paa regnr 10 = 100.0
        leaf = result_df[~result_df["sumpost"]]
        r10 = leaf[leaf["regnr"] == 10].iloc[0]
        assert r10["eliminering"] == pytest.approx(100.0)
        assert r10["konsolidert"] == pytest.approx(50.0 + 100.0)
