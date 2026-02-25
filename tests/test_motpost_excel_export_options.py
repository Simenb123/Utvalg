import pandas as pd
from openpyxl import Workbook

from motpost.excel_sheets.sheet_outlier_full_bilag import (
    build_outlier_frames,
    write_outlier_transactions_sheet,
)
from motpost.konto_core import build_motpost_data
from motpost_konto_core import build_motpost_excel_workbook


def test_excel_export_can_omit_outlier_transactions_and_remove_hyperlinks():
    df = pd.DataFrame(
        [
            {
                "Bilag": 1,
                "Dato": "2025-01-01",
                "Tekst": "Utgående faktura",
                "Konto": "3000",
                "Kontonavn": "Salg",
                "Beløp": -1000.0,
            },
            {
                "Bilag": 1,
                "Dato": "2025-01-01",
                "Tekst": "Utgående faktura",
                "Konto": "1500",
                "Kontonavn": "Kundefordringer",
                "Beløp": 1000.0,
            },
        ]
    )

    data = build_motpost_data(df, {"3000"}, selected_direction="Kredit")

    wb = build_motpost_excel_workbook(
        data,
        combo_status_map={"1500": "outlier"},
        include_outlier_transactions=False,
    )

    # Outlier-arket skal finnes, men inneholde et notat om at bilagslinjer er utelatt.
    ws_out = wb["Outlier - alle transaksjoner"]
    assert isinstance(ws_out["A2"].value, str)
    assert "utelatt" in ws_out["A2"].value.lower()

    # Detaljarket skal ikke ha hyperlink til outlier-arket når bilagslinjer er utelatt.
    ws_detail = wb["#2"]
    assert isinstance(ws_detail["B11"].value, str)
    assert not ws_detail["B11"].value.startswith("=HYPERLINK")
    assert "utelatt" in ws_detail["B11"].value.lower()

    # Bilagsoppsummering: Bilagslinjer-kolonnen skal være tekst, ikke hyperlink.
    # Finn kolonnen dynamisk (layout kan endre seg når antall tabeller/kolonner endres).
    bilagslinjer_cell = None
    for row in ws_detail.iter_rows():
        for cell in row:
            if cell.value == "Bilagslinjer":
                bilagslinjer_cell = cell
                break
        if bilagslinjer_cell is not None:
            break

    assert bilagslinjer_cell is not None
    value_cell = ws_detail.cell(row=bilagslinjer_cell.row + 1, column=bilagslinjer_cell.column)
    assert value_cell.value == "Utelatt i eksport"


def test_outlier_frames_auto_omits_transactions_when_exceeding_excel_row_limit():
    df_scope = pd.DataFrame(
        [
            {
                "Bilag": 1,
                "Dato": "2025-01-01",
                "Tekst": "Utgående faktura",
                "Konto": "3000",
                "Kontonavn": "Salg",
                "Beløp": -1000.0,
            },
            {
                "Bilag": 1,
                "Dato": "2025-01-01",
                "Tekst": "Utgående faktura",
                "Konto": "1500",
                "Kontonavn": "Kundefordringer",
                "Beløp": 1000.0,
            },
        ]
    )

    data = build_motpost_data(df_scope, {"3000"}, selected_direction="Kredit")

    # Tving "for mange rader" uten å lage en enorm DataFrame ved å sette max_data_rows=0.
    frames = build_outlier_frames(
        data.df_scope,
        selected_accounts=list(data.selected_accounts),
        outlier_combos=["1500"],
        combo_name_map={"1500": "1500 - Kundefordringer"},
        include_transactions=True,
        max_data_rows=0,
    )

    assert frames.transactions_included is False
    assert frames.outlier_lines_count > 0
    assert "excel" in frames.transactions_omitted_reason.lower()

    wb = Workbook()
    write_outlier_transactions_sheet(wb, frames=frames, sheet_name="Outlier - alle transaksjoner")

    ws_out = wb["Outlier - alle transaksjoner"]
    assert isinstance(ws_out["A2"].value, str)
    assert "for mange rader" in ws_out["A2"].value.lower()
