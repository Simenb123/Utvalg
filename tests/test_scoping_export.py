"""Tests for scoping_export."""

from __future__ import annotations

from pathlib import Path

import pytest

from src.pages.scoping.backend.engine import ScopingLine, ScopingResult
from src.pages.scoping.backend.export import export_scoping


@pytest.fixture()
def sample_result():
    return ScopingResult(
        lines=[
            ScopingLine(
                regnr="10", regnskapslinje="Salgsinntekt", line_type="PL",
                amount=12_000_000, amount_prior=10_000_000,
                change_amount=2_000_000,
                change_pct=20.0, pct_of_pm=3200.0,
                classification="vesentlig", auto_classification="vesentlig",
                scoping="inn", action_count=7,
            ),
            ScopingLine(
                regnr="19", regnskapslinje="Sum driftsinntekter", line_type="PL",
                amount=12_000_000, is_summary=True,
            ),
            ScopingLine(
                regnr="70", regnskapslinje="Annen driftskostnad", line_type="PL",
                amount=18_000, classification="ikke_vesentlig",
                auto_classification="ikke_vesentlig",
                scoping="ut", rationale="Under SUM",
            ),
        ],
        om=500_000, pm=375_000, sum_threshold=25_000,
        scoped_out_total=18_000, aggregation_ok=True,
    )


def test_export_creates_file(sample_result, tmp_path):
    path = tmp_path / "scoping_test.xlsx"
    result = export_scoping(sample_result, path, client_name="Test AS", year="2025")
    assert result.exists()
    assert result.stat().st_size > 0


def test_export_has_four_sheets(sample_result, tmp_path):
    """Eksport skal ha Oversikt + Alle + Inn + Ut faner."""
    import openpyxl

    path = tmp_path / "scoping_test.xlsx"
    export_scoping(sample_result, path, client_name="Test AS", year="2025")

    wb = openpyxl.load_workbook(str(path))
    assert wb.sheetnames == ["Oversikt", "Alle", "Inn", "Ut"]


def test_oversikt_sheet_has_klient_info(sample_result, tmp_path):
    import openpyxl

    path = tmp_path / "scoping_test.xlsx"
    export_scoping(sample_result, path, client_name="Test AS", year="2025")

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Oversikt"]
    assert ws["A1"].value == "SCOPING — OVERSIKT"

    # Skann alle celler for klient-navn og år
    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, 30) for c in range(1, 5)
    ]
    assert "Test AS" in all_values
    assert "2025" in all_values


def test_alle_sheet_has_correct_headers_with_prior_year(sample_result, tmp_path):
    import openpyxl

    path = tmp_path / "scoping_test.xlsx"
    export_scoping(sample_result, path, client_name="Test AS", year="2025")

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Alle"]

    # Headers ligger på rad 5 (tittel rad 1, beskrivelse rad 2,
    # antall-linje rad 3, blank rad 4, headers rad 5).
    headers = [ws.cell(row=5, column=c).value for c in range(1, 14)]
    assert headers[:7] == [
        "Regnr",
        "Regnskapslinje",
        "Type",
        "UB 2025",
        "UB 2024",
        "Endring",
        "Endring %",
    ]
    assert headers[7:] == [
        "% av PM",
        "Klassifisering",
        "Scoping",
        "Revisjonshandling",
        "Begrunnelse",
        "Handl.",
    ]


def test_alle_sheet_data_row_values(sample_result, tmp_path):
    """Første data-rad (rad 6) skal ha forventet verdier."""
    import openpyxl

    path = tmp_path / "scoping_test.xlsx"
    export_scoping(sample_result, path, client_name="Test AS", year="2025")

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Alle"]

    # Rad 6 er første linje (Regnr 10, Salgsinntekt)
    assert ws.cell(row=6, column=4).value == 12_000_000
    assert ws.cell(row=6, column=5).value == 10_000_000
    assert ws.cell(row=6, column=6).value == 2_000_000
    assert ws.cell(row=6, column=7).value == "+20.0%"


def test_inn_sheet_only_contains_scoped_in(sample_result, tmp_path):
    import openpyxl

    path = tmp_path / "scoping_test.xlsx"
    export_scoping(sample_result, path, client_name="Test AS", year="2025")

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Inn"]

    # Sample-result har bare 1 linje scopet inn (regnr 10)
    # Sjekk at Salgsinntekt er der, men ikke Annen driftskostnad (scopet ut)
    all_values = [
        ws.cell(row=r, column=2).value for r in range(1, 20)
    ]
    assert "Salgsinntekt" in all_values
    assert "Annen driftskostnad" not in all_values


def test_ut_sheet_only_contains_scoped_out(sample_result, tmp_path):
    import openpyxl

    path = tmp_path / "scoping_test.xlsx"
    export_scoping(sample_result, path, client_name="Test AS", year="2025")

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Ut"]

    all_values = [
        ws.cell(row=r, column=2).value for r in range(1, 20)
    ]
    assert "Annen driftskostnad" in all_values
    assert "Salgsinntekt" not in all_values


def test_export_hides_prior_year_columns_when_absent(tmp_path):
    """Når ingen linje har amount_prior, skal UB i fjor-kolonnene utelates."""
    import openpyxl

    result = ScopingResult(
        lines=[
            ScopingLine(
                regnr="70",
                regnskapslinje="Annen driftskostnad",
                line_type="PL",
                amount=18_000,
                classification="ikke_vesentlig",
                auto_classification="ikke_vesentlig",
            )
        ],
        om=500_000,
        pm=375_000,
        sum_threshold=25_000,
    )

    path = tmp_path / "scoping_no_prev.xlsx"
    export_scoping(result, path, client_name="Test AS", year="2025")

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Alle"]
    headers = [ws.cell(row=5, column=c).value for c in range(1, 11)]
    assert headers == [
        "Regnr",
        "Regnskapslinje",
        "Type",
        "UB 2025",
        "% av PM",
        "Klassifisering",
        "Scoping",
        "Revisjonshandling",
        "Begrunnelse",
        "Handl.",
    ]
