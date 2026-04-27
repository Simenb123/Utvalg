"""Smoke tests for the Vaak theme layer (GUI + Excel)."""
from __future__ import annotations

import re
import tkinter as tk

import pytest
from openpyxl import Workbook

import src.shared.ui.excel_theme as vxt
import src.shared.ui.tokens as vt
from theme import apply_theme, style_treeview_tags, tree_tag


_HEX_RE = re.compile(r"^[0-9A-Fa-f]{6}$")


def test_color_tokens_are_6digit_hex() -> None:
    names = [
        "BG_SAND", "BG_SAND_SOFT", "BG_NEUTRAL", "BG_DATA", "BG_ZEBRA",
        "SAGE", "SAGE_DARK", "FOREST", "FOREST_HOVER", "OLIVE",
        "SELECT_BG", "SELECT_FG",
        "TEXT_PRIMARY", "TEXT_MUTED", "TEXT_ON_SAND", "TEXT_ON_FOREST",
        "BORDER", "BORDER_SOFT",
        "POS_TEXT", "NEG_TEXT", "WARN_TEXT",
        "POS_SOFT", "NEG_SOFT", "WARN_SOFT",
    ]
    for name in names:
        value = getattr(vt, name)
        assert _HEX_RE.match(value), f"{name} invalid: {value!r}"


def test_hex_gui_adds_hash_prefix() -> None:
    assert vt.hex_gui("DAC79E") == "#DAC79E"
    assert vt.hex_gui("#DAC79E") == "#DAC79E"


def test_treeview_selection_token_is_distinct_from_status_sand() -> None:
    assert vt.SELECT_BG != vt.BG_SAND
    assert vt.SELECT_FG == vt.TEXT_ON_FOREST


def test_apply_theme_does_not_raise() -> None:
    try:
        root = tk.Tk()
    except tk.TclError:
        pytest.skip("No display available")
    try:
        apply_theme(root)
    finally:
        root.destroy()


def test_style_treeview_tags_applies() -> None:
    try:
        root = tk.Tk()
    except tk.TclError:
        pytest.skip("No display available")
    try:
        apply_theme(root)
        from tkinter import ttk

        tree = ttk.Treeview(root)
        style_treeview_tags(tree, "sumline", "sumline_major", "commented")
        assert tree.tag_configure("sumline", "background") != ""
    finally:
        root.destroy()


def test_tree_tag_returns_copy() -> None:
    a = tree_tag("sumline")
    a["background"] = "xxxx"
    b = tree_tag("sumline")
    assert b["background"] != "xxxx"


def test_register_vaak_styles_adds_all_names() -> None:
    wb = Workbook()
    vxt.register_vaak_styles(wb)
    expected = {
        "vaak_title", "vaak_h1", "vaak_h2", "vaak_header", "vaak_subheader",
        "vaak_body", "vaak_muted", "vaak_sum", "vaak_sum_major",
        "vaak_total", "vaak_pos", "vaak_neg", "vaak_zebra",
    }
    assert expected.issubset(set(wb.named_styles))


def test_register_vaak_styles_is_idempotent() -> None:
    wb = Workbook()
    vxt.register_vaak_styles(wb)
    before = list(wb.named_styles)
    vxt.register_vaak_styles(wb)
    assert list(wb.named_styles) == before


def test_named_style_can_be_applied_to_cell() -> None:
    wb = Workbook()
    vxt.register_vaak_styles(wb)
    ws = wb.active
    ws["A1"] = "Test"
    ws["A1"].style = "vaak_title"
    assert ws["A1"].fill.fgColor.rgb.endswith(vt.BG_SAND)
