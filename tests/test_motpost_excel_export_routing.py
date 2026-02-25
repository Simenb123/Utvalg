import pandas as pd
from openpyxl import load_workbook

from motpost.konto_core import build_motpost_data
from motpost_konto_core import build_motpost_excel_workbook


def _sample_df() -> pd.DataFrame:
    """Liten, deterministisk datasett som gir én motkonto-kombinasjon."""

    return pd.DataFrame(
        [
            {
                "Bilag": 1,
                "Dato": "2025-01-01",
                "Tekst": "Utgående faktura",
                "Konto": "3000",
                "Kontonavn": "Salg",
                "Beløp": -1000.0,
            },
            {
                "Bilag": 1,
                "Dato": "2025-01-01",
                "Tekst": "Utgående faktura",
                "Konto": "2700",
                "Kontonavn": "Utgående mva",
                "Beløp": 250.0,
            },
            {
                "Bilag": 1,
                "Dato": "2025-01-01",
                "Tekst": "Utgående faktura",
                "Konto": "1500",
                "Kontonavn": "Kundefordringer",
                "Beløp": 750.0,
            },
            {
                "Bilag": 2,
                "Dato": "2025-01-02",
                "Tekst": "Utgående faktura",
                "Konto": "3000",
                "Kontonavn": "Salg",
                "Beløp": -500.0,
            },
            {
                "Bilag": 2,
                "Dato": "2025-01-02",
                "Tekst": "Utgående faktura",
                "Konto": "1500",
                "Kontonavn": "Kundefordringer",
                "Beløp": 500.0,
            },
        ]
    )


def test_export_creates_expected_sheets_and_outlier_details(tmp_path):
    df = _sample_df()
    data = build_motpost_data(df, {"3000"}, selected_direction="Kredit")

    # Kombinasjon i datasettet blir "1500, 2700" (komma+space, sortert)
    wb = build_motpost_excel_workbook(data, combo_status_map={"1500, 2700": "outlier"})

    assert wb.sheetnames == [
        "Oversikt",
        "#2",
        "Outlier - alle transaksjoner",
        "Data",
    ]

    # Oversikt: enkel navigasjon
    ws_ov = wb["Oversikt"]
    assert str(ws_ov["A1"].value).startswith("Motpostanalyse")
    assert "'Data'!A1" in str(ws_ov["E3"].value)
    assert "'Outlier - alle transaksjoner'!A1" in str(ws_ov["E4"].value)

    # Outlier-detaljark: skal ha handling/resultat + bilagsoppsummering (B)
    ws_detail = wb["#2"]
    assert any(ws_detail.cell(row=r, column=1).value == "Handling" for r in range(1, ws_detail.max_row + 1))
    assert any(ws_detail.cell(row=r, column=1).value == "Resultat" for r in range(1, ws_detail.max_row + 1))
    # Bilagsoppsummering (B) skal finnes (plasseringen kan variere litt)
    bilag_title_rows = [
        r
        for r in range(1, ws_detail.max_row + 1)
        if ws_detail.cell(row=r, column=1).value == "Bilagsoppsummering (B)"
    ]
    assert bilag_title_rows, "Fant ikke 'Bilagsoppsummering (B)' i detaljfane"

    # Skal kunne lagres/åpnes av openpyxl uten feil
    out_path = tmp_path / "motpost.xlsx"
    wb.save(out_path)
    wb2 = load_workbook(out_path)
    assert wb2.sheetnames == wb.sheetnames


def test_outlier_sheet_exists_even_when_no_outliers():
    df = _sample_df()
    data = build_motpost_data(df, {"3000"}, selected_direction="Kredit")

    wb = build_motpost_excel_workbook(data, combo_status_map={})

    # Ingen detaljark (#2/#4/..), men outlier-arket skal alltid finnes.
    assert wb.sheetnames == [
        "Oversikt",
        "Outlier - alle transaksjoner",
        "Data",
    ]
