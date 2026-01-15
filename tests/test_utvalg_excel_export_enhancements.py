from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import excel_formatting
from controller_export import export_to_excel


def _make_minimal_grunnlag_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Bilag": [100, 100, 101, 101, 101],
            "Konto": [3000, 1920, 3000, 1920, 2740],
            "Dato": [
                datetime(2025, 1, 1),
                datetime(2025, 1, 1),
                datetime(2025, 1, 2),
                datetime(2025, 1, 2),
                datetime(2025, 1, 2),
            ],
            "Beløp": [1000.0, -1000.0, 2000.0, -2000.0, 0.0],
            "Tekst": ["A", "A", "B", "B", "B"],
            "Pos": [1, 2, 1, 2, 3],
        }
    )


def _make_minimal_utvalg_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Bilag": [100, 101],
            "Dato": [datetime(2025, 1, 1), datetime(2025, 1, 2)],
            "Tekst": ["A", "B"],
            "SumBeløp": [1000.0, 2000.0],
            "Gruppe": ["1", "2"],
            "Intervall": ["0-1500", "1500-2500"],
        }
    )


def test_utvalg_export_generates_report_sheets_and_formats(tmp_path):
    grunnlag_df = _make_minimal_grunnlag_df()
    utvalg_df = _make_minimal_utvalg_df()

    outfile = tmp_path / "utvalg.xlsx"
    out_path = export_to_excel(
        str(outfile),
        Utvalg=utvalg_df,
        Grunnlag=grunnlag_df,
        # Assumptions as kwargs (how the GUI typically does it)
        Risiko="Middels",
        Sikkerhet="90%",
        **{
            "Tolererbar feil": 1_300_000,
            "Metode": "quantile",
            "Antall grupper (k)": 3,
            "Utvalgsstørrelse": 2,
            "Retning": "Debet",
            "Bruk absolutt beløp": "Ja",
        },
        # Keep deterministic for the test
        auto_filename=False,
        open_folder=False,
    )

    assert out_path == str(outfile)
    assert outfile.exists()

    wb = load_workbook(outfile)
    assert set(wb.sheetnames) == {
        "Oppsummering",
        "Forutsetninger",
        "Utvalg",
        "Bilagtransaksjoner",
        "Bilagsgrunnlag",
        "Grunnlag",
    }

    ws_utvalg = wb["Utvalg"]
    headers = [ws_utvalg.cell(row=1, column=c).value for c in range(1, ws_utvalg.max_column + 1)]
    assert headers[:2] == ["UtvalgNr", "Bilag"]

    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    assert ws_utvalg.cell(row=2, column=col_idx["UtvalgNr"]).number_format == excel_formatting.FMT_INT
    assert ws_utvalg.cell(row=2, column=col_idx["SumBeløp"]).number_format == excel_formatting.FMT_AMOUNT
    assert ws_utvalg.cell(row=2, column=col_idx["Dato"]).number_format == excel_formatting.FMT_DATE

    ws_sum = wb["Oppsummering"]
    # Find the rows we care about
    rows = {
        ws_sum.cell(row=r, column=1).value: r
        for r in range(2, ws_sum.max_row + 1)
        if ws_sum.cell(row=r, column=1).value
    }

    assert ws_sum.cell(row=rows["Sum beløp i grunnlag"], column=2).number_format == excel_formatting.FMT_AMOUNT
    assert ws_sum.cell(row=rows["Antall rader i grunnlag"], column=2).number_format == excel_formatting.FMT_INT
    assert ws_sum.cell(row=rows["Utvalgsandel (bilag)"], column=2).number_format == excel_formatting.FMT_PERCENT


def test_auto_filename_generates_generic_name_in_selected_directory(tmp_path):
    grunnlag_df = _make_minimal_grunnlag_df()
    utvalg_df = _make_minimal_utvalg_df()

    chosen = tmp_path / "user_selected_name.xlsx"
    out_path = export_to_excel(
        str(chosen),
        Utvalg=utvalg_df,
        Grunnlag=grunnlag_df,
        auto_filename=True,
        open_folder=False,
        filename_prefix="Eksport utvalg",
    )

    out = Path(out_path)
    assert out.exists()
    assert out != chosen

    # "Eksport utvalg dd.mm.yyyy HH.MM.xlsx" (no colon)
    assert re.match(r"^Eksport utvalg \d{2}\.\d{2}\.\d{4} \d{2}\.\d{2}\.xlsx$", out.name)
