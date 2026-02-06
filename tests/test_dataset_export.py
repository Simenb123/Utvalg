from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from dataset_export import build_hovedbok_excel_sheets, export_hovedbok_to_excel


def test_build_hovedbok_excel_sheets_happy_path_no_split() -> None:
    df = pd.DataFrame(
        {
            "Konto": [1000, 2000],
            "Bilag": ["A1", "A2"],
            "Beløp": [100.0, -50.0],
        }
    )
    sheets = build_hovedbok_excel_sheets(df, sheet_name="Hovedbok", max_rows_per_sheet=10)
    assert list(sheets.keys()) == ["Hovedbok"]
    assert len(sheets["Hovedbok"]) == 2


def test_build_hovedbok_excel_sheets_splits_when_over_limit() -> None:
    # Bruk lav grense for test (for å unngå store DataFrames)
    df = pd.DataFrame({"Konto": list(range(12)), "Beløp": list(range(12))})
    sheets = build_hovedbok_excel_sheets(df, sheet_name="Hovedbok", max_rows_per_sheet=5)

    assert list(sheets.keys()) == ["Hovedbok_1", "Hovedbok_2", "Hovedbok_3"]
    assert len(sheets["Hovedbok_1"]) == 5
    assert len(sheets["Hovedbok_2"]) == 5
    assert len(sheets["Hovedbok_3"]) == 2


def test_export_hovedbok_to_excel_writes_xlsx(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "Konto": [1500, 1550],
            "Kontonavn": ["Kundefordringer", "Kundefordringer konsern"],
            "Bilag": ["OCR#1", "OCR#2"],
            "Beløp": [-100.0, -200.5],
        }
    )

    out_file = tmp_path / "hovedbok.xlsx"
    out_path = export_hovedbok_to_excel(out_file, df, sheet_name="Hovedbok", max_rows_per_sheet=10)

    p = Path(out_path)
    assert p.exists()
    wb = load_workbook(p)
    assert "Hovedbok" in wb.sheetnames

    ws = wb["Hovedbok"]
    # Header-rad
    assert ws.cell(row=1, column=1).value == "Konto"
    assert ws.cell(row=1, column=2).value == "Kontonavn"
    assert ws.cell(row=1, column=3).value == "Bilag"
    assert ws.cell(row=1, column=4).value == "Beløp"

    # En data-rad
    assert ws.cell(row=2, column=1).value == 1500
    assert ws.cell(row=2, column=3).value == "OCR#1"


def test_export_hovedbok_to_excel_raises_on_empty_df(tmp_path: Path) -> None:
    df = pd.DataFrame()
    try:
        export_hovedbok_to_excel(tmp_path / "x.xlsx", df)
    except ValueError as e:
        assert "Ingen data" in str(e)
    else:
        raise AssertionError("Forventet ValueError for tomt datasett")
