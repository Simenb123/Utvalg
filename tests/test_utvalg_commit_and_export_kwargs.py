# tests/test_utvalg_commit_and_export_kwargs.py
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from controller_export import export_to_excel


def _make_minimal_utvalg_and_grunnlag() -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Lager et minimalt, men realistisk datasett:
    - Grunnlag: transaksjonslinjer (bilag kan ha flere linjer)
    - Utvalg: bilag-nivå (en rad pr bilag)
    """
    grunnlag = pd.DataFrame(
        [
            # Bilag 1001 (to linjer)
            {"Bilag": 1001, "Dato": "01.01.2025", "Konto": 6800, "Tekst": "Kostnad A", "Beløp": 1000.0},
            {"Bilag": 1001, "Dato": "01.01.2025", "Konto": 2400, "Tekst": "Motpost A", "Beløp": -1000.0},
            # Bilag 1002 (en linje)
            {"Bilag": 1002, "Dato": "02.01.2025", "Konto": 6800, "Tekst": "Kostnad B", "Beløp": 2000.0},
        ]
    )

    utvalg = pd.DataFrame(
        [
            # Vi velger bare bilag 1001 i utvalget
            {"Bilag": 1001, "Dato": "01.01.2025", "Tekst": "Kostnad A", "SumBeløp": 1000.0, "Gruppe": "1", "Intervall": "0-2000"},
        ]
    )
    return utvalg, grunnlag


def _read_sheet_headers(ws) -> list[str]:
    headers = []
    for cell in ws[1]:
        headers.append("" if cell.value is None else str(cell.value))
    return headers


def _sheet_col_index(headers: list[str], name: str) -> int:
    if name not in headers:
        raise AssertionError(f"Fant ikke kolonnen '{name}'. Headers var: {headers}")
    # openpyxl er 1-indeksert
    return headers.index(name) + 1


def test_export_to_excel_accepts_sheet_name_kwargs_and_adds_utvalg_report_sheets(tmp_path: Path) -> None:
    utvalg_df, grunnlag_df = _make_minimal_utvalg_and_grunnlag()
    out_path = tmp_path / "utvalg_export.xlsx"

    export_to_excel(
        str(out_path),
        Utvalg=utvalg_df,
        Grunnlag=grunnlag_df,
        open_folder=False,
        auto_filename=False,
    )

    wb = load_workbook(out_path)
    sheetnames = set(wb.sheetnames)

    # Utvalg-rapporten skal alltid inneholde disse arkene
    expected = {"Oppsummering", "Forutsetninger", "Utvalg", "Bilagtransaksjoner", "Grunnlag"}
    assert expected.issubset(sheetnames)

    # Forutsetninger har norsk header (Felt/Verdi)
    ws = wb["Forutsetninger"]
    headers = _read_sheet_headers(ws)
    assert set(headers[:2]) == {"Felt", "Verdi"}


def test_export_to_excel_utvalg_report_works_without_all_df(tmp_path: Path) -> None:
    """
    Tidligere forventet vi at Bilagtransaksjoner ikke ble laget uten all_df.
    Nå lager vi arket uansett (best-effort) og filtrerer fra det vi har (grunnlag).
    """
    utvalg_df, grunnlag_df = _make_minimal_utvalg_and_grunnlag()
    out_path = tmp_path / "utvalg_export_no_all_df.xlsx"

    export_to_excel(
        str(out_path),
        Utvalg=utvalg_df,
        Grunnlag=grunnlag_df,
        open_folder=False,
        auto_filename=False,
    )

    wb = load_workbook(out_path)
    assert "Bilagtransaksjoner" in wb.sheetnames

    ws = wb["Bilagtransaksjoner"]
    headers = _read_sheet_headers(ws)

    # Skal ha bilag og utvalgsnummer i transaksjonsarket
    bilag_col = _sheet_col_index(headers, "Bilag")
    utvalg_nr_col = _sheet_col_index(headers, "UtvalgNr")

    # Bilagtransaksjoner skal kun inneholde bilag som er i utvalget (her: 1001)
    bilag_values: set[int] = set()
    utvalg_nr_values: set[int] = set()

    for r in range(2, ws.max_row + 1):
        b = ws.cell(row=r, column=bilag_col).value
        u = ws.cell(row=r, column=utvalg_nr_col).value
        if b is not None:
            bilag_values.add(int(b))
        if u is not None:
            utvalg_nr_values.add(int(u))

    assert bilag_values == {1001}
    # UtvalgNr skal være 1 for dette ene utvalgsbilaget
    assert utvalg_nr_values == {1}
