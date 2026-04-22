"""Tester for StatistikkPage — missing-konto-fix, motpost-RL-toggle og kombinasjoner."""

from __future__ import annotations

from types import SimpleNamespace

import pandas as pd
import pytest


def _fake_page(
    *,
    intervals: pd.DataFrame | None = None,
    regnskapslinjer: pd.DataFrame | None = None,
    sb_df: pd.DataFrame | None = None,
    sb_prev_df: pd.DataFrame | None = None,
) -> object:
    page = SimpleNamespace()
    page._rl_intervals = intervals
    page._rl_regnskapslinjer = regnskapslinjer
    page._rl_sb_df = sb_df
    page._rl_sb_prev_df = sb_prev_df

    def _get_effective_sb_df() -> pd.DataFrame | None:
        return sb_df

    page._get_effective_sb_df = _get_effective_sb_df  # type: ignore[attr-defined]
    return page


def test_compute_kontoer_inkluderer_sb_kontoer_uten_bevegelse() -> None:
    """Regresjonstest: konti i SB innenfor RL-range skal vises selv uten tx."""
    from src.pages.statistikk import page_statistikk

    sb_df = pd.DataFrame(
        [
            {"konto": "3000", "kontonavn": "Salg", "ib": 0.0, "ub": -100.0},
            {"konto": "3010", "kontonavn": "Annet salg", "ib": 50.0, "ub": 50.0},
            {"konto": "3020", "kontonavn": "Tomt salg", "ib": 0.0, "ub": 0.0},
        ]
    )
    df_rl = pd.DataFrame(
        [{"Konto": "3000", "Kontonavn": "Salg", "Beløp": -100.0, "Bilag": "1"}]
    )
    page = _fake_page(sb_df=sb_df)
    grp, label = page_statistikk._compute_kontoer(df_rl, page, ranges=[(3000, 3099)])

    kontoer = set(grp["Konto"].astype(str))
    assert "3000" in kontoer
    assert "3010" in kontoer, "SB-only konto med IB må inkluderes"
    assert "3020" not in kontoer, "Helt tom konto (IB=UB=0) skal utelates"

    rad_3010 = grp[grp["Konto"] == "3010"].iloc[0]
    assert float(rad_3010["Bevegelse"]) == 0.0
    assert int(rad_3010["Antall"]) == 0
    assert str(rad_3010["Kontonavn"]) == "Annet salg"
    assert label == "IB"


def test_compute_kontoer_bruker_sb_prev_som_ib() -> None:
    """Når sb_prev finnes skal UB fjor brukes som IB-kolonnen."""
    from src.pages.statistikk import page_statistikk

    sb_df = pd.DataFrame([{"konto": "3000", "kontonavn": "Salg", "ib": 0.0, "ub": -100.0}])
    sb_prev = pd.DataFrame([{"konto": "3000", "kontonavn": "Salg", "ib": 0.0, "ub": -50.0}])
    df_rl = pd.DataFrame(
        [{"Konto": "3000", "Kontonavn": "Salg", "Beløp": -100.0, "Bilag": "1"}]
    )
    page = _fake_page(sb_df=sb_df, sb_prev_df=sb_prev)
    grp, label = page_statistikk._compute_kontoer(df_rl, page, ranges=[(3000, 3099)])

    assert label == "UB fjor"
    rad = grp[grp["Konto"] == "3000"].iloc[0]
    assert float(rad["IB"]) == -50.0


@pytest.mark.skipif(
    __import__("importlib").util.find_spec("tkinter") is None,
    reason="tkinter ikke tilgjengelig",
)
def test_build_motpost_rl_df_aggregerer_paa_regnr(monkeypatch) -> None:
    """_build_motpost_rl_df skal aggregere kontoer til regnr via RL-context."""
    from src.pages.statistikk import page_statistikk
    import regnskapslinje_mapping_service as rl_svc

    grp = pd.DataFrame(
        [
            {"Konto": "2400", "Kontonavn": "Kundefordr", "Beløp": -500.0, "Andel": 50.0, "AntallBilag": 3},
            {"Konto": "2401", "Kontonavn": "Kundefordr 2", "Beløp": -200.0, "Andel": 20.0, "AntallBilag": 1},
            {"Konto": "2710", "Kontonavn": "Mva", "Beløp": -100.0, "Andel": 10.0, "AntallBilag": 1},
        ]
    )

    def fake_context_from_page(page):  # type: ignore[no-untyped-def]
        return None  # ikke brukt av fake resolve

    def fake_resolve(accounts, *, context):  # type: ignore[no-untyped-def]
        rows = []
        map_ = {"2400": (1500, "Kundefordringer"), "2401": (1500, "Kundefordringer"),
                "2710": (2700, "Mva")}
        for k in accounts:
            regnr, navn = map_.get(str(k), (None, ""))
            rows.append(
                {
                    "konto": str(k),
                    "regnr": regnr if regnr is not None else pd.NA,
                    "regnskapslinje": navn,
                    "mapping_status": "interval" if regnr else "unmapped",
                    "source": "interval" if regnr else "",
                }
            )
        return pd.DataFrame(rows)

    monkeypatch.setattr(rl_svc, "context_from_page", fake_context_from_page)
    monkeypatch.setattr(rl_svc, "resolve_accounts_to_rl", fake_resolve)

    page = page_statistikk.StatistikkPage(None)
    page._analyse_page = object()

    result = page._build_motpost_rl_df(grp)

    assert set(result.columns) == {"Regnr", "Regnskapslinje", "Beløp", "Andel", "AntallBilag"}
    assert len(result) == 2

    kundef = result[result["Regnr"] == "1500"].iloc[0]
    assert float(kundef["Beløp"]) == -700.0
    assert int(kundef["AntallBilag"]) == 4
    assert str(kundef["Regnskapslinje"]) == "Kundefordringer"

    mva = result[result["Regnr"] == "2700"].iloc[0]
    assert float(mva["Beløp"]) == -100.0
    # Andel = |beløp| / totalt|beløp|  = 100/800 = 12.5
    assert float(mva["Andel"]) == pytest.approx(12.5, abs=0.1)


@pytest.mark.skipif(
    __import__("importlib").util.find_spec("tkinter") is None,
    reason="tkinter ikke tilgjengelig",
)
def test_compute_kombinasjoner_bygger_kombinasjoner() -> None:
    """Smoke-test at _compute_kombinasjoner returnerer forventede rader + bilag-map."""
    from src.pages.statistikk import page_statistikk

    df_all = pd.DataFrame(
        [
            # Bilag 1: RL-konto 3000 mot 2400
            {"Bilag": "1", "Konto": "3000", "Beløp": -100.0},
            {"Bilag": "1", "Konto": "2400", "Beløp": 100.0},
            # Bilag 2: RL-konto 3000 mot 2400 (samme kombinasjon)
            {"Bilag": "2", "Konto": "3000", "Beløp": -200.0},
            {"Bilag": "2", "Konto": "2400", "Beløp": 200.0},
            # Bilag 3: RL-konto 3000 mot 2710 (ny kombinasjon)
            {"Bilag": "3", "Konto": "3000", "Beløp": -50.0},
            {"Bilag": "3", "Konto": "2710", "Beløp": 50.0},
        ]
    )
    page = page_statistikk.StatistikkPage(None)
    combos, bilag_map = page._compute_kombinasjoner(df_all, {"3000"})

    assert not combos.empty
    assert "Kombinasjon #" in combos.columns
    assert "Kombinasjon" in combos.columns
    assert len(combos) == 2

    # "2400"-kombinasjonen skal ha flest bilag (2)
    top = combos.iloc[0]
    assert top["Kombinasjon"] == "2400"
    assert int(top["Antall bilag"]) == 2

    # bilag_map mapper bilag til kombinasjon-streng
    assert bilag_map.get("1") == "2400"
    assert bilag_map.get("3") == "2710"


def test_compute_kombinasjoner_tomt_input_returnerer_tomt() -> None:
    """Robusthet: ingen RL-kontoer → tom DataFrame og tom bilag-map."""
    from src.pages.statistikk import page_statistikk

    page = page_statistikk.StatistikkPage(None)
    combos, bilag_map = page._compute_kombinasjoner(pd.DataFrame(), set())
    assert combos.empty
    assert bilag_map == {}


def test_get_konto_set_respekterer_override(monkeypatch) -> None:
    """_get_konto_set_for_regnr må droppe konti som override flytter ut av regnr."""
    from src.pages.statistikk import page_statistikk
    import regnskapslinje_mapping_service as rl_svc

    intervals = pd.DataFrame([{"regnr": 15, "fra": 3600, "til": 3999}])
    regnskapslinjer = pd.DataFrame(
        [{"regnr": 15, "regnskapslinje": "Annen driftsinntekt", "sumpost": False}]
    )
    df_all = pd.DataFrame(
        [
            {"Konto": "3600", "Beløp": -100.0},
            {"Konto": "3620", "Beløp": -50.0},  # flyttet via override
            {"Konto": "3900", "Beløp": 10.0},   # flyttet via override
        ]
    )

    page = SimpleNamespace(
        _rl_intervals=intervals,
        _rl_regnskapslinjer=regnskapslinjer,
        _rl_sb_df=None,
        _rl_sb_prev_df=None,
    )
    page._get_effective_sb_df = lambda: None  # type: ignore[attr-defined]

    def fake_context(_page):  # type: ignore[no-untyped-def]
        return None

    def fake_resolve(accounts, *, context):  # type: ignore[no-untyped-def]
        # 3600 blir i regnr 15; 3620 og 3900 flyttes via override til andre regnr.
        override = {"3620": 17, "3900": 19}
        rows = []
        for k in accounts:
            key = str(k)
            if key in override:
                rows.append({"konto": key, "regnr": override[key]})
            elif 3600 <= int(key) <= 3999:
                rows.append({"konto": key, "regnr": 15})
            else:
                rows.append({"konto": key, "regnr": pd.NA})
        return pd.DataFrame(rows)

    monkeypatch.setattr(rl_svc, "context_from_page", fake_context)
    monkeypatch.setattr(rl_svc, "resolve_accounts_to_rl", fake_resolve)

    konto_set = page_statistikk._get_konto_set_for_regnr(
        page, 15, [(3600, 3999)],
        df_all=df_all, sb_df=None, sb_prev_df=None,
    )
    assert konto_set == {"3600"}, f"forventet kun 3600, fikk {konto_set}"


def test_compute_kombinasjoner_export_wrapper() -> None:
    """Eksport-wrapperen _compute_kombinasjoner_export skal bygge samme tabell."""
    from src.pages.statistikk import page_statistikk

    df_all = pd.DataFrame(
        [
            {"Bilag": "1", "Konto": "3000", "Beløp": -100.0},
            {"Bilag": "1", "Konto": "2400", "Beløp": 100.0},
            {"Bilag": "2", "Konto": "3000", "Beløp": -200.0},
            {"Bilag": "2", "Konto": "2400", "Beløp": 200.0},
        ]
    )
    df_rl = df_all[df_all["Konto"] == "3000"]
    combos = page_statistikk._compute_kombinasjoner_export(df_all, df_rl)
    assert not combos.empty
    assert "Kombinasjon #" in combos.columns
    assert int(combos.iloc[0]["Antall bilag"]) == 2


def test_write_workbook_har_sum_rader_og_nye_ark(tmp_path, monkeypatch) -> None:
    """Excel-eksporten skal inneholde summeringslinjer og begge nye seksjoner."""
    from src.pages.statistikk import page_statistikk
    import regnskapslinje_mapping_service as rl_svc
    from openpyxl import load_workbook

    df_all = pd.DataFrame(
        [
            {"Dato": "01.01.2024", "Bilag": "1", "Konto": "3000", "Kontonavn": "Salg",
             "Tekst": "Test", "Beløp": -100.0, "MVA-kode": ""},
            {"Dato": "01.01.2024", "Bilag": "1", "Konto": "2400", "Kontonavn": "Kundefordr",
             "Tekst": "Test", "Beløp": 100.0, "MVA-kode": ""},
            {"Dato": "02.01.2024", "Bilag": "2", "Konto": "3000", "Kontonavn": "Salg",
             "Tekst": "Test 2", "Beløp": -200.0, "MVA-kode": ""},
            {"Dato": "02.01.2024", "Bilag": "2", "Konto": "2710", "Kontonavn": "Mva",
             "Tekst": "Test 2", "Beløp": 200.0, "MVA-kode": ""},
        ]
    )
    df_rl = df_all[df_all["Konto"] == "3000"].copy()

    def fake_context_from_page(page):  # type: ignore[no-untyped-def]
        return None

    def fake_resolve(accounts, *, context):  # type: ignore[no-untyped-def]
        m = {"2400": (1500, "Kundefordringer"), "2710": (2700, "Mva")}
        rows = []
        for k in accounts:
            regnr, navn = m.get(str(k), (None, ""))
            rows.append(
                {
                    "konto": str(k),
                    "regnr": regnr if regnr is not None else pd.NA,
                    "regnskapslinje": navn,
                    "mapping_status": "interval" if regnr else "unmapped",
                    "source": "interval" if regnr else "",
                }
            )
        return pd.DataFrame(rows)

    monkeypatch.setattr(rl_svc, "context_from_page", fake_context_from_page)
    monkeypatch.setattr(rl_svc, "resolve_accounts_to_rl", fake_resolve)

    page = SimpleNamespace(
        _pivot_df_last=pd.DataFrame(),
        _rl_intervals=None,
        _rl_regnskapslinjer=None,
        _rl_sb_df=None,
        _rl_sb_prev_df=None,
    )
    page._get_effective_sb_df = lambda: None  # type: ignore[attr-defined]

    out = tmp_path / "statistikk.xlsx"
    page_statistikk._write_workbook(
        str(out), regnr=3000, rl_name="Salg",
        df_rl=df_rl, df_all=df_all, page=page, client="Test", year="2024",
    )
    assert out.exists()

    wb = load_workbook(out)
    assert "Motpostfordeling" in wb.sheetnames
    assert "Kombinasjoner" in wb.sheetnames

    ws_mp = wb["Motpostfordeling"]
    # Sjekk at det finnes en "Sum"-celle et sted (konto-seksjon)
    sum_cells = [c.value for c in ws_mp["A"] if isinstance(c.value, str) and c.value.startswith("Sum")]
    assert sum_cells, "Motpost-arket skal ha en sum-rad"
    # Sjekk RL-seksjon tilstede
    rl_hdr = [c.value for c in ws_mp["A"] if c.value == "Motpost pr regnskapslinje"]
    assert rl_hdr, "RL-seksjon må stå på Motpostfordeling-arket"

    ws_k = wb["Kombinasjoner"]
    sum_cells_k = [c.value for c in ws_k["A"] if isinstance(c.value, str) and c.value.startswith("Sum")]
    assert sum_cells_k, "Kombinasjoner-arket skal ha en sum-rad"
