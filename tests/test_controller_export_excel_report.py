# -*- coding: utf-8 -*-
import openpyxl
import pandas as pd
import pytest

from controller_export import export_to_excel


def _sheet_kv_to_dict(wb: openpyxl.Workbook, sheet_name: str) -> dict:
    ws = wb[sheet_name]
    data = {}
    # Expect headers in row 1: Felt, Verdi
    for r in range(2, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k is None:
            continue
        data[str(k)] = v
    return data


def test_excel_export_includes_specific_rest_random_breakdown(tmp_path):
    # Grunnlag (transaksjoner)
    df_grunnlag = pd.DataFrame(
        [
            # Bilag 1: net 40
            {"Bilag": 1, "Konto": 6000, "Beløp": 50.0, "Dato": "2025-01-01", "Tekst": "Bilag 1"},
            {"Bilag": 1, "Konto": 6000, "Beløp": -10.0, "Dato": "2025-01-01", "Tekst": "Bilag 1"},
            # Bilag 2: net 200 (spesifikk)
            {"Bilag": 2, "Konto": 3000, "Beløp": 200.0, "Dato": "2025-01-02", "Tekst": "Bilag 2"},
            # Bilag 3: net -150 (spesifikk)
            {"Bilag": 3, "Konto": 3000, "Beløp": -150.0, "Dato": "2025-01-03", "Tekst": "Bilag 3"},
            # Bilag 4: net 30 (rest)
            {"Bilag": 4, "Konto": 6000, "Beløp": 30.0, "Dato": "2025-01-04", "Tekst": "Bilag 4"},
        ]
    )
    df_grunnlag["Dato"] = pd.to_datetime(df_grunnlag["Dato"])

    # Utvalg (bilag-nivå)
    df_utvalg = pd.DataFrame(
        [
            {"UtvalgNr": 1, "Bilag": 2, "Dato": pd.Timestamp("2025-01-02"), "Tekst": "Bilag 2", "SumBeløp": 200.0, "Gruppe": "Spesifikk", "Intervall": ">= 100"},
            {"UtvalgNr": 2, "Bilag": 3, "Dato": pd.Timestamp("2025-01-03"), "Tekst": "Bilag 3", "SumBeløp": -150.0, "Gruppe": "Spesifikk", "Intervall": ">= 100"},
            {"UtvalgNr": 3, "Bilag": 1, "Dato": pd.Timestamp("2025-01-01"), "Tekst": "Bilag 1", "SumBeløp": 40.0, "Gruppe": "1", "Intervall": "0,0 – 99,9"},
        ]
    )

    out_path = tmp_path / "utvalg.xlsx"
    export_to_excel(str(out_path), Utvalg=df_utvalg, Grunnlag=df_grunnlag)

    wb = openpyxl.load_workbook(out_path, data_only=True)

    # Ark finnes
    assert "Oppsummering" in wb.sheetnames
    assert "Forutsetninger" in wb.sheetnames
    assert "Utvalg" in wb.sheetnames
    assert "Bilagtransaksjoner" in wb.sheetnames
    assert "Bilagsgrunnlag" in wb.sheetnames
    assert "Grunnlag" in wb.sheetnames

    opp = _sheet_kv_to_dict(wb, "Oppsummering")
    forut = _sheet_kv_to_dict(wb, "Forutsetninger")

    # Populasjon
    assert opp["Antall bilag i grunnlag"] == 4

    # Spesifikk
    assert opp["Antall bilag spesifikk utvelgelse"] == 2

    # Restpopulasjon etter spesifikk: bilag 1 og 4 => 2
    assert opp["Antall bilag restpopulasjon (etter spesifikk)"] == 2

    # |Netto rest|: 40 + 30 = 70
    assert opp["|Netto restpopulasjon| (basis)"] == pytest.approx(70.0)

    # Tilfeldig: bilag 1 => 1
    assert opp["Antall bilag tilfeldig utvalg"] == 1

    # Total utvalg: 3 bilag
    assert opp["Antall bilag i utvalg (totalt)"] == 3

    # TE fra utvalg
    assert opp["Tolererbar feil (fra utvalg)"] == pytest.approx(100.0)
    assert forut["Tolererbar feil"] == pytest.approx(100.0)

    # Bilagsgrunnlag skal markere valgte bilag
    ws_bg = wb["Bilagsgrunnlag"]
    headers = [ws_bg.cell(row=1, column=c).value for c in range(1, ws_bg.max_column + 1)]
    assert "IUtvalg" in headers
    assert "UtvalgType" in headers

    # Sjekk at bilag 2 er merket som Spesifikk
    bilag_col = headers.index("Bilag") + 1
    iutvalg_col = headers.index("IUtvalg") + 1
    type_col = headers.index("UtvalgType") + 1

    found_b2 = False
    for r in range(2, ws_bg.max_row + 1):
        bilag_val = ws_bg.cell(row=r, column=bilag_col).value
        if bilag_val == 2:
            found_b2 = True
            assert ws_bg.cell(row=r, column=iutvalg_col).value == "Ja"
            assert ws_bg.cell(row=r, column=type_col).value == "Spesifikk"
            break
    assert found_b2
