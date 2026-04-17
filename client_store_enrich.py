# -*- coding: utf-8 -*-
"""client_store_enrich.py – berik klientdata med orgnr/knr fra Visena XLSX.

3-trinns matching:
  1. Eksakt Knr-match (Utvalg display_name har Knr-prefiks)
  2. Eksakt normalisert navnematch
  3. Fuzzy navnematch (SequenceMatcher >= 0.85)
"""

from __future__ import annotations

import difflib
import re
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, List, Optional, Sequence

import client_store

# ---------------------------------------------------------------------------
# Dataklasser
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class VisenaRow:
    firma: str
    org_number: str
    client_number: str
    project: str
    responsible: str
    manager: str = ""          # "Har som manager"
    team_members: str = ""     # "Har som prosjektmedlem"


@dataclass(frozen=True)
class EnrichmentMatch:
    display_name: str
    visena_row: VisenaRow
    match_type: str        # "exact_knr" | "exact_name" | "fuzzy_name"
    match_score: float     # 1.0 for eksakt, 0.85–1.0 for fuzzy


@dataclass(frozen=True)
class EnrichmentPlan:
    matched: List[EnrichmentMatch]
    unmatched_utvalg: List[str]
    unmatched_visena: List[VisenaRow]
    already_enriched: List[str]


# ---------------------------------------------------------------------------
# Normalisering (gjenbruker mønster fra crmsystem_materiality)
# ---------------------------------------------------------------------------

_CLIENT_PREFIX_RE = re.compile(r"^\s*\d{2,12}\s+")
_NONWORD_RE = re.compile(r"[^0-9a-zæøå]+", re.IGNORECASE)
_COMPANY_SUFFIXES = {"as", "asa", "ans", "da", "sa", "ba", "ks", "nuf", "enk"}
_KNR_PREFIX_RE = re.compile(r"^\s*(\d{2,12})\s+")


def _normalize_name(value: str) -> str:
    """Normaliser klientnavn for sammenligning."""
    raw = _CLIENT_PREFIX_RE.sub("", (value or "").strip())
    if not raw:
        return ""
    raw = raw.casefold().replace("&", " og ")
    raw = _NONWORD_RE.sub(" ", raw)
    tokens = [t for t in raw.split() if t]
    while tokens and tokens[-1] in _COMPANY_SUFFIXES:
        tokens.pop()
    return " ".join(tokens)


def _extract_knr(display_name: str) -> Optional[str]:
    """Hent klientnummer-prefiks fra display_name (f.eks. '144187 Spor AS' -> '144187')."""
    m = _KNR_PREFIX_RE.match(display_name or "")
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# Visena XLSX-leser
# ---------------------------------------------------------------------------

def read_enrichment_data_xlsx(path: Path) -> List[VisenaRow]:
    """Les Visena-prosessliste og returner VisenaRow-liste."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # Finn header
        header = None
        for _ in range(20):
            header = next(rows_iter, None)
            if header is None:
                return []
            if any(v is not None and str(v).strip() for v in header):
                break
        if header is None:
            return []

        headers_l = [(str(h).strip().lower() if h else "") for h in header]

        # Finn kolonneindekser
        col_map = {}
        for i, h in enumerate(headers_l):
            if h == "firma":
                col_map["firma"] = i
            elif h == "org.nr":
                col_map["orgnr"] = i
            elif h in {"knr.", "knr", "klientnr", "klientnr."}:
                col_map["knr"] = i
            elif h == "prosjekt":
                col_map["project"] = i
            elif h == "prosjektansvarlig":
                col_map["responsible"] = i
            elif h == "har som manager":
                col_map["manager"] = i
            elif h == "har som prosjektmedlem":
                col_map["team_members"] = i

        if "firma" not in col_map:
            return []

        def _cell(row, key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        def _clean_digits(val: str) -> str:
            return "".join(ch for ch in val if ch.isdigit())

        result: List[VisenaRow] = []
        for row in rows_iter:
            if row is None:
                continue
            firma = _cell(row, "firma")
            if not firma:
                continue
            result.append(VisenaRow(
                firma=firma,
                org_number=_clean_digits(_cell(row, "orgnr")),
                client_number=_clean_digits(_cell(row, "knr")),
                project=_cell(row, "project"),
                manager=_cell(row, "manager"),
                team_members=_cell(row, "team_members"),
                responsible=_cell(row, "responsible"),
            ))
        return result
    finally:
        try:
            wb.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Matching-motor
# ---------------------------------------------------------------------------

_FUZZY_THRESHOLD = 0.85


def plan_enrichment(
    visena_rows: Sequence[VisenaRow],
    existing_clients: Sequence[str],
) -> EnrichmentPlan:
    """Match Visena-rader mot eksisterende klienter.

    3-trinns prioritert matching:
      1. Eksakt Knr-match (display_name har Knr-prefiks)
      2. Eksakt normalisert navnematch
      3. Fuzzy navnematch (>= 0.85)
    """

    # Bygg oppslagsmaps for Utvalg-klienter
    knr_to_client: dict[str, str] = {}       # knr -> display_name
    norm_to_client: dict[str, str] = {}      # normalized name -> display_name

    for dn in existing_clients:
        knr = _extract_knr(dn)
        if knr:
            knr_to_client.setdefault(knr, dn)
        norm = _normalize_name(dn)
        if norm:
            norm_to_client.setdefault(norm, dn)

    # Sjekk allerede berikede klienter
    already_enriched: list[str] = []
    enrichable_clients: set[str] = set()
    for dn in existing_clients:
        meta = client_store.read_client_meta(dn)
        if meta.get("org_number"):
            already_enriched.append(dn)
        else:
            enrichable_clients.add(dn)

    matched: list[EnrichmentMatch] = []
    matched_utvalg: set[str] = set()
    matched_visena: set[int] = set()       # index i visena_rows

    # --- Pass 1: Eksakt Knr-match ---
    visena_by_knr: dict[str, tuple[int, VisenaRow]] = {}
    for i, vr in enumerate(visena_rows):
        if vr.client_number:
            visena_by_knr.setdefault(vr.client_number, (i, vr))

    for knr, dn in knr_to_client.items():
        if dn not in enrichable_clients:
            continue
        hit = visena_by_knr.get(knr)
        if hit is None:
            continue
        idx, vr = hit
        matched.append(EnrichmentMatch(
            display_name=dn,
            visena_row=vr,
            match_type="exact_knr",
            match_score=1.0,
        ))
        matched_utvalg.add(dn)
        matched_visena.add(idx)

    # --- Pass 2: Eksakt navnematch ---
    visena_by_norm: dict[str, tuple[int, VisenaRow]] = {}
    for i, vr in enumerate(visena_rows):
        if i in matched_visena:
            continue
        norm = _normalize_name(vr.firma)
        if norm:
            visena_by_norm.setdefault(norm, (i, vr))

    for norm, dn in norm_to_client.items():
        if dn in matched_utvalg or dn not in enrichable_clients:
            continue
        hit = visena_by_norm.get(norm)
        if hit is None:
            continue
        idx, vr = hit
        matched.append(EnrichmentMatch(
            display_name=dn,
            visena_row=vr,
            match_type="exact_name",
            match_score=1.0,
        ))
        matched_utvalg.add(dn)
        matched_visena.add(idx)

    # --- Pass 3: Fuzzy navnematch ---
    remaining_visena = [
        (i, vr, _normalize_name(vr.firma))
        for i, vr in enumerate(visena_rows)
        if i not in matched_visena and _normalize_name(vr.firma)
    ]
    remaining_utvalg = [
        (dn, _normalize_name(dn))
        for dn in enrichable_clients
        if dn not in matched_utvalg and _normalize_name(dn)
    ]

    for dn, dn_norm in remaining_utvalg:
        best_score = 0.0
        best_hit: Optional[tuple[int, VisenaRow]] = None
        for idx, vr, vr_norm in remaining_visena:
            if idx in matched_visena:
                continue
            score = difflib.SequenceMatcher(None, dn_norm, vr_norm).ratio()
            if score >= _FUZZY_THRESHOLD and score > best_score:
                best_score = score
                best_hit = (idx, vr)

        if best_hit is not None:
            idx, vr = best_hit
            matched.append(EnrichmentMatch(
                display_name=dn,
                visena_row=vr,
                match_type="fuzzy_name",
                match_score=round(best_score, 3),
            ))
            matched_utvalg.add(dn)
            matched_visena.add(idx)

    # Umulige
    unmatched_utvalg = [
        dn for dn in enrichable_clients if dn not in matched_utvalg
    ]
    unmatched_visena = [
        vr for i, vr in enumerate(visena_rows) if i not in matched_visena
    ]

    return EnrichmentPlan(
        matched=matched,
        unmatched_utvalg=sorted(unmatched_utvalg),
        unmatched_visena=unmatched_visena,
        already_enriched=sorted(already_enriched),
    )


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------

_PROGRESS_CB = Callable[[int, int, str], None]


def apply_enrichment(
    matches: Sequence[EnrichmentMatch],
    *,
    progress_cb: Optional[_PROGRESS_CB] = None,
    cancel_event: Optional[threading.Event] = None,
) -> dict:
    """Skriv org_number og client_number til meta.json for matchede klienter.

    Returnerer statistikk-dict.
    """
    enriched = 0
    skipped = 0
    total = len(matches)

    for i, m in enumerate(matches):
        if cancel_event and cancel_event.is_set():
            return {"enriched": enriched, "skipped": skipped, "cancelled": True}

        updates: dict = {}
        if m.visena_row.org_number:
            updates["org_number"] = m.visena_row.org_number
        if m.visena_row.client_number:
            updates["client_number"] = m.visena_row.client_number
        if m.visena_row.project:
            updates["visena_project"] = m.visena_row.project
        if m.visena_row.responsible:
            updates["visena_responsible"] = m.visena_row.responsible
        if m.visena_row.manager:
            updates["visena_manager"] = m.visena_row.manager
        if m.visena_row.team_members:
            updates["visena_team_members"] = m.visena_row.team_members
        updates["enrichment_match_type"] = m.match_type
        updates["enrichment_match_score"] = m.match_score

        if updates:
            ok = client_store.update_client_meta(m.display_name, updates)
            if ok:
                enriched += 1
            else:
                skipped += 1
        else:
            skipped += 1

        if progress_cb:
            progress_cb(i + 1, total, m.display_name)

    # Bygg lokal teamindeks for rask filtrering (ingen nettverks-I/O)
    try:
        _save_team_index(matches)
    except Exception:
        pass

    # Oppdater lokal metadata-indeks
    try:
        import client_meta_index
        for m in matches:
            vr = m.visena_row
            client_meta_index.update_entry(m.display_name, {
                "org_number": vr.org_number,
                "client_number": vr.client_number,
                "responsible": vr.responsible,
                "manager": vr.manager,
                "team_members": vr.team_members,
            })
    except Exception:
        pass

    return {"enriched": enriched, "skipped": skipped, "cancelled": False}


# ---------------------------------------------------------------------------
# Teamfiltrering
# ---------------------------------------------------------------------------

def is_my_client(meta: dict, visena_initials: str, full_name: str) -> bool:
    """Sjekk om en klient tilhører brukeren basert på Visena-data i meta.json.

    Matcher på:
    - visena_responsible == initials
    - visena_manager inneholder full_name
    - visena_team_members inneholder full_name
    """
    if not visena_initials and not full_name:
        return False

    initials_lower = visena_initials.lower()
    name_lower = full_name.lower()

    # Prosjektansvarlig (initialer)
    if initials_lower and (meta.get("visena_responsible") or "").lower() == initials_lower:
        return True

    # Manager (fullt navn)
    if name_lower:
        manager = (meta.get("visena_manager") or "").lower()
        if name_lower in manager:
            return True

        # Teammedlem (fullt navn, kan være flere med newline/komma)
        members = (meta.get("visena_team_members") or "").lower()
        if name_lower in members:
            return True

    return False


_TEAM_INDEX_FILE = "team_index.json"


def _team_index_path() -> Path:
    """Lokal teamindeks — lagres i app data, ikke på nettverksdisk."""
    import app_paths
    return app_paths.data_dir() / _TEAM_INDEX_FILE


def _save_team_index(matches: Sequence[EnrichmentMatch]) -> None:
    """Skriv lokal teamindeks under berikelse — unngår å lese meta.json fra nett."""
    import json
    index: dict[str, dict] = {}

    # Les eksisterende indeks
    idx_path = _team_index_path()
    try:
        if idx_path.exists():
            index = json.loads(idx_path.read_text(encoding="utf-8"))
    except Exception:
        index = {}

    for m in matches:
        vr = m.visena_row
        index[m.display_name] = {
            "responsible": vr.responsible,
            "manager": vr.manager,
            "team_members": vr.team_members,
        }

    try:
        idx_path.parent.mkdir(parents=True, exist_ok=True)
        idx_path.write_text(json.dumps(index, ensure_ascii=False, indent=1), encoding="utf-8")
    except Exception:
        pass


def _load_team_index() -> dict[str, dict]:
    """Les lokal teamindeks (instant, ingen nettverks-I/O)."""
    import json
    idx_path = _team_index_path()
    try:
        if idx_path.exists():
            return json.loads(idx_path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}


def _rebuild_team_index_from_meta(all_clients: list[str]) -> dict[str, dict]:
    """Bygg teamindeks fra meta.json (treg, men kun ved første gangs bruk)."""
    import json
    index: dict[str, dict] = {}
    for dn in all_clients:
        meta = client_store.read_client_meta(dn)
        resp = meta.get("visena_responsible", "")
        mgr = meta.get("visena_manager", "")
        team = meta.get("visena_team_members", "")
        if resp or mgr or team:
            index[dn] = {"responsible": resp, "manager": mgr, "team_members": team}

    # Lagre lokalt for neste gang
    try:
        idx_path = _team_index_path()
        idx_path.parent.mkdir(parents=True, exist_ok=True)
        idx_path.write_text(json.dumps(index, ensure_ascii=False, indent=1), encoding="utf-8")
    except Exception:
        pass

    return index


def get_my_clients(all_clients: list[str]) -> list[str]:
    """Filtrer klientliste til kun 'mine klienter' basert på lokal teamindeks.

    Bruker team_index.json (lokal fil) — ingen nettverks-I/O.
    Hvis indeksen ikke finnes, bygges den fra meta.json (tregt, men kun én gang).
    """
    try:
        import team_config
        user = team_config.current_user()
    except Exception:
        return all_clients

    if user is None:
        return all_clients

    index = _load_team_index()
    if not index:
        # Første gang: bygg fra meta.json (kan ta tid over nettverk)
        index = _rebuild_team_index_from_meta(all_clients)

    if not index:
        return all_clients

    result = []
    for dn in all_clients:
        entry = index.get(dn)
        if entry and is_my_client(entry, user.visena_initials, user.full_name):
            result.append(dn)
    return result
