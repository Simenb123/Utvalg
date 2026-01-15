"""excel_formatting.py

Small helpers for making exported Excel sheets more readable.

The goal is to keep the formatting rules *simple and predictable*:

* Header row is bold, frozen, and has an autofilter.
* Amount columns ("Beløp", "SumBeløp", "Valutabeløp", ...) get a 2-decimal
  thousands-separated format.
* ID / counter columns ("Bilag", "Konto", "UtvalgNr", ...) are integers (no decimals).
* Date columns ("Dato", "Date") are formatted as dd.mm.yyyy.
* Key/value summary sheets ("Felt"/"Verdi") get per-row formatting so that
  counts look like counts, sums look like currency amounts, and shares look like %.

This module is intentionally dependency-light and does not try to be a full Excel
styling system; it just applies the most helpful readability tweaks.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Public formats (used in tests)
FMT_INT = "0"
FMT_AMOUNT = "#,##0.00"
FMT_AMOUNT_0 = "#,##0"
FMT_PERCENT = "0.00%"
FMT_DATE = "dd.mm.yyyy"
FMT_NUMBER_2 = "0.00"

# Header heuristics (normalized headers)
_INT_HEADERS = {
    "utvalgnr",
    "bilag",
    "konto",
    "kundenr",
    "leverandornr",
    "leverandørnr",
    "pos",
    "antall",
}
_AMOUNT_KEYWORDS = (
    "belop",
    "beløp",
    "sumbelop",
    "sumbeløp",
    "amount",
    "sumamount",
    "valutabelop",
    "valutabeløp",
    "mva-belop",
    "mva-beløp",
)
_DATE_KEYWORDS = ("dato", "date")

# For key/value summary sheets
_KV_COUNT_KEYWORDS = ("antall", "grupper", "(k)", "utvalgsstorrelse", "utvalgsstørrelse")
_KV_AMOUNT_KEYWORDS = ("sum", "belop", "beløp", "feil")
_KV_PERCENT_KEYWORDS = ("andel", "%")
_KV_AMOUNT_0_KEYWORDS = ("tolererbar feil",)


def polish_sheet(ws: Worksheet) -> None:
    """Apply basic formatting to an openpyxl worksheet (in-place)."""
    if ws.max_row < 1 or ws.max_column < 1:
        return

    _bold_headers(ws)
    _freeze_panes(ws)
    _set_auto_filter(ws)
    _format_sheet(ws)
    _autosize_columns(ws)


def _bold_headers(ws: Worksheet) -> None:
    """Make the header row bold."""
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _freeze_panes(ws: Worksheet) -> None:
    """Freeze first row."""
    ws.freeze_panes = "A2"


def _set_auto_filter(ws: Worksheet) -> None:
    """Enable autofilter across the used range."""
    if ws.max_row >= 1 and ws.max_column >= 1:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def _format_sheet(ws: Worksheet) -> None:
    """Decide formatting strategy based on sheet shape."""
    if _is_key_value_sheet(ws):
        _format_key_value_sheet(ws)
    else:
        _apply_number_formats_by_header(ws)


def _is_key_value_sheet(ws: Worksheet) -> bool:
    """Detect a simple 2-column key/value sheet (e.g. Felt/Verdi)."""
    if ws.max_column != 2 or ws.max_row < 2:
        return False
    h1 = _normalize_header(ws.cell(row=1, column=1).value)
    h2 = _normalize_header(ws.cell(row=1, column=2).value)
    return (h1 in {"felt", "parameter"} and h2 in {"verdi", "value"}) or (
        h1 in {"key"} and h2 in {"value"}
    )


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    # Normalize Norwegian chars to improve keyword matching
    text = text.replace("ø", "o").replace("å", "a").replace("æ", "ae")
    return text


def _apply_number_formats_by_header(ws: Worksheet) -> None:
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    for col_idx, header in enumerate(headers, start=1):
        h = _normalize_header(header)
        if not h:
            continue

        fmt: str | None = None
        align: Alignment | None = None

        # Dates
        if any(k in h for k in _DATE_KEYWORDS):
            fmt = FMT_DATE
            align = Alignment(horizontal="left")

        # Integers (IDs/counters)
        elif h in _INT_HEADERS:
            fmt = FMT_INT
            align = Alignment(horizontal="right")

        # Amounts
        elif any(k in h for k in _AMOUNT_KEYWORDS):
            fmt = FMT_AMOUNT
            align = Alignment(horizontal="right")

        # Tax percentage column is usually "25.00", not "25%" in underlying data
        elif "prosent" in h or "percent" in h:
            fmt = FMT_NUMBER_2
            align = Alignment(horizontal="right")

        # Shares
        elif "andel" in h or "share" in h:
            # Only treat as % if data looks like 0..1
            sample_vals = _sample_numeric_values(ws, col_idx, limit=50)
            if sample_vals and max(abs(v) for v in sample_vals) <= 1.0:
                fmt = FMT_PERCENT
            else:
                fmt = FMT_NUMBER_2
            align = Alignment(horizontal="right")

        if fmt is None:
            continue

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is None:
                continue

            # Only apply numeric formats to numeric cells (avoid messing with IDs stored as strings)
            if fmt in {FMT_INT, FMT_AMOUNT, FMT_AMOUNT_0, FMT_PERCENT, FMT_NUMBER_2} and not isinstance(
                cell.value, (int, float)
            ):
                continue

            cell.number_format = fmt
            if align is not None:
                cell.alignment = align


def _sample_numeric_values(ws: Worksheet, col_idx: int, limit: int = 50) -> list[float]:
    values: list[float] = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_idx).value
        if isinstance(v, (int, float)):
            values.append(float(v))
            if len(values) >= limit:
                break
    return values


def _format_key_value_sheet(ws: Worksheet) -> None:
    """Format the 'Verdi' column per-row using the 'Felt' column as hint."""
    for row in range(2, ws.max_row + 1):
        key_cell = ws.cell(row=row, column=1)
        val_cell = ws.cell(row=row, column=2)

        key = _normalize_header(key_cell.value)
        val = val_cell.value

        if not isinstance(val, (int, float)):
            continue

        # Decide format
        if any(k in key for k in _KV_PERCENT_KEYWORDS):
            # "andel" is typically 0..1
            if 0 <= float(val) <= 1:
                val_cell.number_format = FMT_PERCENT
            else:
                val_cell.number_format = FMT_NUMBER_2
            val_cell.alignment = Alignment(horizontal="right")
            continue

        if any(k in key for k in _KV_COUNT_KEYWORDS):
            val_cell.number_format = FMT_INT
            val_cell.alignment = Alignment(horizontal="right")
            continue

        if any(k in key for k in _KV_AMOUNT_0_KEYWORDS):
            val_cell.number_format = FMT_AMOUNT_0
            val_cell.alignment = Alignment(horizontal="right")
            continue

        if any(k in key for k in _KV_AMOUNT_KEYWORDS):
            val_cell.number_format = FMT_AMOUNT
            val_cell.alignment = Alignment(horizontal="right")
            continue

        # fallback
        val_cell.number_format = FMT_NUMBER_2
        val_cell.alignment = Alignment(horizontal="right")


def _autosize_columns(ws: Worksheet, max_width: int = 60) -> None:
    """Autosize columns based on cell content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)
