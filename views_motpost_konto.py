"""views_motpost_konto

Motpostanalyse for konto-utvalg i Analyse-fanen.

Funksjonalitet (kort):
  - Bygger et bilagsgrunnlag basert på valgte kontoer.
  - Viser pivot over motkontoer (andre kontoer på samme bilag).
  - Viser bilagsliste for valgt motkonto.
  - Lar bruker markere motkontoer som *outliers* (for testing).
  - Eksporterer en Excel-arbeidsbok med flere faner, inkl. outliers.

Merk:
  - Andel (%) beregnes mot *sum valgte kontoer (netto)*.
  - Prosent vises som prosentpoeng (25.0 = 25 %).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Any, Iterable, Optional

import numbers

import pandas as pd
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from konto_utils import konto_to_str
from formatting import fmt_amount


# -----------------------------
# Data modeller
# -----------------------------


@dataclass(frozen=True)
class MotpostData:
    selected_accounts: tuple[str, ...]
    bilag_count: int
    selected_sum: float
    control_sum: float
    df_motkonto: pd.DataFrame
    df_selected: pd.DataFrame
    df_scope: pd.DataFrame

    # Lazy cache for bilagsdetaljer (1 rad per bilag per motkonto) brukt i eksport / tester.
    _df_details_cache: Optional[pd.DataFrame] = field(default=None, init=False, repr=False)

    @property
    def df_summary(self) -> pd.DataFrame:
        """Bakoverkompatibel alias for pivoten over motkonto."""
        df = self.df_motkonto
        if df is None:
            return pd.DataFrame()
        # Legg på evt. alias-kolonner brukt i eldre tester, uten å påvirke UI/Excel.
        if "SumBeløp" not in df.columns and "Sum" in df.columns:
            df = df.copy()
            df["SumBeløp"] = df["Sum"]
        if "AntallBilag" not in df.columns and "Antall bilag" in df.columns:
            df = df.copy()
            df["AntallBilag"] = df["Antall bilag"]
        return df

    @property
    def df_details(self) -> pd.DataFrame:
        """Bakoverkompatibel: bilagsdetaljer per (bilag, motkonto) i scope."""
        if self._df_details_cache is None:
            df = build_bilag_details_all(self)
            object.__setattr__(self, "_df_details_cache", df)
        return self._df_details_cache


def _konto_str(value: Any) -> str:
    """Normaliser konto/bilag til streng uten trailing '.0'."""

    return konto_to_str(value)


def _norm(value: Any) -> str:
    """Normaliser tekst (brukes for å matche kolonnenavn i Excel-eksport)."""

    return str(value).strip().lower()


def _safe_float(value: Any) -> float:
    try:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip().replace(" ", "")
        if s == "":
            return 0.0
        # Tillat komma som desimal
        s = s.replace(",", ".")
        return float(s)
    except Exception:
        return 0.0


def _to_datetime(value: Any) -> Optional[datetime]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value
    try:
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return None
        if isinstance(ts, pd.Timestamp):
            return ts.to_pydatetime()
        if isinstance(ts, datetime):
            return ts
        return None
    except Exception:
        return None


def _fmt_date_ddmmyyyy(value: Any) -> str:
    dt = _to_datetime(value)
    if not dt:
        return ""
    return dt.strftime("%d.%m.%Y")


def _fmt_percent_points(value: Any, decimals: int = 1) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    try:
        v = float(value)
    except Exception:
        return ""
    fmt = f"{{:.{decimals}f}}".format(v)
    # Norsk: komma som desimal
    fmt = fmt.replace(".", ",")
    return f"{fmt} %"


def _first_non_empty(series: pd.Series) -> Any:
    for v in series.tolist():
        if v is None:
            continue
        if isinstance(v, float) and pd.isna(v):
            continue
        s = str(v).strip()
        if s != "" and s.lower() != "nan":
            return v
    return None


def _unique_join(values: Iterable[str]) -> str:
    out: list[str] = []
    seen: set[str] = set()
    for v in values:
        s = str(v).strip()
        if s == "" or s.lower() == "nan":
            continue
        if s in seen:
            continue
        seen.add(s)
        out.append(s)
    return ", ".join(out)


# -----------------------------
# Bygg analysegrunnlag
# -----------------------------


def build_motpost_data(df_all: pd.DataFrame, selected_accounts: set[str] | Iterable[str]) -> MotpostData:
    """Bygger datagrunnlag for motpostanalyse.

    Forventede kolonner i df_all (minimum):
      - Bilag
      - Konto
      - Beløp

    Ekstra (valgfritt):
      - Kontonavn, Dato, Tekst
    """

    selected_set = {_konto_str(k) for k in selected_accounts}
    selected_tuple = tuple(sorted(selected_set))

    if df_all is None or df_all.empty:
        empty = pd.DataFrame()
        return MotpostData(
            selected_accounts=selected_tuple,
            bilag_count=0,
            selected_sum=0.0,
            control_sum=0.0,
            df_motkonto=empty,
            df_selected=empty,
            df_scope=empty,
        )

    required = {"Bilag", "Konto", "Beløp"}
    missing = required - set(df_all.columns)
    if missing:
        # Returner tomt grunnlag hvis essensielle kolonner mangler
        empty = pd.DataFrame()
        return MotpostData(
            selected_accounts=selected_tuple,
            bilag_count=0,
            selected_sum=0.0,
            control_sum=0.0,
            df_motkonto=empty,
            df_selected=empty,
            df_scope=empty,
        )

    df = df_all.copy()

    # Normaliser nøkler
    df["Bilag_str"] = df["Bilag"].map(_konto_str)
    df["Konto_str"] = df["Konto"].map(_konto_str)

    # Beløp som float
    df["Beløp_num"] = df["Beløp"].map(_safe_float)

    # Dato (kan være tom)
    if "Dato" in df.columns:
        df["Dato_dt"] = pd.to_datetime(df["Dato"], errors="coerce")
    else:
        df["Dato_dt"] = pd.NaT

    # Scope: bilag som inneholder minst én valgt konto
    df_sel = df[df["Konto_str"].isin(selected_set)].copy()
    bilag_scope = sorted(df_sel["Bilag_str"].dropna().unique().tolist())

    if not bilag_scope:
        empty = pd.DataFrame()
        return MotpostData(
            selected_accounts=selected_tuple,
            bilag_count=0,
            selected_sum=0.0,
            control_sum=0.0,
            df_motkonto=empty,
            df_selected=empty,
            df_scope=empty,
        )

    df_scope = df[df["Bilag_str"].isin(set(bilag_scope))].copy()

    selected_sum = float(df_sel["Beløp_num"].sum())
    control_sum = float(df_scope["Beløp_num"].sum())
    bilag_count = int(len(bilag_scope))

    # Pivot for valgte kontoer
    df_selected_pivot = (
        df_sel.groupby("Konto_str", dropna=False)
        .agg(
            Kontonavn=("Kontonavn", _first_non_empty) if "Kontonavn" in df_sel.columns else ("Konto_str", lambda s: ""),
            Sum=("Beløp_num", "sum"),
            Antall_bilag=("Bilag_str", pd.Series.nunique),
        )
        .reset_index()
        .rename(columns={"Konto_str": "Konto", "Antall_bilag": "Antall bilag"})
    )
    df_selected_pivot["% andel"] = (
        (df_selected_pivot["Sum"] / selected_sum * 100.0) if selected_sum != 0 else 0.0
    )
    df_selected_pivot = df_selected_pivot[["Konto", "Kontonavn", "Sum", "% andel", "Antall bilag"]]
    df_selected_pivot = df_selected_pivot.sort_values(by="Sum", key=lambda s: s.abs(), ascending=False)

    # Pivot for motkontoer (alle andre kontoer i scope)
    df_mot = df_scope[~df_scope["Konto_str"].isin(selected_set)].copy()

    df_mot_pivot = (
        df_mot.groupby("Konto_str", dropna=False)
        .agg(
            Kontonavn=("Kontonavn", _first_non_empty) if "Kontonavn" in df_mot.columns else ("Konto_str", lambda s: ""),
            Sum=("Beløp_num", "sum"),
            Antall_bilag=("Bilag_str", pd.Series.nunique),
        )
        .reset_index()
        .rename(columns={"Konto_str": "Motkonto", "Antall_bilag": "Antall bilag"})
    )
    df_mot_pivot["% andel"] = ((df_mot_pivot["Sum"] / selected_sum * 100.0) if selected_sum != 0 else 0.0)
    df_mot_pivot = df_mot_pivot[["Motkonto", "Kontonavn", "Sum", "% andel", "Antall bilag"]]
    df_mot_pivot = df_mot_pivot.sort_values(by="Sum", key=lambda s: s.abs(), ascending=False)

    # df_scope: behold standard kolonnenavn for videre bruk.
    # NB: Dersom kildedata allerede har en "Beløp"-kolonne (ofte tekst), vil en rename
    # gi duplikatnavn. Vi overstyrer derfor eksplisitt og fjerner hjelpekollonnen.
    df_scope["Beløp"] = df_scope["Beløp_num"]
    df_scope = df_scope.drop(columns=["Beløp_num"], errors="ignore")

    return MotpostData(
        selected_accounts=selected_tuple,
        bilag_count=bilag_count,
        selected_sum=selected_sum,
        control_sum=control_sum,
        df_motkonto=df_mot_pivot.reset_index(drop=True),
        df_selected=df_selected_pivot.reset_index(drop=True),
        df_scope=df_scope.reset_index(drop=True),
    )


def build_bilag_details(data: MotpostData, motkonto: str) -> pd.DataFrame:
    """Bygger bilagsliste for en gitt motkonto."""

    if data.df_scope is None or data.df_scope.empty:
        return pd.DataFrame()

    motkonto = _konto_str(motkonto)
    selected_set = set(data.selected_accounts)

    df = data.df_scope.copy()
    df["Bilag_str"] = df["Bilag"].map(_konto_str)
    df["Konto_str"] = df["Konto"].map(_konto_str)

    # Bilag som inneholder motkonto
    df_m = df[df["Konto_str"] == motkonto]
    if df_m.empty:
        return pd.DataFrame()
    bilag_set = set(df_m["Bilag_str"].unique().tolist())

    rows: list[dict[str, Any]] = []
    for bilag in sorted(bilag_set):
        df_b = df[df["Bilag_str"] == bilag]
        selected_sum = float(df_b[df_b["Konto_str"].isin(selected_set)]["Beløp"].map(_safe_float).sum())
        mot_sum = float(df_b[df_b["Konto_str"] == motkonto]["Beløp"].map(_safe_float).sum())
        kontoer = _unique_join(sorted({_konto_str(x) for x in df_b["Konto"].tolist()}))
        dato = _first_non_empty(df_b["Dato"].astype(object)) if "Dato" in df_b.columns else None
        tekst = _first_non_empty(df_b["Tekst"].astype(object)) if "Tekst" in df_b.columns else None
        rows.append(
            {
                "Bilag": bilag,
                "Dato": _to_datetime(dato),
                "Tekst": tekst or "",
                "Beløp (valgte kontoer)": selected_sum,
                "Motbeløp": mot_sum,
                "Kontoer i bilag": kontoer,
            }
        )

    return pd.DataFrame(rows)



def build_bilag_details_all(data: MotpostData) -> pd.DataFrame:
    """Bygger bilagsdetaljer for *alle* motkontoer i scope.

    Returnerer 1 rad per (bilag, motkonto) med bl.a. sum for valgte kontoer og motbeløp.
    Denne brukes i Excel-eksport når ingen motkonto er valgt, og som bakoverkompatibel `df_details`.
    """
    df = data.df_scope
    if df is None or df.empty:
        return pd.DataFrame(
            columns=[
                "Bilag_key",
                "Bilag",
                "Dato",
                "Tekst",
                "Motkonto",
                "Motkontonavn",
                "Beløp (valgte kontoer)",
                "Motbeløp",
                "Kontoer i bilag",
            ]
        )

    selected_set = set(_konto_str(k) for k in data.selected_accounts)

    # Sikre at hjelpekolonnene finnes (build_motpost_data legger disse på, men vær robust).
    df_work = df.copy()
    if "Bilag_str" not in df_work.columns:
        df_work["Bilag_str"] = df_work.get("Bilag", "").map(_konto_str)
    if "Konto_str" not in df_work.columns:
        df_work["Konto_str"] = df_work.get("Konto", "").map(_konto_str)

    # Beløp er forventet numerisk i df_scope. Hvis ikke, forsøk å konvertere.
    if "Beløp" in df_work.columns and not pd.api.types.is_numeric_dtype(df_work["Beløp"]):
        df_work["Beløp"] = df_work["Beløp"].map(_safe_float)

    sel_mask = df_work["Konto_str"].isin(selected_set)

    # Sum valgte kontoer per bilag
    sel_sum = (
        df_work.loc[sel_mask]
        .groupby("Bilag_str")["Beløp"]
        .sum()
        .rename("Beløp (valgte kontoer)")
    )

    # Motkonto summer per bilag
    df_mot = df_work.loc[~sel_mask].copy()
    if df_mot.empty:
        return pd.DataFrame(
            columns=[
                "Bilag_key",
                "Bilag",
                "Dato",
                "Tekst",
                "Motkonto",
                "Motkontonavn",
                "Beløp (valgte kontoer)",
                "Motbeløp",
                "Kontoer i bilag",
            ]
        )

    def _first_date(series: pd.Series):
        s = series.dropna()
        return s.iloc[0] if not s.empty else None

    mot_agg = (
        df_mot.groupby(["Bilag_str", "Konto_str"])
        .agg(
            Motbeløp=("Beløp", "sum"),
            Motkontonavn=("Kontonavn", _first_non_empty),
        )
        .reset_index()
        .rename(columns={"Konto_str": "Motkonto"})
    )

    # Meta per bilag
    # Bilag_key kan mangle; bruk da Bilag_str
    if "Bilag_key" in df_work.columns:
        bilag_key = df_work.groupby("Bilag_str")["Bilag_key"].agg(lambda s: s.dropna().iloc[0] if not s.dropna().empty else "")
    else:
        bilag_key = pd.Series(df_work["Bilag_str"].unique(), index=df_work["Bilag_str"].unique())

    meta = (
        df_work.groupby("Bilag_str")
        .agg(
            Bilag=("Bilag", lambda s: s.dropna().iloc[0] if not s.dropna().empty else ""),
            Dato=("Dato", _first_date),
            Tekst=("Tekst", _first_non_empty),
            **{"Kontoer i bilag": ("Konto_str", lambda s: ", ".join(sorted(set(x for x in s.dropna().tolist() if str(x).strip()))))},
        )
        .reset_index()
    )
    meta["Bilag_key"] = meta["Bilag_str"].map(bilag_key.to_dict()).fillna(meta["Bilag_str"])

    details = mot_agg.merge(sel_sum.reset_index(), on="Bilag_str", how="left").merge(meta, on="Bilag_str", how="left")
    details["Beløp (valgte kontoer)"] = details["Beløp (valgte kontoer)"].fillna(0.0)

    # Kolonneordre
    out_cols = [
        "Bilag_key",
        "Bilag",
        "Dato",
        "Tekst",
        "Motkonto",
        "Motkontonavn",
        "Beløp (valgte kontoer)",
        "Motbeløp",
        "Kontoer i bilag",
    ]
    details = details[out_cols]

    # Alias brukt i eldre tester
    if "Beløp valgte kontoer" not in details.columns:
        details["Beløp valgte kontoer"] = details["Beløp (valgte kontoer)"]

    return details


def build_outlier_bilag_transactions(data: MotpostData, outliers: set[str]) -> pd.DataFrame:
    """Alle transaksjoner for bilag som inneholder outlier-motkonto(er)."""

    if data.df_scope is None or data.df_scope.empty or not outliers:
        return pd.DataFrame()

    out_set = {_konto_str(x) for x in outliers}
    selected_set = set(data.selected_accounts)

    df = data.df_scope.copy()
    df["Bilag_str"] = df["Bilag"].map(_konto_str)
    df["Konto_str"] = df["Konto"].map(_konto_str)
    df["Beløp_num"] = df["Beløp"].map(_safe_float)

    df_out = df[df["Konto_str"].isin(out_set)]
    if df_out.empty:
        return pd.DataFrame()

    bilag_out = sorted(df_out["Bilag_str"].unique().tolist())
    df_all = df[df["Bilag_str"].isin(set(bilag_out))].copy()

    # Agg for hver bilag
    sum_selected_per_bilag = (
        df_all[df_all["Konto_str"].isin(selected_set)]
        .groupby("Bilag_str")["Beløp_num"]
        .sum()
        .to_dict()
    )
    outliers_per_bilag = (
        df_all[df_all["Konto_str"].isin(out_set)]
        .groupby("Bilag_str")["Konto_str"]
        .apply(lambda s: _unique_join(sorted(set(s.tolist()))))
        .to_dict()
    )

    df_all["Beløp (valgte kontoer)"] = df_all["Bilag_str"].map(lambda b: float(sum_selected_per_bilag.get(b, 0.0)))
    df_all["Outlier motkontoer i bilag"] = df_all["Bilag_str"].map(lambda b: outliers_per_bilag.get(b, ""))

    # Normaliser dato
    if "Dato" in df_all.columns:
        df_all["Dato_dt"] = pd.to_datetime(df_all["Dato"], errors="coerce")
    else:
        df_all["Dato_dt"] = pd.NaT

    # Kolonneutvalg
    cols = [
        "Bilag_str",
        "Dato_dt",
        "Tekst" if "Tekst" in df_all.columns else None,
        "Konto_str",
        "Kontonavn" if "Kontonavn" in df_all.columns else None,
        "Beløp_num",
        "Beløp (valgte kontoer)",
        "Outlier motkontoer i bilag",
    ]
    cols = [c for c in cols if c is not None]
    df_outlier = df_all[cols].copy()

    rename_map = {
        "Bilag_str": "Bilag",
        "Dato_dt": "Dato",
        "Konto_str": "Konto",
        "Beløp_num": "Beløp",
    }
    df_outlier = df_outlier.rename(columns=rename_map)

    # Sorter pent
    sort_cols = ["Bilag", "Dato"] if "Dato" in df_outlier.columns else ["Bilag"]
    df_outlier = df_outlier.sort_values(by=sort_cols, ascending=True)
    return df_outlier.reset_index(drop=True)


# -----------------------------
# Excel eksport (bygg workbook)
# -----------------------------


_FILL_TITLE = PatternFill("solid", fgColor="C6EFCE")  # lys grønn
_FILL_HEADER = PatternFill("solid", fgColor="E2EFDA")
_FILL_OUTLIER = PatternFill("solid", fgColor="FFF2CC")  # lys gul

_BORDER_THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _set_cell(ws, row: int, col: int, value: Any, *, bold: bool = False, fill: PatternFill | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    if bold:
        cell.font = Font(bold=True)
    if fill is not None:
        cell.fill = fill
    return cell


def _autosize_columns(ws, *, min_width: int = 10, max_width: int = 60):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        best = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            best = max(best, len(s))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, best + 2))


def _apply_table_style(
    ws,
    *,
    header_row: int,
    data_first_row: int,
    data_last_row: int,
    percent_headers: set[str] | None = None,
    outlier_col_name: str | None = None,
    outlier_yes_value: str = "Ja",
):
    """Enkel, robust formatering for tabeller der header ikke nødvendigvis er på rad 1."""

    percent_headers = percent_headers or {"% andel"}
    headers: list[str] = []
    for c in range(1, ws.max_column + 1):
        hv = ws.cell(row=header_row, column=c).value
        headers.append(str(hv) if hv is not None else "")

    # Header stil
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = Font(bold=True)
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER_THIN

    # Finn outlier-kolonne
    outlier_col_idx: Optional[int] = None
    if outlier_col_name:
        for idx, h in enumerate(headers, start=1):
            if str(h).strip().lower() == outlier_col_name.strip().lower():
                outlier_col_idx = idx
                break

    # Data-celler
    amount_headers = {"sum", "beløp", "motbeløp", "beløp (valgte kontoer)"}
    date_headers = {"dato"}
    int_headers = {"antall bilag"}

    for r in range(data_first_row, data_last_row + 1):
        outlier_row = False
        if outlier_col_idx is not None:
            v = ws.cell(row=r, column=outlier_col_idx).value
            outlier_row = (str(v).strip() == outlier_yes_value)

        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(vertical="top")

            h_norm = str(h).strip().lower()
            # Beløp
            if h_norm in {x.lower() for x in amount_headers}:
                if isinstance(cell.value, numbers.Number) and not isinstance(cell.value, bool):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="top")
            # Dato
            elif h_norm in {x.lower() for x in date_headers}:
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = "dd.mm.yyyy"
                    cell.alignment = Alignment(horizontal="left", vertical="top")
            # Antall
            elif h_norm in {x.lower() for x in int_headers}:
                if isinstance(cell.value, numbers.Number) and not isinstance(cell.value, bool):
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="right", vertical="top")
            # Prosentpoeng
            elif h_norm in {x.lower() for x in percent_headers}:
                if isinstance(cell.value, numbers.Number) and not isinstance(cell.value, bool):
                    # Verdien er prosentpoeng (25.0 betyr 25 %)
                    cell.number_format = '0.0"%"'
                    cell.alignment = Alignment(horizontal="right", vertical="top")

            if outlier_row:
                cell.fill = _FILL_OUTLIER


def _write_df_table(ws, df: pd.DataFrame, *, title: str, summary: str, start_row: int = 1) -> tuple[int, int]:
    """Skriver et 'ark med tittel + sammendrag + tabell'.

    Returnerer (header_row, last_row).
    """

    # Tittel (rad 1)
    _set_cell(ws, start_row, 1, title, bold=True, fill=_FILL_TITLE)

    # Sammendrag (rad 2)
    _set_cell(ws, start_row + 1, 1, summary)

    # Tabell header (rad 3)
    header_row = start_row + 2
    if df is None or df.empty:
        _set_cell(ws, header_row, 1, "Ingen data")
        ws.freeze_panes = ws["A4"]
        return header_row, header_row

    headers = list(df.columns)
    for col_idx, h in enumerate(headers, start=1):
        _set_cell(ws, header_row, col_idx, h, bold=True, fill=_FILL_HEADER)

    # Data starter rad 4
    r = header_row + 1
    for _, row in df.iterrows():
        for col_idx, h in enumerate(headers, start=1):
            v = row[h]
            # Normaliser manglende verdier (NaN/NaT) til tom celle i Excel
            try:
                if pd.isna(v):
                    v = None
            except Exception:
                pass

            # Skriv dato som datetime hvis mulig (gir riktig Excel-format)
            if isinstance(v, pd.Timestamp):
                v = v.to_pydatetime()
            elif _norm(h) == "dato" and isinstance(v, str) and v.strip():
                # Forsok å tolke streng som dato (aksepterer både YYYY-MM-DD og dd.mm.yyyy)
                try:
                    v_dt = pd.to_datetime(v, errors="raise", dayfirst=True)
                    if isinstance(v_dt, pd.Timestamp):
                        v = v_dt.to_pydatetime()
                except Exception:
                    pass

            ws.cell(row=r, column=col_idx, value=v)
        r += 1

    ws.freeze_panes = ws["A4"]
    return header_row, r - 1


def build_motpost_excel_workbook(
    data: MotpostData,
    outlier_motkonto: Optional[Iterable[str]] = None,
    *,
    selected_motkonto: Optional[str] = None,
    df_details_view: Optional[pd.DataFrame] = None,
    outliers: Optional[Iterable[str]] = None,
    outlier_accounts: Optional[Iterable[str]] = None,
) -> Workbook:
    """Bygger openpyxl Workbook for motpostanalyse.

    Merk: signaturen er litt fleksibel (synonymer for outliers) for å være robust
    mot ulike kall i GUI/tester.
    """

    out_set: set[str] = set()
    for src in (outlier_motkonto, outliers, outlier_accounts):
        if src:
            out_set |= {_konto_str(x) for x in src}

    wb = Workbook()

    # Default sheet -> Oversikt
    ws_overview = wb.active
    ws_overview.title = "Oversikt"
    ws_overview["A1"] = "Oversikt"
    ws_overview["A1"].font = Font(bold=True)
    ws_overview["A3"] = "Denne arbeidsboken er generert fra motpostanalysen."

    ws_overview["A5"] = "Fane"
    ws_overview["B5"] = "Beskrivelse"
    ws_overview["A5"].font = Font(bold=True)
    ws_overview["B5"].font = Font(bold=True)

    rows = [
        (
            "Motkonto",
            "Oppsummering av motkontoer (andre kontoer på samme bilag som valgte kontoer).",
        ),
        (
            "Valgte kontoer",
            "Oppsummering av valgte kontoer (sum og andel).",
        ),
        (
            "Bilag",
            "Bilagsliste for valgt motkonto. Hvis ingen motkonto er valgt ved eksport, viser fanen alle bilag/motkontoer i grunnlaget.",
        ),
        (
            "Outliers",
            "Motkontoer som er markert som outliers (typisk gjenstand for testing).",
        ),
        (
            "OutlierBilag",
            "Alle transaksjoner på bilag som inneholder outlier-motkontoer.",
        ),
    ]
    r = 6
    for a, b in rows:
        ws_overview.cell(row=r, column=1, value=a)
        ws_overview.cell(row=r, column=2, value=b)
        r += 1
    _autosize_columns(ws_overview)

    # Felles sammendrag
    summary_line = (
        f"Valgte kontoer: {', '.join(data.selected_accounts)} | "
        f"Bilag i grunnlag: {data.bilag_count} | "
        f"Sum valgte kontoer (netto): {fmt_amount(data.selected_sum)} | "
        f"Kontroll (valgt + mot): {fmt_amount(data.control_sum)}"
    )

    # Motkonto
    ws_mot = wb.create_sheet("Motkonto")
    df_mot = data.df_motkonto.copy() if data.df_motkonto is not None else pd.DataFrame()
    if not df_mot.empty:
        df_mot["Outlier"] = df_mot["Motkonto"].map(lambda k: "Ja" if _konto_str(k) in out_set else "")
    header_row, last_row = _write_df_table(ws_mot, df_mot, title="Motkonto (pivot)", summary=summary_line)
    if last_row >= header_row + 1:
        _apply_table_style(
            ws_mot,
            header_row=header_row,
            data_first_row=header_row + 1,
            data_last_row=last_row,
            outlier_col_name="Outlier",
        )
        _autosize_columns(ws_mot)

    # Valgte kontoer
    ws_sel = wb.create_sheet("Valgte kontoer")
    df_sel = data.df_selected.copy() if data.df_selected is not None else pd.DataFrame()
    header_row, last_row = _write_df_table(ws_sel, df_sel, title="Valgte kontoer (pivot)", summary=summary_line)
    if last_row >= header_row + 1:
        _apply_table_style(ws_sel, header_row=header_row, data_first_row=header_row + 1, data_last_row=last_row)
        _autosize_columns(ws_sel)

    # Bilag (for valgt motkonto)
    ws_bilag = wb.create_sheet("Bilag")
    if selected_motkonto:
        df_b = build_bilag_details(data, selected_motkonto)
        title = f"Bilag for valgt motkonto: {selected_motkonto}"
        header_row, last_row = _write_df_table(ws_bilag, df_b, title=title, summary=summary_line)
        if last_row >= header_row + 1:
            _apply_table_style(ws_bilag, header_row=header_row, data_first_row=header_row + 1, data_last_row=last_row)
            _autosize_columns(ws_bilag)
    else:
        # Ingen motkonto valgt: eksporter bilagsdetaljer for alle motkontoer i scope.
        df_bilag_all = (df_details_view if df_details_view is not None else data.df_details).copy()
        if "Beløp valgte kontoer" in df_bilag_all.columns:
            df_bilag_all = df_bilag_all.drop(columns=["Beløp valgte kontoer"])
        header_row, last_row = _write_df_table(
            ws_bilag,
            df_bilag_all,
            title="Bilag (alle motkontoer)",
            summary=summary_line,
        )
        _apply_table_style(ws_bilag, header_row=header_row, data_first_row=header_row + 1, data_last_row=last_row)
        _autosize_columns(ws_bilag)

    # Outliers (subset av motkonto)
    ws_out = wb.create_sheet("Outliers")
    df_out = pd.DataFrame()
    if out_set and df_mot is not None and not df_mot.empty:
        df_out = df_mot[df_mot["Motkonto"].map(lambda k: _konto_str(k) in out_set)].copy()
    header_row, last_row = _write_df_table(ws_out, df_out, title="Outliers (motkonto)", summary=summary_line)
    if last_row >= header_row + 1:
        _apply_table_style(
            ws_out,
            header_row=header_row,
            data_first_row=header_row + 1,
            data_last_row=last_row,
            outlier_col_name="Outlier",
        )
        _autosize_columns(ws_out)

    # OutlierBilag (alle transaksjoner)
    ws_out_bilag = wb.create_sheet("OutlierBilag")
    df_out_bilag = build_outlier_bilag_transactions(data, out_set)
    header_row, last_row = _write_df_table(
        ws_out_bilag,
        df_out_bilag,
        title="Bilag relatert til outlier-motkontoer (alle transaksjoner)",
        summary=summary_line,
    )
    if last_row >= header_row + 1:
        _apply_table_style(ws_out_bilag, header_row=header_row, data_first_row=header_row + 1, data_last_row=last_row)
        _autosize_columns(ws_out_bilag)

    return wb


# -----------------------------
# GUI
# -----------------------------


class MotpostKontoView(tk.Toplevel):
    def __init__(self, master: tk.Misc, df_transactions: pd.DataFrame, konto_list: list[str] | set[str] | tuple[str, ...]):
        super().__init__(master)
        self.title("Motpostanalyse")
        self.geometry("1100x700")

        self._df_all = df_transactions
        self._selected_accounts = {_konto_str(k) for k in konto_list}
        self._data = build_motpost_data(self._df_all, self._selected_accounts)

        self._outliers: set[str] = set()
        self._selected_motkonto: Optional[str] = None

        self._details_limit_var = tk.IntVar(value=200)

        self._build_ui()
        self._render_summary()

    # --- UI bygging ---
    def _build_ui(self) -> None:
        top = ttk.Frame(self)
        top.pack(side=tk.TOP, fill=tk.X, padx=10, pady=8)

        info = (
            f"Valgte kontoer: {', '.join(self._data.selected_accounts)}  |  "
            f"Bilag i grunnlag: {self._data.bilag_count}  |  "
            f"Sum valgte kontoer (netto): {fmt_amount(self._data.selected_sum)}  |  "
            f"Kontroll (valgt + mot): {fmt_amount(self._data.control_sum)}"
        )
        self._info_label = ttk.Label(top, text=info)
        self._info_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        btn_frame = ttk.Frame(top)
        btn_frame.pack(side=tk.RIGHT)

        ttk.Button(btn_frame, text="Merk outlier", command=self._mark_outlier).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_frame, text="Nullstill outliers", command=self._clear_outliers).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_frame, text="Eksporter Excel", command=self._export_excel).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_frame, text="Lukk", command=self.destroy).pack(side=tk.LEFT)

        # Mid: motkonto pivot
        mid = ttk.Frame(self)
        mid.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10)

        ttk.Label(mid, text="Motkonto (pivot)").pack(anchor=tk.W)

        columns = ("Motkonto", "Kontonavn", "Sum", "% andel", "Antall bilag", "Outlier")
        self._tree_summary = ttk.Treeview(mid, columns=columns, show="headings", selectmode="extended")
        for c in columns:
            self._tree_summary.heading(c, text=c)
            self._tree_summary.column(c, width=120 if c != "Tekst" else 300, anchor=tk.W)

        self._tree_summary.column("Sum", anchor=tk.E, width=140)
        self._tree_summary.column("% andel", anchor=tk.E, width=90)
        self._tree_summary.column("Antall bilag", anchor=tk.E, width=90)
        self._tree_summary.column("Outlier", anchor=tk.W, width=70)

        yscroll = ttk.Scrollbar(mid, orient=tk.VERTICAL, command=self._tree_summary.yview)
        self._tree_summary.configure(yscrollcommand=yscroll.set)

        self._tree_summary.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)

        self._tree_summary.tag_configure("neg", foreground="red")
        self._tree_summary.tag_configure("outlier", background="#FFF2CC")
        self._tree_summary.bind("<<TreeviewSelect>>", self._on_select_motkonto)

        # Bottom: bilag-liste for valgt motkonto
        bottom = ttk.Frame(self)
        bottom.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=(8, 10))

        header = ttk.Frame(bottom)
        header.pack(side=tk.TOP, fill=tk.X)

        ttk.Label(header, text="Bilag for valgt motkonto").pack(side=tk.LEFT)
        ttk.Label(header, text="Vis:").pack(side=tk.LEFT, padx=(10, 2))
        sp = ttk.Spinbox(header, from_=50, to=5000, increment=50, width=7, textvariable=self._details_limit_var, command=self._refresh_details)
        sp.pack(side=tk.LEFT)

        ttk.Button(header, text="Drilldown", command=self._drilldown).pack(side=tk.RIGHT)

        columns2 = ("Bilag", "Dato", "Tekst", "Beløp (valgte kontoer)", "Motbeløp", "Kontoer i bilag")
        self._tree_details = ttk.Treeview(bottom, columns=columns2, show="headings", selectmode="browse")
        for c in columns2:
            self._tree_details.heading(c, text=c)
            self._tree_details.column(c, width=120, anchor=tk.W)

        self._tree_details.column("Tekst", width=350)
        self._tree_details.column("Beløp (valgte kontoer)", anchor=tk.E, width=160)
        self._tree_details.column("Motbeløp", anchor=tk.E, width=120)
        self._tree_details.column("Kontoer i bilag", width=180)

        yscroll2 = ttk.Scrollbar(bottom, orient=tk.VERTICAL, command=self._tree_details.yview)
        self._tree_details.configure(yscrollcommand=yscroll2.set)

        self._tree_details.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll2.pack(side=tk.RIGHT, fill=tk.Y)

        self._tree_details.tag_configure("neg", foreground="red")

    # --- Rendering ---
    def _render_summary(self) -> None:
        self._tree_summary.delete(*self._tree_summary.get_children())

        df = self._data.df_motkonto
        if df is None or df.empty:
            return

        for _, row in df.iterrows():
            motkonto = _konto_str(row.get("Motkonto"))
            kontonavn = row.get("Kontonavn", "")
            s = float(row.get("Sum", 0.0))
            share = float(row.get("% andel", 0.0))
            cnt = int(row.get("Antall bilag", 0))
            out = "Ja" if motkonto in self._outliers else ""

            tags: list[str] = []
            if s < 0:
                tags.append("neg")
            if motkonto in self._outliers:
                tags.append("outlier")

            self._tree_summary.insert(
                "",
                tk.END,
                values=(motkonto, kontonavn, fmt_amount(s), _fmt_percent_points(share), cnt, out),
                tags=tuple(tags),
            )

    def _refresh_details(self) -> None:
        self._tree_details.delete(*self._tree_details.get_children())

        if not self._selected_motkonto:
            return

        df_b = build_bilag_details(self._data, self._selected_motkonto)
        if df_b is None or df_b.empty:
            return

        limit = int(self._details_limit_var.get() or 200)
        df_b = df_b.head(limit)

        for _, row in df_b.iterrows():
            bilag = row.get("Bilag", "")
            dato = _fmt_date_ddmmyyyy(row.get("Dato"))
            tekst = row.get("Tekst", "")
            bel_sel = float(row.get("Beløp (valgte kontoer)", 0.0))
            motb = float(row.get("Motbeløp", 0.0))
            kontoer = row.get("Kontoer i bilag", "")

            tags: list[str] = []
            if bel_sel < 0 or motb < 0:
                tags.append("neg")

            self._tree_details.insert(
                "",
                tk.END,
                values=(bilag, dato, tekst, fmt_amount(bel_sel), fmt_amount(motb), kontoer),
                tags=tuple(tags),
            )

    # --- Events / actions ---
    def _on_select_motkonto(self, _event=None) -> None:
        sel = self._tree_summary.selection()
        if not sel:
            self._selected_motkonto = None
            self._refresh_details()
            return
        # Bruk første valgte som "aktiv" motkonto for bilagsvisning
        item = sel[0]
        motkonto = self._tree_summary.item(item, "values")[0]
        self._selected_motkonto = _konto_str(motkonto)
        self._refresh_details()

    def _mark_outlier(self) -> None:
        sel = self._tree_summary.selection()
        if not sel:
            messagebox.showinfo("Motpostanalyse", "Velg en eller flere motkontoer for å markere som outlier.")
            return
        for item in sel:
            motkonto = self._tree_summary.item(item, "values")[0]
            self._outliers.add(_konto_str(motkonto))
        self._render_summary()

    def _clear_outliers(self) -> None:
        self._outliers.clear()
        self._render_summary()

    def _export_excel(self) -> None:
        default_name = "motpostanalyse.xlsx"
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Lagre Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        try:
            wb = build_motpost_excel_workbook(
                self._data,
                outlier_motkonto=self._outliers,
                selected_motkonto=self._selected_motkonto,
            )
            wb.save(path)
            messagebox.showinfo("Motpostanalyse", f"Eksportert til Excel:\n{path}")
        except Exception as e:
            messagebox.showerror("Motpostanalyse", f"Kunne ikke eksportere til Excel:\n{e}")

    def _drilldown(self) -> None:
        sel = self._tree_details.selection()
        if not sel:
            messagebox.showinfo("Motpostanalyse", "Velg et bilag i listen for å åpne drilldown.")
            return
        bilag = self._tree_details.item(sel[0], "values")[0]
        bilag = _konto_str(bilag)
        try:
            from views_bilag_drill import BilagDrillDialog

            dlg = BilagDrillDialog(self, self._df_all)
            dlg.preset_and_show(bilag)
        except Exception as e:
            messagebox.showerror("Motpostanalyse", f"Kunne ikke åpne drilldown:\n{e}")


def show_motpost_konto(master: tk.Misc, df_transactions: pd.DataFrame, konto_list: list[str] | set[str] | tuple[str, ...]) -> None:
    """Entry-point brukt fra Analyse-fanen."""

    MotpostKontoView(master, df_transactions, list(konto_list))


# Bakoverkompatibilitet (noen steder kan ha importert underscorenavnet)
_show_motpost_konto = show_motpost_konto
