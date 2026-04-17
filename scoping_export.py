"""
Eksporterer scoping-resultat til en formatert Excel-fil med:
  - Oppsett (klient, år, vesentlighetsgrenser)
  - Scoping-tabell (alle regnskapslinjer med klassifisering og beslutning)
  - Aggregeringskontroll
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scoping_engine import ScopingResult

log = logging.getLogger(__name__)

# Farger
_FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
_FILL_VESENTLIG = PatternFill("solid", fgColor="FDE8E8")
_FILL_MODERAT = PatternFill("solid", fgColor="FEF3C7")
_FILL_IKKE = PatternFill("solid", fgColor="D1FAE5")
_FILL_MANUELL = PatternFill("solid", fgColor="DBEAFE")
_FILL_SUMMARY = PatternFill("solid", fgColor="F0F0F0")
_FILL_OK = PatternFill("solid", fgColor="D1FAE5")
_FILL_WARN = PatternFill("solid", fgColor="FDE8E8")

_FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
_FONT_BOLD = Font(bold=True, size=10)
_FONT_NORMAL = Font(size=10)
_FONT_SUMMARY = Font(size=10, color="999999")
_FONT_TITLE = Font(bold=True, size=14)

_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _fmt_amount(value: float) -> str:
    if abs(value) < 0.5:
        return "0"
    return f"{value:,.0f}".replace(",", " ")


def _class_label(c: str) -> str:
    return {
        "vesentlig": "Vesentlig",
        "moderat": "Moderat",
        "ikke_vesentlig": "Ikke vesentlig",
        "manuell": "Manuell",
    }.get(c, "")


def _scope_label(s: str) -> str:
    return {"inn": "Inn", "ut": "Ut"}.get(s, "")


def export_scoping(
    result: ScopingResult,
    path: str | Path,
    *,
    client_name: str = "",
    year: str = "",
) -> Path:
    """Eksporter scoping-resultat til Excel-arbeidspapir."""
    path = Path(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scoping"

    try:
        year_int = int(year) if year else None
    except (TypeError, ValueError):
        year_int = None
    has_prior = any(line.amount_prior is not None for line in result.lines)

    # Oppsett-header
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = "Scoping regnskapslinjer"
    c.font = _FONT_TITLE

    row = 3
    info = [
        ("Klient:", client_name or "—"),
        ("År:", year or "—"),
        ("Overall materiality (OM):", _fmt_amount(result.om)),
        ("Performance materiality (PM):", _fmt_amount(result.pm)),
        ("Clearly trivial (SUM):", _fmt_amount(result.sum_threshold)),
    ]
    for label, val in info:
        ws.cell(row=row, column=1, value=label).font = _FONT_BOLD
        ws.cell(row=row, column=2, value=val).font = _FONT_NORMAL
        row += 1

    row += 1

    ub_label = f"UB {year_int}" if year_int is not None else "UB"
    ub_fjor_label = f"UB {year_int - 1}" if year_int is not None else "UB i fjor"
    headers = [
        ("Regnr", 8),
        ("Regnskapslinje", 30),
        ("Type", 6),
        (ub_label, 15),
    ]
    if has_prior:
        headers.extend(
            [
                (ub_fjor_label, 15),
                ("Endring", 15),
                ("Endring %", 10),
            ]
        )
    headers.extend(
        [
            ("% av PM", 10),
            ("Klassifisering", 16),
            ("Scoping", 10),
            ("Revisjonshandling", 30),
            ("Begrunnelse", 35),
            ("Handl.", 8),
        ]
    )

    for col_idx, (hdr, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row += 1

    amount_cols = {4}
    if has_prior:
        amount_cols.update({5, 6})
    pct_cols = {7} if has_prior else {5}

    for line in result.lines:
        fill_map = {
            "vesentlig": _FILL_VESENTLIG,
            "moderat": _FILL_MODERAT,
            "ikke_vesentlig": _FILL_IKKE,
            "manuell": _FILL_MANUELL,
        }
        if line.is_summary:
            fill = _FILL_SUMMARY
            font = _FONT_SUMMARY
        else:
            fill = fill_map.get(line.classification, PatternFill())
            font = _FONT_NORMAL

        values: list[object] = [
            int(line.regnr) if line.regnr.isdigit() else line.regnr,
            line.regnskapslinje,
            line.line_type,
            line.amount,
        ]
        if has_prior:
            values.extend(
                [
                    line.amount_prior,
                    line.change_amount,
                    f"{line.change_pct:+.1f}%"
                    if line.change_pct is not None and not line.is_summary
                    else "",
                ]
            )
        values.extend(
            [
                "" if line.is_summary else f"{line.pct_of_pm:.0f}%",
                _class_label(line.classification),
                _scope_label(line.scoping),
                line.audit_action if not line.is_summary else "",
                line.rationale,
                line.action_count if line.action_count else "",
            ]
        )

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = font
            cell.fill = fill
            cell.border = _THIN_BORDER
            if col_idx in amount_cols:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in pct_cols:
                cell.alignment = Alignment(horizontal="right")

        row += 1

    row += 1
    non_sum = [l for l in result.lines if not l.is_summary]
    scoped_out = sum(abs(l.amount) for l in non_sum if l.scoping == "ut")
    ok = scoped_out < result.om if result.om > 0 else True

    total_saldo = sum(abs(l.amount) for l in non_sum)
    pct_om = round(scoped_out / result.om * 100, 1) if result.om > 0 else 0
    pct_total = round(scoped_out / total_saldo * 100, 1) if total_saldo > 0 else 0

    ws.cell(row=row, column=1, value="AGGREGERINGSRISIKO — KONTROLL").font = _FONT_BOLD
    row += 1

    vesentlige = sum(1 for l in non_sum if l.classification == "vesentlig")
    moderate = sum(1 for l in non_sum if l.classification == "moderat")
    ikke_ves = sum(1 for l in non_sum if l.classification == "ikke_vesentlig")
    scoped_inn = sum(1 for l in non_sum if l.scoping == "inn")
    scoped_ut_n = sum(1 for l in non_sum if l.scoping == "ut")
    ikke_besluttet = sum(1 for l in non_sum if not l.scoping)

    agg_rows = [
        ("Sum scopet ut (abs. beløp):", _fmt_amount(scoped_out)),
        ("Scopet ut som % av OM:", f"{pct_om}%"),
        ("Scopet ut som % av total saldo:", f"{pct_total}%"),
        ("Antall linjer scopet inn:", str(scoped_inn)),
        ("Antall linjer scopet ut:", str(scoped_ut_n)),
        ("Vurdering aggregeringsrisiko:", "OK" if ok else "ADVARSEL — overskrider OM!"),
    ]
    for label, val in agg_rows:
        ws.cell(row=row, column=1, value=label).font = _FONT_NORMAL
        cell = ws.cell(row=row, column=2, value=val)
        cell.font = _FONT_BOLD
        if label.startswith("Vurdering"):
            cell.fill = _FILL_OK if ok else _FILL_WARN
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Oppsummering").font = _FONT_BOLD
    row += 1
    summary_lines = [
        f"Regnskapslinjer totalt: {len(non_sum)}",
        f"Vesentlige: {vesentlige}",
        f"Moderate: {moderate}",
        f"Ikke vesentlige: {ikke_ves}",
        f"Scopet inn: {scoped_inn}",
        f"Scopet ut: {scoped_ut_n}",
        f"Ikke besluttet: {ikke_besluttet}",
    ]
    for txt in summary_lines:
        ws.cell(row=row, column=1, value=txt).font = _FONT_NORMAL
        row += 1

    wb.save(str(path))
    return path
