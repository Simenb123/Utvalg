"""analyse_regnskapsoppstilling_excel.py — Excel-eksport av regnskapsoppstilling.

Ark 1: Regnskapsoppstilling (alle RL med IB, Endring, UB, UB i fjor,
       Endring vs i fjor kr, Endring vs i fjor %, Antall poster)
       + Nøkkeltall (lønnsomhet, likviditet, soliditet, effektivitet)
         plassert 3 rader under regnskapsoppstillingen
Ark 2: Beregningsgrunnlag (formelreferanse for alle nøkkeltall)
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import vaak_excel_theme as vxt
import vaak_tokens as vt
from workpaper_forside import build_forside_sheet


_TITLE_FILL     = vxt.FILL_TITLE
_SUBTITLE_FILL  = vxt.FILL_NEUTRAL
_HEADER_FILL    = vxt.FILL_SUBHEADER
_SUM_FILL       = vxt.FILL_SUMLINE
_SUM_MAJOR_FILL = vxt.FILL_SUMLINE_MAJOR
_ZEBRA_FILL     = vxt.FILL_ZEBRA
_CAT_FILL       = vxt.FILL_HEADER
_NK_HEADER_FILL = vxt.FILL_TITLE
_POS_FONT       = vxt.FONT_POS
_NEG_FONT       = vxt.FONT_NEG
_THIN_SIDE      = Side(style="thin", color=vt.BORDER_SOFT)
_MEDIUM_SIDE    = Side(style="medium", color=vt.FOREST)
_BORDER         = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                         top=_THIN_SIDE, bottom=_THIN_SIDE)
_BORDER_SUM_TOP = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                         top=_MEDIUM_SIDE, bottom=_THIN_SIDE)
_AMOUNT_FMT   = '#,##0;[Red]-#,##0'
_INT_FMT      = '#,##0'
_PCT_FMT      = '0.0"%"'
_DECIMAL_FMT  = '0.00'

# Sumpost-nr som alltid markeres med ekstra uthevning
_MAJOR_SUM_REGNR = {80, 160, 280, 665, 715, 820, 850}


def _active_col_specs(rl_df: pd.DataFrame, *, include_antall: bool = True
                      ) -> list[tuple[str, str, str, float]]:
    """Bygg kolonnespec dynamisk etter hva som faktisk finnes i rl_df.

    Returnerer liste med (df-kolonne, header-tekst, formattype, kolonnebredde).
    Format-typer: 'int', 'text', 'amount', 'pct'.

    IB og periodens Bevegelse er bevisst utelatt — fokus i en
    regnskapsoppstilling er UB pr. periode + sammenligning mot fjor.
    """
    specs: list[tuple[str, str, str, float]] = [
        ("regnr",          "Nr",                  "int",    7.0),
        ("regnskapslinje", "Regnskapslinje",       "text",  42.0),
        ("UB",             "I år",                "amount", 17.0),
    ]
    if "UB_fjor" in rl_df.columns:
        specs.append(("UB_fjor",        "I fjor",      "amount", 17.0))
        specs.append(("_endr_fjor_kr",  "Endring",     "amount", 17.0))
        specs.append(("_endr_fjor_pct", "Endring (%)", "pct",    14.0))
    if include_antall and "Antall" in rl_df.columns:
        specs.append(("Antall",     "Antall poster",       "int",    13.0))
    return specs


def build_regnskapsoppstilling_workbook(
    rl_df: pd.DataFrame,
    *,
    regnskapslinjer: Optional[pd.DataFrame] = None,
    transactions_df: Optional[pd.DataFrame] = None,
    client: str | None = None,
    year: str | int | None = None,
) -> Workbook:
    """Bygg Excel-arbeidsbok med regnskapsoppstilling, nøkkeltall og beregningsgrunnlag.

    Ark 1: Regnskapsoppstilling + nøkkeltall under.
    Ark 2: Beregningsgrunnlag (formelreferanse).
    """
    wb = Workbook()

    # --- Ark 1: Regnskapsoppstilling + Nøkkeltall ---
    _build_regnskapsoppstilling_sheet(wb, rl_df, regnskapslinjer=regnskapslinjer,
                                     client=client, year=year)

    # --- Ark 2: Beregningsgrunnlag ---
    _build_formula_sheet(wb)

    build_forside_sheet(wb, workpaper_navn="Regnskapsoppstilling")

    return wb


# ---------------------------------------------------------------------------
# Ark 1: Regnskapsoppstilling + Nøkkeltall
# ---------------------------------------------------------------------------

def _build_regnskapsoppstilling_sheet(
    wb: Workbook,
    rl_df: pd.DataFrame,
    *,
    regnskapslinjer: Optional[pd.DataFrame] = None,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.active
    ws.title = "Regnskapsoppstilling"

    col_specs = _active_col_specs(rl_df)
    n_cols = len(col_specs)
    last_col_letter = get_column_letter(n_cols)

    # --- Tittelbanner ---
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "Regnskapsoppstilling"
    ws["A1"].font = Font(name=vt.FONT_FAMILY_BODY, size=16, bold=True, color=vt.TEXT_PRIMARY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["A1"].fill = _TITLE_FILL
    ws.row_dimensions[1].height = 32

    # Undertittel: klient / år / generert — lysere bakgrunn
    sub_parts: list[str] = []
    if client:
        sub_parts.append(str(client))
    if year not in {None, ""}:
        sub_parts.append(str(year))
    sub_parts.append(f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = "   ·   ".join(sub_parts)
    ws["A2"].font = Font(name=vt.FONT_FAMILY_BODY, size=10, color=vt.TEXT_MUTED)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["A2"].fill = _SUBTITLE_FILL
    ws.row_dimensions[2].height = 22

    # --- Kolonneoverskrifter ---
    header_row = 4
    for col_idx, (_, header, _, _) in enumerate(col_specs, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(name=vt.FONT_FAMILY_BODY, bold=True, size=10, color=vt.TEXT_PRIMARY)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        # Tallkolonner høyre-justert, tekst venstre
        align_h = "left" if header in {"Nr", "Regnskapslinje"} else "right"
        cell.alignment = Alignment(
            horizontal=align_h,
            vertical="center",
            wrap_text=True,
            indent=1 if align_h == "left" else 0,
        )
    ws.row_dimensions[header_row].height = 30

    # --- Datalinjer ---
    sum_regnr = _sumline_regnr(regnskapslinjer)
    has_fjor = "UB_fjor" in rl_df.columns
    data_row = header_row + 1
    zebra_toggle = False

    for _, row in rl_df.iterrows():
        regnr = _safe_int(row.get("regnr"))
        is_major = regnr in _MAJOR_SUM_REGNR
        is_sum   = regnr in sum_regnr

        # Beregn endring vs i fjor
        endr_kr = None
        endr_pct = None
        if has_fjor:
            ub_val = _safe_float(row.get("UB"))
            fjor_val = _safe_float(row.get("UB_fjor"))
            if ub_val != 0.0 or fjor_val != 0.0:
                endr_kr = ub_val - fjor_val
                if abs(fjor_val) > 1e-9:
                    endr_pct = (endr_kr / abs(fjor_val)) * 100

        # Zebra-stripe brukes kun på ikke-sumrader
        use_zebra = (not is_sum and not is_major) and zebra_toggle
        if not is_sum and not is_major:
            zebra_toggle = not zebra_toggle
        else:
            zebra_toggle = False  # reset etter sumrad så neste datarad starter frisk

        # Sumrader (spesielt major) skal ha tydelig topp-border
        cell_border = _BORDER_SUM_TOP if is_major else _BORDER

        for col_idx, (df_col, _, fmt_type, _) in enumerate(col_specs, start=1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.border = cell_border

            # Hent verdi — spesialbehandling for beregnede kolonner
            if df_col == "_endr_fjor_kr":
                raw = endr_kr
            elif df_col == "_endr_fjor_pct":
                raw = endr_pct
            else:
                raw = row.get(df_col) if df_col in row.index else None

            if fmt_type == "int":
                cell.value = _safe_int(raw)
                cell.number_format = _INT_FMT
                cell.alignment = Alignment(horizontal="right", indent=1)
            elif fmt_type == "amount":
                val = _safe_float(raw)
                cell.value = val if val != 0.0 else None
                cell.number_format = _AMOUNT_FMT
                cell.alignment = Alignment(horizontal="right", indent=1)
            elif fmt_type == "pct":
                val = _safe_float(raw)
                cell.value = val if val != 0.0 else None
                cell.number_format = _PCT_FMT
                cell.alignment = Alignment(horizontal="right", indent=1)
                # Fargelegg retning — grønn = vekst, rød = nedgang
                if val > 0.05:
                    cell.font = _POS_FONT
                elif val < -0.05:
                    cell.font = _NEG_FONT
            else:  # text
                cell.value = str(raw or "")
                cell.alignment = Alignment(horizontal="left", indent=1)

            if is_major:
                cell.font = Font(name=vt.FONT_FAMILY_BODY, bold=True, size=11, color=vt.TEXT_PRIMARY)
                cell.fill = _SUM_MAJOR_FILL
            elif is_sum:
                cell.font = Font(name=vt.FONT_FAMILY_BODY, bold=True, size=10, color=vt.TEXT_PRIMARY)
                cell.fill = _SUM_FILL
            elif use_zebra:
                cell.fill = _ZEBRA_FILL

        # Litt mer luft — høyere rad for sumposter
        ws.row_dimensions[data_row].height = 20 if is_major else (18 if is_sum else 16)
        data_row += 1

    # --- Formatering ---
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{max(data_row - 1, header_row)}"
    ws.sheet_view.showGridLines = True

    for col_idx, (_, _, _, width) in enumerate(col_specs, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Nøkkeltall under regnskapsoppstillingen ---
    _append_nokkeltall_section(ws, rl_df, start_row=data_row + 3,
                               client=client, year=year)


# ---------------------------------------------------------------------------
# Nøkkeltall-seksjon (på samme ark under regnskapsoppstillingen)
# ---------------------------------------------------------------------------

def _append_nokkeltall_section(
    ws,
    rl_df: pd.DataFrame,
    *,
    start_row: int,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    """Legg nøkkeltall inn på arket, startende fra start_row."""
    try:
        from nokkeltall_engine import compute_nokkeltall
    except ImportError:
        return

    result = compute_nokkeltall(rl_df, client=client or "", year=year or "")
    if not result.metrics:
        return

    row = start_row

    # Kolonne-layout matcher regnskapsoppstillingen:
    #   col 1: (tom, Nr brukes ikke for nøkkeltall)
    #   col 2: Nøkkeltall
    #   col 3: I år
    #   col 4: I fjor
    #   col 5: Endring
    #   col 6: Endring (%)

    # --- Seksjonstittel ---
    ws.merge_cells(f"A{row}:F{row}")
    title_cell = ws.cell(row=row, column=1, value="Nøkkeltall")
    title_cell.font = Font(name=vt.FONT_FAMILY_BODY, size=13, bold=True, color=vt.TEXT_PRIMARY)
    title_cell.fill = _NK_HEADER_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26
    row += 1

    # --- Header ---
    nk_headers = ["", "Nøkkeltall", "I år", "I fjor", "Endring", "Endring (%)"]
    for col_idx, h in enumerate(nk_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(name=vt.FONT_FAMILY_BODY, bold=True, size=10, color=vt.TEXT_PRIMARY)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        align_h = "left" if col_idx <= 2 else "right"
        cell.alignment = Alignment(
            horizontal=align_h, vertical="center",
            indent=1 if align_h == "left" else 0,
        )
    ws.row_dimensions[row].height = 20
    row += 1

    # --- Nøkkeltall-rader ---
    current_cat = ""

    for m in result.metrics:
        if m.value is None:
            continue

        # Kategori-header
        if m.category != current_cat:
            current_cat = m.category
            cat_cell = ws.cell(row=row, column=2, value=current_cat)
            cat_cell.font = Font(name=vt.FONT_FAMILY_BODY, bold=True, size=10, color=vt.FOREST)
            cat_cell.fill = _CAT_FILL
            cat_cell.border = _BORDER
            for ci in (1, 3, 4, 5, 6):
                c = ws.cell(row=row, column=ci)
                c.fill = _CAT_FILL
                c.border = _BORDER
            row += 1

        # Label (kol 2)
        ws.cell(row=row, column=2, value=m.label).border = _BORDER

        # I år (kol 3)
        fmt = _excel_fmt(m.fmt)
        val_cell = ws.cell(row=row, column=3, value=_display_val(m.value, m.fmt))
        val_cell.number_format = fmt
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.border = _BORDER

        # I fjor (kol 4)
        prev_cell = ws.cell(row=row, column=4,
                            value=_display_val(m.prev_value, m.fmt) if m.prev_value is not None else None)
        prev_cell.number_format = fmt
        prev_cell.alignment = Alignment(horizontal="right")
        prev_cell.border = _BORDER

        # Endring (kol 5) — prosentpoeng for %-nøkkeltall, absolutt for beløp/desimal
        endr_cell = ws.cell(row=row, column=5)
        endr_cell.alignment = Alignment(horizontal="right")
        endr_cell.border = _BORDER
        if m.prev_value is not None:
            diff = m.value - m.prev_value
            if m.fmt == "pct":
                endr_cell.value = diff
                endr_cell.number_format = '0.0" pp"'
            elif m.fmt == "decimal":
                endr_cell.value = diff
                endr_cell.number_format = _DECIMAL_FMT
            elif m.fmt == "amount":
                endr_cell.value = diff
                endr_cell.number_format = _AMOUNT_FMT

        # Endring (%) (kol 6)
        pct_cell = ws.cell(row=row, column=6)
        pct_cell.alignment = Alignment(horizontal="right", indent=1)
        pct_cell.border = _BORDER
        if m.change_pct is not None:
            pct_cell.value = m.change_pct
            pct_cell.number_format = _PCT_FMT
            if m.change_pct > 0.05:
                pct_cell.font = _POS_FONT
            elif m.change_pct < -0.05:
                pct_cell.font = _NEG_FONT

        row += 1


# ---------------------------------------------------------------------------
# Ark 2: Beregningsgrunnlag
# ---------------------------------------------------------------------------

_FORMULA_REF = [
    ("Lønnsomhet", [
        ("Bruttofortjeneste", "(Salgsinntekt \u2212 Varekostnad) / Salgsinntekt \u00d7 100", "RL 10, 20"),
        ("Driftsmargin", "Driftsresultat / Sum driftsinntekter \u00d7 100", "RL 80, 19"),
        ("Nettoresultatmargin", "\u00c5rsresultat / Sum driftsinntekter \u00d7 100", "RL 280, 19"),
        ("EBITDA-margin", "(Driftsinntekter \u2212 (Driftskostnader \u2212 Avskrivning)) / Driftsinntekter \u00d7 100", "RL 19, 79, 50"),
        ("Resultat f\u00f8r skatt i % av inntekter", "Resultat f\u00f8r skattekostnad / Sum driftsinntekter \u00d7 100", "RL 160, 19"),
    ]),
    ("Likviditet", [
        ("Likviditetsgrad 1", "Sum oml\u00f8psmidler / Sum kortsiktig gjeld", "RL 660, 810"),
        ("Likviditetsgrad 2", "(Sum oml\u00f8psmidler \u2212 Varelager) / Sum kortsiktig gjeld", "RL 660, 605, 810"),
        ("Arbeidskapital", "Sum oml\u00f8psmidler \u2212 Sum kortsiktig gjeld", "RL 660, 810"),
    ]),
    ("Soliditet", [
        ("Egenkapitalandel", "Sum egenkapital / Sum eiendeler \u00d7 100", "RL 715, 665"),
        ("Gjeldsgrad", "Sum gjeld / Sum egenkapital", "RL 820, 715"),
    ]),
    ("Effektivitet", [
        ("Kundefordringer i % av salg", "Kundefordringer / Salgsinntekt \u00d7 100", "RL 610, 10"),
        ("Varelager i % av varekostnad", "Varelager / Varekostnad \u00d7 100", "RL 605, 20"),
        ("Leverand\u00f8rgjeld i % av driftskostnader", "Leverand\u00f8rgjeld / (Varekostnad + Annen driftskostnad) \u00d7 100", "RL 780, 20, 70"),
        ("L\u00f8nnskostnad i % av driftsinntekter", "L\u00f8nnskostnad / Sum driftsinntekter \u00d7 100", "RL 40, 19"),
        ("Annen driftskostnad i % av driftsinntekter", "Annen driftskostnad / Sum driftsinntekter \u00d7 100", "RL 70, 19"),
    ]),
]


def _build_formula_sheet(wb: Workbook) -> None:
    """Bygg beregningsgrunnlag-ark med formelreferanser for alle nøkkeltall."""
    ws = wb.create_sheet("Beregningsgrunnlag")

    n_cols = 3
    last_col = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "Beregningsgrunnlag — formler og regnskapslinjer"
    ws["A1"].font = Font(name=vt.FONT_FAMILY_BODY, size=16, bold=True, color=vt.TEXT_PRIMARY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["A1"].fill = _TITLE_FILL
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = "Alle beløp hentes fra UB (utgående balanse / akkumulert).   ·   RL = regnskapslinjenummer."
    ws["A2"].font = Font(name=vt.FONT_FAMILY_BODY, size=10, color=vt.TEXT_MUTED)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws["A2"].fill = _SUBTITLE_FILL
    ws.row_dimensions[2].height = 22

    # Header
    header_row = 4
    for col_idx, (h, w) in enumerate([("Nøkkeltall", 38.0), ("Formel", 58.0), ("Regnskapslinjer", 20.0)], start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(name=vt.FONT_FAMILY_BODY, bold=True, size=10, color=vt.TEXT_PRIMARY)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[header_row].height = 20

    # Data
    row = header_row + 1
    for cat, items in _FORMULA_REF:
        # Kategori-header
        cat_cell = ws.cell(row=row, column=1, value=cat)
        cat_cell.font = Font(bold=True, size=10, color="4472C4")
        cat_cell.fill = _CAT_FILL
        cat_cell.border = _BORDER
        for ci in range(2, n_cols + 1):
            c = ws.cell(row=row, column=ci)
            c.fill = _CAT_FILL
            c.border = _BORDER
        row += 1

        for label, formula, rl in items:
            ws.cell(row=row, column=1, value=label).border = _BORDER
            formula_cell = ws.cell(row=row, column=2, value=formula)
            formula_cell.border = _BORDER
            formula_cell.font = Font(size=10, name=vt.FONT_FAMILY_MONO, color=vt.TEXT_PRIMARY)
            rl_cell = ws.cell(row=row, column=3, value=rl)
            rl_cell.border = _BORDER
            rl_cell.alignment = Alignment(horizontal="left")
            row += 1

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Hjelpefunksjoner
# ---------------------------------------------------------------------------

def _display_val(value: float | None, fmt: str) -> float | None:
    """Konverter verdi for Excel-visning (beløp i hele kroner)."""
    if value is None:
        return None
    if fmt == "amount":
        return round(value)
    return value


def _excel_fmt(fmt: str) -> str:
    """Map nøkkeltall-format til Excel tallformat."""
    if fmt == "pct":
        return '0.0"%"'
    if fmt == "decimal":
        return _DECIMAL_FMT
    if fmt == "amount":
        return _AMOUNT_FMT
    return "0"


def save_regnskapsoppstilling_workbook(
    path: str | Path,
    *,
    rl_df: pd.DataFrame,
    regnskapslinjer: Optional[pd.DataFrame] = None,
    transactions_df: Optional[pd.DataFrame] = None,
    client: str | None = None,
    year: str | int | None = None,
) -> str:
    out = Path(path)
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_regnskapsoppstilling_workbook(
        rl_df,
        regnskapslinjer=regnskapslinjer,
        transactions_df=transactions_df,
        client=client,
        year=year,
    )
    wb.save(out)
    return str(out)


def _sumline_regnr(regnskapslinjer: Optional[pd.DataFrame]) -> set[int]:
    if regnskapslinjer is None or regnskapslinjer.empty:
        return set()
    try:
        from regnskap_mapping import normalize_regnskapslinjer
        regn = normalize_regnskapslinjer(regnskapslinjer)
        return {int(v) for v in regn.loc[regn["sumpost"], "regnr"].astype(int).tolist()}
    except Exception:
        return set()


def _safe_float(value: object) -> float:
    try:
        v = float(value)  # type: ignore[arg-type]
        return v if v == v else 0.0   # NaN → 0
    except Exception:
        return 0.0


def _safe_int(value: object) -> int:
    try:
        return int(float(value))  # type: ignore[arg-type]
    except Exception:
        return 0
