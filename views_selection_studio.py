from __future__ import annotations
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from typing import Optional
from .stratifiering import beregn_strata, trekk_sample, summer_per_bilag

class SelectionStudio(tk.Toplevel):
    """
    Veiviser for delutvalg/stratifisering.
    Krever df_base (grunnlag) og valgfritt df_all (for totalsummer).
    """

    def __init__(self, master, df: pd.DataFrame, on_commit=None, df_all: Optional[pd.DataFrame] = None) -> None:
        super().__init__(master)
        self.title("Delutvalg og stratifisering")
        self.geometry("1200x750")
        self.on_commit = on_commit

        try:
            style = ttk.Style(self)
            style.configure("small.TLabel", font=("Segoe UI", 8), wraplength=160)
        except Exception:
            pass

        # Data
        self.df_base = df.copy()
        self.df_all = df_all.copy() if df_all is not None else None
        self.df_work = self.df_base.copy()

        # Interne variabler
        self.summary: Optional[pd.DataFrame] = None
        self.bilag_df: Optional[pd.DataFrame] = None
        self.interval_map: dict[int, str] = {}
        self.custom_counts: dict[int, int] = {}
        self.sample_df: pd.DataFrame = pd.DataFrame()

        # UI-variabler
        self.var_dir = tk.StringVar(value="Alle")
        self.var_min = tk.StringVar(value="")
        self.var_max = tk.StringVar(value="")
        self.var_abs = tk.BooleanVar(value=True)
        self.var_mode = tk.StringVar(value="quantile")
        self.var_k = tk.IntVar(value=5)
        self.var_n_per = tk.IntVar(value=5)
        self.var_total = tk.IntVar(value=0)
        self.var_auto = tk.BooleanVar(value=False)
        self.var_show_sum_bilag_int = tk.BooleanVar(value=False)
        self.var_show_sum_rows_int = tk.BooleanVar(value=False)
        self.var_show_sum_bilag_all = tk.BooleanVar(value=False)

        # Bygg UI
        self._build_ui()
        # Filtrer grunnlag første gang
        self._apply_filters()
        self._update_summary()
        self._update_sample_view()

    # ---------------- UI ----------------
    def _build_ui(self) -> None:
        # Venstre panel (kontroller)
        left = ttk.Frame(self); left.pack(side="left", fill="y", padx=6, pady=6)
        right = ttk.Frame(self); right.pack(side="left", fill="both", expand=True, padx=6, pady=6)

        # Filtrering (Steg 1)
        ttk.Label(left, text="Steg 1: Filtre (grunnlag)", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0,2))
        ttk.Label(left, text="Retning:").pack(anchor="w")
        ttk.Combobox(left, values=("Alle","Debet","Kredit"),
                     textvariable=self.var_dir, state="readonly", width=10).pack(anchor="w")
        ttk.Label(left, text="Beløp fra/til:").pack(anchor="w", pady=(6,0))
        row = ttk.Frame(left); row.pack(anchor="w")
        ttk.Entry(row, textvariable=self.var_min, width=8).pack(side="left")
        ttk.Label(row, text=" til ").pack(side="left")
        ttk.Entry(row, textvariable=self.var_max, width=8).pack(side="left")
        ttk.Checkbutton(left, text="Bruk absolutt beløp", variable=self.var_abs).pack(anchor="w", pady=(4,4))
        ttk.Button(left, text="Oppdater grunnlag", command=self._on_update_filters).pack(anchor="w")

        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=8)

        # Stratifisering (Steg 2)
        ttk.Label(left, text="Steg 2: Stratifisering", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0,2))
        ttk.Combobox(left, values=("quantile","equal"), textvariable=self.var_mode,
                     state="readonly", width=12).pack(anchor="w")
        ttk.Label(
            left,
            text="quantile deler bilagene i like store\n"
                 "grupper basert på beløpssum.\n"
                 "equal deler beløpsområdet i like intervaller.",
            style="small.TLabel"
        ).pack(anchor="w", pady=(2,4))
        ttk.Label(left, text="Antall grupper (k):").pack(anchor="w")
        ttk.Spinbox(left, from_=2, to=50, textvariable=self.var_k, width=6).pack(anchor="w")

        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=8)

        # Trekk (Steg 3)
        ttk.Label(left, text="Steg 3: Trekk", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0,2))
        ttk.Label(left, text="Antall bilag per gruppe:").pack(anchor="w")
        ttk.Spinbox(left, from_=0, to=1000, textvariable=self.var_n_per, width=6).pack(anchor="w")
        ttk.Label(left, text="Totalt antall bilag i utvalg:").pack(anchor="w", pady=(4,0))
        ttk.Spinbox(left, from_=0, to=10000, textvariable=self.var_total, width=8).pack(anchor="w")
        ttk.Checkbutton(left, text="Auto-fordel etter sumandel", variable=self.var_auto).pack(anchor="w", pady=(2,2))
        ttk.Button(left, text="Tilpass per gruppe", command=self._custom_counts).pack(anchor="w")

        ttk.Separator(left, orient="horizontal"). pack(fill="x", pady=8)

        # Visningsalternativer
        ttk.Label(left, text="Visningsalternativer", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0,2))
        ttk.Checkbutton(left, text="Sum bilag (kontointervallet)",
                        variable=self.var_show_sum_bilag_int,
                        command=self._update_sample_view).pack(anchor="w")
        ttk.Checkbutton(left, text="Sum rader (kontointervallet)",
                        variable=self.var_show_sum_rows_int,
                        command=self._update_sample_view).pack(anchor="w")
        ttk.Checkbutton(left, text="Sum bilag (alle kontoer)",
                        variable=self.var_show_sum_bilag_all,
                        command=self._update_sample_view).pack(anchor="w")

        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=8)

        # Handlingsknapper
        btn_row = ttk.Frame(left); btn_row.pack(anchor="w", pady=(0,0))
        ttk.Button(btn_row, text="Generer grupper", command=self._on_build_groups).pack(side="left")
        ttk.Button(btn_row, text="Trekk utvalg", command=self._on_draw_sample).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Legg i utvalg", command=self._on_commit).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Eksporter Excel", command=self._on_export).pack(side="left", padx=4)

        # Høyresiden: oversikt og tabeller
        self.lbl_summary = ttk.Label(
            right,
            text="Grunnlag: 0 rader | Sum: 0,00 (Fjernet: 0 rader, 0,00)"
        )
        self.lbl_summary.pack(anchor="w")
        ttk.Label(right, text="Grupper (strata)").pack(anchor="w", pady=(4,0))
        self.tree_strata = ttk.Treeview(
            right,
            columns=(
                "Gruppe","Antall bilag","Sum Beløp",
                "Min Beløp","Median Beløp","Maks Beløp","Intervall"
            ),
            show="headings",
            height=6
        )
        for c, w in zip(
            self.tree_strata["columns"],
            (70,80,120,100,100,100,220)
        ):
            self.tree_strata.heading(c, text=c)
            self.tree_strata.column(c, width=w, stretch=(c == "Intervall"))
        self.tree_strata.pack(fill="x")
        ttk.Label(right, text="Trekk (sample)").pack(anchor="w", pady=(8,0))
        self.tree_sample = ttk.Treeview(right, show="headings")
        self.tree_sample.pack(fill="both", expand=True)

    # ------------- Filtre -------------
    def _on_update_filters(self) -> None:
        self._apply_filters()
        self._update_summary()

    def _apply_filters(self) -> None:
        df = self.df_base.copy()
        bel = pd.to_numeric(
            df.get("Beløp", pd.Series(dtype="float64")),
            errors="coerce"
        ).fillna(0.0)
        direction = self.var_dir.get()
        if direction == "Debet":
            df = df[bel > 0]; bel = bel.loc[df.index]
        elif direction == "Kredit":
            df = df[pd.to_numeric(df["Beløp"], errors="coerce") < 0]
            bel = bel.loc[df.index]
        bel_for_limit = bel.abs() if self.var_abs.get() else bel
        try:
            min_val = float(
                str(self.var_min.get()).replace(" ","").replace(",",".")
            ) if self.var_min.get() else float("nan")
        except Exception:
            min_val = float("nan")
        try:
            max_val = float(
                str(self.var_max.get()).replace(" ","").replace(",",".")
            ) if self.var_max.get() else float("nan")
        except Exception:
            max_val = float("nan")
        if not pd.isna(min_val):
            df = df[bel_for_limit >= min_val]
            bel_for_limit = bel_for_limit.loc[df.index]
        if not pd.isna(max_val):
            df = df[bel_for_limit <= max_val]
            bel_for_limit = bel_for_limit.loc[df.index]
        self.df_work = df

    def _update_summary(self) -> None:
        N = len(self.df_work)
        S = pd.to_numeric(
            self.df_work.get("Beløp", pd.Series(dtype="float64")),
            errors="coerce"
        ).fillna(0.0).sum()
        N0 = len(self.df_base)
        S0 = pd.to_numeric(
            self.df_base.get("Beløp", pd.Series(dtype="float64")),
            errors="coerce"
        ).fillna(0.0).sum()
        removed_n = N0 - N; removed_s = S0 - S
        txt = (
            f"Grunnlag: {N:,} rader | Sum: {S:,.2f} "
            f"(Fjernet: {removed_n:,} rader, {removed_s:,.2f})"
        ).replace(",", " ").replace(".", ",")
        self.lbl_summary.config(text=txt)

    # ------------- Grupper (generer) -------------
    def _on_build_groups(self) -> None:
        if self.df_work.empty:
            messagebox.showinfo("Stratifisering", "Ingen rader i grunnlaget. Oppdater filter først.")
            return
        try:
            summary, bilag_df, interval_map = beregn_strata(
                self.df_work,
                k=self.var_k.get(),
                mode=self.var_mode.get(),
                abs_belop=self.var_abs.get()
            )
            self.summary, self.bilag_df, self.interval_map = summary, bilag_df, interval_map
            self.custom_counts = {}
            # Oppdater treet
            self.tree_strata.delete(*self.tree_strata.get_children())
            for _, row in summary.iterrows():
                self.tree_strata.insert("", "end", values=[
                    row["Gruppe"],
                    row["Antall_bilag"],
                    f"{row['SumBeløp']:.2f}".replace(".", ","),
                    f"{row['Min_Beløp']:.2f}".replace(".", ","),
                    f"{row['Median_Beløp']:.2f}".replace(".", ","),
                    f"{row['Max_Beløp']:.2f}".replace(".", ","),
                    row["Intervall"]
                ])
            self.sample_df = pd.DataFrame()
            self._update_sample_view()
        except Exception as e:
            messagebox.showerror("Stratifisering", f"Feil ved beregning av grupper: {e}")

    # ------------- Tilpass antall -------------
    def _custom_counts(self) -> None:
        if self.summary is None or self.summary.empty:
            messagebox.showinfo("Tilpass", "Generer grupper først.")
            return
        top = tk.Toplevel(self)
        top.title("Tilpass antall bilag per gruppe")
        frm = ttk.Frame(top, padding=8); frm.pack(fill="both", expand=True)
        ttk.Label(
            frm,
            text="Antall bilag per gruppe",
            font=("Segoe UI", 10, "bold")
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0,6))
        ttk.Label(frm, text="Gruppe").grid(row=1, column=0, sticky="w")
        ttk.Label(frm, text="Maks").grid(row=1, column=1, sticky="w")
        ttk.Label(frm, text="Antall").grid(row=1, column=2, sticky="w")
        vars_dict = {}
        i = 2
        for _, row in self.summary.iterrows():
            g = int(row["Gruppe"])
            max_n = int(row["Antall_bilag"])
            ttk.Label(frm, text=str(g)).grid(row=i, column=0, sticky="w")
            ttk.Label(frm, text=str(max_n)).grid(row=i, column=1, sticky="w")
            v = tk.IntVar(value=min(self.var_n_per.get(), max_n))
            ttk.Spinbox(frm, from_=0, to=max_n, textvariable=v, width=6).grid(row=i, column=2, sticky="w")
            vars_dict[g] = v
            i += 1

        def apply_counts() -> None:
            self.custom_counts = {g: max(0, v.get()) for g, v in vars_dict.items()}
            top.destroy()

        ttk.Button(frm, text="Bruk", command=apply_counts).grid(row=i, column=0, pady=(8,0))
        ttk.Button(frm, text="Avbryt", command=top.destroy).grid(row=i, column=2, pady=(8,0))

    # ------------- Trekk sample -------------
    def _on_draw_sample(self) -> None:
        if self.summary is None or self.summary.empty:
            messagebox.showinfo("Trekk", "Generer grupper først.")
            return
        try:
            selected_bilags = trekk_sample(
                self.bilag_df,
                self.summary,
                custom_counts=self.custom_counts if self.custom_counts else None,
                n_per_group=self.var_n_per.get(),
                total_n=self.var_total.get(),
                auto_fordel=self.var_auto.get()
            )
            if selected_bilags:
                sample_df = self.df_base[self.df_base["Bilag"].isin(selected_bilags)].copy()
                sample_df = sample_df.merge(
                    self.bilag_df[["Bilag","__grp__"]],
                    on="Bilag",
                    how="left"
                )
                sample_df["Stratum"] = sample_df["__grp__"].map(self.interval_map)
                sample_df.drop(columns=["__grp__"], inplace=True)
            else:
                sample_df = pd.DataFrame()
            if not sample_df.empty:
                sums_df = summer_per_bilag(self.df_base, self.df_all, selected_bilags)
                sample_df = sample_df.merge(sums_df, on="Bilag", how="left")
            self.sample_df = sample_df
            self._update_sample_view()
        except Exception as e:
            messagebox.showerror("Trekk", f"Feil under trekking: {e}")

    # ------------- Oppdater samplevisning -------------
    def _update_sample_view(self) -> None:
        self.tree_sample.delete(*self.tree_sample.get_children())
        if self.sample_df is None or self.sample_df.empty:
            self.tree_sample["columns"] = ()
            return
        base_order = ["Bilag","Konto","Kontonavn","Dato","Beløp","Tekst"]
        cols = [c for c in base_order if c in self.sample_df.columns]
        viscols = ["Bilag"] + [c for c in cols if c != "Bilag"]
        # Valgfri visning av summekolonner
        if self.var_show_sum_bilag_int.get() and "Sum bilag (kontointervallet)" in self.sample_df.columns:
            viscols.append("Sum bilag (kontointervallet)")
        if self.var_show_sum_rows_int.get() and "Sum rader (kontointervallet)" in self.sample_df.columns:
            viscols.append("Sum rader (kontointervallet)")
        if self.var_show_sum_bilag_all.get() and "Sum bilag (alle kontoer)" in self.sample_df.columns:
            viscols.append("Sum bilag (alle kontoer)")
        if "Stratum" in self.sample_df.columns:
            viscols.append("Stratum")
        self.tree_sample["columns"] = tuple(viscols)
        for c in viscols:
            self.tree_sample.heading(c, text=c)
            w = 120
            if c == "Tekst": w = 200
            if c.startswith("Sum"): w = 140
            self.tree_sample.column(c, width=w, stretch=True)
        for _, row in self.sample_df.iterrows():
            values = [row.get(c, "") for c in viscols]
            self.tree_sample.insert("", "end", values=values)

    # ------------- Legg i utvalg -------------
    def _on_commit(self) -> None:
        if self.on_commit and not self.sample_df.empty:
            self.on_commit(self.sample_df)
            self.destroy()

    # ------------- Eksporter -------------
    def _on_export(self) -> None:
        if self.sample_df.empty:
            messagebox.showinfo("Eksport", "Ingen rader i utvalget.")
            return
        p = filedialog.asksaveasfilename(
            title="Lagre til Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel-filer","*.xlsx")]
        )
        if not p:
            return
        try:
            import openpyxl
        except Exception:
            messagebox.showerror("Eksport", "openpyxl er ikke installert.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            wb = Workbook()
            # Ark for grupper
            strata_df = (
                self.summary.copy()
                if self.summary is not None
                else pd.DataFrame(columns=self.tree_strata["columns"])
            )
            strata_df = strata_df.rename(columns={
                "Antall_bilag": "Antall bilag",
                "SumBeløp": "Sum Beløp",
                "Min_Beløp": "Min Beløp",
                "Median_Beløp": "Median Beløp",
                "Max_Beløp": "Maks Beløp"
            })
            ws = wb.active; ws.title = "Grupper"
            for r in dataframe_to_rows(strata_df, index=False, header=True):
                ws.append(r)
            # Ark for sample
            ws2 = wb.create_sheet("Trekk")
            for r in dataframe_to_rows(self.sample_df, index=False, header=True):
                ws2.append(r)
            # Ark for summer dersom valgt
            if "Sum bilag (kontointervallet)" in self.sample_df.columns:
                df = self.sample_df[["Bilag","Sum bilag (kontointervallet)"]].drop_duplicates()
                ws3 = wb.create_sheet("Sum bilag kontointervall")
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws3.append(r)
            if "Sum rader (kontointervallet)" in self.sample_df.columns:
                df = self.sample_df[["Bilag","Sum rader (kontointervallet)"]].drop_duplicates()
                ws4 = wb.create_sheet("Sum rader kontointervall")
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws4.append(r)
            if "Sum bilag (alle kontoer)" in self.sample_df.columns:
                df = self.sample_df[["Bilag","Sum bilag (alle kontoer)"]].drop_duplicates()
                ws5 = wb.create_sheet("Sum bilag alle kontoer")
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws5.append(r)
            wb.save(p)
            messagebox.showinfo("Eksport", f"Utvalget er lagret til {p}")
        except Exception as e:
            messagebox.showerror("Eksport", f"Feil under eksport: {e}")
