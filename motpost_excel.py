"""Motpost: Excel-export (openpyxl).

Flyttet ut fra motpost_konto_core.py for å holde koden mer oversiktlig og
redusere størrelsen på kjernelogikken (data/pivot).

Dette er i hovedsak "presentation"-logikk:
- Ark/kolonner/formatering
- Tabell-stiler
- Kombinasjonsfaner
"""

from __future__ import annotations

from typing import Iterable, Optional
import numbers

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from formatting import fmt_amount
from motpost_combinations import (
    build_motkonto_combinations,
    build_motkonto_combinations_per_selected_account,
)
from motpost_utils import _konto_str

# Import kun for type-hint / struktur (ingen GUI)
from motpost_konto_core import MotpostData


def _norm(s: str) -> str:
    return "".join(ch for ch in str(s).strip().lower() if ch.isalnum())


# -----------------------------
# Excel export (openpyxl)
# -----------------------------

# Basic styles used across sheets
_FILL_TITLE = PatternFill("solid", fgColor="C6EFCE")  # light green
_FILL_HEADER = PatternFill("solid", fgColor="BDD7EE")  # light blue
_FILL_INFO = PatternFill("solid", fgColor="FCE4D6")  # light orange

_BORDER_THIN = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)


def _set_cell(ws, row: int, col: int, value, *, bold: bool = False, fill=None, number_format: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    if bold:
        cell.font = Font(bold=True)
    if fill is not None:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    cell.border = _BORDER_THIN
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    return cell


def _autosize_columns(ws, min_width: int = 10, max_width: int = 55):
    """Grov auto-width basert på max len i kolonnen."""

    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for c in col_cells:
            v = c.value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def _apply_table_style(
    ws,
    header_row: int,
    last_row: int,
    *,
    money_headers: Optional[set[str]] = None,
    percent_headers: Optional[set[str]] = None,
):
    """Gjør området om til en Excel-tabell og legger på format.

    money_headers/percent_headers angir kolonner (header-tekst) som skal ha
    passende number_format.
    """

    if money_headers is None:
        money_headers = set()
    if percent_headers is None:
        percent_headers = set()

    # Header styling
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = _FILL_HEADER
        c.font = Font(bold=True)
        c.border = _BORDER_THIN
        c.alignment = Alignment(vertical="top", wrap_text=True)

    # Number formats
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    for col, h in enumerate(headers, start=1):
        if not h:
            continue
        h_str = str(h)
        if h_str in money_headers:
            for r in range(header_row + 1, last_row + 1):
                ws.cell(row=r, column=col).number_format = "#,##0.00"
        elif h_str in percent_headers:
            for r in range(header_row + 1, last_row + 1):
                ws.cell(row=r, column=col).number_format = "0.0%"

    # Create Excel Table
    start_cell = f"A{header_row}"
    end_cell = f"{get_column_letter(ws.max_column)}{last_row}"
    table = Table(displayName=f"T{ws.title.replace(' ', '')}", ref=f"{start_cell}:{end_cell}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def _write_df_table(ws, df: pd.DataFrame, *, title: str, summary: str, start_row: int = 1) -> tuple[int, int]:
    """Skriver et 'ark med tittel + sammendrag + tabell'.

    Returnerer (header_row, last_row).
    """

    # Tittel (rad 1)
    _set_cell(ws, start_row, 1, title, bold=True, fill=_FILL_TITLE)

    # Sammendrag (rad 2)
    _set_cell(ws, start_row + 1, 1, summary, fill=_FILL_INFO)

    # Tom rad (rad 3)
    start_table_row = start_row + 3

    if df is None or df.empty:
        _set_cell(ws, start_table_row, 1, "Ingen data")
        _autosize_columns(ws)
        return start_table_row, start_table_row

    # Skriv dataframe
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_table_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Grov type-basert formatting
            if r_idx > start_table_row and isinstance(value, numbers.Real):
                # money-like headers
                header = ws.cell(row=start_table_row, column=c_idx).value
                if header in {"Sum", "Beløp", "Sum valgte kontoer", "Motbeløp"}:
                    cell.number_format = "#,##0.00"

            # Date formatting: hvis header ligner "dato"
            if r_idx > start_table_row:
                header = ws.cell(row=start_table_row, column=c_idx).value
                if header and _norm(header) in {"dato", "transaksjonsdato", "bokforingsdato"}:
                    # Hvis openpyxl ser dette som datetime, gi norsk format
                    if hasattr(value, "strftime"):
                        cell.number_format = "dd.mm.yyyy"

    header_row = start_table_row
    last_row = start_table_row + len(df)

    # Tabell-style
    money_headers = {"Sum", "Beløp", "Sum valgte kontoer", "Motbeløp"}
    percent_headers = {"% andel", "% andel bilag"}
    _apply_table_style(ws, header_row, last_row, money_headers=money_headers, percent_headers=percent_headers)

    _autosize_columns(ws)
    return header_row, last_row


def build_motpost_excel_workbook(
    data: MotpostData,
    outlier_motkonto: Optional[Iterable[str]] = None,
    *,
    selected_motkonto: Optional[str] = None,
    df_details_view: Optional[pd.DataFrame] = None,
    outliers: Optional[Iterable[str]] = None,
    outlier_accounts: Optional[Iterable[str]] = None,
) -> Workbook:
    """Bygger openpyxl Workbook for motpostanalyse.

    Merk: signaturen er litt fleksibel (synonymer for outliers) for å være robust
    mot ulike kall i GUI/tester.
    """

    out_set: set[str] = set()
    for src in (outlier_motkonto, outliers, outlier_accounts):
        if src:
            out_set |= {_konto_str(x) for x in src}

    wb = Workbook()

    # Default sheet -> Oversikt
    ws_overview = wb.active
    ws_overview.title = "Oversikt"
    ws_overview["A1"] = "Oversikt"
    ws_overview["A1"].font = Font(bold=True)
    ws_overview["A3"] = "Denne arbeidsboken er generert fra motpostanalysen." 

    ws_overview["A5"] = "Fane"
    ws_overview["B5"] = "Beskrivelse"
    ws_overview["A5"].font = Font(bold=True)
    ws_overview["B5"].font = Font(bold=True)

    rows = [
        (
            "Motkonto",
            "Oppsummering av motkontoer (andre kontoer på samme bilag som valgte kontoer).",
        ),
        (
            "Kombinasjoner",
            "Oversikt over vanlige kombinasjoner av motkontoer per bilag (sett av motkontoer som forekommer sammen).",
        ),
        (
            "Kombinasjoner pr konto",
            "Samme som 'Kombinasjoner', men gruppert pr hver valgt konto.",
        ),
        (
            "Valgte kontoer",
            "Oppsummering av valgte kontoer (sum og andel).",
        ),
        (
            "Bilag",
            "Bilagsliste for valgt motkonto. Hvis ingen motkonto er valgt ved eksport, viser fanen alle bilag/motkontoer i grunnlaget.",
        ),
        (
            "Outliers",
            "Motkontoer som er markert som outliers (typisk gjenstand for testing).",
        ),
        (
            "OutlierBilag",
            "Alle transaksjoner på bilag som inneholder outlier-motkontoer.",
        ),
    ]
    r = 6
    for a, b in rows:
        ws_overview.cell(row=r, column=1, value=a)
        ws_overview.cell(row=r, column=2, value=b)
        r += 1
    _autosize_columns(ws_overview)

    # ------------------ Motkonto sheet ------------------
    ws_mot = wb.create_sheet("Motkonto")
    df_mot = data.df_motkonto.copy() if data.df_motkonto is not None else pd.DataFrame()
    summary_line = f"Valgte kontoer: {', '.join(data.selected_accounts)} | Bilag i grunnlag: {data.bilag_count} | Sum valgte kontoer: {fmt_amount(data.selected_sum)}"
    _write_df_table(ws_mot, df_mot, title="Motkonto (pivot)", summary=summary_line)

    # ------------------ Kombinasjoner sheet ------------------
    ws_combo = wb.create_sheet("Kombinasjoner")
    combos = build_motkonto_combinations(
        df_scope=data.df_scope,
        selected_accounts=set(data.selected_accounts),
    )
    df_combos = getattr(combos, "df", combos)
    bilag_count_combo = getattr(
        combos,
        "bilag_count",
        int(data.df_scope["Bilag_str"].nunique()) if "Bilag_str" in data.df_scope.columns else data.bilag_count,
    )
    summary_line = f"Motkonto-kombinasjoner | Antall kombinasjoner: {len(df_combos)} | Bilag i grunnlag: {bilag_count_combo}"
    _write_df_table(ws_combo, df_combos, title="Motkonto-kombinasjoner", summary=summary_line)

    # ------------------ Kombinasjoner pr konto sheet ------------------
    ws_combo_acc = wb.create_sheet("Kombinasjoner pr konto")
    combos_per = build_motkonto_combinations_per_selected_account(
        df_scope=data.df_scope,
        selected_accounts=set(data.selected_accounts),
    )
    df_combos_per = getattr(combos_per, "df", combos_per)
    bilag_count_combo_per = getattr(combos_per, "bilag_count", bilag_count_combo)
    summary_line = f"Motkonto-kombinasjoner pr konto | Antall rader: {len(df_combos_per)} | Bilag i grunnlag: {bilag_count_combo_per}"
    _write_df_table(ws_combo_acc, df_combos_per, title="Motkonto-kombinasjoner pr konto", summary=summary_line)

    # ------------------ Valgte kontoer sheet ------------------
    ws_sel = wb.create_sheet("Valgte kontoer")
    df_sel = pd.DataFrame(
        {
            "Konto": list(data.selected_accounts),
            "Sum": [data.selected_sum] * len(data.selected_accounts),
        }
    )
    _write_df_table(
        ws_sel,
        df_sel,
        title="Valgte kontoer",
        summary=f"Valgte kontoer: {', '.join(data.selected_accounts)} | Sum: {fmt_amount(data.selected_sum)}",
    )

    # ------------------ Bilag sheet ------------------
    ws_bilag = wb.create_sheet("Bilag")
    df_details = df_details_view if df_details_view is not None else data.df_details
    if selected_motkonto:
        df_details = df_details[df_details["Motkonto"] == _konto_str(selected_motkonto)].copy()
    bilag_col = (
        "Bilag_str"
        if "Bilag_str" in df_details.columns
        else "Bilag_key"
        if "Bilag_key" in df_details.columns
        else "Bilag"
        if "Bilag" in df_details.columns
        else None
    )
    bilag_count_details = int(df_details[bilag_col].nunique()) if (bilag_col and not df_details.empty) else 0

    _write_df_table(
        ws_bilag,
        df_details,
        title="Bilag",
        summary=f"Rader: {len(df_details)} | Bilag: {bilag_count_details}",
    )

    # ------------------ Outliers sheet ------------------
    ws_out = wb.create_sheet("Outliers")
    df_out = data.df_motkonto.copy() if data.df_motkonto is not None else pd.DataFrame()
    if not df_out.empty:
        df_out["Outlier"] = df_out["Motkonto"].map(lambda k: "Ja" if _konto_str(k) in out_set else "")
        df_out = df_out[df_out["Outlier"] == "Ja"].copy()
    _write_df_table(ws_out, df_out, title="Outliers", summary=f"Antall outliers: {len(out_set)}")

    # ------------------ OutlierBilag sheet ------------------
    ws_outbilag = wb.create_sheet("OutlierBilag")
    df_outbilag = data.df_scope.copy() if data.df_scope is not None else pd.DataFrame()
    if not df_outbilag.empty and out_set:
        # Alle transaksjoner på bilag som inneholder outlier-motkonto
        out_bilag = set(
            df_outbilag.loc[df_outbilag["Konto_str"].isin(out_set), "Bilag_str"].astype(str).unique()
        )
        df_outbilag = df_outbilag[df_outbilag["Bilag_str"].isin(out_bilag)].copy()
    _write_df_table(
        ws_outbilag,
        df_outbilag,
        title="OutlierBilag",
        summary=f"Bilag med outliers: {len(df_outbilag['Bilag_str'].unique()) if not df_outbilag.empty else 0}",
    )

    return wb
