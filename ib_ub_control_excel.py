"""Excel-arbeidspapir for SB/HB-avstemming og IB/UB-kontinuitetskontroll.

SB/HB-arbeidsboken har fire ark:
  - Oppsummering: totaler og nøkkeltall
  - Avstemming pr konto: full kontovisning
  - Avstemming pr RL: aggregert til regnskapslinje (hvis mapping finnes)
  - Avvik: kun kontoer med differanse

IB/UB-arbeidsboken har tre ark:
  - Oppsummering: totaler og nøkkeltall
  - IB/UB pr konto: IB(i år) vs UB(fjor) per konto
  - Avvik: kontoer der IB ≠ UB fjor
"""

from __future__ import annotations

from datetime import datetime
from typing import Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import formatting
from src.shared.workpapers.forside import build_forside_sheet

_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
_HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
_SUM_FILL = PatternFill("solid", fgColor="F3F6F9")
_AVVIK_FILL = PatternFill("solid", fgColor="FFF2CC")
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_AMOUNT_FMT = '#,##0.00;[Red]-#,##0.00'


def build_ib_ub_workpaper(
    account_recon: pd.DataFrame,
    *,
    rl_recon: Optional[pd.DataFrame] = None,
    summary: Optional[Dict[str, object]] = None,
    client: str | None = None,
    year: str | int | None = None,
) -> Workbook:
    """Bygg komplett IB/UB-kontroll arbeidspapir."""
    wb = Workbook()

    _build_summary_sheet(wb, summary or {}, client=client, year=year)
    _build_account_sheet(wb, account_recon, client=client, year=year)
    if rl_recon is not None and not rl_recon.empty:
        _build_rl_sheet(wb, rl_recon, client=client, year=year)
    _build_discrepancy_sheet(wb, account_recon, client=client, year=year)

    build_forside_sheet(wb, workpaper_navn="SB/HB Avstemming")

    # Slett standard-arket hvis det er tomt
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    return wb


# ---------------------------------------------------------------------------
# Oppsummering
# ---------------------------------------------------------------------------

def _build_summary_sheet(
    wb: Workbook,
    summary: Dict[str, object],
    *,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.create_sheet("Oppsummering")

    title_parts = ["SB/HB Avstemming"]
    if client:
        title_parts.append(str(client))
    if year not in {None, ""}:
        title_parts.append(str(year))
    title = " — ".join(title_parts)

    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = _TITLE_FILL

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    row = 4
    items = [
        ("Sum SB IB", summary.get("total_sb_ib", 0)),
        ("Sum SB UB", summary.get("total_sb_ub", 0)),
        ("Sum SB netto (UB − IB)", summary.get("total_sb_netto", 0)),
        ("Sum HB posteringer", summary.get("total_hb_sum", 0)),
        ("Total differanse", summary.get("total_differanse", 0)),
        ("", ""),
        ("Antall kontoer", summary.get("antall_kontoer", 0)),
        ("Antall avvik", summary.get("antall_avvik", 0)),
    ]
    for label, value in items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value if value != "" else None)
        if isinstance(value, float):
            cell.number_format = _AMOUNT_FMT
        if label == "Total differanse" and isinstance(value, (int, float)) and abs(float(value)) > 0.01:
            cell.fill = _AVVIK_FILL
            cell.font = Font(bold=True, color="CC0000")
        if label == "Antall avvik" and isinstance(value, int) and value > 0:
            cell.fill = _AVVIK_FILL
            cell.font = Font(bold=True, color="CC0000")
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    ws.sheet_properties.tabColor = "4472C4"


# ---------------------------------------------------------------------------
# Avstemming pr konto
# ---------------------------------------------------------------------------

def _build_account_sheet(
    wb: Workbook,
    account_recon: pd.DataFrame,
    *,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.create_sheet("Avstemming pr konto")

    title = "Avstemming SB/HB pr konto"
    if client:
        title += f" — {client}"
    if year not in {None, ""}:
        title += f" {year}"

    headers = ["Konto", "Kontonavn", "SB IB", "SB UB", "SB Netto", "HB Sum", "Differanse"]
    cols = ["konto", "kontonavn", "sb_ib", "sb_ub", "sb_netto", "hb_sum", "differanse"]
    amount_cols = {2, 3, 4, 5, 6}  # 0-indexed

    _write_title_and_header(ws, title, headers, span=len(headers))
    _write_data_rows(ws, account_recon, cols, amount_cols, highlight_col=6)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    for c in "CDEFG":
        ws.column_dimensions[c].width = 16
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{max(5, 4 + len(account_recon))}"

    ws.sheet_properties.tabColor = "4472C4"


# ---------------------------------------------------------------------------
# Avstemming pr regnskapslinje
# ---------------------------------------------------------------------------

def _build_rl_sheet(
    wb: Workbook,
    rl_recon: pd.DataFrame,
    *,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.create_sheet("Avstemming pr RL")

    title = "Avstemming SB/HB pr regnskapslinje"
    if client:
        title += f" — {client}"
    if year not in {None, ""}:
        title += f" {year}"

    headers = ["Nr", "Regnskapslinje", "SB IB", "SB UB", "SB Netto", "HB Sum", "Differanse"]
    cols = ["regnr", "regnskapslinje", "sb_ib", "sb_ub", "sb_netto", "hb_sum", "differanse"]
    amount_cols = {2, 3, 4, 5, 6}

    _write_title_and_header(ws, title, headers, span=len(headers))
    _write_data_rows(ws, rl_recon, cols, amount_cols, highlight_col=6)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    for c in "CDEFG":
        ws.column_dimensions[c].width = 16
    ws.freeze_panes = "A5"

    ws.sheet_properties.tabColor = "4472C4"


# ---------------------------------------------------------------------------
# Avvik
# ---------------------------------------------------------------------------

def _build_discrepancy_sheet(
    wb: Workbook,
    account_recon: pd.DataFrame,
    *,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.create_sheet("Avvik")
    discrepancies = account_recon.loc[account_recon["har_avvik"]].copy()

    title = "Kontoer med avvik"
    if client:
        title += f" — {client}"
    if year not in {None, ""}:
        title += f" {year}"

    headers = ["Konto", "Kontonavn", "SB Netto", "HB Sum", "Differanse"]
    cols = ["konto", "kontonavn", "sb_netto", "hb_sum", "differanse"]
    amount_cols = {2, 3, 4}

    _write_title_and_header(ws, title, headers, span=len(headers))

    if discrepancies.empty:
        ws.cell(row=5, column=1, value="Ingen avvik funnet ✓")
        ws["A5"].font = Font(color="006100")
    else:
        _write_data_rows(ws, discrepancies, cols, amount_cols, highlight_col=4)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    for c in "CDE":
        ws.column_dimensions[c].width = 16
    ws.freeze_panes = "A5"

    ws.sheet_properties.tabColor = "FF0000" if not discrepancies.empty else "70AD47"


# ---------------------------------------------------------------------------
# Hjelpere
# ---------------------------------------------------------------------------

def _write_title_and_header(ws, title: str, headers: list[str], *, span: int) -> None:
    last_col_letter = chr(ord("A") + span - 1)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = _TITLE_FILL

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")


def _write_data_rows(
    ws,
    df: pd.DataFrame,
    cols: list[str],
    amount_cols: set[int],
    *,
    highlight_col: int | None = None,
) -> None:
    start_row = 5
    for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row):
        for col_idx, col_name in enumerate(cols):
            raw = row.get(col_name, "")
            if col_idx in amount_cols:
                value = float(raw) if raw is not None and raw == raw else 0.0
            else:
                value = raw if raw is not None and raw == raw else ""
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
            cell.border = _BORDER
            if col_idx in amount_cols:
                cell.number_format = _AMOUNT_FMT
                cell.alignment = Alignment(horizontal="right")

            # Marker avvik-rader
            if highlight_col is not None and col_idx == highlight_col:
                diff_val = float(row.get(cols[highlight_col], 0) or 0)
                if abs(diff_val) > 0.01:
                    for ci in range(len(cols)):
                        c = ws.cell(row=row_idx, column=ci + 1)
                        c.fill = _AVVIK_FILL


# ===========================================================================
# IB/UB-kontinuitetskontroll (IB i år == UB fjor)
# ===========================================================================

def build_continuity_workpaper(
    account_df: pd.DataFrame,
    *,
    summary: Optional[Dict[str, object]] = None,
    client: str | None = None,
    year: str | int | None = None,
) -> Workbook:
    """Bygg IB/UB-kontinuitetskontroll arbeidspapir."""
    wb = Workbook()

    _build_continuity_summary_sheet(wb, summary or {}, client=client, year=year)
    _build_continuity_account_sheet(wb, account_df, client=client, year=year)
    _build_continuity_discrepancy_sheet(wb, account_df, client=client, year=year)

    build_forside_sheet(wb, workpaper_navn="IB/UB Kontinuitetskontroll")

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    return wb


def _build_continuity_summary_sheet(
    wb: Workbook,
    summary: Dict[str, object],
    *,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.create_sheet("Oppsummering")

    title_parts = ["IB/UB Kontinuitetskontroll"]
    if client:
        title_parts.append(str(client))
    if year not in {None, ""}:
        title_parts.append(str(year))
    title = " — ".join(title_parts)

    prev_year = ""
    try:
        prev_year = str(int(year) - 1) if year else ""
    except (ValueError, TypeError):
        pass

    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = _TITLE_FILL

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generert {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    ws.merge_cells("A3:D3")
    ws["A3"] = f"Kontroll: IB {year or 'i år'} skal stemme med UB {prev_year or 'fjor'}"
    ws["A3"].font = Font(italic=True, color="333333")

    row = 5
    items = [
        (f"Sum IB {year or 'i år'}", summary.get("total_ib_current", 0)),
        (f"Sum UB {prev_year or 'fjor'}", summary.get("total_ub_previous", 0)),
        ("Total differanse", summary.get("total_differanse", 0)),
        ("", ""),
        ("Antall kontoer med saldo", summary.get("antall_kontoer", 0)),
        ("Antall avvik", summary.get("antall_avvik", 0)),
        (f"Kun i {year or 'i år'} (ny konto)", summary.get("kun_i_current", 0)),
        (f"Kun i {prev_year or 'fjor'} (fjernet konto)", summary.get("kun_i_previous", 0)),
    ]
    for label, value in items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value if value != "" else None)
        if isinstance(value, float):
            cell.number_format = _AMOUNT_FMT
        if label == "Total differanse" and isinstance(value, (int, float)) and abs(float(value)) > 0.01:
            cell.fill = _AVVIK_FILL
            cell.font = Font(bold=True, color="CC0000")
        if label == "Antall avvik" and isinstance(value, int) and value > 0:
            cell.fill = _AVVIK_FILL
            cell.font = Font(bold=True, color="CC0000")
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.sheet_properties.tabColor = "4472C4"


def _build_continuity_account_sheet(
    wb: Workbook,
    account_df: pd.DataFrame,
    *,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.create_sheet("IB UB pr konto")

    prev_year = ""
    try:
        prev_year = str(int(year) - 1) if year else ""
    except (ValueError, TypeError):
        pass

    title = "IB/UB-kontinuitet pr konto"
    if client:
        title += f" — {client}"
    if year not in {None, ""}:
        title += f" {year}"

    has_kontotype = "kontotype" in account_df.columns
    if has_kontotype:
        headers = ["Konto", "Kontonavn", "Type", f"IB {year or 'i år'}", f"UB {prev_year or 'fjor'}", "Differanse"]
        cols = ["konto", "kontonavn", "kontotype", "ib_current", "ub_previous", "differanse"]
        amount_cols = {3, 4, 5}
        highlight = 5
        last_col = "F"
    else:
        headers = ["Konto", "Kontonavn", f"IB {year or 'i år'}", f"UB {prev_year or 'fjor'}", "Differanse"]
        cols = ["konto", "kontonavn", "ib_current", "ub_previous", "differanse"]
        amount_cols = {2, 3, 4}
        highlight = 4
        last_col = "E"

    _write_title_and_header(ws, title, headers, span=len(headers))
    _write_data_rows(ws, account_df, cols, amount_cols, highlight_col=highlight)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    if has_kontotype:
        ws.column_dimensions["C"].width = 12
        for c in "DEF":
            ws.column_dimensions[c].width = 18
    else:
        for c in "CDE":
            ws.column_dimensions[c].width = 18
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_col}{max(5, 4 + len(account_df))}"
    ws.sheet_properties.tabColor = "4472C4"


def _build_continuity_discrepancy_sheet(
    wb: Workbook,
    account_df: pd.DataFrame,
    *,
    client: str | None = None,
    year: str | int | None = None,
) -> None:
    ws = wb.create_sheet("Avvik")
    discrepancies = account_df.loc[account_df["har_avvik"]].copy()

    prev_year = ""
    try:
        prev_year = str(int(year) - 1) if year else ""
    except (ValueError, TypeError):
        pass

    title = "IB/UB-avvik"
    if client:
        title += f" — {client}"
    if year not in {None, ""}:
        title += f" {year}"

    has_kontotype = "kontotype" in account_df.columns
    if has_kontotype:
        headers = ["Konto", "Kontonavn", "Type", f"IB {year or 'i år'}", f"UB {prev_year or 'fjor'}", "Differanse"]
        cols = ["konto", "kontonavn", "kontotype", "ib_current", "ub_previous", "differanse"]
        amount_cols = {3, 4, 5}
        highlight = 5
    else:
        headers = ["Konto", "Kontonavn", f"IB {year or 'i år'}", f"UB {prev_year or 'fjor'}", "Differanse"]
        cols = ["konto", "kontonavn", "ib_current", "ub_previous", "differanse"]
        amount_cols = {2, 3, 4}
        highlight = 4

    _write_title_and_header(ws, title, headers, span=len(headers))

    if discrepancies.empty:
        ws.cell(row=5, column=1, value="Ingen avvik — IB stemmer med UB fjor ✓")
        ws["A5"].font = Font(color="006100")
    else:
        _write_data_rows(ws, discrepancies, cols, amount_cols, highlight_col=highlight)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    if has_kontotype:
        ws.column_dimensions["C"].width = 12
        for c in "DEF":
            ws.column_dimensions[c].width = 18
    else:
        for c in "CDE":
            ws.column_dimensions[c].width = 18
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "FF0000" if not discrepancies.empty else "70AD47"
